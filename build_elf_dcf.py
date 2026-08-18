#!/usr/bin/env python3
"""
build_elf_dcf.py
Builds institutional-grade DCF valuation model for e.l.f. Beauty, Inc. (NYSE: ELF)
Follows Wall Street Investment Banking standards:
- Dynamic scenario switching (Bear / Base / Bull) via CHOOSE / INDEX logic
- 100% dynamic formulas over hardcodes
- Embedded cell comments with source document references
- 3 symmetric 5x5 sensitivity analysis grids with base cases anchored at center
- Comprehensive WACC schedule and mid-year discounting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

def build_model(filename="ELF_DCF_Model_Gemini-3.7-Flash_20260818.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Valuation"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Color Palette (Classic Institutional Navy)
    NAVY = "1F4E79"       # Section Headers
    MED_BLUE = "2F5597"   # Secondary Headers
    SOFT_BLUE = "D9E1F2"  # Subheaders / Table Headers
    ACCENT_BLUE = "BDD7EE"# Active Case / Base Case highlight
    ZEBRA_FILL = "F9FBFD" # Alternating row fill
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"

    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color="000000")
    font_subhdr = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=10, bold=False, color="000000")
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)

    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")

    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_bottom = Side(border_style="medium", color=NAVY)
    double_bottom = Side(border_style="double", color=NAVY)
    top_thin = Side(border_style="thin", color=NAVY)

    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)
    border_total = Border(top=top_thin, bottom=double_bottom, left=thin_line, right=thin_line)

    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # --- Title Block ---
    ws["A1"] = "e.l.f. Beauty, Inc. (NYSE: ELF)"
    ws["A1"].font = font_title
    ws["A2"] = "Institutional DCF Valuation Model | Report Date: August 18, 2026 | Fiscal Year Ends March 31 | Source: SEC Form 10-K & Company Disclosures"
    ws["A2"].font = font_subtitle

    # --- Scenario Selector ---
    ws["A4"] = "Active Scenario Selector:"
    ws["A4"].font = font_bold
    ws["B4"] = 2
    ws["B4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["B4"].fill = fill_navy
    ws["B4"].alignment = align_center
    ws["B4"].comment = Comment("Case Selector: 1 = Bear Case, 2 = Base Case, 3 = Bull Case", "Model Builder")

    ws["C4"] = '=IF(B4=1,"[1] BEAR CASE",IF(B4=2,"[2] BASE CASE (Active)","[3] BULL CASE"))'
    ws["C4"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws["C4"].alignment = align_left

    # Helper for section headers
    def add_section_header(row, title, max_col="H"):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_sec_hdr
        ws[f"A{row}"].fill = fill_navy
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_navy

    # --- Section I: Market Data & Capital Structure Inputs ---
    add_section_header(6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS")

    market_inputs = [
        ("Current Stock Price ($)", 91.54, "$#,##0.00", "Market close price as of August 2026", "B7"),
        ("Diluted Shares Outstanding (M)", 59.06, "#,##0.00", "SEC Form 10-K FY2026 Diluted share count (59.06M shares)", "B8"),
        ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
        ("Cash and Cash Equivalents ($M)", 289.70, "$#,##0.00", "SEC Form 10-K Balance sheet as of March 31, 2026 ($289.70M)", "B10"),
        ("Total Debt ($M)", 841.70, "$#,##0.00", "SEC Form 10-K Balance sheet as of March 31, 2026 ($841.70M)", "B11"),
        ("Net Debt ($M)", "=B11-B10", "$#,##0.00", None, "B12"),
        ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
        ("Effective Corporate Tax Rate (%)", 0.250, "0.0%", "Normalized corporate effective tax rate", "B14"),
        ("Risk-Free Rate (10-Yr US Treasury) (%)", 0.0473, "0.00%", "10-Year US Treasury Benchmark Yield as of August 2026", "B15"),
        ("Equity Beta (5-Year Monthly)", 1.56, "0.00", "5-Year Monthly Beta vs S&P 500", "B16"),
        ("Market Equity Risk Premium (ERP) (%)", 0.0550, "0.00%", "Standard institutional market risk premium", "B17"),
        ("Pre-Tax Cost of Debt (%)", 0.0650, "0.00%", "Estimated corporate borrowing rate / credit spread", "B18"),
    ]

    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs, start=7):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws[f"A{idx}"].border = border_cell
        
        ws[f"B{idx}"] = val
        ws[f"B{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws[f"B{idx}"].number_format = num_fmt
        ws[f"B{idx}"].alignment = align_right
        ws[f"B{idx}"].border = border_cell
        if comment_text:
            ws[f"B{idx}"].comment = Comment(comment_text, "ELF 10-K / Market Data")

    # --- Section II: Scenario Assumptions & Driver Blocks ---
    add_section_header(20, "II. SCENARIO ASSUMPTIONS & DRIVER BLOCKS")

    headers_s2 = ["Driver / Assumption", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E", "Terminal / Target"]
    for col_idx, text in enumerate(headers_s2, start=1):
        c = ws.cell(row=21, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header

    scenario_blocks = [
        ("Bear Case Assumptions (Selector = 1)", 22, [
            ("  Revenue Growth (%)", [0.100, 0.080, 0.060, 0.050, 0.040], 0.025, "0.0%"),
            ("  EBIT Margin (%)", [0.070, 0.085, 0.095, 0.105, 0.110], 0.110, "0.0%"),
            ("  D&A (% of Revenue)", [0.040, 0.036, 0.032, 0.030, 0.028], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.015, 0.015, 0.015, 0.015, 0.015], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.020, 0.020, 0.020, 0.020, 0.020], None, "0.0%"),
        ]),
        ("Base Case Assumptions (Selector = 2)", 28, [
            ("  Revenue Growth (%)", [0.160, 0.135, 0.110, 0.090, 0.070], 0.030, "0.0%"),
            ("  EBIT Margin (%)", [0.100, 0.120, 0.135, 0.145, 0.150], 0.150, "0.0%"),
            ("  D&A (% of Revenue)", [0.040, 0.036, 0.032, 0.030, 0.028], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.015, 0.015, 0.015, 0.015, 0.015], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.020, 0.020, 0.020, 0.020, 0.020], None, "0.0%"),
        ]),
        ("Bull Case Assumptions (Selector = 3)", 34, [
            ("  Revenue Growth (%)", [0.220, 0.180, 0.150, 0.120, 0.090], 0.035, "0.0%"),
            ("  EBIT Margin (%)", [0.120, 0.140, 0.155, 0.165, 0.170], 0.170, "0.0%"),
            ("  D&A (% of Revenue)", [0.040, 0.036, 0.032, 0.030, 0.028], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.015, 0.015, 0.015, 0.015, 0.015], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.020, 0.020, 0.020, 0.020, 0.020], None, "0.0%"),
        ]),
    ]

    for title, sub_hdr_row, rows in scenario_blocks:
        ws[f"A{sub_hdr_row}"] = title
        ws[f"A{sub_hdr_row}"].font = font_subhdr
        ws[f"A{sub_hdr_row}"].fill = fill_soft_blue
        ws.merge_cells(f"A{sub_hdr_row}:H{sub_hdr_row}")
        for c_idx in range(1, 9):
            ws.cell(row=sub_hdr_row, column=c_idx).fill = fill_soft_blue
            ws.cell(row=sub_hdr_row, column=c_idx).border = border_cell
            
        for r_offset, (label, vals, term_val, num_fmt) in enumerate(rows, start=1):
            curr_r = sub_hdr_row + r_offset
            ws[f"A{curr_r}"] = label
            ws[f"A{curr_r}"].font = font_regular
            ws[f"A{curr_r}"].border = border_cell
            # Col B (Historical FY2026A reference)
            ws.cell(row=curr_r, column=2, value="—").alignment = align_center
            ws.cell(row=curr_r, column=2).border = border_cell
            
            for c_idx, val in enumerate(vals, start=3):
                cell = ws.cell(row=curr_r, column=c_idx, value=val)
                cell.font = font_regular
                cell.number_format = num_fmt
                cell.alignment = align_right
                cell.border = border_cell
            if term_val is not None:
                cell_term = ws.cell(row=curr_r, column=8, value=term_val)
                cell_term.font = font_regular
                cell_term.number_format = num_fmt
                cell_term.alignment = align_right
                cell_term.border = border_cell
            else:
                cell_term = ws.cell(row=curr_r, column=8, value="")
                cell_term.border = border_cell

    # Consolidated Active Drivers (Dynamic via CHOOSE formulas)
    ws["A40"] = "Active Scenario Consolidated Drivers (Dynamic)"
    ws["A40"].font = font_bold
    ws["A40"].fill = fill_accent_blue
    ws.merge_cells("A40:H40")
    for c_idx in range(1, 9):
        ws.cell(row=40, column=c_idx).fill = fill_accent_blue
        ws.cell(row=40, column=c_idx).border = border_cell

    active_driver_rows = [
        ("  Active Revenue Growth (%)", 23, 29, 35, "0.0%"),
        ("  Active EBIT Margin (%)", 24, 30, 36, "0.0%"),
        ("  Active D&A (% of Revenue)", 25, 31, 37, "0.0%"),
        ("  Active CapEx (% of Revenue)", 26, 32, 38, "0.0%"),
        ("  Active Δ NWC (% of Δ Rev)", 27, 33, 39, "0.0%"),
    ]

    for idx, (label, r_bear, r_base, r_bull, num_fmt) in enumerate(active_driver_rows, start=41):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = font_bold
        ws[f"A{idx}"].border = border_cell
        ws.cell(row=idx, column=2, value="—").alignment = align_center
        ws.cell(row=idx, column=2).border = border_cell
        
        for c_idx in range(3, 8):
            col_let = get_column_letter(c_idx)
            cell = ws.cell(row=idx, column=c_idx)
            cell.value = f"=CHOOSE($B$4, {col_let}{r_bear}, {col_let}{r_base}, {col_let}{r_bull})"
            cell.font = font_bold
            cell.number_format = num_fmt
            cell.alignment = align_right
            cell.border = border_cell
        
        # Terminal growth in column H (Col 8)
        if idx == 41: # Terminal g
            cell_g = ws.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H23, H29, H35)")
            cell_g.font = font_bold
            cell_g.number_format = "0.0%"
            cell_g.alignment = align_right
            cell_g.border = border_cell
        elif idx == 42: # Target Margin
            cell_g = ws.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H24, H30, H36)")
            cell_g.font = font_bold
            cell_g.number_format = "0.0%"
            cell_g.alignment = align_right
            cell_g.border = border_cell
        else:
            cell_g = ws.cell(row=idx, column=8, value="")
            cell_g.border = border_cell

    # --- Section III: Cost of Capital (WACC) Schedule ---
    add_section_header(48, "III. COST OF CAPITAL (WACC) PARAMETERS & SCHEDULE")

    wacc_items = [
        ("Risk-Free Rate (Rf)", "=$B$15", "0.00%", "Row 15: 10-Yr US Treasury Yield", False),
        ("Equity Beta (β)", "=$B$16", "0.00", "Row 16: 5-Year Monthly Beta", False),
        ("Equity Risk Premium (ERP)", "=$B$17", "0.00%", "Row 17: Market Risk Premium", False),
        ("Cost of Equity (Ke) [Rf + β * ERP]", "=B49+B50*B51", "0.00%", "CAPM Formula: Cost of Equity", True),
        ("Pre-Tax Cost of Debt (Kd)", "=$B$18", "0.00%", "Row 18: Weighted average borrowing cost", False),
        ("Effective Corporate Tax Rate (t)", "=$B$14", "0.0%", "Row 14: Corporate tax rate", False),
        ("After-Tax Cost of Debt [Kd * (1 - t)]", "=B53*(1-B54)", "0.00%", "Effective after-tax cost of debt", False),
        ("Market Capitalization (Equity Value) ($M)", "=$B$9", "$#,##0.00", "Row 9: Diluted Market Cap", False),
        ("Total Debt ($M)", "=$B$11", "$#,##0.00", "Row 11: Total Debt", False),
        ("Total Capital ($M)", "=B56+B57", "$#,##0.00", "Market Cap + Total Debt", False),
        ("Weight of Equity (We)", "=B56/B58", "0.0%", "Equity Value / Total Capital", False),
        ("Weight of Debt (Wd)", "=B57/B58", "0.0%", "Total Debt / Total Capital", False),
        ("Calculated WACC [(Ke * We) + (Kd_after * Wd)]", "=(B52*B59)+(B55*B60)", "0.00%", "Weighted Average Cost of Capital", True),
    ]

    for idx, (label, val, num_fmt, comment_t, is_key) in enumerate(wacc_items, start=49):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = font_bold if is_key else font_regular
        ws[f"A{idx}"].border = border_cell
        if is_key:
            ws[f"A{idx}"].fill = fill_accent_blue
            
        ws[f"B{idx}"] = val
        ws[f"B{idx}"].font = font_bold if is_key else font_regular
        ws[f"B{idx}"].number_format = num_fmt
        ws[f"B{idx}"].alignment = align_right
        ws[f"B{idx}"].border = border_cell
        if is_key:
            ws[f"B{idx}"].fill = fill_accent_blue
        if comment_t:
            ws[f"B{idx}"].comment = Comment(comment_t, "WACC Build")

    # --- Section IV: Unlevered Free Cash Flow (UFCF) Projections ---
    add_section_header(63, "IV. UNLEVERED FREE CASH FLOW (UFCF) PROJECTIONS")

    headers_s4 = ["Line Item ($M)", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
    for col_idx, text in enumerate(headers_s4, start=1):
        c = ws.cell(row=64, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header

    # Row 65: Revenue Growth Rate
    ws["A65"] = "Revenue Growth Rate (%)"
    ws["A65"].font = font_regular
    ws["A65"].border = border_cell
    ws["B65"] = 0.246
    ws["B65"].font = font_regular
    ws["B65"].number_format = "0.0%"
    ws["B65"].alignment = align_right
    ws["B65"].border = border_cell
    ws["B65"].comment = Comment("FY2026 Actual Net Sales YoY Growth (+24.6%)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=65, column=c_idx, value=f"={col_let}41")
        cell.font = font_regular
        cell.number_format = "0.0%"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 66: Net Sales (Revenue)
    ws["A66"] = "Net Sales (Revenue) ($M)"
    ws["A66"].font = font_bold
    ws["A66"].border = border_cell
    ws["B66"] = 1636.50
    ws["B66"].font = font_bold
    ws["B66"].number_format = "$#,##0.00"
    ws["B66"].alignment = align_right
    ws["B66"].border = border_cell
    ws["B66"].comment = Comment("FY2026 Actual Net Sales ($1,636.50M)", "SEC Form 10-K")

    ws["C66"] = "=B66*(1+C65)"
    ws["D66"] = "=C66*(1+D65)"
    ws["E66"] = "=D66*(1+E65)"
    ws["F66"] = "=E66*(1+F65)"
    ws["G66"] = "=F66*(1+G65)"
    for c_idx in range(3, 8):
        c = ws.cell(row=66, column=c_idx)
        c.font = font_bold
        c.number_format = "$#,##0.00"
        c.alignment = align_right
        c.border = border_cell

    # Row 67: EBIT Margin (%)
    ws["A67"] = "EBIT Margin (%)"
    ws["A67"].font = font_regular
    ws["A67"].border = border_cell
    ws["B67"] = "=B68/B66"
    ws["B67"].font = font_regular
    ws["B67"].number_format = "0.0%"
    ws["B67"].alignment = align_right
    ws["B67"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=67, column=c_idx, value=f"={col_let}42")
        cell.font = font_regular
        cell.number_format = "0.0%"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 68: Operating Income (EBIT)
    ws["A68"] = "Operating Income (EBIT) ($M)"
    ws["A68"].font = font_bold
    ws["A68"].border = border_cell
    ws["B68"] = 74.00
    ws["B68"].font = font_bold
    ws["B68"].number_format = "$#,##0.00"
    ws["B68"].alignment = align_right
    ws["B68"].border = border_cell
    ws["B68"].comment = Comment("FY2026 Actual Operating Income ($74.00M)", "SEC Form 10-K")

    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=68, column=c_idx, value=f"={col_let}66*{col_let}67")
        cell.font = font_bold
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 69: (-) Taxes
    ws["A69"] = "(-) Provision for Taxes ($M)"
    ws["A69"].font = font_regular
    ws["A69"].border = border_cell
    ws["B69"] = "—"
    ws["B69"].alignment = align_center
    ws["B69"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=69, column=c_idx, value=f"={col_let}68*$B$14")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 70: NOPAT
    ws["A70"] = "Net Operating Profit After Tax (NOPAT) ($M)"
    ws["A70"].font = font_bold
    ws["A70"].border = border_cell
    ws["B70"] = "—"
    ws["B70"].alignment = align_center
    ws["B70"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=70, column=c_idx, value=f"={col_let}68-{col_let}69")
        cell.font = font_bold
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 71: (+) D&A
    ws["A71"] = "(+) Depreciation & Amortization (D&A) ($M)"
    ws["A71"].font = font_regular
    ws["A71"].border = border_cell
    ws["B71"] = 79.36
    ws["B71"].font = font_regular
    ws["B71"].number_format = "$#,##0.00"
    ws["B71"].alignment = align_right
    ws["B71"].border = border_cell
    ws["B71"].comment = Comment("FY2026 Actual D&A ($79.36M)", "SEC Form 10-K")

    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=71, column=c_idx, value=f"={col_let}66*{col_let}43")
        cell.font = font_regular
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 72: (-) CapEx
    ws["A72"] = "(-) Capital Expenditures (CapEx) ($M)"
    ws["A72"].font = font_regular
    ws["A72"].border = border_cell
    ws["B72"] = 22.45
    ws["B72"].font = font_regular
    ws["B72"].number_format = "($#,##0.00);($#,##0.00);$0.00"
    ws["B72"].alignment = align_right
    ws["B72"].border = border_cell
    ws["B72"].comment = Comment("FY2026 Actual CapEx ($22.45M)", "SEC Form 10-K")

    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=72, column=c_idx, value=f"={col_let}66*{col_let}44")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 73: (-) Change in NWC
    ws["A73"] = "(-) Change in Net Working Capital (Δ NWC) ($M)"
    ws["A73"].font = font_regular
    ws["A73"].border = border_cell
    ws["B73"] = 10.50
    ws["B73"].font = font_regular
    ws["B73"].number_format = "($#,##0.00);($#,##0.00);$0.00"
    ws["B73"].alignment = align_right
    ws["B73"].border = border_cell
    ws["B73"].comment = Comment("FY2026 Actual Working Capital change ($10.50M)", "SEC Form 10-K")

    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        prev_col = get_column_letter(c_idx - 1)
        cell = ws.cell(row=73, column=c_idx, value=f"=({col_let}66-{prev_col}66)*{col_let}45")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 74: Unlevered Free Cash Flow (UFCF)
    ws["A74"] = "Unlevered Free Cash Flow (UFCF) ($M)"
    ws["A74"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws["A74"].fill = fill_accent_blue
    ws["A74"].border = border_total
    ws["B74"] = "—"
    ws["B74"].alignment = align_center
    ws["B74"].fill = fill_accent_blue
    ws["B74"].border = border_total
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=74, column=c_idx, value=f"={col_let}70+{col_let}71-{col_let}72-{col_let}73")
        cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        cell.fill = fill_accent_blue
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_total

    # --- Section V: Discounting & Present Value ---
    add_section_header(76, "V. DISCOUNTING & PRESENT VALUE (MID-YEAR CONVENTION)")

    headers_s5 = ["Discount Parameters", "", "Year 1 (FY27E)", "Year 2 (FY28E)", "Year 3 (FY29E)", "Year 4 (FY30E)", "Year 5 (FY31E)"]
    for col_idx, text in enumerate(headers_s5, start=1):
        c = ws.cell(row=77, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 2 else align_left
        c.border = border_header

    # Row 78: Discount Period
    ws["A78"] = "Discount Period (Years - Mid-Year Convention)"
    ws["A78"].font = font_regular
    ws["A78"].border = border_cell
    ws["B78"] = ""
    ws["B78"].border = border_cell
    periods = [0.5, 1.5, 2.5, 3.5, 4.5]
    for c_idx, p in enumerate(periods, start=3):
        cell = ws.cell(row=78, column=c_idx, value=p)
        cell.font = font_regular
        cell.number_format = "0.0"
        cell.alignment = align_center
        cell.border = border_cell

    # Row 79: Discount Factor
    ws["A79"] = "Discount Factor [ 1 / (1 + WACC)^t ]"
    ws["A79"].font = font_regular
    ws["A79"].border = border_cell
    ws["B79"] = ""
    ws["B79"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=79, column=c_idx, value=f"=1/(1+$B$61)^{col_let}78")
        cell.font = font_regular
        cell.number_format = "0.0000"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 80: Present Value of UFCF
    ws["A80"] = "Present Value of UFCF ($M)"
    ws["A80"].font = font_bold
    ws["A80"].border = border_cell
    ws["B80"] = ""
    ws["B80"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=80, column=c_idx, value=f"={col_let}74*{col_let}79")
        cell.font = font_bold
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 81: Sum of PV of Explicit Cash Flows
    ws["A81"] = "Cumulative PV of 5-Year Explicit Cash Flows ($M)"
    ws["A81"].font = font_bold
    ws["A81"].fill = fill_soft_blue
    ws["A81"].border = border_total
    ws["B81"] = "=SUM(C80:G80)"
    ws["B81"].font = font_bold
    ws["B81"].fill = fill_soft_blue
    ws["B81"].number_format = "$#,##0.00"
    ws["B81"].alignment = align_right
    ws["B81"].border = border_total
    ws.merge_cells("C81:G81")
    for c_idx in range(3, 8):
        ws.cell(row=81, column=c_idx).fill = fill_soft_blue
        ws.cell(row=81, column=c_idx).border = border_total

    # --- Section VI: Terminal Value & Valuation Summary ---
    add_section_header(83, "VI. TERMINAL VALUE & VALUATION SUMMARY")

    val_summary_items = [
        ("Terminal Value Assumptions & Calculation:", None, None, None, True),
        ("  Final Year Projected UFCF (FY2031E) ($M)", "=G74", "$#,##0.00", None, False),
        ("  Perpetual Terminal Growth Rate (g)", "=$H$41", "0.0%", None, False),
        ("  WACC (Discount Rate)", "=$B$61", "0.00%", None, False),
        ("  Terminal Year Cash Flow ($M)", "=B85*(1+B86)", "$#,##0.00", None, False),
        ("  Terminal Value at FY2031E ($M)", "=B88/(B87-B86)", "$#,##0.00", None, False),
        ("  PV of Terminal Value ($M)", "=B89/(1+B87)^5", "$#,##0.00", None, False),
        ("  Terminal Value as % of Enterprise Value", "=B90/B95", "0.0%", None, False),
        ("Enterprise Value to Equity Value Bridge ($M):", None, None, None, True),
        ("  PV of Explicit 5-Year Cash Flows ($M)", "=B81", "$#,##0.00", None, False),
        ("  (+) PV of Terminal Value ($M)", "=B90", "$#,##0.00", None, False),
        ("Enterprise Value (EV) ($M)", "=B93+B94", "$#,##0.00", None, False),
        ("  (-) Total Debt ($M)", "=$B$11", "($#,##0.00);($#,##0.00);$0.00", None, False),
        ("  (+) Total Cash & Equivalents ($M)", "=$B$10", "$#,##0.00", None, False),
        ("  (-) Net Debt ($M)", "=$B$12", "($#,##0.00);($#,##0.00);$0.00", None, False),
        ("Implied Equity Value ($M)", "=B95-B98", "$#,##0.00", None, False),
        ("  Diluted Shares Outstanding (M)", "=$B$8", "#,##0.00", None, False),
        ("Implied Price per Share ($)", "=B99/B100", "$#,##0.00", None, False),
        ("Current Market Share Price ($)", "=$B$7", "$#,##0.00", None, False),
        ("Implied Upside / (Downside) (%)", "=(B101/B102)-1", "+0.0%;-0.0%;0.0%", None, False),
    ]

    current_r = 84
    for item in val_summary_items:
        label, formula_val, num_fmt, comment_t, is_header = item
        if is_header:
            ws[f"A{current_r}"] = label
            ws[f"A{current_r}"].font = font_subhdr
            ws[f"A{current_r}"].fill = fill_soft_blue
            ws.merge_cells(f"A{current_r}:H{current_r}")
            for c_idx in range(1, 9):
                ws.cell(row=current_r, column=c_idx).fill = fill_soft_blue
                ws.cell(row=current_r, column=c_idx).border = border_cell
        else:
            ws[f"A{current_r}"] = label
            is_highlight = label in ["Enterprise Value (EV) ($M)", "Implied Equity Value ($M)", "Implied Price per Share ($)", "Implied Upside / (Downside) (%)"]
            ws[f"A{current_r}"].font = font_bold if is_highlight else font_regular
            ws[f"A{current_r}"].border = border_cell
            if is_highlight:
                ws[f"A{current_r}"].fill = fill_accent_blue
                
            ws[f"B{current_r}"] = formula_val
            ws[f"B{current_r}"].font = font_bold if is_highlight else font_regular
            ws[f"B{current_r}"].number_format = num_fmt
            ws[f"B{current_r}"].alignment = align_right
            ws[f"B{current_r}"].border = border_cell
            if is_highlight:
                ws[f"B{current_r}"].fill = fill_accent_blue
        current_r += 1

    # --- Section VII: Sensitivity Analysis (5x5 Institutional Grids) ---
    add_section_header(106, "VII. SENSITIVITY ANALYSIS (5x5 INSTITUTIONAL VALUATION GRIDS)")

    # Table 1: WACC vs Terminal Growth Rate
    ws["A107"] = "Sensitivity Table 1: Implied Share Price ($) vs. WACC and Perpetual Terminal Growth Rate (g)"
    ws["A107"].font = font_subhdr
    ws.merge_cells("A107:H107")

    ws["A108"] = "WACC \\ g"
    ws["A108"].font = font_tbl_hdr
    ws["A108"].fill = fill_soft_blue
    ws["A108"].alignment = align_center
    ws["A108"].border = border_header

    t1_g_values = [0.020, 0.025, 0.030, 0.035, 0.040]
    t1_wacc_values = [0.1017, 0.1117, 0.1217, 0.1317, 0.1417]

    for col_idx, g_val in enumerate(t1_g_values, start=2):
        c = ws.cell(row=108, column=col_idx, value=g_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "0.0%"
        c.alignment = align_center
        c.border = border_header

    for row_idx, w_val in enumerate(t1_wacc_values, start=109):
        ws.cell(row=row_idx, column=1, value=w_val)
        ws.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 111 else fill_soft_blue
        ws.cell(row=row_idx, column=1).number_format = "0.00%"
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            # Full DCF recalculation formula:
            formula = (
                f"=(($C$74/(1+$A{row_idx})^0.5 + $D$74/(1+$A{row_idx})^1.5 + $E$74/(1+$A{row_idx})^2.5 + "
                f"$F$74/(1+$A{row_idx})^3.5 + $G$74/(1+$A{row_idx})^4.5 + "
                f"($G$74*(1+{col_let}$108)/($A{row_idx}-{col_let}$108))/(1+$A{row_idx})^5) - $B$12)/$B$8"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 111 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 111 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!

    # Table 2: FY27E Revenue Growth vs Target FY31E EBIT Margin
    ws["A116"] = "Sensitivity Table 2: Implied Share Price ($) vs. FY27E Revenue Growth & FY31E Target EBIT Margin"
    ws["A116"].font = font_subhdr
    ws.merge_cells("A116:H116")

    ws["A117"] = "EBIT Margin \\ Rev Growth"
    ws["A117"].font = font_tbl_hdr
    ws["A117"].fill = fill_soft_blue
    ws["A117"].alignment = align_center
    ws["A117"].border = border_header

    t2_rev_growth = [0.120, 0.140, 0.160, 0.180, 0.200]
    t2_ebit_margins = [0.110, 0.130, 0.150, 0.170, 0.190]

    for col_idx, rev_val in enumerate(t2_rev_growth, start=2):
        c = ws.cell(row=117, column=col_idx, value=rev_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "0.0%"
        c.alignment = align_center
        c.border = border_header

    for row_idx, ebit_val in enumerate(t2_ebit_margins, start=118):
        ws.cell(row=row_idx, column=1, value=ebit_val)
        ws.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 120 else fill_soft_blue
        ws.cell(row=row_idx, column=1).number_format = "0.0%"
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            # Full DCF recalculation scaling explicit FCFs & terminal value by revenue growth & margin:
            formula = (
                f"=((($C$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67))/(1+$B$61)^0.5 + "
                f"  ($D$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67))/(1+$B$61)^1.5 + "
                f"  ($E$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67))/(1+$B$61)^2.5 + "
                f"  ($F$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67))/(1+$B$61)^3.5 + "
                f"  ($G$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67))/(1+$B$61)^4.5 + "
                f"  (($G$74*(1+({col_let}$117-$C$65))*($A{row_idx}/$G$67)*(1+$H$41)/($B$61-$H$41))/(1+$B$61)^5)) - $B$12)/$B$8"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 120 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 120 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!

    # Table 3: Equity Beta vs Risk-Free Rate
    ws["A125"] = "Sensitivity Table 3: Implied Share Price ($) vs. Equity Beta (β) and Risk-Free Rate (Rf)"
    ws["A125"].font = font_subhdr
    ws.merge_cells("A125:H125")

    ws["A126"] = "Beta \\ Risk-Free Rate"
    ws["A126"].font = font_tbl_hdr
    ws["A126"].fill = fill_soft_blue
    ws["A126"].alignment = align_center
    ws["A126"].border = border_header

    t3_rf_values = [0.0373, 0.0423, 0.0473, 0.0523, 0.0573]
    t3_beta_values = [1.16, 1.36, 1.56, 1.76, 1.96]

    for col_idx, rf_val in enumerate(t3_rf_values, start=2):
        c = ws.cell(row=126, column=col_idx, value=rf_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "0.00%"
        c.alignment = align_center
        c.border = border_header

    for row_idx, b_val in enumerate(t3_beta_values, start=127):
        ws.cell(row=row_idx, column=1, value=b_val)
        ws.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 129 else fill_soft_blue
        ws.cell(row=row_idx, column=1).number_format = "0.00"
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            custom_wacc = f"(($A{row_idx}*$B$17 + {col_let}$126)*$B$59 + $B$55*$B$60)"
            formula = (
                f"=((($C$74/(1+{custom_wacc})^0.5 + $D$74/(1+{custom_wacc})^1.5 + $E$74/(1+{custom_wacc})^2.5 + "
                f"$F$74/(1+{custom_wacc})^3.5 + $G$74/(1+{custom_wacc})^4.5 + "
                f"($G$74*(1+$H$41)/({custom_wacc}-$H$41))/(1+{custom_wacc})^5) - $B$12)/$B$8)"
            )
            cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 129 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 129 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!

    # --- Column Width Adjustments ---
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 16

    wb.save(filename)
    print(f"Model saved successfully to {filename}")

if __name__ == "__main__":
    build_model()
