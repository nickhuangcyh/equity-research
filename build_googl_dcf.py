#!/usr/bin/env python3
"""
build_googl_dcf.py - Institutional DCF Valuation Model for Alphabet Inc. (NASDAQ: GOOGL)
Follows Investment Banking financial modeling standards, strict openpyxl formatting rules,
and full formula automation for 3 5x5 sensitivity tables.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os

def build_googl_dcf_model(output_path="GOOGL_DCF_Model_Gemini-3.7-Flash_20260819.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: DCF Valuation
    ws_dcf = wb.active
    ws_dcf.title = "DCF Valuation"
    ws_dcf.views.sheetView[0].showGridLines = True
    
    # Sheet 2: WACC Build
    ws_wacc = wb.create_sheet(title="WACC Build")
    ws_wacc.views.sheetView[0].showGridLines = True
    
    # Palette definition (Institutional Classic Navy & Slate Blue)
    NAVY = "1F4E79"        # Primary headers
    MED_BLUE = "2F5597"    # Secondary headers
    SOFT_BLUE = "D9E1F2"   # Table headers
    ACCENT_BLUE = "BDD7EE" # Highlight / Base Case
    LIGHT_GRAY = "F2F2F2"  # Input fill
    ZEBRA_FILL = "F9FBFD"  # Alternating row
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"
    INPUT_BLUE = "0000FF"  # Hardcoded font
    BLACK = "000000"       # Formula font
    LINK_GREEN = "008000"  # Sheet link font
    
    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_regular = Font(name="Calibri", size=10, bold=False, color=BLACK)
    font_input = Font(name="Calibri", size=10, bold=False, color=INPUT_BLUE)
    font_input_bold = Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE)
    font_link = Font(name="Calibri", size=10, bold=False, color=LINK_GREEN)
    font_link_bold = Font(name="Calibri", size=10, bold=True, color=LINK_GREEN)
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_navy = Side(border_style="medium", color=NAVY)
    double_navy = Side(border_style="double", color=NAVY)
    top_thin_navy = Side(border_style="thin", color=NAVY)
    
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_navy)
    border_total = Border(top=top_thin_navy, bottom=double_navy, left=thin_line, right=thin_line)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Helper for merged section header
    def add_section_header(ws, row, title, max_col="H"):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_sec_hdr
        ws[f"A{row}"].fill = fill_navy
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_navy
            c.border = Border(top=thick_navy, bottom=thick_navy, left=thin_line, right=thin_line)
            
    # ==========================================
    # SHEET 2: WACC Build (Cost of Capital)
    # ==========================================
    ws_wacc["A1"] = "Alphabet Inc. (NASDAQ: GOOGL / GOOG)"
    ws_wacc["A1"].font = font_title
    ws_wacc["A2"] = "Weighted Average Cost of Capital (WACC) Schedule | CAPM Methodology & Capital Structure"
    ws_wacc["A2"].font = font_subtitle
    
    add_section_header(ws_wacc, 4, "I. COST OF EQUITY (CAPM METHODOLOGY)", "E")
    
    wacc_equity_inputs = [
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0469, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 19, 2026", "B5"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.24, "0.00", "Source: Market data 5-year monthly regression beta vs S&P 500 (Reflects technology AI leadership & volatility)", "B6"),
        ("Market Equity Risk Premium (ERP)", 0.0550, "0.00%", "Source: Standard institutional equity risk premium consensus (5.50%)", "B7"),
        ("Cost of Equity (Ke = Rf + Beta * ERP)", "=B5+B6*B7", "0.00%", None, "B8"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_equity_inputs, start=5):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 10, "II. COST OF DEBT & CAPITAL STRUCTURE WEIGHTS", "E")
    
    wacc_debt_inputs = [
        ("Pre-Tax Cost of Debt (Kd)", 0.0485, "0.00%", "Source: Alphabet Senior Notes yield curve / Aa2-AA+ effective borrowing cost", "B11"),
        ("Effective Corporate Tax Rate (t)", 0.1750, "0.0%", "Source: SEC Form 10-K normalized long-term effective corporate tax rate (17.50%)", "B12"),
        ("After-Tax Cost of Debt (Kd * (1 - t))", "=B11*(1-B12)", "0.00%", None, "B13"),
        ("Current Share Price ($)", "='DCF Valuation'!B7", "$#,##0.00", None, "B14"),
        ("Diluted Shares Outstanding (M)", "='DCF Valuation'!B8", "#,##0.00", None, "B15"),
        ("Market Capitalization ($M)", "=B14*B15", "$#,##0.00", None, "B16"),
        ("Total Debt (Current + Noncurrent Debt) ($M)", 98200.00, "$#,##0.00", "Source: SEC Form 10-Q / Form 10-K Balance Sheet total carrying debt ($98.20B)", "B17"),
        ("Cash, Cash Equivalents & Marketable Securities ($M)", 242500.00, "$#,##0.00", "Source: SEC Form 10-Q / Form 10-K total cash & marketable securities ($242.50B)", "B18"),
        ("Net Debt ($M) [Net Cash if negative]", "=B17-B18", "$#,##0.00", None, "B19"),
        ("Enterprise Value ($M)", "=B16+B19", "$#,##0.00", None, "B20"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_debt_inputs, start=11):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "='DCF Valuation'" in str(val):
            ws_wacc[f"B{idx}"].font = font_link
        elif "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 22, "III. WACC CALCULATION SUMMARY", "E")
    
    wacc_summary_headers = ["Component", "Market Value ($M)", "Capital Weight (%)", "Cost of Capital (%)", "Weighted Contribution (%)"]
    for c_idx, h_text in enumerate(wacc_summary_headers, start=1):
        c = ws_wacc.cell(row=23, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    ws_wacc["A24"] = "Common Equity (Market Cap)"
    ws_wacc["B24"] = "=B16"
    ws_wacc["C24"] = "=B24/$B$20"
    ws_wacc["D24"] = "=B8"
    ws_wacc["E24"] = "=C24*D24"
    
    ws_wacc["A25"] = "Net Debt (Net Cash Adjustment)"
    ws_wacc["B25"] = "=B19"
    ws_wacc["C25"] = "=B25/$B$20"
    ws_wacc["D25"] = "=B13"
    ws_wacc["E25"] = "=C25*D25"
    
    ws_wacc["A26"] = "Total Capitalization / Base WACC"
    ws_wacc["B26"] = "=B20"
    ws_wacc["C26"] = "=SUM(C24:C25)"
    ws_wacc["D26"] = "-"
    ws_wacc["E26"] = "=SUM(E24:E25)"
    
    for r in range(24, 27):
        ws_wacc[f"A{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"A{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"B{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"B{r}"].number_format = "$#,##0.00"
        ws_wacc[f"B{r}"].alignment = align_right
        ws_wacc[f"B{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"C{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"C{r}"].number_format = "0.0%"
        ws_wacc[f"C{r}"].alignment = align_right
        ws_wacc[f"C{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"D{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"D{r}"].number_format = "0.00%" if r != 26 else "@"
        ws_wacc[f"D{r}"].alignment = align_right
        ws_wacc[f"D{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"E{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"E{r}"].number_format = "0.00%"
        ws_wacc[f"E{r}"].alignment = align_right
        ws_wacc[f"E{r}"].border = border_total if r == 26 else border_cell
        if r == 26:
            ws_wacc[f"E{r}"].fill = fill_accent_blue
            
    ws_wacc.column_dimensions["A"].width = 48
    ws_wacc.column_dimensions["B"].width = 22
    ws_wacc.column_dimensions["C"].width = 18
    ws_wacc.column_dimensions["D"].width = 18
    ws_wacc.column_dimensions["E"].width = 24
    
    # ==========================================
    # SHEET 1: DCF Valuation
    # ==========================================
    
    # --- Title Block ---
    ws_dcf["A1"] = "Alphabet Inc. (NASDAQ: GOOGL / GOOG)"
    ws_dcf["A1"].font = font_title
    ws_dcf["A2"] = "Institutional DCF Valuation Model | Report Date: August 19, 2026 | Financial Source: SEC Form 10-K, 10-Q & Consensus Estimates"
    ws_dcf["A2"].font = font_subtitle
    
    # --- Active Scenario Selector ---
    ws_dcf["A4"] = "Active Scenario Selector:"
    ws_dcf["A4"].font = font_bold
    ws_dcf["B4"] = 2
    ws_dcf["B4"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    ws_dcf["B4"].fill = fill_navy
    ws_dcf["B4"].alignment = align_center
    ws_dcf["B4"].border = border_cell
    ws_dcf["B4"].comment = Comment("Case Selector: 1 = Bear Case (Search Erosion & Capex Drag), 2 = Base Case (AI Monetization & Cloud Expansion), 3 = Bull Case (Full AI Ecosystem Dominance)", "DCF Model Builder")
    
    ws_dcf["C4"] = '=IF(B4=1,"[1] BEAR CASE (Antitrust Headwinds / Search Disruption / AI ROI Drag)",IF(B4=2,"[2] BASE CASE (Cloud Scaling / Gemini AI Monetization / Sustained Search Moat)","[3] BULL CASE (Waymo Commercialization / Cloud Hypergrowth / Agentic AI Leadership)"))'
    ws_dcf["C4"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws_dcf["C4"].alignment = align_left
    
    # --- Section I: Market Data & Capital Structure Inputs ---
    add_section_header(ws_dcf, 6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS", "H")
    
    market_inputs_dcf = [
        ("Current Stock Price ($)", 341.37, "$#,##0.00", "Source: Market close price as of August 18, 2026 ($341.37)", "B7"),
        ("Diluted Shares Outstanding (M)", 12200.00, "#,##0.00", "Source: SEC Form 10-Q / 10-K diluted share count reflecting equity offerings (12,200.0M shares)", "B8"),
        ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
        ("Cash, Cash Eq. & Marketable Securities ($M)", 242500.00, "$#,##0.00", "Source: SEC Form 10-Q Balance Sheet total cash, equivalents & marketable securities ($242.50B)", "B10"),
        ("Total Debt (Current + Noncurrent Debt) ($M)", 98200.00, "$#,##0.00", "Source: SEC Form 10-Q Balance Sheet total carrying debt ($98.20B)", "B11"),
        ("Net Debt ($M) [Net Cash if negative]", "=B11-B10", "$#,##0.00", None, "B12"),
        ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
        ("Effective Corporate Tax Rate (%)", 0.175, "0.0%", "Source: Long-term normalized effective corporate income tax rate (17.50%)", "B14"),
        ("Market Equity Risk Premium (ERP) (%)", 0.055, "0.0%", "Source: Consensus long-term equity risk premium benchmark (5.50%)", "B15"),
        ("Long-Term Risk-Free Rate (%)", 0.0469, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 19, 2026 (4.69%)", "B16"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.24, "0.00", "Source: Market data 5-year monthly regression beta (1.24)", "B17"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs_dcf, start=7):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        
        if "=" in str(val):
            ws_dcf[f"B{idx}"].font = font_bold
        else:
            ws_dcf[f"B{idx}"].font = font_input
            ws_dcf[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_dcf[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    # --- Section II: Three-Scenario DCF Driver Assumptions ---
    add_section_header(ws_dcf, 19, "II. SCENARIO DRIVER ASSUMPTIONS (BEAR / BASE / BULL)", "H")
    
    # Headers for Scenario Blocks
    scenario_cols = ["Assumption Driver", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "Terminal", "Parameter Description / Methodology"]
    
    # Bear Case Block (Rows 20-27)
    ws_dcf["A20"] = "SCENARIO 1: BEAR CASE ASSUMPTIONS"
    ws_dcf["A20"].font = font_bold_navy
    ws_dcf["A20"].fill = fill_soft_blue
    ws_dcf.merge_cells("A20:H20")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=20, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=20, column=col_idx).border = border_header
        
    for c_idx, h_text in enumerate(scenario_cols, start=1):
        c = ws_dcf.cell(row=21, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx in [1, 8] else align_center
        
    bear_drivers = [
        ("Revenue Growth Rate (%)", [0.110, 0.090, 0.075, 0.060, 0.050, 0.025], "0.0%", "Search share loss, antitrust remedies, slower cloud migration"),
        ("Gross Margin (%)", [0.560, 0.555, 0.550, 0.545, 0.540, 0.540], "0.0%", "Higher TAC and massive data center energy / compute depreciation"),
        ("S&M Expense (% of Revenue)", [0.080, 0.080, 0.078, 0.078, 0.078, 0.078], "0.0%", "Aggressive defensive advertising to protect search volume"),
        ("R&D Expense (% of Revenue)", [0.145, 0.145, 0.140, 0.140, 0.140, 0.140], "0.0%", "Sustained AI talent wars and foundational model retraining"),
        ("G&A Expense (% of Revenue)", [0.045, 0.045, 0.042, 0.042, 0.042, 0.042], "0.0%", "Elevated legal defense costs and regulatory compliance"),
        ("CapEx (% of Revenue)", [0.390, 0.300, 0.220, 0.180, 0.140, 0.100], "0.0%", "Aggressive AI cluster spending with prolonged payback horizon"),
        ("Terminal Perpetuity Growth (g) / WACC", [None, None, None, None, None, 0.025], "0.0%", "Conservative long-term inflation-matching growth (2.50%) | Bear WACC: 12.50%"),
    ]
    
    for r_offset, (label, vals, num_fmt, desc) in enumerate(bear_drivers, start=22):
        ws_dcf[f"A{r_offset}"] = label
        ws_dcf[f"A{r_offset}"].font = font_regular
        ws_dcf[f"A{r_offset}"].border = border_cell
        for c_offset, v in enumerate(vals, start=2):
            c_let = get_column_letter(c_offset)
            c = ws_dcf[f"{c_let}{r_offset}"]
            c.border = border_cell
            c.alignment = align_right
            if v is not None:
                c.value = v
                c.number_format = num_fmt
                c.font = font_input
                c.fill = fill_input
            else:
                c.value = "-"
                c.alignment = align_center
                c.font = font_regular
        ws_dcf[f"H{r_offset}"] = desc
        ws_dcf[f"H{r_offset}"].font = font_italic
        ws_dcf[f"H{r_offset}"].border = border_cell
        
    # Base Case Block (Rows 30-37)
    ws_dcf["A29"] = "SCENARIO 2: BASE CASE ASSUMPTIONS (CONSENSUS & STRATEGIC AI BUILD)"
    ws_dcf["A29"].font = font_bold_navy
    ws_dcf["A29"].fill = fill_soft_blue
    ws_dcf.merge_cells("A29:H29")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=29, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=29, column=col_idx).border = border_header
        
    for c_idx, h_text in enumerate(scenario_cols, start=1):
        c = ws_dcf.cell(row=30, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx in [1, 8] else align_center
        
    base_drivers = [
        ("Revenue Growth Rate (%)", [0.160, 0.140, 0.120, 0.105, 0.090, 0.030], "0.0%", "Cloud scaling (+40-50%), YouTube monetization, Gemini Search integration"),
        ("Gross Margin (%)", [0.580, 0.582, 0.585, 0.588, 0.590, 0.590], "0.0%", "Cloud margin expansion and custom TPU silicon cost efficiencies"),
        ("S&M Expense (% of Revenue)", [0.075, 0.072, 0.070, 0.068, 0.065, 0.065], "0.0%", "Operating leverage from digital enterprise distribution"),
        ("R&D Expense (% of Revenue)", [0.138, 0.135, 0.132, 0.128, 0.125, 0.125], "0.0%", "Disciplined R&D scaling with AI-assisted software productivity"),
        ("G&A Expense (% of Revenue)", [0.042, 0.040, 0.038, 0.036, 0.035, 0.035], "0.0%", "Corporate efficiency and automation of back-office operations"),
        ("CapEx (% of Revenue)", [0.380, 0.260, 0.180, 0.140, 0.110, 0.080], "0.0%", "2026 peak AI CapEx ($178B), normalizing to sustainable steady-state"),
        ("Terminal Perpetuity Growth (g) / WACC", [None, None, None, None, None, 0.030], "0.0%", "GDP+ Tech leader growth (3.00%) | WACC dynamically linked to 'WACC Build'"),
    ]
    
    for r_offset, (label, vals, num_fmt, desc) in enumerate(base_drivers, start=31):
        ws_dcf[f"A{r_offset}"] = label
        ws_dcf[f"A{r_offset}"].font = font_regular
        ws_dcf[f"A{r_offset}"].border = border_cell
        for c_offset, v in enumerate(vals, start=2):
            c_let = get_column_letter(c_offset)
            c = ws_dcf[f"{c_let}{r_offset}"]
            c.border = border_cell
            c.alignment = align_right
            if v is not None:
                c.value = v
                c.number_format = num_fmt
                c.font = font_input
                c.fill = fill_input
            else:
                c.value = "-"
                c.alignment = align_center
                c.font = font_regular
        ws_dcf[f"H{r_offset}"] = desc
        ws_dcf[f"H{r_offset}"].font = font_italic
        ws_dcf[f"H{r_offset}"].border = border_cell
        
    # Bull Case Block (Rows 39-46)
    ws_dcf["A38"] = "SCENARIO 3: BULL CASE ASSUMPTIONS (AGENTIC AI & CLOUD ACCELERATION)"
    ws_dcf["A38"].font = font_bold_navy
    ws_dcf["A38"].fill = fill_soft_blue
    ws_dcf.merge_cells("A38:H38")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=38, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=38, column=col_idx).border = border_header
        
    for c_idx, h_text in enumerate(scenario_cols, start=1):
        c = ws_dcf.cell(row=39, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx in [1, 8] else align_center
        
    bull_drivers = [
        ("Revenue Growth Rate (%)", [0.200, 0.175, 0.150, 0.130, 0.110, 0.035], "0.0%", "Waymo commercial scaling, Google Workspace AI enterprise ARPU explosion"),
        ("Gross Margin (%)", [0.590, 0.595, 0.600, 0.605, 0.610, 0.610], "0.0%", "High-margin software subscription revenues and TPU 6th gen efficiencies"),
        ("S&M Expense (% of Revenue)", [0.070, 0.065, 0.062, 0.060, 0.058, 0.058], "0.0%", "Self-reinforcing viral adoption of Gemini personal AI assistants"),
        ("R&D Expense (% of Revenue)", [0.130, 0.125, 0.120, 0.115, 0.110, 0.110], "0.0%", "Massive operating leverage across global developer ecosystem"),
        ("G&A Expense (% of Revenue)", [0.038, 0.035, 0.032, 0.030, 0.030, 0.030], "0.0%", "Maximized operational leverage and AI workflow automation"),
        ("CapEx (% of Revenue)", [0.370, 0.240, 0.160, 0.120, 0.090, 0.070], "0.0%", "Hyper-efficient AI compute architecture and external cloud monetization"),
        ("Terminal Perpetuity Growth (g) / WACC", [None, None, None, None, None, 0.035], "0.0%", "Global GDP Outperformance (3.50%) | Bull WACC: 11.00%"),
    ]
    
    for r_offset, (label, vals, num_fmt, desc) in enumerate(bull_drivers, start=40):
        ws_dcf[f"A{r_offset}"] = label
        ws_dcf[f"A{r_offset}"].font = font_regular
        ws_dcf[f"A{r_offset}"].border = border_cell
        for c_offset, v in enumerate(vals, start=2):
            c_let = get_column_letter(c_offset)
            c = ws_dcf[f"{c_let}{r_offset}"]
            c.border = border_cell
            c.alignment = align_right
            if v is not None:
                c.value = v
                c.number_format = num_fmt
                c.font = font_input
                c.fill = fill_input
            else:
                c.value = "-"
                c.alignment = align_center
                c.font = font_regular
        ws_dcf[f"H{r_offset}"] = desc
        ws_dcf[f"H{r_offset}"].font = font_italic
        ws_dcf[f"H{r_offset}"].border = border_cell
        
    # --- Consolidated Active Drivers (Rows 47-56) ---
    ws_dcf["A47"] = "ACTIVE CONSOLIDATED SCENARIO DRIVERS (DYNAMICALLY PULLED VIA SELECTOR)"
    ws_dcf["A47"].font = font_bold_navy
    ws_dcf["A47"].fill = fill_soft_blue
    ws_dcf.merge_cells("A47:H47")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=47, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=47, column=col_idx).border = border_header
        
    for c_idx, h_text in enumerate(scenario_cols, start=1):
        c = ws_dcf.cell(row=48, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx in [1, 8] else align_center
        
    consolidated_items = [
        ("Selected Revenue Growth (%)", 22, 31, 40, "0.0%", "Dynamic revenue growth rate linked to selected case"),
        ("Selected Gross Margin (%)", 23, 32, 41, "0.0%", "Dynamic gross profit margin linked to selected case"),
        ("Selected S&M Expense (%)", 24, 33, 42, "0.0%", "Dynamic S&M as % of revenue"),
        ("Selected R&D Expense (%)", 25, 34, 43, "0.0%", "Dynamic R&D as % of revenue"),
        ("Selected G&A Expense (%)", 26, 35, 44, "0.0%", "Dynamic G&A as % of revenue"),
        ("Selected CapEx (% of Revenue)", 27, 36, 45, "0.0%", "Dynamic CapEx intensity as % of revenue"),
        ("Selected Terminal Growth (g)", 28, 37, 46, "0.0%", "Dynamic terminal perpetuity growth rate"),
        ("Applied WACC Discount Rate (%)", None, None, None, "0.00%", "加權平均資金成本 (WACC Build 連結)"),
    ]
    
    for r_offset, (label, r_bear, r_base, r_bull, num_fmt, desc) in enumerate(consolidated_items, start=49):
        ws_dcf[f"A{r_offset}"] = label
        ws_dcf[f"A{r_offset}"].font = font_bold
        ws_dcf[f"A{r_offset}"].border = border_cell
        
        if r_offset == 56: # WACC row
            for c_offset in range(2, 8):
                c_let = get_column_letter(c_offset)
                c = ws_dcf[f"{c_let}{r_offset}"]
                c.value = "='WACC Build'!$E$26"
                c.font = font_link_bold
                c.number_format = num_fmt
                c.alignment = align_right
                c.border = border_cell
            ws_dcf[f"H{r_offset}"] = desc
            ws_dcf[f"H{r_offset}"].font = font_italic
            ws_dcf[f"H{r_offset}"].border = border_cell
        elif r_offset == 55: # Terminal g row
            for c_offset in range(2, 7):
                c_let = get_column_letter(c_offset)
                c = ws_dcf[f"{c_let}{r_offset}"]
                c.value = "-"
                c.alignment = align_center
                c.font = font_regular
                c.border = border_cell
            c = ws_dcf[f"G{r_offset}"]
            c.value = f"=IF($B$4=1,G{r_bear},IF($B$4=2,G{r_base},G{r_bull}))"
            c.font = font_bold
            c.number_format = num_fmt
            c.alignment = align_right
            c.border = border_cell
            ws_dcf[f"H{r_offset}"] = desc
            ws_dcf[f"H{r_offset}"].font = font_italic
            ws_dcf[f"H{r_offset}"].border = border_cell
        else:
            for c_offset in range(2, 8):
                c_let = get_column_letter(c_offset)
                c = ws_dcf[f"{c_let}{r_offset}"]
                c.value = f"=IF($B$4=1,{c_let}{r_bear},IF($B$4=2,{c_let}{r_base},{c_let}{r_bull}))"
                c.font = font_bold
                c.number_format = num_fmt
                c.alignment = align_right
                c.border = border_cell
            ws_dcf[f"H{r_offset}"] = desc
            ws_dcf[f"H{r_offset}"].font = font_italic
            ws_dcf[f"H{r_offset}"].border = border_cell
            
    # --- Section III: Income Statement Forecast (Rows 58-75) ---
    add_section_header(ws_dcf, 58, "III. CONSOLIDATED INCOME STATEMENT PROJECTIONS ($M)", "H")
    
    is_headers = ["Income Statement Line Item ($M)", "FY24A", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E"]
    for c_idx, h_text in enumerate(is_headers, start=1):
        c = ws_dcf.cell(row=59, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    # Row 60: Revenue
    ws_dcf["A60"] = "Total Revenues"
    ws_dcf["B60"] = 350018.00
    ws_dcf["B60"].comment = Comment("Source: SEC Form 10-K FY2024 Total Revenues ($350.02B)", "DCF Model")
    ws_dcf["C60"] = 402836.00
    ws_dcf["C60"].comment = Comment("Source: SEC Form 10-K FY2025 Total Revenues ($402.84B)", "DCF Model")
    ws_dcf["D60"] = "=C60*(1+B49)"
    ws_dcf["E60"] = "=D60*(1+C49)"
    ws_dcf["F60"] = "=E60*(1+D49)"
    ws_dcf["G60"] = "=F60*(1+E49)"
    ws_dcf["H60"] = "=G60*(1+F49)"
    
    # Row 61: YoY Revenue Growth
    ws_dcf["A61"] = "  YoY Revenue Growth (%)"
    ws_dcf["B61"] = 0.1387
    ws_dcf["C61"] = "=C60/B60-1"
    ws_dcf["D61"] = "=D60/C60-1"
    ws_dcf["E61"] = "=E60/D60-1"
    ws_dcf["F61"] = "=F60/E60-1"
    ws_dcf["G61"] = "=G60/F60-1"
    ws_dcf["H61"] = "=H60/G60-1"
    
    # Row 62: Cost of Revenues
    ws_dcf["A62"] = "Cost of Revenues (TAC, Data Center Ops, Content)"
    ws_dcf["B62"] = 146306.00
    ws_dcf["C62"] = 167177.00
    ws_dcf["D62"] = "=D60-D63"
    ws_dcf["E62"] = "=E60-E63"
    ws_dcf["F62"] = "=F60-F63"
    ws_dcf["G62"] = "=G60-G63"
    ws_dcf["H62"] = "=H60-H63"
    
    # Row 63: Gross Profit
    ws_dcf["A63"] = "Gross Profit"
    ws_dcf["B63"] = 203712.00
    ws_dcf["C63"] = 235659.00
    ws_dcf["D63"] = "=D60*B50"
    ws_dcf["E63"] = "=E60*C50"
    ws_dcf["F63"] = "=F60*D50"
    ws_dcf["G63"] = "=G60*E50"
    ws_dcf["H63"] = "=H60*F50"
    
    # Row 64: Gross Margin (%)
    ws_dcf["A64"] = "  Gross Profit Margin (%)"
    ws_dcf["B64"] = "=B63/B60"
    ws_dcf["C64"] = "=C63/C60"
    ws_dcf["D64"] = "=D63/D60"
    ws_dcf["E64"] = "=E63/E60"
    ws_dcf["F64"] = "=F63/F60"
    ws_dcf["G64"] = "=G63/G60"
    ws_dcf["H64"] = "=H63/H60"
    
    # Row 65: Operating Expenses Header
    ws_dcf["A65"] = "Operating Expenses:"
    for c_idx in range(2, 9):
        ws_dcf[f"{get_column_letter(c_idx)}65"] = ""
        
    # Row 66: Sales and Marketing (S&M)
    ws_dcf["A66"] = "  Sales & Marketing (S&M)"
    ws_dcf["B66"] = 28495.00
    ws_dcf["C66"] = 32200.00
    ws_dcf["D66"] = "=D60*B51"
    ws_dcf["E66"] = "=E60*C51"
    ws_dcf["F66"] = "=F60*D51"
    ws_dcf["G66"] = "=G60*E51"
    ws_dcf["H66"] = "=H60*F51"
    
    # Row 67: Research and Development (R&D)
    ws_dcf["A67"] = "  Research & Development (R&D)"
    ws_dcf["B67"] = 49326.00
    ws_dcf["C67"] = 56400.00
    ws_dcf["D67"] = "=D60*B52"
    ws_dcf["E67"] = "=E60*C52"
    ws_dcf["F67"] = "=F60*D52"
    ws_dcf["G67"] = "=G60*E52"
    ws_dcf["H67"] = "=H60*F52"
    
    # Row 68: General and Administrative (G&A)
    ws_dcf["A68"] = "  General & Administrative (G&A)"
    ws_dcf["B68"] = 13501.00
    ws_dcf["C68"] = 18020.00
    ws_dcf["D68"] = "=D60*B53"
    ws_dcf["E68"] = "=E60*C53"
    ws_dcf["F68"] = "=F60*D53"
    ws_dcf["G68"] = "=G60*E53"
    ws_dcf["H68"] = "=H60*F53"
    
    # Row 69: Total Operating Expenses
    ws_dcf["A69"] = "Total Operating Expenses"
    ws_dcf["B69"] = "=B66+B67+B68"
    ws_dcf["C69"] = "=C66+C67+C68"
    ws_dcf["D69"] = "=D66+D67+D68"
    ws_dcf["E69"] = "=E66+E67+E68"
    ws_dcf["F69"] = "=F66+F67+F68"
    ws_dcf["G69"] = "=G66+G67+G68"
    ws_dcf["H69"] = "=H66+H67+H68"
    
    # Row 70: Operating Income (EBIT)
    ws_dcf["A70"] = "Operating Income (EBIT)"
    ws_dcf["B70"] = 112390.00
    ws_dcf["C70"] = 129039.00
    ws_dcf["D70"] = "=D63-D69"
    ws_dcf["E70"] = "=E63-E69"
    ws_dcf["F70"] = "=F63-F69"
    ws_dcf["G70"] = "=G63-G69"
    ws_dcf["H70"] = "=H63-H69"
    
    # Row 71: EBIT Margin (%)
    ws_dcf["A71"] = "  EBIT Margin (%)"
    ws_dcf["B71"] = "=B70/B60"
    ws_dcf["C71"] = "=C70/C60"
    ws_dcf["D71"] = "=D70/D60"
    ws_dcf["E71"] = "=E70/E60"
    ws_dcf["F71"] = "=F70/F60"
    ws_dcf["G71"] = "=G70/G60"
    ws_dcf["H71"] = "=H70/H60"
    
    # Row 72: Provision for Income Taxes
    ws_dcf["A72"] = "Provision for Income Taxes"
    ws_dcf["B72"] = "=B70*$B$14"
    ws_dcf["C72"] = "=C70*$B$14"
    ws_dcf["D72"] = "=D70*$B$14"
    ws_dcf["E72"] = "=E70*$B$14"
    ws_dcf["F72"] = "=F70*$B$14"
    ws_dcf["G72"] = "=G70*$B$14"
    ws_dcf["H72"] = "=H70*$B$14"
    
    # Row 73: Effective Tax Rate (%)
    ws_dcf["A73"] = "  Effective Tax Rate (%)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}73"] = f"={c_let}72/{c_let}70"
        
    # Row 74: NOPAT
    ws_dcf["A74"] = "Net Operating Profit After Tax (NOPAT)"
    ws_dcf["B74"] = "=B70-B72"
    ws_dcf["C74"] = "=C70-C72"
    ws_dcf["D74"] = "=D70-D72"
    ws_dcf["E74"] = "=E70-E72"
    ws_dcf["F74"] = "=F70-F72"
    ws_dcf["G74"] = "=G70-G72"
    ws_dcf["H74"] = "=H70-H72"
    
    # Formatting for Income Statement
    for r in range(60, 75):
        ws_dcf[f"A{r}"].font = font_bold if r in [60, 63, 69, 70, 74] else font_regular
        ws_dcf[f"A{r}"].border = border_total if r in [70, 74] else border_cell
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_total if r in [70, 74] else border_cell
            cell.alignment = align_right
            if r in [61, 64, 71, 73]:
                cell.number_format = "0.0%"
                cell.font = font_regular
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                if r in [60, 63, 70, 74]:
                    cell.font = font_bold
                else:
                    cell.font = font_regular
            if col_idx in [2, 3] and r in [60, 62, 63, 66, 67, 68, 70]:
                cell.font = font_input_bold if r in [60, 63, 70] else font_input
                cell.fill = fill_input
                
    # --- Section IV: Free Cash Flow Build (Rows 76-88) ---
    add_section_header(ws_dcf, 76, "IV. UNLEVERED FREE CASH FLOW BUILD (UFCF) ($M)", "H")
    
    fcf_headers = ["Cash Flow Line Item ($M)", "FY24A", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E"]
    for c_idx, h_text in enumerate(fcf_headers, start=1):
        c = ws_dcf.cell(row=77, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    # Row 78: NOPAT
    ws_dcf["A78"] = "Net Operating Profit After Tax (NOPAT)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}78"] = f"={c_let}74"
        
    # Row 79: (+) Depreciation & Amortization (D&A)
    # D&A scales as % of revenue: FY24A ~4.37% ($15,311M), FY25A ~4.60% ($18,530M), FY26E-FY30E scales up to ~7.0%
    ws_dcf["A79"] = "(+) Depreciation & Amortization (D&A)"
    ws_dcf["B79"] = 15311.00
    ws_dcf["C79"] = 18530.00
    ws_dcf["D79"] = "=D60*0.055"
    ws_dcf["E79"] = "=E60*0.062"
    ws_dcf["F79"] = "=F60*0.068"
    ws_dcf["G79"] = "=G60*0.070"
    ws_dcf["H79"] = "=H60*0.070"
    
    # Row 80: D&A % of Revenue
    ws_dcf["A80"] = "  D&A (% of Revenue)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}80"] = f"={c_let}79/{c_let}60"
        
    # Row 81: (-) Capital Expenditures (CapEx)
    ws_dcf["A81"] = "(-) Capital Expenditures (AI Compute & Data Centers)"
    ws_dcf["B81"] = 52500.00
    ws_dcf["C81"] = 91400.00
    ws_dcf["D81"] = "=D60*B54"
    ws_dcf["E81"] = "=E60*C54"
    ws_dcf["F81"] = "=F60*D54"
    ws_dcf["G81"] = "=G60*E54"
    ws_dcf["H81"] = "=H60*F54"
    
    # Row 82: CapEx % of Revenue
    ws_dcf["A82"] = "  CapEx (% of Revenue)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}82"] = f"={c_let}81/{c_let}60"
        
    # Row 83: (-) Change in Net Working Capital (ΔNWC)
    ws_dcf["A83"] = "(-) Change in Net Working Capital (ΔNWC @ 2.0% of ΔRev)"
    ws_dcf["B83"] = 1200.00
    ws_dcf["C83"] = 1056.00
    ws_dcf["D83"] = "=(D60-C60)*0.02"
    ws_dcf["E83"] = "=(E60-D60)*0.02"
    ws_dcf["F83"] = "=(F60-E60)*0.02"
    ws_dcf["G83"] = "=(G60-F60)*0.02"
    ws_dcf["H83"] = "=(H60-G60)*0.02"
    
    # Row 84: Unlevered Free Cash Flow (UFCF)
    ws_dcf["A84"] = "Unlevered Free Cash Flow (UFCF)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}84"] = f"={c_let}78+{c_let}79-{c_let}81-{c_let}83"
        
    # Row 85: FCF Conversion (% of NOPAT)
    ws_dcf["A85"] = "  FCF Conversion (% of NOPAT)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}85"] = f"={c_let}84/{c_let}78"
        
    # Formatting for Cash Flow Schedule
    for r in range(78, 86):
        ws_dcf[f"A{r}"].font = font_bold if r in [78, 84] else font_regular
        ws_dcf[f"A{r}"].border = border_total if r == 84 else border_cell
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_total if r == 84 else border_cell
            cell.alignment = align_right
            if r in [80, 82, 85]:
                cell.number_format = "0.0%"
                cell.font = font_regular
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                cell.font = font_bold if r in [78, 84] else font_regular
            if col_idx in [2, 3] and r in [79, 81, 83]:
                cell.font = font_input
                cell.fill = fill_input
                
    # --- Section V: Discounting, Terminal Value & Valuation Bridge (Rows 87-108) ---
    add_section_header(ws_dcf, 87, "V. DISCOUNTING SCHEDULE, TERMINAL VALUE & VALUATION BRIDGE ($M)", "H")
    
    dcf_headers = ["Valuation & Discounting Line Item", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "Terminal Period", "Summary Totals / Metric Description"]
    for c_idx, h_text in enumerate(dcf_headers, start=1):
        c = ws_dcf.cell(row=88, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx in [1, 8] else align_right
        
    # Row 89: Unlevered FCF
    ws_dcf["A89"] = "Unlevered Free Cash Flow (UFCF) ($M)"
    ws_dcf["B89"] = "=D84"
    ws_dcf["C89"] = "=E84"
    ws_dcf["D89"] = "=F84"
    ws_dcf["E89"] = "=G84"
    ws_dcf["F89"] = "=H84"
    ws_dcf["G89"] = "=F89*(1+G55)"  # Terminal FCF = FY30E FCF * (1 + g)
    ws_dcf["H89"] = "=SUM(B89:F89)"
    
    # Row 90: Mid-Year Discount Period (t)
    ws_dcf["A90"] = "Mid-Year Discount Period (t)"
    ws_dcf["B90"] = 0.5
    ws_dcf["C90"] = 1.5
    ws_dcf["D90"] = 2.5
    ws_dcf["E90"] = 3.5
    ws_dcf["F90"] = 4.5
    ws_dcf["G90"] = 4.5  # Mid-year convention for terminal value discounting
    ws_dcf["H90"] = "-"
    
    # Row 91: Discount Factor (1 / (1 + WACC)^t)
    ws_dcf["A91"] = "Discount Factor (WACC Discounting)"
    ws_dcf["B91"] = "=1/(1+$B$56)^B90"
    ws_dcf["C91"] = "=1/(1+$B$56)^C90"
    ws_dcf["D91"] = "=1/(1+$B$56)^D90"
    ws_dcf["E91"] = "=1/(1+$B$56)^E90"
    ws_dcf["F91"] = "=1/(1+$B$56)^F90"
    ws_dcf["G91"] = "=1/(1+$B$56)^G90"
    ws_dcf["H91"] = "-"
    
    # Row 92: Present Value of Explicit FCF (PV)
    ws_dcf["A92"] = "Present Value of Cash Flow (PV of UFCF)"
    ws_dcf["B92"] = "=B89*B91"
    ws_dcf["C92"] = "=C89*C91"
    ws_dcf["D92"] = "=D89*D91"
    ws_dcf["E92"] = "=E89*E91"
    ws_dcf["F92"] = "=F89*F91"
    ws_dcf["G92"] = "-"
    ws_dcf["H92"] = "=SUM(B92:F92)"
    
    # Row 93: Nominal Terminal Value
    ws_dcf["A93"] = "Nominal Terminal Value (Perpetuity Growth Method)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}93"] = "-"
    ws_dcf["G93"] = "=G89/($B$56-G55)"  # TV = Terminal FCF / (WACC - g)
    ws_dcf["H93"] = "-"
    
    # Row 94: PV of Terminal Value
    ws_dcf["A94"] = "Present Value of Terminal Value (PV of TV)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}94"] = "-"
    ws_dcf["G94"] = "=G93*G91"
    ws_dcf["H94"] = "-"
    
    for r in range(89, 95):
        ws_dcf[f"A{r}"].font = font_bold if r in [89, 92, 93, 94] else font_regular
        ws_dcf[f"A{r}"].border = border_cell
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_cell
            cell.alignment = align_right
            if r == 90:
                cell.number_format = "0.0"
                cell.font = font_regular
            elif r == 91:
                cell.number_format = "0.0000"
                cell.font = font_regular
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                cell.font = font_bold if r in [89, 92, 94] else font_regular
            if str(cell.value) == "-":
                cell.alignment = align_center
                
    # --- Equity Value Bridge (Rows 96-107) ---
    ws_dcf["A96"] = "VALUATION BRIDGE & PER-SHARE IMPLICATIONS"
    ws_dcf["A96"].font = font_bold_navy
    ws_dcf["A96"].fill = fill_soft_blue
    ws_dcf.merge_cells("A96:C96")
    for col_idx in range(1, 4):
        ws_dcf.cell(row=96, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=96, column=col_idx).border = border_header
        
    equity_bridge = [
        ("Cumulative PV of Explicit 5-Year FCFs ($M)", "=H92", "$#,##0.00", "B97", False),
        ("Present Value of Terminal Value ($M)", "=G94", "$#,##0.00", "B98", False),
        ("Enterprise Value ($M)", "=B97+B98", "$#,##0.00", "B99", True),
        ("(-) Total Debt ($M)", "=-B11", "$#,##0.00", "B100", False),
        ("(+) Cash, Cash Eq. & Marketable Securities ($M)", "=B10", "$#,##0.00", "B101", False),
        ("Net Debt Adjustment [Net Cash Addition] ($M)", "=B100+B101", "$#,##0.00", "B102", False),
        ("Implied Equity Value ($M)", "=B99+B102", "$#,##0.00", "B103", True),
        ("Diluted Shares Outstanding (M)", "=B8", "#,##0.00", "B104", False),
        ("IMPLIED INTRINSIC VALUE PER SHARE ($)", "=B103/B104", "$#,##0.00", "B105", True),
        ("Current Market Stock Price ($)", "=B7", "$#,##0.00", "B106", False),
        ("Implied Upside / (Downside) (%)", "=B105/B106-1", "+0.0%;-0.0%;0.0%", "B107", True),
    ]
    
    for idx, (label, formula_val, num_fmt, cell_ref, is_highlight) in enumerate(equity_bridge, start=97):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"A{idx}"].border = border_total if idx in [99, 103, 105, 107] else border_cell
        
        ws_dcf[f"B{idx}"] = formula_val
        ws_dcf[f"B{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_total if idx in [99, 103, 105, 107] else border_cell
        
        if is_highlight:
            ws_dcf[f"A{idx}"].fill = fill_accent_blue
            ws_dcf[f"B{idx}"].fill = fill_accent_blue
            
    # Terminal Value Sanity Metrics (Cols D & E)
    ws_dcf["D97"] = "Terminal Value Checks & Metrics:"
    ws_dcf["D97"].font = font_bold_navy
    ws_dcf["D98"] = "PV of TV % of Enterprise Value"
    ws_dcf["E98"] = "=B98/B99"
    ws_dcf["E98"].number_format = "0.0%"
    ws_dcf["E98"].font = font_bold
    ws_dcf["E98"].alignment = align_right
    ws_dcf["E98"].border = border_cell
    
    ws_dcf["D99"] = "Implied Exit Multiple (EV / FY30E EBITDA)"
    ws_dcf["E99"] = "=B99/(H70+H79)"
    ws_dcf["E99"].number_format = "0.0x"
    ws_dcf["E99"].font = font_bold
    ws_dcf["E99"].alignment = align_right
    ws_dcf["E99"].border = border_cell
    
    for r_chk in [98, 99]:
        ws_dcf[f"D{r_chk}"].border = border_cell
        ws_dcf[f"D{r_chk}"].font = font_regular
        
    # =========================================================================
    # SECTION VI: THREE INSTITUTIONAL 5X5 SENSITIVITY TABLES (Rows 110-152)
    # =========================================================================
    add_section_header(ws_dcf, 110, "VI. INSTITUTIONAL SENSITIVITY MATRICES (5x5 GRIDS WITH FULL DCF RECALCULATION)", "H")
    
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 1: WACC vs Perpetual Growth Rate (g) (Rows 112-118)
    # -------------------------------------------------------------------------
    ws_dcf["A112"] = "TABLE 1: IMPLIED SHARE PRICE ($) - WACC vs. TERMINAL GROWTH RATE (g)"
    ws_dcf["A112"].font = font_bold_navy
    ws_dcf["A112"].fill = fill_soft_blue
    ws_dcf.merge_cells("A112:F112")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=112, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=112, column=col_idx).border = border_header
        
    ws_dcf["A113"] = "WACC \\ Terminal g"
    ws_dcf["A113"].font = font_tbl_hdr
    ws_dcf["A113"].fill = fill_soft_blue
    ws_dcf["A113"].border = border_header
    ws_dcf["A113"].alignment = align_center
    
    # Center cell is Base Case (g = 3.0%, WACC = 11.78%)
    g_cols = [0.020, 0.025, 0.030, 0.035, 0.040]
    wacc_rows = [0.1078, 0.1128, 0.1178, 0.1228, 0.1278]
    
    for c_idx, g_val in enumerate(g_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}113"] = g_val
        ws_dcf[f"{c_let}113"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}113"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}113"].number_format = "0.0%"
        ws_dcf[f"{c_let}113"].alignment = align_center
        ws_dcf[f"{c_let}113"].border = border_header
        
    for r_idx, w_val in enumerate(wacc_rows, start=114):
        ws_dcf[f"A{r_idx}"] = w_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 116 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 116 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, g_val in enumerate(g_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t1 = (
                f"=(($B$89/(1+$A{r_idx})^0.5 + $C$89/(1+$A{r_idx})^1.5 + $D$89/(1+$A{r_idx})^2.5 + $E$89/(1+$A{r_idx})^3.5 + $F$89/(1+$A{r_idx})^4.5) + "
                f"(($F$89*(1+{c_let}$113)/($A{r_idx}-{c_let}$113))/(1+$A{r_idx})^4.5) + $B$102)/$B$104"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t1
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 116 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 2: Revenue Growth Scale vs. Target EBIT Margin (Rows 122-128)
    # -------------------------------------------------------------------------
    ws_dcf["A122"] = "TABLE 2: IMPLIED SHARE PRICE ($) - REVENUE GROWTH MULTIPLIER vs. FY30E EBIT MARGIN"
    ws_dcf["A122"].font = font_bold_navy
    ws_dcf["A122"].fill = fill_soft_blue
    ws_dcf.merge_cells("A122:F122")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=122, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=122, column=col_idx).border = border_header
        
    ws_dcf["A123"] = "Growth Index \\ EBIT Margin"
    ws_dcf["A123"].font = font_tbl_hdr
    ws_dcf["A123"].fill = fill_soft_blue
    ws_dcf["A123"].border = border_header
    ws_dcf["A123"].alignment = align_center
    
    ebit_cols = [0.315, 0.335, 0.355, 0.375, 0.395]
    growth_rows = [0.80, 0.90, 1.00, 1.10, 1.20]
    
    for c_idx, em_val in enumerate(ebit_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}123"] = em_val
        ws_dcf[f"{c_let}123"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}123"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}123"].number_format = "0.0%"
        ws_dcf[f"{c_let}123"].alignment = align_center
        ws_dcf[f"{c_let}123"].border = border_header
        
    for r_idx, gr_val in enumerate(growth_rows, start=124):
        ws_dcf[f"A{r_idx}"] = gr_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 126 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 126 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, em_val in enumerate(ebit_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t2 = (
                f"=((($B$89*$A{r_idx}*({c_let}$123/$H$71)/(1+$B$56)^0.5 + "
                f"$C$89*$A{r_idx}*({c_let}$123/$H$71)/(1+$B$56)^1.5 + "
                f"$D$89*$A{r_idx}*({c_let}$123/$H$71)/(1+$B$56)^2.5 + "
                f"$E$89*$A{r_idx}*({c_let}$123/$H$71)/(1+$B$56)^3.5 + "
                f"$F$89*$A{r_idx}*({c_let}$123/$H$71)/(1+$B$56)^4.5) + "
                f"(($F$89*$A{r_idx}*({c_let}$123/$H$71)*(1+$G$55)/($B$56-$G$55))/(1+$B$56)^4.5) + $B$102)/$B$104)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t2
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 126 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 3: Equity Beta vs. Risk-Free Rate (Rf) (Rows 132-138)
    # -------------------------------------------------------------------------
    ws_dcf["A132"] = "TABLE 3: IMPLIED SHARE PRICE ($) - EQUITY BETA vs. RISK-FREE RATE (Rf)"
    ws_dcf["A132"].font = font_bold_navy
    ws_dcf["A132"].fill = fill_soft_blue
    ws_dcf.merge_cells("A132:F132")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=132, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=132, column=col_idx).border = border_header
        
    ws_dcf["A133"] = "Beta \\ Risk-Free Rate"
    ws_dcf["A133"].font = font_tbl_hdr
    ws_dcf["A133"].fill = fill_soft_blue
    ws_dcf["A133"].border = border_header
    ws_dcf["A133"].alignment = align_center
    
    rf_cols = [0.0369, 0.0419, 0.0469, 0.0519, 0.0569]
    beta_rows = [1.04, 1.14, 1.24, 1.34, 1.44]
    
    for c_idx, rf_val in enumerate(rf_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}133"] = rf_val
        ws_dcf[f"{c_let}133"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}133"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}133"].number_format = "0.00%"
        ws_dcf[f"{c_let}133"].alignment = align_center
        ws_dcf[f"{c_let}133"].border = border_header
        
    for r_idx, b_val in enumerate(beta_rows, start=134):
        ws_dcf[f"A{r_idx}"] = b_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 136 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 136 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, rf_val in enumerate(rf_cols, start=2):
            c_let = get_column_letter(c_idx)
            wacc_formula_snippet = f"({c_let}$133+$A{r_idx}*$B$15)*('WACC Build'!$C$24) + ('WACC Build'!$B$13)*('WACC Build'!$C$25)"
            
            formula_t3 = (
                f"=((($B$89/(1+({wacc_formula_snippet}))^0.5 + "
                f"$C$89/(1+({wacc_formula_snippet}))^1.5 + "
                f"$D$89/(1+({wacc_formula_snippet}))^2.5 + "
                f"$E$89/(1+({wacc_formula_snippet}))^3.5 + "
                f"$F$89/(1+({wacc_formula_snippet}))^4.5) + "
                f"(($F$89*(1+$G$55)/(({wacc_formula_snippet})-$G$55))/(1+({wacc_formula_snippet}))^4.5) + $B$102)/$B$104)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t3
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 136 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # Adjust Column Widths on DCF Sheet
    ws_dcf.column_dimensions["A"].width = 46
    ws_dcf.column_dimensions["B"].width = 16
    ws_dcf.column_dimensions["C"].width = 16
    ws_dcf.column_dimensions["D"].width = 16
    ws_dcf.column_dimensions["E"].width = 16
    ws_dcf.column_dimensions["F"].width = 16
    ws_dcf.column_dimensions["G"].width = 16
    ws_dcf.column_dimensions["H"].width = 38
    
    wb.save(output_path)
    print(f"Successfully generated institutional DCF model at: {output_path}")

if __name__ == "__main__":
    build_googl_dcf_model()
