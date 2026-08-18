#!/usr/bin/env python3
"""
INGR DCF Model Builder
Ingredion Incorporated — 5-Year DCF Model (Bear / Base / Bull)
Generated: 2026-08-18
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import date

# ── Output path ──────────────────────────────────────────────────────────────
OUTPUT_PATH = "reports/dcf/INGR_DCF_Model_2026-08-18.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C_HDR_DARK   = "1F4E79"   # dark blue  — section headers
C_HDR_LIGHT  = "D9E1F2"   # light blue — sub-headers / column labels
C_INPUT_FILL = "F2F2F2"   # light grey — input cells
C_OUTPUT     = "BDD7EE"   # medium blue — key output rows
C_WHITE      = "FFFFFF"
C_FONT_BLUE  = "0000FF"   # hardcoded inputs
C_FONT_BLACK = "000000"   # formulas
C_FONT_GREEN = "008000"   # cross-sheet links
C_FONT_WHITE = "FFFFFF"

# ── Border helpers ────────────────────────────────────────────────────────────
def thin_side():   return Side(style="thin")
def medium_side(): return Side(style="medium")
def thick_side():  return Side(style="thick")

BORDER_THIN   = Border(left=thin_side(),   right=thin_side(),   top=thin_side(),   bottom=thin_side())
BORDER_MEDIUM = Border(left=medium_side(), right=medium_side(), top=medium_side(), bottom=medium_side())
BORDER_THICK  = Border(left=thick_side(),  right=thick_side(),  top=thick_side(),  bottom=thick_side())
BORDER_BOX_T  = Border(left=thick_side(),  right=thick_side(),  top=thick_side(),  bottom=medium_side())
BORDER_BOX_M  = Border(left=thick_side(),  right=thick_side(),  top=None,          bottom=None)
BORDER_BOX_B  = Border(left=thick_side(),  right=thick_side(),  top=medium_side(), bottom=thick_side())

# ── Fill helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Font helpers ──────────────────────────────────────────────────────────────
def font(color=C_FONT_BLACK, bold=False, size=10, name="Calibri"):
    return Font(color=color, bold=bold, size=size, name=name)

# ── Number formats ────────────────────────────────────────────────────────────
FMT_DOLLAR   = '$#,##0;($#,##0);"-"'
FMT_DOLLAR2  = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT1     = '0.0%;(0.0%);"-"'
FMT_NUM      = '#,##0;(#,##0);"-"'
FMT_TEXT     = "@"

# ── Cell writer helpers ───────────────────────────────────────────────────────
def write_val(ws, row, col, value, fmt=None, fg=C_FONT_BLACK, bold=False,
              fill_color=None, border=None, align="right", comment_text=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(fg, bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fmt:      c.number_format = fmt
    if fill_color: c.fill = fill(fill_color)
    if border:   c.border = border
    if comment_text:
        c.comment = Comment(comment_text, "DCF Model")
    return c

def write_formula(ws, row, col, formula, fmt=None, bold=False,
                  fill_color=None, border=None, align="right", green=False):
    fg = C_FONT_GREEN if green else C_FONT_BLACK
    return write_val(ws, row, col, formula, fmt=fmt, fg=fg, bold=bold,
                     fill_color=fill_color, border=border, align=align)

def write_header(ws, row, col, text, span_end_col=None, sub=False):
    """Write a section header; merges cols if span_end_col given."""
    if sub:
        fc, tc = C_FONT_BLACK, C_HDR_LIGHT
    else:
        fc, tc = C_FONT_WHITE, C_HDR_DARK
    write_val(ws, row, col, text, fg=fc, bold=True, fill_color=tc,
              align="left", border=BORDER_THIN)
    if span_end_col and span_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=span_end_col)
        # re-apply style to merged cell (openpyxl needs top-left only)
        c = ws.cell(row=row, column=col)
        c.font      = font(fc, bold=True)
        c.fill      = fill(tc)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = BORDER_THIN

def blank_row(ws, row, n_cols=12):
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).value = None

# ── Column widths helper ──────────────────────────────────────────────────────
def set_col_widths(ws, widths: dict):
    """widths = {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

print("✓ Part 1: helpers loaded")

# ═══════════════════════════════════════════════════════════════════════════════
# MODEL ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════

# ── Market data ───────────────────────────────────────────────────────────────
STOCK_PRICE   = 103.59   # USD, Aug 17 2026 close
SHARES_DILUT  = 64.0     # million diluted shares
TOTAL_DEBT    = 1971.0   # $M, FY2025
CASH          = 1030.0   # $M, FY2025
NET_DEBT      = TOTAL_DEBT - CASH   # 938.0

# ── WACC inputs ───────────────────────────────────────────────────────────────
RF            = 0.0470   # risk-free rate, 10Y UST, FRED 2026-08-11
BETA          = 0.70     # 60-month beta, Barchart
ERP           = 0.0500   # equity risk premium
KE            = RF + BETA * ERP            # 0.0820
KD_PRETAX     = 0.0450   # pre-tax cost of debt (market rate)
TAX_RATE      = 0.25
KD_AT         = KD_PRETAX * (1 - TAX_RATE) # 0.03375
MKTCAP        = STOCK_PRICE * SHARES_DILUT  # 6,629.76
EV_WACC       = MKTCAP + NET_DEBT           # 7,567.76
W_E           = MKTCAP / EV_WACC            # 0.8760
W_D           = NET_DEBT / EV_WACC          # 0.1240
WACC          = W_E * KE + W_D * KD_AT      # 0.07601

# ── Historical data (FY2021-FY2025, $M) ──────────────────────────────────────
HIST_YEARS  = ["FY2021A", "FY2022A", "FY2023A", "FY2024A", "FY2025A"]
HIST_REV    = [6894, 7946, 8160, 7430, 7219]
HIST_GP     = [1331, 1494, 1749, 1791, 1828]
HIST_EBIT   = [681,  771,  968, 1012, 1040]
HIST_DA     = [220,  215,  219,  214,  222]
HIST_CAPEX  = [300,  300,  314,  295,  433]
HIST_NI     = [117,  492,  643,  647,  729]

# ── Projection years ──────────────────────────────────────────────────────────
PROJ_YEARS  = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
N_PROJ      = 5
BASE_REV    = 7219.0   # FY2025A

# ── Scenario growth rates ─────────────────────────────────────────────────────
BEAR_GRW = [0.01, 0.01, 0.02, 0.02, 0.02]
BASE_GRW = [0.02, 0.03, 0.04, 0.03, 0.03]
BULL_GRW = [0.04, 0.05, 0.06, 0.05, 0.04]

# ── Scenario EBIT margins ─────────────────────────────────────────────────────
BEAR_EBIT_MRG = [0.135, 0.130, 0.130, 0.125, 0.125]
BASE_EBIT_MRG = [0.140, 0.145, 0.145, 0.150, 0.150]
BULL_EBIT_MRG = [0.145, 0.150, 0.155, 0.160, 0.160]

# ── Shared FCF parameters ─────────────────────────────────────────────────────
DA_PCT    = 0.030                          # D&A % of revenue
CAPEX_PCT = [0.055, 0.050, 0.045, 0.045, 0.045]  # CapEx % by year
NWC_PCT   = 0.080                          # ΔNwc % of ΔRev

# ── Terminal growth rates ─────────────────────────────────────────────────────
TERM_G = {"Bear": 0.020, "Base": 0.025, "Bull": 0.030}

# ── Mid-year convention discount periods ─────────────────────────────────────
DISC_PERIODS = [0.5, 1.5, 2.5, 3.5, 4.5]

print("✓ Part 2: assumptions loaded")

# ═══════════════════════════════════════════════════════════════════════════════
# WACC SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def build_wacc_sheet(wb):
    ws = wb.create_sheet("WACC")
    set_col_widths(ws, {"A": 38, "B": 16, "C": 16, "D": 16})
    ws.row_dimensions[1].height = 22

    r = 1
    # Title
    write_val(ws, r, 1, "Ingredion (INGR) — WACC Calculation",
              fg=C_FONT_WHITE, bold=True, fill_color=C_HDR_DARK, align="left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1)
    c.font = font(C_FONT_WHITE, bold=True, size=12)
    c.fill = fill(C_HDR_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    r += 1
    write_val(ws, r, 1, "As of: 2026-08-18  |  Source: StockAnalysis, FRED, Barchart",
              fg="595959", align="left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    # ── Cost of Equity ────────────────────────────────────────────────────────
    r += 2
    write_header(ws, r, 1, "COST OF EQUITY (CAPM)", span_end_col=4)

    r += 1
    write_val(ws, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws, r, 2, "Value", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    rows_ke = [
        ("Risk-Free Rate (10Y UST)",  RF,      FMT_PCT1,    "Source: FRED DGS10, 2026-08-11, https://fred.stlouisfed.org/series/DGS10"),
        ("Beta (60-Month Monthly)",   BETA,    "0.00",      "Source: Barchart.com, 2026-08-18, 60-month beta vs S&P 500"),
        ("Equity Risk Premium",       ERP,     FMT_PCT1,    "Source: Damodaran long-run ERP, 5.0% market standard"),
        ("Cost of Equity (Ke)",       "=B{r0}+B{r1}*B{r2}", FMT_PCT1, "= Rf + Beta × ERP"),
    ]
    ke_start = r + 1
    for i, (label, val, fmt, cmt) in enumerate(rows_ke):
        rr = ke_start + i
        write_val(ws, rr, 1, label, align="left", fill_color=C_INPUT_FILL if i < 3 else None)
        if i < 3:
            write_val(ws, rr, 2, val, fmt=fmt, fg=C_FONT_BLUE,
                      fill_color=C_INPUT_FILL, comment_text=cmt)
        else:
            formula = f"=B{ke_start}+B{ke_start+1}*B{ke_start+2}"
            write_formula(ws, rr, 2, formula, fmt=FMT_PCT1, bold=True,
                          fill_color=C_OUTPUT)
        write_val(ws, rr, 3, cmt if i < 3 else "= Rf + Beta × ERP",
                  align="left", fg="595959")
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    r = ke_start + 3

    # ── Cost of Debt ──────────────────────────────────────────────────────────
    r += 2
    write_header(ws, r, 1, "COST OF DEBT", span_end_col=4)
    r += 1
    write_val(ws, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws, r, 2, "Value", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    kd_start = r + 1
    rows_kd = [
        ("Pre-Tax Cost of Debt (Kd)", KD_PRETAX, FMT_PCT1,
         "Source: Market rate estimate; BBB+ equivalent + ~80bps spread over 10Y UST"),
        ("Tax Rate",                  TAX_RATE,  FMT_PCT1,
         "Source: Normalised tax rate; FY2022-2025 avg ~25.4% ex-FY2021 distortion"),
        ("After-Tax Cost of Debt",    None,      FMT_PCT1, "= Kd × (1 − Tax Rate)"),
    ]
    for i, (label, val, fmt, cmt) in enumerate(rows_kd):
        rr = kd_start + i
        write_val(ws, rr, 1, label, align="left",
                  fill_color=C_INPUT_FILL if i < 2 else None)
        if i < 2:
            write_val(ws, rr, 2, val, fmt=fmt, fg=C_FONT_BLUE,
                      fill_color=C_INPUT_FILL, comment_text=cmt)
        else:
            write_formula(ws, rr, 2, f"=B{kd_start}*(1-B{kd_start+1})",
                          fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)
        write_val(ws, rr, 3, cmt, align="left", fg="595959")
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    r = kd_start + 2

    # ── Capital Structure ─────────────────────────────────────────────────────
    r += 2
    write_header(ws, r, 1, "CAPITAL STRUCTURE", span_end_col=4)
    r += 1
    write_val(ws, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws, r, 2, "Value ($M)", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    cs_start = r + 1
    cs_rows = [
        ("Current Stock Price",        STOCK_PRICE, FMT_DOLLAR2, "Source: StockAnalysis.com, NYSE close 2026-08-17"),
        ("Diluted Shares Outstanding (M)", SHARES_DILUT, "#,##0.0", "Source: StockAnalysis.com TTM diluted shares, Aug 2026"),
        ("Market Capitalisation ($M)", None,        FMT_DOLLAR, "= Price × Shares"),
        ("Total Debt ($M)",            TOTAL_DEBT,  FMT_DOLLAR, "Source: StockAnalysis.com balance sheet FY2025"),
        ("Cash & Equivalents ($M)",    CASH,        FMT_DOLLAR, "Source: StockAnalysis.com balance sheet FY2025"),
        ("Net Debt ($M)",              None,        FMT_DOLLAR, "= Total Debt − Cash"),
        ("Enterprise Value ($M)",      None,        FMT_DOLLAR, "= Market Cap + Net Debt"),
    ]
    for i, (label, val, fmt, cmt) in enumerate(cs_rows):
        rr = cs_start + i
        write_val(ws, rr, 1, label, align="left",
                  fill_color=C_INPUT_FILL if val is not None else None)
        if val is not None:
            write_val(ws, rr, 2, val, fmt=fmt, fg=C_FONT_BLUE,
                      fill_color=C_INPUT_FILL, comment_text=cmt)
        elif i == 2:  # Mkt Cap
            write_formula(ws, rr, 2, f"=B{cs_start}*B{cs_start+1}", fmt=FMT_DOLLAR)
        elif i == 5:  # Net Debt
            write_formula(ws, rr, 2, f"=B{cs_start+3}-B{cs_start+4}", fmt=FMT_DOLLAR)
        else:         # EV
            write_formula(ws, rr, 2, f"=B{cs_start+2}+B{cs_start+5}",
                          fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)
        write_val(ws, rr, 3, cmt, align="left", fg="595959")
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    r = cs_start + 6

    # ── WACC Calculation ──────────────────────────────────────────────────────
    r += 2
    write_header(ws, r, 1, "WACC CALCULATION", span_end_col=4)
    r += 1
    for hdr, col in [("Component", 1), ("Weight", 2), ("Cost", 3), ("Contribution", 4)]:
        write_val(ws, r, col, hdr, bold=True, fill_color=C_HDR_LIGHT,
                  align="left" if col == 1 else "right")

    # Row references for weights
    ev_row   = cs_start + 6   # EV row
    mkt_row  = cs_start + 2   # Mkt Cap row
    nd_row   = cs_start + 5   # Net Debt row
    ke_row   = ke_start + 3   # Ke result row
    kdat_row = kd_start + 2   # Kd after-tax result row

    wc_start = r + 1
    # Equity row
    rr = wc_start
    write_val(ws, rr, 1, "Equity", align="left")
    write_formula(ws, rr, 2, f"=B{mkt_row}/B{ev_row}", fmt=FMT_PCT1)
    write_formula(ws, rr, 3, f"=B{ke_row}", fmt=FMT_PCT1, green=True)
    write_formula(ws, rr, 4, f"=B{rr}*C{rr}", fmt=FMT_PCT1)
    # Debt row
    rr = wc_start + 1
    write_val(ws, rr, 1, "Net Debt (after-tax)", align="left")
    write_formula(ws, rr, 2, f"=B{nd_row}/B{ev_row}", fmt=FMT_PCT1)
    write_formula(ws, rr, 3, f"=B{kdat_row}", fmt=FMT_PCT1, green=True)
    write_formula(ws, rr, 4, f"=B{rr}*C{rr}", fmt=FMT_PCT1)
    # WACC total
    rr = wc_start + 3
    write_val(ws, rr, 1, "WACC", bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws, rr, 4, f"=D{wc_start}+D{wc_start+1}",
                  fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    c = ws.cell(row=rr, column=1)
    c.font = font(C_FONT_BLACK, bold=True)
    c.fill = fill(C_OUTPUT)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Store WACC row for cross-sheet reference
    ws["A1"].comment = Comment(f"WACC result cell: D{rr}", "DCF Model")

    print("✓ Part 3: WACC sheet built")
    return ws, wc_start + 3   # return ws and WACC result row

# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET — ROW LAYOUT (define ALL rows FIRST before writing any formula)
# ═══════════════════════════════════════════════════════════════════════════════
# Section 1: Header
R_TITLE          = 1
R_SUBTITLE       = 2
R_BLANK1         = 3
R_CASE_LABEL     = 4
R_CASE_VAL       = 5
R_BLANK2         = 6

# Section 2: Market Data
R_MKT_HDR        = 7
R_MKT_COL        = 8
R_PRICE          = 9
R_SHARES         = 10
R_MKTCAP         = 11
R_TOTALDEBT      = 12
R_CASH           = 13
R_NETDEBT        = 14
R_BLANK3         = 15

# Section 3: Scenario Assumption Blocks
# Bear block
R_BEAR_HDR       = 16
R_BEAR_COL       = 17
R_BEAR_GRW       = 18   # revenue growth row (FY26..FY30)
R_BEAR_EBIT_MRG  = 19
R_BEAR_TERM_G    = 20
R_BLANK4         = 21

# Base block
R_BASE_HDR       = 22
R_BASE_COL       = 23
R_BASE_GRW       = 24
R_BASE_EBIT_MRG  = 25
R_BASE_TERM_G    = 26
R_BLANK5         = 27

# Bull block
R_BULL_HDR       = 28
R_BULL_COL       = 29
R_BULL_GRW       = 30
R_BULL_EBIT_MRG  = 31
R_BULL_TERM_G    = 32
R_BLANK6         = 33

# Shared FCF parameters
R_SHARED_HDR     = 34
R_SHARED_COL     = 35
R_TAX            = 36
R_DA_PCT         = 37
R_CAPEX_PCT      = 38
R_NWC_PCT        = 39
R_WACC_REF       = 40
R_BLANK7         = 41

# Consolidation column (col H) — INDEX formulas picking from scenario blocks
R_SEL_HDR        = 42
R_SEL_COL        = 43
R_SEL_GRW        = 44   # per-year growth (one row per proj year: 44-48)
R_SEL_EBIT_MRG   = 49   # per-year EBIT margin (49-53)
R_SEL_TERM_G     = 54
R_BLANK8         = 55

# Section 4: Historical & Projected Financials
R_IS_HDR         = 56
R_IS_COL         = 57
R_IS_REV         = 58
R_IS_REV_GRW     = 59
R_BLANK9         = 60
R_IS_GP          = 61
R_IS_GP_MRG      = 62
R_BLANK10        = 63
R_IS_EBIT        = 64
R_IS_EBIT_MRG    = 65
R_IS_TAX         = 66
R_IS_NOPAT       = 67
R_BLANK11        = 68

# Section 5: FCF Build
R_FCF_HDR        = 69
R_FCF_COL        = 70
R_FCF_NOPAT      = 71
R_FCF_DA         = 72
R_FCF_DA_PCT     = 73
R_FCF_CAPEX      = 74
R_FCF_CAPEX_PCT  = 75
R_FCF_NWC        = 76
R_FCF_NWC_PCT    = 77
R_BLANK12        = 78
R_FCF_UFCF       = 79
R_FCF_UFCF_MRG   = 80
R_BLANK13        = 81

# Section 6: Discounting & Valuation
R_DCF_HDR        = 82
R_DCF_COL        = 83
R_DCF_UFCF       = 84
R_DCF_PERIOD     = 85
R_DCF_DISC       = 86
R_DCF_PV         = 87
R_BLANK14        = 88
R_TV_FCF         = 89
R_TV_VAL         = 90
R_TV_PV          = 91
R_BLANK15        = 92
R_SUM_HDR        = 93
R_SUM_PV_FCF     = 94
R_SUM_PV_TV      = 95
R_SUM_EV         = 96
R_SUM_NETDEBT    = 97
R_SUM_EQ         = 98
R_BLANK16        = 99
R_SUM_SHARES     = 100
R_SUM_PRICE      = 101
R_SUM_CUR_PRICE  = 102
R_SUM_UPSIDE     = 103
R_BLANK17        = 104

# Sensitivity tables start row
R_SENS1_HDR      = 105

# ── Column layout ─────────────────────────────────────────────────────────────
# Col A: labels
# Col B: FY2021A  Col C: FY2022A  Col D: FY2023A  Col E: FY2024A  Col F: FY2025A
# Col G: FY2026E  Col H: FY2027E  Col I: FY2028E  Col J: FY2029E  Col K: FY2030E
# Col L: Terminal / Consolidation / Notes
# Scenario assumption blocks also use B-F for Bear/Base/Bull years

# Assumption blocks column mapping (Bear/Base/Bull all use same col layout):
# Col B=FY2026E, C=FY2027E, D=FY2028E, E=FY2029E, F=FY2030E
# Consolidation column = Col H (index formula by case selector)

COL_LABEL   = 1   # A
COL_Y1A     = 2   # B  (FY2021A in IS; FY2026E in assumption blocks)
COL_Y2A     = 3   # C
COL_Y3A     = 4   # D
COL_Y4A     = 5   # E
COL_Y5A     = 6   # F  (FY2025A in IS)
COL_Y1P     = 7   # G  (FY2026E in IS/FCF)
COL_Y2P     = 8   # H
COL_Y3P     = 9   # I
COL_Y4P     = 10  # J
COL_Y5P     = 11  # K  (FY2030E)
COL_TERM    = 12  # L  (Terminal)

# Assumption blocks use cols B-F for proj years 1-5
ASSUMP_COLS = [COL_Y1A, COL_Y2A, COL_Y3A, COL_Y4A, COL_Y5A]  # B,C,D,E,F
PROJ_COLS   = [COL_Y1P, COL_Y2P, COL_Y3P, COL_Y4P, COL_Y5P]  # G,H,I,J,K
HIST_COLS   = [COL_Y1A, COL_Y2A, COL_Y3A, COL_Y4A, COL_Y5A]  # B-F (FY2021-2025)

# Consolidation rows: one row per proj year within the SEL block
# R_SEL_GRW   = 44 (FY2026), 45(FY2027), 46(FY2028), 47(FY2029), 48(FY2030)
# R_SEL_EBIT_MRG = 49..53
SEL_GRW_ROWS  = list(range(R_SEL_GRW,  R_SEL_GRW + 5))   # 44-48
SEL_MRG_ROWS  = list(range(R_SEL_EBIT_MRG, R_SEL_EBIT_MRG + 5))  # 49-53

# Case selector cell
CASE_CELL = "B5"   # 1=Bear, 2=Base, 3=Bull
CASE_COL  = 2      # column B

print("✓ Part 4: row layout constants defined")

# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET BUILDER — Section A: Header + Market Data + Case Selector
# ═══════════════════════════════════════════════════════════════════════════════
def build_dcf_header(ws):
    """Rows 1-15: title, case selector, market data."""
    set_col_widths(ws, {
        "A": 36, "B": 12, "C": 12, "D": 12, "E": 12,
        "F": 12, "G": 12, "H": 12, "I": 12, "J": 12,
        "K": 12, "L": 14,
    })

    # Row 1: Title
    ws.cell(row=R_TITLE, column=1).value = "Ingredion Incorporated (INGR) — DCF Valuation Model"
    ws.cell(row=R_TITLE, column=1).font = font(C_FONT_WHITE, bold=True, size=13)
    ws.cell(row=R_TITLE, column=1).fill = fill(C_HDR_DARK)
    ws.cell(row=R_TITLE, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=R_TITLE, start_column=1, end_row=R_TITLE, end_column=COL_TERM)
    ws.row_dimensions[R_TITLE].height = 24

    # Row 2: Subtitle
    ws.cell(row=R_SUBTITLE, column=1).value = (
        "Ticker: INGR  |  Date: 2026-08-18  |  Fiscal Year End: December  "
        "|  Currency: USD ($M unless noted)"
    )
    ws.cell(row=R_SUBTITLE, column=1).font = font("595959", size=9)
    ws.cell(row=R_SUBTITLE, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=R_SUBTITLE, start_column=1, end_row=R_SUBTITLE, end_column=COL_TERM)

    # Row 4-5: Case Selector
    write_val(ws, R_CASE_LABEL, COL_LABEL,
              "▶  CASE SELECTOR  (1 = Bear  |  2 = Base  |  3 = Bull)",
              bold=True, fg=C_FONT_WHITE, fill_color=C_HDR_DARK, align="left")
    ws.merge_cells(start_row=R_CASE_LABEL, start_column=1,
                   end_row=R_CASE_LABEL, end_column=COL_TERM)
    c = ws.cell(row=R_CASE_LABEL, column=1)
    c.font = font(C_FONT_WHITE, bold=True); c.fill = fill(C_HDR_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    write_val(ws, R_CASE_VAL, COL_LABEL, "Selected Case:", bold=True, align="left")
    write_val(ws, R_CASE_VAL, CASE_COL, 2, fmt="0",
              fg=C_FONT_BLUE, fill_color=C_INPUT_FILL, bold=True,
              comment_text="Source: User input. 1=Bear, 2=Base, 3=Bull")
    write_formula(ws, R_CASE_VAL, COL_Y2A,
                  f'=IF({CASE_CELL}=1,"Bear",IF({CASE_CELL}=2,"Base","Bull"))',
                  bold=True, fill_color=C_OUTPUT)
    ws.merge_cells(start_row=R_CASE_VAL, start_column=COL_Y2A,
                   end_row=R_CASE_VAL, end_column=COL_Y4A)
    c = ws.cell(row=R_CASE_VAL, column=COL_Y2A)
    c.font = font(C_FONT_BLACK, bold=True); c.fill = fill(C_OUTPUT)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Row 7-14: Market Data
    write_header(ws, R_MKT_HDR, 1, "MARKET DATA & KEY INPUTS", span_end_col=COL_TERM)

    write_val(ws, R_MKT_COL, COL_LABEL, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws, R_MKT_COL, COL_Y1A, "Value", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws, R_MKT_COL, COL_Y2A, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws.merge_cells(start_row=R_MKT_COL, start_column=COL_Y2A,
                   end_row=R_MKT_COL, end_column=COL_TERM)
    c = ws.cell(row=R_MKT_COL, column=COL_Y2A)
    c.font = font(C_FONT_BLACK, bold=True); c.fill = fill(C_HDR_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center")

    mkt_data = [
        (R_PRICE,    "Current Stock Price",             STOCK_PRICE, FMT_DOLLAR2,
         "Source: StockAnalysis.com, NYSE close 2026-08-17"),
        (R_SHARES,   "Diluted Shares Outstanding (M)",  SHARES_DILUT, "#,##0.0",
         "Source: StockAnalysis.com TTM diluted shares, Aug 2026"),
        (R_MKTCAP,   "Market Capitalisation ($M)",       None,       FMT_DOLLAR,
         "= Price × Diluted Shares"),
        (R_TOTALDEBT,"Total Debt ($M)",                  TOTAL_DEBT, FMT_DOLLAR,
         "Source: StockAnalysis.com balance sheet FY2025A"),
        (R_CASH,     "Cash & Equivalents ($M)",          CASH,       FMT_DOLLAR,
         "Source: StockAnalysis.com balance sheet FY2025A"),
        (R_NETDEBT,  "Net Debt ($M)",                    None,       FMT_DOLLAR,
         "= Total Debt − Cash"),
    ]
    for row, label, val, fmt, note in mkt_data:
        write_val(ws, row, COL_LABEL, label, align="left",
                  fill_color=C_INPUT_FILL if val is not None else None)
        if val is not None:
            write_val(ws, row, COL_Y1A, val, fmt=fmt, fg=C_FONT_BLUE,
                      fill_color=C_INPUT_FILL, comment_text=note)
        elif row == R_MKTCAP:
            write_formula(ws, row, COL_Y1A,
                          f"=B{R_PRICE}*B{R_SHARES}", fmt=FMT_DOLLAR)
        else:
            write_formula(ws, row, COL_Y1A,
                          f"=B{R_TOTALDEBT}-B{R_CASH}",
                          fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)
        note_c = ws.cell(row=row, column=COL_Y2A)
        note_c.value = note
        note_c.font = font("595959", size=9)
        note_c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=COL_Y2A,
                       end_row=row, end_column=COL_TERM)

    print("✓ Part 5a: header + market data written")


# ─────────────────────────────────────────────────────────────────────────────
def build_scenario_assumptions(ws):
    """Rows 16-54: Bear/Base/Bull assumption blocks + shared params + consolidation."""

    # Helper: write one assumption block
    def write_assump_block(hdr_row, col_row, grw_row, mrg_row, term_row,
                            scenario_name, grw_vals, mrg_vals, term_val, fill_hex):
        # Header
        write_header(ws, hdr_row, 1,
                     f"{scenario_name.upper()} CASE ASSUMPTIONS", span_end_col=COL_TERM)

        # Column header row
        write_val(ws, col_row, COL_LABEL, "Assumption",
                  bold=True, fill_color=C_HDR_LIGHT, align="left")
        for i, yr in enumerate(PROJ_YEARS):
            write_val(ws, col_row, ASSUMP_COLS[i], yr,
                      bold=True, fill_color=C_HDR_LIGHT)
        for c in range(ASSUMP_COLS[-1]+1, COL_TERM+1):
            ws.cell(row=col_row, column=c).fill = fill(C_HDR_LIGHT)

        # Revenue growth row
        write_val(ws, grw_row, COL_LABEL, "Revenue Growth (%)", align="left",
                  fill_color=C_INPUT_FILL)
        for i, v in enumerate(grw_vals):
            write_val(ws, grw_row, ASSUMP_COLS[i], v, fmt=FMT_PCT1,
                      fg=C_FONT_BLUE, fill_color=C_INPUT_FILL,
                      comment_text=f"Source: DCF model assumption, {scenario_name} case FY{2026+i}")

        # EBIT margin row
        write_val(ws, mrg_row, COL_LABEL, "EBIT Margin (%)", align="left",
                  fill_color=C_INPUT_FILL)
        for i, v in enumerate(mrg_vals):
            write_val(ws, mrg_row, ASSUMP_COLS[i], v, fmt=FMT_PCT1,
                      fg=C_FONT_BLUE, fill_color=C_INPUT_FILL,
                      comment_text=f"Source: DCF model assumption, {scenario_name} case FY{2026+i}")

        # Terminal growth (single value in col B)
        write_val(ws, term_row, COL_LABEL, "Terminal Growth Rate (%)", align="left",
                  fill_color=C_INPUT_FILL)
        write_val(ws, term_row, ASSUMP_COLS[0], term_val, fmt=FMT_PCT1,
                  fg=C_FONT_BLUE, fill_color=C_INPUT_FILL,
                  comment_text=f"Source: DCF model assumption, {scenario_name} terminal g")

    write_assump_block(R_BEAR_HDR, R_BEAR_COL, R_BEAR_GRW, R_BEAR_EBIT_MRG, R_BEAR_TERM_G,
                       "Bear", BEAR_GRW, BEAR_EBIT_MRG, TERM_G["Bear"], "FFF2CC")
    write_assump_block(R_BASE_HDR, R_BASE_COL, R_BASE_GRW, R_BASE_EBIT_MRG, R_BASE_TERM_G,
                       "Base", BASE_GRW, BASE_EBIT_MRG, TERM_G["Base"], "E2EFDA")
    write_assump_block(R_BULL_HDR, R_BULL_COL, R_BULL_GRW, R_BULL_EBIT_MRG, R_BULL_TERM_G,
                       "Bull", BULL_GRW, BULL_EBIT_MRG, TERM_G["Bull"], "DDEBF7")

    # ── Shared FCF parameters ─────────────────────────────────────────────────
    write_header(ws, R_SHARED_HDR, 1, "SHARED FCF PARAMETERS (ALL SCENARIOS)", span_end_col=COL_TERM)
    write_val(ws, R_SHARED_COL, COL_LABEL, "Assumption", bold=True,
              fill_color=C_HDR_LIGHT, align="left")
    for i, yr in enumerate(PROJ_YEARS):
        write_val(ws, R_SHARED_COL, ASSUMP_COLS[i], yr, bold=True, fill_color=C_HDR_LIGHT)

    write_val(ws, R_TAX, COL_LABEL, "Tax Rate (%)", align="left", fill_color=C_INPUT_FILL)
    write_val(ws, R_TAX, ASSUMP_COLS[0], TAX_RATE, fmt=FMT_PCT1, fg=C_FONT_BLUE,
              fill_color=C_INPUT_FILL,
              comment_text="Source: Normalised tax rate; FY2022-2025 avg ~25.4%")

    write_val(ws, R_DA_PCT, COL_LABEL, "D&A (% of Revenue)", align="left",
              fill_color=C_INPUT_FILL)
    write_val(ws, R_DA_PCT, ASSUMP_COLS[0], DA_PCT, fmt=FMT_PCT1, fg=C_FONT_BLUE,
              fill_color=C_INPUT_FILL,
              comment_text="Source: FY2021-2025 D&A avg ~2.9% of revenue; rounded to 3.0%")

    write_val(ws, R_CAPEX_PCT, COL_LABEL, "CapEx (% of Revenue)", align="left",
              fill_color=C_INPUT_FILL)
    for i, v in enumerate(CAPEX_PCT):
        write_val(ws, R_CAPEX_PCT, ASSUMP_COLS[i], v, fmt=FMT_PCT1, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: DCF assumption FY{2026+i}; FY2025 peak 6.0%, normalising to 4.5%")

    write_val(ws, R_NWC_PCT, COL_LABEL, "ΔNwc (% of ΔRevenue)", align="left",
              fill_color=C_INPUT_FILL)
    write_val(ws, R_NWC_PCT, ASSUMP_COLS[0], NWC_PCT, fmt=FMT_PCT1, fg=C_FONT_BLUE,
              fill_color=C_INPUT_FILL,
              comment_text="Source: NWC ~40% of revenue; incremental NWC need ~8% of revenue change")

    write_val(ws, R_WACC_REF, COL_LABEL, "WACC", align="left")
    write_formula(ws, R_WACC_REF, ASSUMP_COLS[0], "=WACC!D24",
                  fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT, green=True)

    # ── Consolidation column (col B, rows 44-53) ──────────────────────────────
    write_header(ws, R_SEL_HDR, 1,
                 "SELECTED CASE — CONSOLIDATION (INDEX formula → used by projections)",
                 span_end_col=COL_TERM)
    write_val(ws, R_SEL_COL, COL_LABEL, "Assumption",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for i, yr in enumerate(PROJ_YEARS):
        write_val(ws, R_SEL_COL, ASSUMP_COLS[i], yr, bold=True, fill_color=C_HDR_LIGHT)

    # Revenue growth consolidation: one row per year
    for i in range(N_PROJ):
        rr = SEL_GRW_ROWS[i]
        label = f"Revenue Growth — {PROJ_YEARS[i]}"
        write_val(ws, rr, COL_LABEL, label, align="left")
        bear_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BEAR_GRW}"
        base_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BASE_GRW}"
        bull_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BULL_GRW}"
        formula = (f"=INDEX({bear_cell}:{bull_cell},1,{CASE_CELL})"
                   if False else
                   f"=IF({CASE_CELL}=1,{bear_cell},IF({CASE_CELL}=2,{base_cell},{bull_cell}))")
        write_formula(ws, rr, ASSUMP_COLS[0], formula, fmt=FMT_PCT1, fill_color=C_OUTPUT)

    # EBIT margin consolidation: one row per year
    for i in range(N_PROJ):
        rr = SEL_MRG_ROWS[i]
        label = f"EBIT Margin — {PROJ_YEARS[i]}"
        write_val(ws, rr, COL_LABEL, label, align="left")
        bear_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BEAR_EBIT_MRG}"
        base_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BASE_EBIT_MRG}"
        bull_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_BULL_EBIT_MRG}"
        formula = (
            f"=IF({CASE_CELL}=1,{bear_cell},IF({CASE_CELL}=2,{base_cell},{bull_cell}))")
        write_formula(ws, rr, ASSUMP_COLS[0], formula, fmt=FMT_PCT1, fill_color=C_OUTPUT)

    # Terminal growth consolidation
    write_val(ws, R_SEL_TERM_G, COL_LABEL, "Terminal Growth Rate", align="left")
    write_formula(ws, R_SEL_TERM_G, ASSUMP_COLS[0],
                  f"=IF({CASE_CELL}=1,B{R_BEAR_TERM_G},IF({CASE_CELL}=2,B{R_BASE_TERM_G},B{R_BULL_TERM_G}))",
                  fmt=FMT_PCT1, fill_color=C_OUTPUT)

    print("✓ Part 5b: scenario assumption blocks written")

# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET — Section B: Historical + Projected Income Statement
# ═══════════════════════════════════════════════════════════════════════════════
def build_income_statement(ws):
    """Rows 56-67: IS header, historical actuals, projected financials."""
    ALL_YEARS = HIST_YEARS + PROJ_YEARS
    ALL_COLS  = HIST_COLS  + PROJ_COLS   # B-F historical, G-K projected

    # Header
    write_header(ws, R_IS_HDR, 1,
                 "HISTORICAL & PROJECTED INCOME STATEMENT ($M)", span_end_col=COL_TERM)

    # Column labels
    write_val(ws, R_IS_COL, COL_LABEL, "Income Statement ($M)",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for i, yr in enumerate(ALL_YEARS):
        write_val(ws, R_IS_COL, ALL_COLS[i], yr, bold=True, fill_color=C_HDR_LIGHT)

    # ── Revenue ───────────────────────────────────────────────────────────────
    write_val(ws, R_IS_REV, COL_LABEL, "Revenue", bold=True, align="left")
    # Historical actuals (blue)
    for i, v in enumerate(HIST_REV):
        write_val(ws, R_IS_REV, HIST_COLS[i], v, fmt=FMT_DOLLAR, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: StockAnalysis.com / S&P Global, {HIST_YEARS[i]} annual revenue")

    # Projected: Rev_n = Rev_{n-1} × (1 + growth_consolidation_row)
    # PROJ_COLS[0]=G (FY2026E), prior year for FY2026 is HIST_COLS[4]=F (FY2025A)
    prev_col = get_column_letter(HIST_COLS[4])  # F = FY2025A
    for i in range(N_PROJ):
        pc  = get_column_letter(PROJ_COLS[i])
        sel_grw_cell = f"B{SEL_GRW_ROWS[i]}"   # consolidation col (col B) for year i growth
        if i == 0:
            formula = f"={prev_col}{R_IS_REV}*(1+{sel_grw_cell})"
        else:
            pp = get_column_letter(PROJ_COLS[i-1])
            formula = f"={pp}{R_IS_REV}*(1+{sel_grw_cell})"
        write_formula(ws, R_IS_REV, PROJ_COLS[i], formula, fmt=FMT_DOLLAR, bold=True)

    # Revenue growth %
    write_val(ws, R_IS_REV_GRW, COL_LABEL, "  % Growth", align="left", fg="595959")
    for i in range(1, 5):   # FY2022-2025 historical growth
        hc = get_column_letter(HIST_COLS[i])
        hp = get_column_letter(HIST_COLS[i-1])
        c = ws.cell(row=R_IS_REV_GRW, column=HIST_COLS[i],
                    value=f"={hc}{R_IS_REV}/{hp}{R_IS_REV}-1")
        c.number_format = FMT_PCT1
        c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        if i == 0:
            pp = get_column_letter(HIST_COLS[4])
        else:
            pp = get_column_letter(PROJ_COLS[i-1])
        c = ws.cell(row=R_IS_REV_GRW, column=PROJ_COLS[i],
                    value=f"={pc}{R_IS_REV}/{pp}{R_IS_REV}-1")
        c.number_format = FMT_PCT1
        c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Gross Profit ──────────────────────────────────────────────────────────
    write_val(ws, R_IS_GP, COL_LABEL, "Gross Profit", align="left")
    for i, v in enumerate(HIST_GP):
        write_val(ws, R_IS_GP, HIST_COLS[i], v, fmt=FMT_DOLLAR, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: StockAnalysis.com / S&P Global, {HIST_YEARS[i]}")
    # Projected GP — use historical gross margin trend; approach: derived from EBIT+opex
    # Simplify: not separately modelled — EBIT margin is the driver; leave GP projected blank
    # (We do not need a separate GP line for FCF; EBIT is what matters)

    # GP margin (historical only)
    write_val(ws, R_IS_GP_MRG, COL_LABEL, "  % Margin", align="left", fg="595959")
    for i in range(5):
        hc = get_column_letter(HIST_COLS[i])
        c = ws.cell(row=R_IS_GP_MRG, column=HIST_COLS[i],
                    value=f"={hc}{R_IS_GP}/{hc}{R_IS_REV}")
        c.number_format = FMT_PCT1
        c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── EBIT ─────────────────────────────────────────────────────────────────
    write_val(ws, R_IS_EBIT, COL_LABEL, "EBIT (Operating Income)", bold=True, align="left")
    for i, v in enumerate(HIST_EBIT):
        write_val(ws, R_IS_EBIT, HIST_COLS[i], v, fmt=FMT_DOLLAR, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: StockAnalysis.com / S&P Global, {HIST_YEARS[i]}")
    for i in range(N_PROJ):
        rc = get_column_letter(PROJ_COLS[i])
        sel_mrg_cell = f"B{SEL_MRG_ROWS[i]}"
        formula = f"={rc}{R_IS_REV}*{sel_mrg_cell}"
        write_formula(ws, R_IS_EBIT, PROJ_COLS[i], formula, fmt=FMT_DOLLAR, bold=True)

    # EBIT margin
    write_val(ws, R_IS_EBIT_MRG, COL_LABEL, "  % Margin", align="left", fg="595959")
    for i in range(5):
        hc = get_column_letter(HIST_COLS[i])
        c = ws.cell(row=R_IS_EBIT_MRG, column=HIST_COLS[i],
                    value=f"={hc}{R_IS_EBIT}/{hc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        c = ws.cell(row=R_IS_EBIT_MRG, column=PROJ_COLS[i],
                    value=f"={pc}{R_IS_EBIT}/{pc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Taxes ─────────────────────────────────────────────────────────────────
    write_val(ws, R_IS_TAX, COL_LABEL, "(-) Income Taxes", align="left")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_IS_TAX, PROJ_COLS[i],
                      f"=-{pc}{R_IS_EBIT}*B{R_TAX}", fmt=FMT_DOLLAR)

    # ── NOPAT ─────────────────────────────────────────────────────────────────
    write_val(ws, R_IS_NOPAT, COL_LABEL, "NOPAT", bold=True, align="left",
              fill_color=C_OUTPUT)
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_IS_NOPAT, PROJ_COLS[i],
                      f"={pc}{R_IS_EBIT}+{pc}{R_IS_TAX}",
                      fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)

    print("✓ Part 6a: income statement written")


# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET — Section C: Free Cash Flow Build
# ═══════════════════════════════════════════════════════════════════════════════
def build_fcf(ws):
    """Rows 69-80: FCF schedule."""
    ALL_YEARS = HIST_YEARS + PROJ_YEARS
    ALL_COLS  = HIST_COLS  + PROJ_COLS

    write_header(ws, R_FCF_HDR, 1,
                 "FREE CASH FLOW BUILD ($M)", span_end_col=COL_TERM)

    # Column headers
    write_val(ws, R_FCF_COL, COL_LABEL, "Cash Flow ($M)",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for i, yr in enumerate(ALL_YEARS):
        write_val(ws, R_FCF_COL, ALL_COLS[i], yr, bold=True, fill_color=C_HDR_LIGHT)

    # ── NOPAT (link from IS) ──────────────────────────────────────────────────
    write_val(ws, R_FCF_NOPAT, COL_LABEL, "NOPAT", bold=True, align="left")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_FCF_NOPAT, PROJ_COLS[i],
                      f"={pc}{R_IS_NOPAT}", fmt=FMT_DOLLAR)

    # ── D&A ───────────────────────────────────────────────────────────────────
    write_val(ws, R_FCF_DA, COL_LABEL, "(+) D&A", align="left")
    # Historical actuals
    for i, v in enumerate(HIST_DA):
        write_val(ws, R_FCF_DA, HIST_COLS[i], v, fmt=FMT_DOLLAR, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: StockAnalysis.com cash flow statement, {HIST_YEARS[i]}")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_FCF_DA, PROJ_COLS[i],
                      f"={pc}{R_IS_REV}*B{R_DA_PCT}", fmt=FMT_DOLLAR)

    write_val(ws, R_FCF_DA_PCT, COL_LABEL, "    % of Revenue", align="left", fg="595959")
    for i in range(5):
        hc = get_column_letter(HIST_COLS[i])
        c = ws.cell(row=R_FCF_DA_PCT, column=HIST_COLS[i],
                    value=f"={hc}{R_FCF_DA}/{hc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        c = ws.cell(row=R_FCF_DA_PCT, column=PROJ_COLS[i],
                    value=f"={pc}{R_FCF_DA}/{pc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── CapEx ─────────────────────────────────────────────────────────────────
    write_val(ws, R_FCF_CAPEX, COL_LABEL, "(-) Capital Expenditures", align="left")
    for i, v in enumerate(HIST_CAPEX):
        write_val(ws, R_FCF_CAPEX, HIST_COLS[i], -v, fmt=FMT_DOLLAR, fg=C_FONT_BLUE,
                  fill_color=C_INPUT_FILL,
                  comment_text=f"Source: StockAnalysis.com cash flow statement, {HIST_YEARS[i]}")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        capex_cell = f"{get_column_letter(ASSUMP_COLS[i])}{R_CAPEX_PCT}"
        write_formula(ws, R_FCF_CAPEX, PROJ_COLS[i],
                      f"=-{pc}{R_IS_REV}*{capex_cell}", fmt=FMT_DOLLAR)

    write_val(ws, R_FCF_CAPEX_PCT, COL_LABEL, "    % of Revenue", align="left", fg="595959")
    for i in range(5):
        hc = get_column_letter(HIST_COLS[i])
        c = ws.cell(row=R_FCF_CAPEX_PCT, column=HIST_COLS[i],
                    value=f"=ABS({hc}{R_FCF_CAPEX})/{hc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        c = ws.cell(row=R_FCF_CAPEX_PCT, column=PROJ_COLS[i],
                    value=f"=ABS({pc}{R_FCF_CAPEX})/{pc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── ΔNwc ──────────────────────────────────────────────────────────────────
    write_val(ws, R_FCF_NWC, COL_LABEL, "(-) Change in NWC (ΔNwc)", align="left")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        if i == 0:
            prev = get_column_letter(HIST_COLS[4])
        else:
            prev = get_column_letter(PROJ_COLS[i-1])
        write_formula(ws, R_FCF_NWC, PROJ_COLS[i],
                      f"=-({pc}{R_IS_REV}-{prev}{R_IS_REV})*B{R_NWC_PCT}",
                      fmt=FMT_DOLLAR)

    write_val(ws, R_FCF_NWC_PCT, COL_LABEL, "    % of ΔRevenue", align="left", fg="595959")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        if i == 0:
            prev = get_column_letter(HIST_COLS[4])
        else:
            prev = get_column_letter(PROJ_COLS[i-1])
        rev_delta = f"({pc}{R_IS_REV}-{prev}{R_IS_REV})"
        c = ws.cell(row=R_FCF_NWC_PCT, column=PROJ_COLS[i],
                    value=f"=IF({rev_delta}=0,0,-{pc}{R_FCF_NWC}/{rev_delta})")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Unlevered FCF ─────────────────────────────────────────────────────────
    write_val(ws, R_FCF_UFCF, COL_LABEL, "Unlevered Free Cash Flow",
              bold=True, align="left", fill_color=C_OUTPUT)
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_FCF_UFCF, PROJ_COLS[i],
                      f"={pc}{R_FCF_NOPAT}+{pc}{R_FCF_DA}+{pc}{R_FCF_CAPEX}+{pc}{R_FCF_NWC}",
                      fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)

    # FCF margin
    write_val(ws, R_FCF_UFCF_MRG, COL_LABEL, "  FCF Margin", align="left", fg="595959")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        c = ws.cell(row=R_FCF_UFCF_MRG, column=PROJ_COLS[i],
                    value=f"={pc}{R_FCF_UFCF}/{pc}{R_IS_REV}")
        c.number_format = FMT_PCT1; c.font = font("595959")
        c.alignment = Alignment(horizontal="right", vertical="center")

    print("✓ Part 6b: FCF build written")

# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET — Section D: Discounting + Terminal Value + Valuation Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_valuation(ws):
    """Rows 82-103: discount factors, terminal value, equity bridge."""

    write_header(ws, R_DCF_HDR, 1,
                 "DCF DISCOUNTING & TERMINAL VALUE", span_end_col=COL_TERM)

    # Column headers
    write_val(ws, R_DCF_COL, COL_LABEL, "Item",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for i, yr in enumerate(PROJ_YEARS):
        write_val(ws, R_DCF_COL, PROJ_COLS[i], yr, bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws, R_DCF_COL, COL_TERM, "Terminal", bold=True, fill_color=C_HDR_LIGHT)

    # ── Unlevered FCF (linked) ─────────────────────────────────────────────────
    write_val(ws, R_DCF_UFCF, COL_LABEL, "Unlevered FCF ($M)", align="left")
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_DCF_UFCF, PROJ_COLS[i],
                      f"={pc}{R_FCF_UFCF}", fmt=FMT_DOLLAR)

    # ── Discount periods (mid-year) ────────────────────────────────────────────
    write_val(ws, R_DCF_PERIOD, COL_LABEL, "Discount Period (mid-year)", align="left")
    for i, p in enumerate(DISC_PERIODS):
        write_val(ws, R_DCF_PERIOD, PROJ_COLS[i], p, fmt="0.0",
                  fg=C_FONT_BLUE, fill_color=C_INPUT_FILL,
                  comment_text="Source: Mid-year convention; period = year - 0.5")
    # Terminal value discount period = 4.5
    write_val(ws, R_DCF_PERIOD, COL_TERM, 4.5, fmt="0.0",
              fg=C_FONT_BLUE, fill_color=C_INPUT_FILL,
              comment_text="Source: Terminal value discounted at same period as final FCF year")

    # ── Discount factors ──────────────────────────────────────────────────────
    write_val(ws, R_DCF_DISC, COL_LABEL, "Discount Factor  [1/(1+WACC)^t]", align="left")
    wacc_cell = f"B{R_WACC_REF}"
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_DCF_DISC, PROJ_COLS[i],
                      f"=1/(1+{wacc_cell})^{pc}{R_DCF_PERIOD}", fmt="0.0000")
    # Terminal discount factor
    tc = get_column_letter(COL_TERM)
    write_formula(ws, R_DCF_DISC, COL_TERM,
                  f"=1/(1+{wacc_cell})^{tc}{R_DCF_PERIOD}", fmt="0.0000")

    # ── PV of FCF ─────────────────────────────────────────────────────────────
    write_val(ws, R_DCF_PV, COL_LABEL, "PV of FCF ($M)", align="left",
              fill_color=C_OUTPUT)
    for i in range(N_PROJ):
        pc = get_column_letter(PROJ_COLS[i])
        write_formula(ws, R_DCF_PV, PROJ_COLS[i],
                      f"={pc}{R_DCF_UFCF}*{pc}{R_DCF_DISC}",
                      fmt=FMT_DOLLAR, fill_color=C_OUTPUT)

    # ── Terminal Value ────────────────────────────────────────────────────────
    tc = get_column_letter(COL_TERM)
    last_pc = get_column_letter(PROJ_COLS[4])   # K = FY2030E

    write_val(ws, R_TV_FCF, COL_LABEL, "Terminal Year FCF ($M)", align="left")
    write_formula(ws, R_TV_FCF, COL_TERM,
                  f"={last_pc}{R_FCF_UFCF}*(1+B{R_SEL_TERM_G})",
                  fmt=FMT_DOLLAR)

    write_val(ws, R_TV_VAL, COL_LABEL, "Terminal Value ($M)", align="left")
    write_formula(ws, R_TV_VAL, COL_TERM,
                  f"={tc}{R_TV_FCF}/({wacc_cell}-B{R_SEL_TERM_G})",
                  fmt=FMT_DOLLAR)

    write_val(ws, R_TV_PV, COL_LABEL, "PV of Terminal Value ($M)", bold=True,
              align="left", fill_color=C_OUTPUT)
    write_formula(ws, R_TV_PV, COL_TERM,
                  f"={tc}{R_TV_VAL}*{tc}{R_DCF_DISC}",
                  fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)

    # ── Valuation Summary ─────────────────────────────────────────────────────
    write_header(ws, R_SUM_HDR, 1, "VALUATION SUMMARY — EV → EQUITY BRIDGE",
                 span_end_col=COL_TERM)

    write_val(ws, R_SUM_PV_FCF, COL_LABEL, "Sum of PV of FCFs ($M)", align="left")
    pv_cells = "+".join(
        f"{get_column_letter(PROJ_COLS[i])}{R_DCF_PV}" for i in range(N_PROJ)
    )
    write_formula(ws, R_SUM_PV_FCF, COL_Y1A,
                  f"={pv_cells}", fmt=FMT_DOLLAR)

    write_val(ws, R_SUM_PV_TV, COL_LABEL, "PV of Terminal Value ($M)", align="left")
    write_formula(ws, R_SUM_PV_TV, COL_Y1A,
                  f"={tc}{R_TV_PV}", fmt=FMT_DOLLAR)

    write_val(ws, R_SUM_EV, COL_LABEL, "Enterprise Value ($M)",
              bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws, R_SUM_EV, COL_Y1A,
                  f"=B{R_SUM_PV_FCF}+B{R_SUM_PV_TV}",
                  fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)

    write_val(ws, R_SUM_NETDEBT, COL_LABEL, "(-) Net Debt ($M)", align="left")
    write_formula(ws, R_SUM_NETDEBT, COL_Y1A,
                  f"=B{R_NETDEBT}", fmt=FMT_DOLLAR)

    write_val(ws, R_SUM_EQ, COL_LABEL, "Equity Value ($M)",
              bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws, R_SUM_EQ, COL_Y1A,
                  f"=B{R_SUM_EV}-B{R_SUM_NETDEBT}",
                  fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)

    write_val(ws, R_SUM_SHARES, COL_LABEL, "Diluted Shares Outstanding (M)", align="left")
    write_formula(ws, R_SUM_SHARES, COL_Y1A,
                  f"=B{R_SHARES}", fmt="#,##0.0")

    write_val(ws, R_SUM_PRICE, COL_LABEL, "IMPLIED PRICE PER SHARE",
              bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws, R_SUM_PRICE, COL_Y1A,
                  f"=B{R_SUM_EQ}/B{R_SUM_SHARES}",
                  fmt=FMT_DOLLAR2, bold=True, fill_color=C_OUTPUT)

    write_val(ws, R_SUM_CUR_PRICE, COL_LABEL, "Current Stock Price", align="left")
    write_formula(ws, R_SUM_CUR_PRICE, COL_Y1A,
                  f"=B{R_PRICE}", fmt=FMT_DOLLAR2)

    write_val(ws, R_SUM_UPSIDE, COL_LABEL, "Implied Upside / (Downside)",
              bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws, R_SUM_UPSIDE, COL_Y1A,
                  f"=B{R_SUM_PRICE}/B{R_SUM_CUR_PRICE}-1",
                  fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)

    print("✓ Part 7: discounting + valuation summary written")

# ═══════════════════════════════════════════════════════════════════════════════
# DCF SHEET — Section E: Sensitivity Tables
# ═══════════════════════════════════════════════════════════════════════════════
def build_sensitivity(ws):
    """
    Three 5×5 sensitivity tables starting at R_SENS1_HDR (row 105).
    Each cell contains a full DCF recalc formula for that assumption combo.
    """

    # ── Helper: full DCF inline recalc formula ────────────────────────────────
    # Given a WACC cell ref and terminal_g cell ref, build the implied price formula.
    # The formula recalculates:
    #   Sum of PV(FCF_i) using wacc_ref as discount rate
    #   + PV(TV) using wacc_ref and tg_ref
    #   - Net Debt
    #   / Shares
    #
    # We reference the FCF rows already on sheet; only WACC and terminal-g change.

    def sens_price_formula(wacc_ref, tg_ref):
        """Build full DCF recalc formula for one sensitivity cell."""
        # PV of each projected FCF
        pv_parts = []
        for i in range(N_PROJ):
            fc = get_column_letter(PROJ_COLS[i])
            period = DISC_PERIODS[i]
            pv_parts.append(
                f"{fc}{R_FCF_UFCF}/(1+{wacc_ref})^{period}"
            )
        pv_sum = "+".join(pv_parts)

        # Terminal value PV
        last_fc = get_column_letter(PROJ_COLS[4])
        tv = (
            f"(({last_fc}{R_FCF_UFCF}*(1+{tg_ref}))/({wacc_ref}-{tg_ref}))"
            f"/(1+{wacc_ref})^4.5"
        )

        return (
            f"=({pv_sum}+{tv}"
            f"-B{R_NETDEBT})/B{R_SHARES}"
        )

    # ── Table 1: WACC vs Terminal Growth Rate ─────────────────────────────────
    r0 = R_SENS1_HDR
    write_header(ws, r0, 1,
                 "SENSITIVITY TABLE 1 — Implied Share Price: WACC vs Terminal Growth Rate",
                 span_end_col=COL_TERM)

    wacc_vals  = [0.065, 0.070, 0.076, 0.082, 0.090]   # base = 0.076 (middle)
    tg_vals    = [0.010, 0.015, 0.025, 0.030, 0.035]   # base = 0.025 (middle col)

    # Column headers (terminal growth)
    write_val(ws, r0+1, COL_LABEL, "WACC  →  Terminal G",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for j, tg in enumerate(tg_vals):
        write_val(ws, r0+1, COL_Y1A+j, tg, fmt=FMT_PCT1,
                  bold=True, fill_color=C_HDR_LIGHT)

    # Row headers (WACC) + formula cells
    for i, wacc_v in enumerate(wacc_vals):
        rr = r0 + 2 + i
        write_val(ws, rr, COL_LABEL, wacc_v, fmt=FMT_PCT1,
                  bold=True, fill_color=C_HDR_LIGHT)
        for j, tg_v in enumerate(tg_vals):
            formula = sens_price_formula(str(wacc_v), str(tg_v))
            is_base = (i == 2 and j == 2)   # center = base case
            write_formula(ws, rr, COL_Y1A+j, formula, fmt=FMT_DOLLAR2,
                          bold=is_base,
                          fill_color=C_OUTPUT if is_base else None)

    # ── Table 2: Revenue Growth (FY2026 base) vs EBIT Margin (FY2026 base) ────
    r1 = r0 + 9
    write_header(ws, r1, 1,
                 "SENSITIVITY TABLE 2 — Implied Share Price: FY2026 Revenue Growth vs FY2026 EBIT Margin",
                 span_end_col=COL_TERM)

    rev_grw_vals  = [0.00, 0.01, 0.02, 0.03, 0.04]   # base = 0.02 (middle)
    ebit_mrg_vals = [0.120, 0.130, 0.140, 0.150, 0.160]  # base = 0.14 (middle)

    write_val(ws, r1+1, COL_LABEL, "Rev Grw  →  EBIT Mrg",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for j, em in enumerate(ebit_mrg_vals):
        write_val(ws, r1+1, COL_Y1A+j, em, fmt=FMT_PCT1,
                  bold=True, fill_color=C_HDR_LIGHT)

    for i, rg in enumerate(rev_grw_vals):
        rr = r1 + 2 + i
        write_val(ws, rr, COL_LABEL, rg, fmt=FMT_PCT1,
                  bold=True, fill_color=C_HDR_LIGHT)
        for j, em in enumerate(ebit_mrg_vals):
            # Override FY2026 growth and margin; keep all other years at base
            # FY2026 Revenue with override growth
            rev26 = f"F{R_IS_REV}*(1+{rg})"   # F = FY2025A col
            ebit26 = f"({rev26})*{em}"
            nopat26 = f"({ebit26})*(1-B{R_TAX})"
            da26    = f"({rev26})*B{R_DA_PCT}"
            capex26 = f"-({rev26})*B{R_CAPEX_PCT}"   # col B = FY2026 CapEx %
            nwc26   = f"-(({rev26})-F{R_IS_REV})*B{R_NWC_PCT}"
            fcf26   = f"({nopat26}+{da26}+{capex26}+{nwc26})"
            pv26    = f"{fcf26}/(1+B{R_WACC_REF})^0.5"

            # Years 2-5: use existing FCF rows
            pv_rest = "+".join(
                f"{get_column_letter(PROJ_COLS[k])}{R_FCF_UFCF}/(1+B{R_WACC_REF})^{DISC_PERIODS[k]}"
                for k in range(1, N_PROJ)
            )

            last_fc = get_column_letter(PROJ_COLS[4])
            tv = (
                f"(({last_fc}{R_FCF_UFCF}*(1+B{R_SEL_TERM_G}))/(B{R_WACC_REF}-B{R_SEL_TERM_G}))"
                f"/(1+B{R_WACC_REF})^4.5"
            )

            formula = (
                f"=({pv26}+{pv_rest}+{tv}"
                f"-B{R_NETDEBT})/B{R_SHARES}"
            )
            is_base = (i == 2 and j == 2)
            write_formula(ws, rr, COL_Y1A+j, formula, fmt=FMT_DOLLAR2,
                          bold=is_base,
                          fill_color=C_OUTPUT if is_base else None)

    # ── Table 3: Beta vs Risk-Free Rate ───────────────────────────────────────
    r2 = r1 + 9
    write_header(ws, r2, 1,
                 "SENSITIVITY TABLE 3 — Implied Share Price: Beta vs Risk-Free Rate",
                 span_end_col=COL_TERM)

    beta_vals = [0.50, 0.60, 0.70, 0.80, 0.90]    # base = 0.70 (middle)
    rf_vals   = [0.035, 0.045, 0.047, 0.055, 0.065]  # base = 0.047 (middle)

    write_val(ws, r2+1, COL_LABEL, "Beta  →  Risk-Free Rate",
              bold=True, fill_color=C_HDR_LIGHT, align="left")
    for j, rf in enumerate(rf_vals):
        write_val(ws, r2+1, COL_Y1A+j, rf, fmt=FMT_PCT1,
                  bold=True, fill_color=C_HDR_LIGHT)

    for i, beta_v in enumerate(beta_vals):
        rr = r2 + 2 + i
        write_val(ws, rr, COL_LABEL, beta_v, fmt="0.00",
                  bold=True, fill_color=C_HDR_LIGHT)
        for j, rf_v in enumerate(rf_vals):
            # Recompute WACC with override beta and rf
            ke_ov   = f"({rf_v}+{beta_v}*{ERP})"
            mktcap  = f"B{R_MKTCAP}"
            nd      = f"B{R_NETDEBT}"
            ev_ov   = f"({mktcap}+{nd})"
            wacc_ov = (
                f"(({mktcap}/{ev_ov})*{ke_ov}"
                f"+(({nd}/{ev_ov})*{KD_AT}))"
            )
            tg_ref  = f"B{R_SEL_TERM_G}"

            pv_parts = []
            for k in range(N_PROJ):
                fc = get_column_letter(PROJ_COLS[k])
                pv_parts.append(
                    f"{fc}{R_FCF_UFCF}/(1+{wacc_ov})^{DISC_PERIODS[k]}"
                )
            pv_sum = "+".join(pv_parts)

            last_fc = get_column_letter(PROJ_COLS[4])
            tv = (
                f"(({last_fc}{R_FCF_UFCF}*(1+{tg_ref}))/({wacc_ov}-{tg_ref}))"
                f"/(1+{wacc_ov})^4.5"
            )
            formula = (
                f"=({pv_sum}+{tv}"
                f"-B{R_NETDEBT})/B{R_SHARES}"
            )
            is_base = (i == 2 and j == 2)
            write_formula(ws, rr, COL_Y1A+j, formula, fmt=FMT_DOLLAR2,
                          bold=is_base,
                          fill_color=C_OUTPUT if is_base else None)

    print("✓ Part 8: three sensitivity tables written")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN — assemble workbook and save
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    import os
    os.makedirs("reports/dcf", exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Build WACC sheet ──────────────────────────────────────────────────────
    # Rename default sheet to WACC
    ws_wacc = wb.active
    ws_wacc.title = "WACC"
    build_wacc_sheet.__globals__['wb'] = wb   # not needed; pass directly
    ws_wacc2, wacc_result_row = build_wacc_sheet(wb)
    # build_wacc_sheet creates a NEW sheet called "WACC" — remove the blank active sheet
    # Actually: let's re-do: create DCF first as active, then WACC
    # Re-approach: use wb.active for DCF, add WACC separately

    # Remove the sheet just made and restart cleanly
    wb2 = openpyxl.Workbook()
    ws_dcf = wb2.active
    ws_dcf.title = "DCF"

    # Build WACC sheet
    ws_wacc = wb2.create_sheet("WACC")

    # ── Populate WACC sheet inline ────────────────────────────────────────────
    set_col_widths(ws_wacc, {"A": 38, "B": 16, "C": 16, "D": 16})
    ws_wacc.row_dimensions[1].height = 22

    r = 1
    ws_wacc.cell(row=r, column=1).value = "Ingredion (INGR) — WACC Calculation"
    ws_wacc.cell(row=r, column=1).font = font(C_FONT_WHITE, bold=True, size=12)
    ws_wacc.cell(row=r, column=1).fill = fill(C_HDR_DARK)
    ws_wacc.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws_wacc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    r += 1
    ws_wacc.cell(row=r, column=1).value = "As of: 2026-08-18  |  Source: StockAnalysis.com, FRED, Barchart"
    ws_wacc.cell(row=r, column=1).font = font("595959", size=9)
    ws_wacc.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws_wacc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    # Cost of Equity
    r += 2
    write_header(ws_wacc, r, 1, "COST OF EQUITY (CAPM)", span_end_col=4)
    r += 1
    write_val(ws_wacc, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws_wacc, r, 2, "Value", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws_wacc, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws_wacc.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c = ws_wacc.cell(row=r, column=3); c.font=font(C_FONT_BLACK,bold=True); c.fill=fill(C_HDR_LIGHT); c.alignment=Alignment(horizontal="left",vertical="center")

    ke_start = r + 1
    ke_data = [
        ("Risk-Free Rate (10Y UST)",   RF,   FMT_PCT1, "Source: FRED DGS10, 2026-08-11, https://fred.stlouisfed.org/series/DGS10"),
        ("Beta (60-Month Monthly)",    BETA, "0.00",   "Source: Barchart.com, 2026-08-18, 60-month beta vs S&P 500"),
        ("Equity Risk Premium (ERP)",  ERP,  FMT_PCT1, "Source: Damodaran long-run ERP, 5.0% standard"),
    ]
    for i, (lbl, v, fmt, cmt) in enumerate(ke_data):
        rr = ke_start + i
        write_val(ws_wacc, rr, 1, lbl, align="left", fill_color=C_INPUT_FILL)
        write_val(ws_wacc, rr, 2, v, fmt=fmt, fg=C_FONT_BLUE, fill_color=C_INPUT_FILL, comment_text=cmt)
        c2 = ws_wacc.cell(row=rr, column=3); c2.value=cmt; c2.font=font("595959",size=9); c2.alignment=Alignment(horizontal="left",vertical="center")
        ws_wacc.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    rr = ke_start + 3
    write_val(ws_wacc, rr, 1, "Cost of Equity (Ke)", align="left")
    write_formula(ws_wacc, rr, 2, f"=B{ke_start}+B{ke_start+1}*B{ke_start+2}", fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)
    c2=ws_wacc.cell(row=rr,column=3); c2.value="= Rf + Beta × ERP"; c2.font=font("595959",size=9); c2.alignment=Alignment(horizontal="left",vertical="center")
    ws_wacc.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    KE_ROW = rr

    # Cost of Debt
    r = ke_start + 5
    write_header(ws_wacc, r, 1, "COST OF DEBT", span_end_col=4)
    r += 1
    write_val(ws_wacc, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws_wacc, r, 2, "Value", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws_wacc, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws_wacc.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c=ws_wacc.cell(row=r,column=3); c.font=font(C_FONT_BLACK,bold=True); c.fill=fill(C_HDR_LIGHT); c.alignment=Alignment(horizontal="left",vertical="center")

    kd_start = r + 1
    kd_data = [
        ("Pre-Tax Cost of Debt (Kd)", KD_PRETAX, FMT_PCT1, "Source: Market rate; BBB+ equiv + ~80bps spread over 10Y UST"),
        ("Tax Rate",                  TAX_RATE,  FMT_PCT1, "Source: Normalised; FY2022-2025 avg ~25.4%"),
    ]
    for i, (lbl, v, fmt, cmt) in enumerate(kd_data):
        rr = kd_start + i
        write_val(ws_wacc, rr, 1, lbl, align="left", fill_color=C_INPUT_FILL)
        write_val(ws_wacc, rr, 2, v, fmt=fmt, fg=C_FONT_BLUE, fill_color=C_INPUT_FILL, comment_text=cmt)
        c2=ws_wacc.cell(row=rr,column=3); c2.value=cmt; c2.font=font("595959",size=9); c2.alignment=Alignment(horizontal="left",vertical="center")
        ws_wacc.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    rr = kd_start + 2
    write_val(ws_wacc, rr, 1, "After-Tax Cost of Debt", align="left")
    write_formula(ws_wacc, rr, 2, f"=B{kd_start}*(1-B{kd_start+1})", fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)
    c2=ws_wacc.cell(row=rr,column=3); c2.value="= Kd × (1 − Tax Rate)"; c2.font=font("595959",size=9); c2.alignment=Alignment(horizontal="left",vertical="center")
    ws_wacc.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
    KDAT_ROW = rr

    # Capital Structure
    r = kd_start + 4
    write_header(ws_wacc, r, 1, "CAPITAL STRUCTURE", span_end_col=4)
    r += 1
    write_val(ws_wacc, r, 1, "Item", bold=True, fill_color=C_HDR_LIGHT, align="left")
    write_val(ws_wacc, r, 2, "Value ($M)", bold=True, fill_color=C_HDR_LIGHT)
    write_val(ws_wacc, r, 3, "Notes", bold=True, fill_color=C_HDR_LIGHT, align="left")
    ws_wacc.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c=ws_wacc.cell(row=r,column=3); c.font=font(C_FONT_BLACK,bold=True); c.fill=fill(C_HDR_LIGHT); c.alignment=Alignment(horizontal="left",vertical="center")

    cs_start = r + 1
    cs_data = [
        ("Current Stock Price",              STOCK_PRICE,  FMT_DOLLAR2, "Source: StockAnalysis.com, NYSE close 2026-08-17"),
        ("Diluted Shares Outstanding (M)",   SHARES_DILUT, "#,##0.0",   "Source: StockAnalysis.com TTM diluted shares"),
        ("Market Capitalisation ($M)",       None,         FMT_DOLLAR,  "= Price × Shares"),
        ("Total Debt ($M)",                  TOTAL_DEBT,   FMT_DOLLAR,  "Source: StockAnalysis.com balance sheet FY2025A"),
        ("Cash & Equivalents ($M)",          CASH,         FMT_DOLLAR,  "Source: StockAnalysis.com balance sheet FY2025A"),
        ("Net Debt ($M)",                    None,         FMT_DOLLAR,  "= Total Debt − Cash"),
        ("Enterprise Value ($M)",            None,         FMT_DOLLAR,  "= Market Cap + Net Debt"),
    ]
    for i, (lbl, v, fmt, cmt) in enumerate(cs_data):
        rr = cs_start + i
        write_val(ws_wacc, rr, 1, lbl, align="left", fill_color=C_INPUT_FILL if v is not None else None)
        if v is not None:
            write_val(ws_wacc, rr, 2, v, fmt=fmt, fg=C_FONT_BLUE, fill_color=C_INPUT_FILL, comment_text=cmt)
        elif i == 2:
            write_formula(ws_wacc, rr, 2, f"=B{cs_start}*B{cs_start+1}", fmt=FMT_DOLLAR)
        elif i == 5:
            write_formula(ws_wacc, rr, 2, f"=B{cs_start+3}-B{cs_start+4}", fmt=FMT_DOLLAR)
        else:
            write_formula(ws_wacc, rr, 2, f"=B{cs_start+2}+B{cs_start+5}", fmt=FMT_DOLLAR, bold=True, fill_color=C_OUTPUT)
        c2=ws_wacc.cell(row=rr,column=3); c2.value=cmt; c2.font=font("595959",size=9); c2.alignment=Alignment(horizontal="left",vertical="center")
        ws_wacc.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)

    EV_ROW   = cs_start + 6
    MKT_ROW  = cs_start + 2
    ND_ROW   = cs_start + 5

    # WACC Calculation
    r = cs_start + 8
    write_header(ws_wacc, r, 1, "WACC CALCULATION", span_end_col=4)
    r += 1
    for hdr, col in [("Component",1),("Weight",2),("Cost",3),("Contribution",4)]:
        write_val(ws_wacc, r, col, hdr, bold=True, fill_color=C_HDR_LIGHT, align="left" if col==1 else "right")

    wc_start = r + 1
    # Equity
    rr = wc_start
    write_val(ws_wacc, rr, 1, "Equity", align="left")
    write_formula(ws_wacc, rr, 2, f"=B{MKT_ROW}/B{EV_ROW}", fmt=FMT_PCT1)
    write_formula(ws_wacc, rr, 3, f"=B{KE_ROW}", fmt=FMT_PCT1, green=True)
    write_formula(ws_wacc, rr, 4, f"=B{rr}*C{rr}", fmt=FMT_PCT1)
    # Debt
    rr = wc_start + 1
    write_val(ws_wacc, rr, 1, "Net Debt (after-tax)", align="left")
    write_formula(ws_wacc, rr, 2, f"=B{ND_ROW}/B{EV_ROW}", fmt=FMT_PCT1)
    write_formula(ws_wacc, rr, 3, f"=B{KDAT_ROW}", fmt=FMT_PCT1, green=True)
    write_formula(ws_wacc, rr, 4, f"=B{rr}*C{rr}", fmt=FMT_PCT1)
    # WACC total — THIS is the cell referenced from DCF sheet as =WACC!D{WACC_ROW}
    WACC_ROW = wc_start + 3
    write_val(ws_wacc, WACC_ROW, 1, "WACC", bold=True, align="left", fill_color=C_OUTPUT)
    write_formula(ws_wacc, WACC_ROW, 4, f"=D{wc_start}+D{wc_start+1}",
                  fmt=FMT_PCT1, bold=True, fill_color=C_OUTPUT)
    ws_wacc.merge_cells(start_row=WACC_ROW, start_column=1, end_row=WACC_ROW, end_column=3)
    c=ws_wacc.cell(row=WACC_ROW,column=1); c.font=font(C_FONT_BLACK,bold=True); c.fill=fill(C_OUTPUT); c.alignment=Alignment(horizontal="left",vertical="center")

    print(f"✓ WACC sheet built — WACC result at WACC!D{WACC_ROW}")

    # ── Patch R_WACC_REF formula in DCF assumptions to match actual WACC row ─
    # The DCF sheet references =WACC!D24 — update if WACC_ROW differs
    global _WACC_SHEET_ROW
    _WACC_SHEET_ROW = WACC_ROW

    # ── Build DCF sheet sections ──────────────────────────────────────────────
    build_dcf_header(ws_dcf)
    build_scenario_assumptions(ws_dcf)
    build_income_statement(ws_dcf)
    build_fcf(ws_dcf)
    build_valuation(ws_dcf)
    build_sensitivity(ws_dcf)

    # Patch WACC cross-reference to use actual row
    wacc_ref_cell = ws_dcf.cell(row=R_WACC_REF, column=ASSUMP_COLS[0])
    wacc_ref_cell.value = f"=WACC!D{WACC_ROW}"
    wacc_ref_cell.font = font(C_FONT_GREEN, bold=True)
    wacc_ref_cell.fill = fill(C_OUTPUT)
    wacc_ref_cell.number_format = FMT_PCT1

    # ── Set DCF sheet as active ───────────────────────────────────────────────
    wb2.active = ws_dcf

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws_dcf.freeze_panes  = "B58"   # freeze label col + rows above IS
    ws_wacc.freeze_panes = "A5"

    # ── Save ─────────────────────────────────────────────────────────────────
    wb2.save(OUTPUT_PATH)
    print(f"\n✅ Model saved → {OUTPUT_PATH}")
    return OUTPUT_PATH

_WACC_SHEET_ROW = 24   # default; patched at runtime

if __name__ == "__main__":
    main()
