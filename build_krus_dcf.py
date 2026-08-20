#!/usr/bin/env python3
"""
build_krus_dcf.py - Institutional DCF Valuation Model for Kura Sushi USA, Inc. (NASDAQ: KRUS)
Follows Investment Banking financial modeling standards, strict openpyxl formatting rules,
and full formula automation for three 5x5 sensitivity tables.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os

def build_krus_dcf_model(output_path="KRUS_DCF_Model_Gemini-3.7-Flash_20260820.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: DCF Valuation
    ws_dcf = wb.active
    ws_dcf.title = "DCF Valuation"
    ws_dcf.views.sheetView[0].showGridLines = True
    
    # Sheet 2: WACC Build
    ws_wacc = wb.create_sheet(title="WACC Build")
    ws_wacc.views.sheetView[0].showGridLines = True
    
    # Palette definition (Institutional Classic Navy & Slate Blue)
    NAVY = "1F4E79"        # Primary headers
    MED_BLUE = "2F5597"    # Secondary headers
    SOFT_BLUE = "D9E1F2"   # Table headers
    ACCENT_BLUE = "BDD7EE" # Highlight / Base Case
    LIGHT_GRAY = "F2F2F2"  # Input fill
    ZEBRA_FILL = "F9FBFD"  # Alternating row
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"
    INPUT_BLUE = "0000FF"  # Hardcoded font
    BLACK = "000000"       # Formula font
    LINK_GREEN = "008000"  # Sheet link font
    
    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_subhdr = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_regular = Font(name="Calibri", size=10, bold=False, color=BLACK)
    font_input = Font(name="Calibri", size=10, bold=False, color=INPUT_BLUE)
    font_input_bold = Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE)
    font_link = Font(name="Calibri", size=10, bold=False, color=LINK_GREEN)
    font_link_bold = Font(name="Calibri", size=10, bold=True, color=LINK_GREEN)
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_navy = Side(border_style="medium", color=NAVY)
    double_navy = Side(border_style="double", color=NAVY)
    top_thin_navy = Side(border_style="thin", color=NAVY)
    
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_navy)
    border_total = Border(top=top_thin_navy, bottom=double_navy, left=thin_line, right=thin_line)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Helper for merged section header
    def add_section_header(ws, row, title, max_col="H"):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_sec_hdr
        ws[f"A{row}"].fill = fill_navy
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_navy
            c.border = Border(top=thick_navy, bottom=thick_navy, left=thin_line, right=thin_line)
            
    # ==========================================
    # SHEET 2: WACC Build (Cost of Capital)
    # ==========================================
    ws_wacc["A1"] = "Kura Sushi USA, Inc. (NASDAQ: KRUS)"
    ws_wacc["A1"].font = font_title
    ws_wacc["A2"] = "Weighted Average Cost of Capital (WACC) Schedule | CAPM Methodology & Capital Structure"
    ws_wacc["A2"].font = font_subtitle
    
    add_section_header(ws_wacc, 4, "I. COST OF EQUITY (CAPM METHODOLOGY)", "E")
    
    wacc_equity_inputs = [
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0464, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 20, 2026 (4.64%)", "B5"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.33, "0.00", "Source: Market data 5-year monthly regression beta vs S&P 500 (1.33)", "B6"),
        ("Market Equity Risk Premium (ERP)", 0.0550, "0.00%", "Source: Standard institutional equity risk premium consensus (5.50%)", "B7"),
        ("Cost of Equity (Ke = Rf + Beta * ERP)", "=B5+B6*B7", "0.00%", None, "B8"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_equity_inputs, start=5):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 10, "II. COST OF DEBT & CAPITAL STRUCTURE WEIGHTS", "E")
    
    wacc_debt_inputs = [
        ("Pre-Tax Cost of Debt (Kd)", 0.0600, "0.00%", "Source: Estimated commercial borrowing rate benchmark (6.00%)", "B11"),
        ("Effective Corporate Tax Rate (t)", 0.2500, "0.0%", "Source: US Federal (21%) + State (~4%) statutory corporate tax rate (25.0%)", "B12"),
        ("After-Tax Cost of Debt (Kd * (1 - t))", "=B11*(1-B12)", "0.00%", None, "B13"),
        ("Current Share Price ($)", "='DCF Valuation'!B7", "$#,##0.00", None, "B14"),
        ("Diluted Shares Outstanding (M)", "='DCF Valuation'!B8", "#,##0.00", None, "B15"),
        ("Market Capitalization ($M)", "=B14*B15", "$#,##0.00", None, "B16"),
        ("Total Debt ($M)", 0.00, "$#,##0.00", "Source: SEC Form 10-K / 10-Q Balance Sheet: KRUS operates with zero funded debt ($0.00M)", "B17"),
        ("Cash and Cash Equivalents ($M)", 50.99, "$#,##0.00", "Source: SEC Form 10-K / 10-Q Balance Sheet total cash and cash equivalents ($50.99M)", "B18"),
        ("Net Debt ($M) [Net Cash if negative]", "=B17-B18", "$#,##0.00;($#,##0.00);$0.00", None, "B19"),
        ("Enterprise Value ($M)", "=B16+B19", "$#,##0.00", None, "B20"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_debt_inputs, start=11):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "='DCF Valuation'" in str(val):
            ws_wacc[f"B{idx}"].font = font_link
        elif "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 22, "III. WACC CALCULATION SUMMARY", "E")
    
    wacc_summary_headers = ["Component", "Market Value ($M)", "Capital Weight (%)", "Cost of Capital (%)", "Weighted Contribution (%)"]
    for c_idx, h_text in enumerate(wacc_summary_headers, start=1):
        c = ws_wacc.cell(row=23, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    ws_wacc["A24"] = "Common Equity (Market Cap)"
    ws_wacc["B24"] = "=B16"
    ws_wacc["C24"] = "=B24/$B$20"
    ws_wacc["D24"] = "=B8"
    ws_wacc["E24"] = "=C24*D24"
    
    ws_wacc["A25"] = "Net Debt (Net Cash Adjustment)"
    ws_wacc["B25"] = "=B19"
    ws_wacc["C25"] = "=B25/$B$20"
    ws_wacc["D25"] = "=B13"
    ws_wacc["E25"] = "=C25*D25"
    
    ws_wacc["A26"] = "Total Capitalization / Base WACC (CAPM Unlevered)"
    ws_wacc["B26"] = "=B20"
    ws_wacc["C26"] = "=SUM(C24:C25)"
    ws_wacc["D26"] = "-"
    ws_wacc["E26"] = "=B8" # For 100% equity / zero debt company, WACC = Ke = 11.96%
    
    for r in range(24, 27):
        ws_wacc[f"A{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"A{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"B{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"B{r}"].number_format = "$#,##0.00"
        ws_wacc[f"B{r}"].alignment = align_right
        ws_wacc[f"B{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"C{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"C{r}"].number_format = "0.0%"
        ws_wacc[f"C{r}"].alignment = align_right
        ws_wacc[f"C{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"D{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"D{r}"].number_format = "0.00%" if r != 26 else "@"
        ws_wacc[f"D{r}"].alignment = align_right
        ws_wacc[f"D{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"E{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"E{r}"].number_format = "0.00%"
        ws_wacc[f"E{r}"].alignment = align_right
        ws_wacc[f"E{r}"].border = border_total if r == 26 else border_cell
        if r == 26:
            ws_wacc[f"E{r}"].fill = fill_accent_blue
            
    ws_wacc.column_dimensions["A"].width = 48
    ws_wacc.column_dimensions["B"].width = 22
    ws_wacc.column_dimensions["C"].width = 18
    ws_wacc.column_dimensions["D"].width = 18
    ws_wacc.column_dimensions["E"].width = 24
    
    # ==========================================
    # SHEET 1: DCF Valuation
    # ==========================================
    
    # --- Title Block ---
    ws_dcf["A1"] = "Kura Sushi USA, Inc. (NASDAQ: KRUS)"
    ws_dcf["A1"].font = font_title
    ws_dcf["A2"] = "Institutional DCF Valuation Model | Report Date: August 20, 2026 | Financial Source: SEC Form 10-K, 10-Q & Consensus Estimates"
    ws_dcf["A2"].font = font_subtitle
    
    # --- Active Scenario Selector ---
    ws_dcf["A4"] = "Active Scenario Selector:"
    ws_dcf["A4"].font = font_bold
    ws_dcf["B4"] = 2
    ws_dcf["B4"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    ws_dcf["B4"].fill = fill_navy
    ws_dcf["B4"].alignment = align_center
    ws_dcf["B4"].border = border_cell
    ws_dcf["B4"].comment = Comment("Case Selector: 1 = Bear Case (Expansion Friction & Labor Drag), 2 = Base Case (Steady Unit Growth & Scale Leverage), 3 = Bull Case (Accelerated Expansion & Automation Synergy)", "DCF Model Builder")
    
    ws_dcf["C4"] = '=IF(B4=1,"[1] BEAR CASE (Traffic Moderation / Wage Inflation / Capex Drag)",IF(B4=2,"[2] BASE CASE (Steady Store Expansion / G&A Operating Leverage / Tech Automation)","[3] BULL CASE (Accelerated Openings / SSSG Outperformance / Rapid Margin Expansion)"))'
    ws_dcf["C4"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws_dcf["C4"].alignment = align_left
    
    # --- Section I: Market Data & Capital Structure Inputs ---
    add_section_header(ws_dcf, 6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS", "H")
    
    market_inputs_dcf = [
        ("Current Stock Price ($)", 50.14, "$#,##0.00", "Source: Market close price as of August 19, 2026 ($50.14)", "B7"),
        ("Diluted Shares Outstanding (M)", 12.15, "#,##0.00", "Source: SEC Form 10-Q / 10-K diluted share count (12.15M shares)", "B8"),
        ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
        ("Cash and Cash Equivalents ($M)", 50.99, "$#,##0.00", "Source: SEC Form 10-K / 10-Q Balance Sheet cash and equivalents ($50.99M)", "B10"),
        ("Total Debt ($M)", 0.00, "$#,##0.00", "Source: SEC Form 10-K / 10-Q Balance Sheet: KRUS operates with zero funded debt ($0.00M)", "B11"),
        ("Net Debt ($M) [Net Cash if negative]", "=B11-B10", "$#,##0.00;($#,##0.00);$0.00", None, "B12"),
        ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
        ("Effective Corporate Tax Rate (%)", 0.2500, "0.0%", "Source: US Federal statutory rate (21%) + State tax (~4%) = 25.0%", "B14"),
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0464, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 20, 2026 (4.64%)", "B15"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.33, "0.00", "Source: Market data 5-year monthly regression beta vs S&P 500 (1.33)", "B16"),
        ("Market Equity Risk Premium (ERP)", 0.0550, "0.00%", "Source: Standard institutional equity risk premium consensus (5.50%)", "B17"),
        ("Pre-Tax Cost of Debt (Kd)", 0.0600, "0.00%", "Source: Estimated commercial borrowing rate benchmark (6.00%)", "B18"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs_dcf, start=7):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_dcf[f"B{idx}"].font = font_bold
        else:
            ws_dcf[f"B{idx}"].font = font_input
            ws_dcf[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_dcf[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    # --- Section II: Scenario Assumptions & Driver Blocks ---
    add_section_header(ws_dcf, 20, "II. SCENARIO ASSUMPTIONS & DRIVER BLOCKS", "H")
    
    headers_s2 = ["Driver / Assumption", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal / WACC"]
    for col_idx, text in enumerate(headers_s2, start=1):
        c = ws_dcf.cell(row=21, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header
        
    scenario_blocks = [
        ("Bear Case Assumptions (Selector = 1)", 22, [
            ("  Revenue Growth (%)", [0.120, 0.110, 0.090, 0.070, 0.050], 0.020, "0.0%"),
            ("  EBIT Margin (%)", [-0.015, 0.000, 0.015, 0.025, 0.035], 0.1250, "0.0%"),
            ("  D&A (% of Revenue)", [0.060, 0.060, 0.060, 0.060, 0.060], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.110, 0.090, 0.075, 0.065, 0.055], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
        ("Base Case Assumptions (Selector = 2)", 29, [
            ("  Revenue Growth (%)", [0.170, 0.180, 0.150, 0.120, 0.090], 0.025, "0.0%"),
            ("  EBIT Margin (%)", [0.005, 0.025, 0.045, 0.060, 0.070], 0.1196, "0.0%"),
            ("  D&A (% of Revenue)", [0.060, 0.060, 0.060, 0.060, 0.060], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.120, 0.100, 0.085, 0.070, 0.060], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
        ("Bull Case Assumptions (Selector = 3)", 36, [
            ("  Revenue Growth (%)", [0.210, 0.220, 0.180, 0.150, 0.120], 0.030, "0.0%"),
            ("  EBIT Margin (%)", [0.020, 0.045, 0.065, 0.080, 0.095], 0.1150, "0.0%"),
            ("  D&A (% of Revenue)", [0.060, 0.060, 0.060, 0.060, 0.060], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.130, 0.110, 0.090, 0.075, 0.065], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
    ]
    
    for title, sub_hdr_row, rows in scenario_blocks:
        ws_dcf[f"A{sub_hdr_row}"] = title
        ws_dcf[f"A{sub_hdr_row}"].font = font_subhdr
        ws_dcf[f"A{sub_hdr_row}"].fill = fill_soft_blue
        ws_dcf.merge_cells(f"A{sub_hdr_row}:H{sub_hdr_row}")
        for c_idx in range(1, 9):
            ws_dcf.cell(row=sub_hdr_row, column=c_idx).fill = fill_soft_blue
            ws_dcf.cell(row=sub_hdr_row, column=c_idx).border = border_cell
            
        for r_offset, (label, vals, term_val, num_fmt) in enumerate(rows, start=1):
            curr_r = sub_hdr_row + r_offset
            ws_dcf[f"A{curr_r}"] = label
            ws_dcf[f"A{curr_r}"].font = font_regular
            ws_dcf[f"A{curr_r}"].border = border_cell
            
            ws_dcf.cell(row=curr_r, column=2, value="—").alignment = align_center
            ws_dcf.cell(row=curr_r, column=2).border = border_cell
            
            for c_idx, val in enumerate(vals, start=3):
                cell = ws_dcf.cell(row=curr_r, column=c_idx, value=val)
                cell.font = font_input
                cell.fill = fill_input
                cell.number_format = num_fmt
                cell.alignment = align_right
                cell.border = border_cell
            if term_val is not None:
                cell_term = ws_dcf.cell(row=curr_r, column=8, value=term_val)
                cell_term.font = font_input
                cell_term.fill = fill_input
                cell_term.number_format = "0.0%" if r_offset == 1 else "0.00%"
                cell_term.alignment = align_right
                cell_term.border = border_cell
            else:
                cell_term = ws_dcf.cell(row=curr_r, column=8, value="")
                cell_term.border = border_cell
                
    # Consolidated Active Drivers (Dynamic via CHOOSE formulas)
    ws_dcf["A43"] = "Active Scenario Consolidated Drivers (Dynamic)"
    ws_dcf["A43"].font = font_bold
    ws_dcf["A43"].fill = fill_accent_blue
    ws_dcf.merge_cells("A43:H43")
    for c_idx in range(1, 9):
        ws_dcf.cell(row=43, column=c_idx).fill = fill_accent_blue
        ws_dcf.cell(row=43, column=c_idx).border = border_cell
        
    active_driver_rows = [
        ("  Active Revenue Growth (%)", 23, 30, 37, "0.0%"),
        ("  Active EBIT Margin (%)", 24, 31, 38, "0.0%"),
        ("  Active D&A (% of Revenue)", 25, 32, 39, "0.0%"),
        ("  Active CapEx (% of Revenue)", 26, 33, 40, "0.0%"),
        ("  Active Δ NWC (% of Δ Rev)", 27, 34, 41, "0.0%"),
    ]
    
    for idx, (label, r_bear, r_base, r_bull, num_fmt) in enumerate(active_driver_rows, start=44):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold
        ws_dcf[f"A{idx}"].border = border_cell
        ws_dcf.cell(row=idx, column=2, value="—").alignment = align_center
        ws_dcf.cell(row=idx, column=2).border = border_cell
        
        for c_idx in range(3, 8):
            col_let = get_column_letter(c_idx)
            cell = ws_dcf.cell(row=idx, column=c_idx)
            cell.value = f"=CHOOSE($B$4, {col_let}{r_bear}, {col_let}{r_base}, {col_let}{r_bull})"
            cell.font = font_bold
            cell.number_format = num_fmt
            cell.alignment = align_right
            cell.border = border_cell
            
        # Terminal / WACC in column H (Col 8)
        if idx == 44: # Terminal g
            cell_g = ws_dcf.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H23, H30, H37)")
            cell_g.font = font_bold
            cell_g.number_format = "0.0%"
            cell_g.alignment = align_right
            cell_g.border = border_cell
        elif idx == 45: # Target WACC
            cell_g = ws_dcf.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H24, H31, H38)")
            cell_g.font = font_bold
            cell_g.number_format = "0.00%"
            cell_g.alignment = align_right
            cell_g.border = border_cell
        else:
            cell_g = ws_dcf.cell(row=idx, column=8, value="")
            cell_g.border = border_cell
            
    # --- Section III: Cost of Capital (WACC) Schedule ---
    add_section_header(ws_dcf, 51, "III. COST OF CAPITAL (WACC) PARAMETERS & SCHEDULE", "H")
    
    wacc_items_dcf = [
        ("Risk-Free Rate (Rf)", "=$B$15", "0.00%", "Row 15: 10-Yr US Treasury Yield", False),
        ("Equity Beta (β)", "=$B$16", "0.00", "Row 16: 5-Year Monthly Beta", False),
        ("Equity Risk Premium (ERP)", "=$B$17", "0.00%", "Row 17: Market Equity Risk Premium", False),
        ("Cost of Equity (Ke) [Rf + β * ERP]", "=B52+B53*B54", "0.00%", "CAPM Formula: Cost of Equity", True),
        ("Pre-Tax Cost of Debt (Kd)", "=$B$18", "0.00%", "Row 18: Borrowing rate (~6.00%)", False),
        ("Effective Corporate Tax Rate (t)", "=$B$14", "0.0%", "Row 14: Corporate tax rate", False),
        ("After-Tax Cost of Debt [Kd * (1 - t)]", "=B56*(1-B57)", "0.00%", "Effective after-tax cost of debt", False),
        ("Market Capitalization (Equity Value) ($M)", "=$B$9", "$#,##0.00", "Row 9: Diluted Market Cap", False),
        ("Total Debt ($M)", "=$B$11", "$#,##0.00", "Row 11: Total Debt", False),
        ("Cash and Cash Equivalents ($M)", "=$B$10", "$#,##0.00", "Row 10: Cash and Marketable Securities", False),
        ("Net Debt ($M)", "=B60-B61", "$#,##0.00;($#,##0.00);$0.00", "Row 12: Total Debt - Cash", False),
        ("Enterprise Value (EV) ($M)", "=B59+B62", "$#,##0.00", "Row 13: Market Cap + Net Debt", False),
        ("Selected DCF Discount Rate (WACC)", "='WACC Build'!E26", "0.00%", "Discount rate aligned with CAPM unlevered capital structure", True),
    ]
    
    for idx, (label, val, num_fmt, comment_t, is_key) in enumerate(wacc_items_dcf, start=52):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if is_key else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        if is_key:
            ws_dcf[f"A{idx}"].fill = fill_accent_blue
            
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].font = font_bold if is_key else font_regular
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        if is_key:
            ws_dcf[f"B{idx}"].fill = fill_accent_blue
        if comment_t:
            ws_dcf[f"B{idx}"].comment = Comment(comment_t, "WACC Schedule")
            
    # --- Section IV: Unlevered Free Cash Flow (UFCF) Projections ---
    add_section_header(ws_dcf, 67, "IV. UNLEVERED FREE CASH FLOW (UFCF) PROJECTIONS", "H")
    
    headers_s4 = ["Line Item ($M)", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for col_idx, text in enumerate(headers_s4, start=1):
        c = ws_dcf.cell(row=68, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header
        
    # Row 69: Revenue Growth Rate
    ws_dcf["A69"] = "Revenue Growth Rate (%)"
    ws_dcf["A69"].font = font_regular
    ws_dcf["A69"].border = border_cell
    ws_dcf["B69"] = 0.189
    ws_dcf["B69"].font = font_input
    ws_dcf["B69"].fill = fill_input
    ws_dcf["B69"].number_format = "0.0%"
    ws_dcf["B69"].alignment = align_right
    ws_dcf["B69"].border = border_cell
    ws_dcf["B69"].comment = Comment("FY2025 Actual Total Sales YoY Growth (+18.9%)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=69, column=c_idx, value=f"={col_let}44")
        cell.font = font_regular
        cell.number_format = "0.0%"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 70: Total Revenue ($M)
    ws_dcf["A70"] = "Total Revenue ($M)"
    ws_dcf["A70"].font = font_bold
    ws_dcf["A70"].border = border_cell
    ws_dcf["B70"] = 282.80
    ws_dcf["B70"].font = font_input_bold
    ws_dcf["B70"].fill = fill_input
    ws_dcf["B70"].number_format = "$#,##0.00"
    ws_dcf["B70"].alignment = align_right
    ws_dcf["B70"].border = border_cell
    ws_dcf["B70"].comment = Comment("FY2025 Actual Total Revenue ($282.80M)", "SEC Form 10-K")
    
    ws_dcf["C70"] = "=B70*(1+C69)"
    ws_dcf["D70"] = "=C70*(1+D69)"
    ws_dcf["E70"] = "=D70*(1+E69)"
    ws_dcf["F70"] = "=E70*(1+F69)"
    ws_dcf["G70"] = "=F70*(1+G69)"
    for c_idx in range(3, 8):
        c = ws_dcf.cell(row=70, column=c_idx)
        c.font = font_bold
        c.number_format = "$#,##0.00"
        c.alignment = align_right
        c.border = border_cell
        
    # Row 71: EBIT Margin (%)
    ws_dcf["A71"] = "EBIT Margin (%)"
    ws_dcf["A71"].font = font_regular
    ws_dcf["A71"].border = border_cell
    ws_dcf["B71"] = -0.023
    ws_dcf["B71"].font = font_input
    ws_dcf["B71"].fill = fill_input
    ws_dcf["B71"].number_format = "-0.0%;0.0%;0.0%"
    ws_dcf["B71"].alignment = align_right
    ws_dcf["B71"].border = border_cell
    ws_dcf["B71"].comment = Comment("FY2025 Actual Operating Loss Margin (-2.3%)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=71, column=c_idx, value=f"={col_let}45")
        cell.font = font_regular
        cell.number_format = "0.0%"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 72: Operating Income / (Loss) (EBIT) ($M)
    ws_dcf["A72"] = "Operating Income / (Loss) (EBIT) ($M)"
    ws_dcf["A72"].font = font_bold
    ws_dcf["A72"].border = border_cell
    ws_dcf["B72"] = -6.50
    ws_dcf["B72"].font = font_input_bold
    ws_dcf["B72"].fill = fill_input
    ws_dcf["B72"].number_format = "$#,##0.00;($#,##0.00);$0.00"
    ws_dcf["B72"].alignment = align_right
    ws_dcf["B72"].border = border_cell
    ws_dcf["B72"].comment = Comment("FY2025 Actual Operating Loss ($(6.50)M)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=72, column=c_idx, value=f"={col_let}70*{col_let}71")
        cell.font = font_bold
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 73: (-) Taxes ($M)
    ws_dcf["A73"] = "(-) Provision for Taxes ($M)"
    ws_dcf["A73"].font = font_regular
    ws_dcf["A73"].border = border_cell
    ws_dcf["B73"] = 0.00
    ws_dcf["B73"].font = font_regular
    ws_dcf["B73"].number_format = "($#,##0.00);($#,##0.00);$0.00"
    ws_dcf["B73"].alignment = align_right
    ws_dcf["B73"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=73, column=c_idx, value=f"=MAX(0, {col_let}72*$B$14)")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 74: NOPAT ($M)
    ws_dcf["A74"] = "Net Operating Profit After Tax (NOPAT) ($M)"
    ws_dcf["A74"].font = font_bold
    ws_dcf["A74"].border = border_cell
    ws_dcf["B74"] = -6.50
    ws_dcf["B74"].font = font_bold
    ws_dcf["B74"].number_format = "$#,##0.00;($#,##0.00);$0.00"
    ws_dcf["B74"].alignment = align_right
    ws_dcf["B74"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=74, column=c_idx, value=f"={col_let}72-{col_let}73")
        cell.font = font_bold
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 75: (+) D&A ($M)
    ws_dcf["A75"] = "(+) Depreciation & Amortization (D&A) ($M)"
    ws_dcf["A75"].font = font_regular
    ws_dcf["A75"].border = border_cell
    ws_dcf["B75"] = 17.00
    ws_dcf["B75"].font = font_input
    ws_dcf["B75"].fill = fill_input
    ws_dcf["B75"].number_format = "$#,##0.00"
    ws_dcf["B75"].alignment = align_right
    ws_dcf["B75"].border = border_cell
    ws_dcf["B75"].comment = Comment("FY2025 Actual D&A (~$17.00M)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=75, column=c_idx, value=f"={col_let}70*{col_let}46")
        cell.font = font_regular
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 76: (-) CapEx ($M)
    ws_dcf["A76"] = "(-) Capital Expenditures (CapEx) ($M)"
    ws_dcf["A76"].font = font_regular
    ws_dcf["A76"].border = border_cell
    ws_dcf["B76"] = 42.00
    ws_dcf["B76"].font = font_input
    ws_dcf["B76"].fill = fill_input
    ws_dcf["B76"].number_format = "($#,##0.00);($#,##0.00);$0.00"
    ws_dcf["B76"].alignment = align_right
    ws_dcf["B76"].border = border_cell
    ws_dcf["B76"].comment = Comment("FY2025 Actual CapEx ($42.00M)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=76, column=c_idx, value=f"={col_let}70*{col_let}47")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 77: (-) Change in NWC ($M)
    ws_dcf["A77"] = "(-) Change in Net Working Capital (Δ NWC) ($M)"
    ws_dcf["A77"].font = font_regular
    ws_dcf["A77"].border = border_cell
    ws_dcf["B77"] = 0.50
    ws_dcf["B77"].font = font_input
    ws_dcf["B77"].fill = fill_input
    ws_dcf["B77"].number_format = "($#,##0.00);($#,##0.00);$0.00"
    ws_dcf["B77"].alignment = align_right
    ws_dcf["B77"].border = border_cell
    ws_dcf["B77"].comment = Comment("FY2025 Working Capital change (~$0.50M)", "SEC Form 10-K")
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        prev_col = get_column_letter(c_idx - 1)
        cell = ws_dcf.cell(row=77, column=c_idx, value=f"=({col_let}70-{prev_col}70)*{col_let}48")
        cell.font = font_regular
        cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 78: Unlevered Free Cash Flow (UFCF) ($M)
    ws_dcf["A78"] = "Unlevered Free Cash Flow (UFCF) ($M)"
    ws_dcf["A78"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws_dcf["A78"].fill = fill_accent_blue
    ws_dcf["A78"].border = border_total
    ws_dcf["B78"] = -32.00
    ws_dcf["B78"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws_dcf["B78"].alignment = align_right
    ws_dcf["B78"].number_format = "$#,##0.00;($#,##0.00);$0.00"
    ws_dcf["B78"].fill = fill_accent_blue
    ws_dcf["B78"].border = border_total
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=78, column=c_idx, value=f"={col_let}74+{col_let}75-{col_let}76-{col_let}77")
        cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        cell.fill = fill_accent_blue
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_total
        
    # --- Section V: Discounting & Present Value ---
    add_section_header(ws_dcf, 81, "V. DISCOUNTING & PRESENT VALUE (MID-YEAR CONVENTION)", "H")
    
    headers_s5 = ["Discount Parameters", "", "Year 1 (FY26E)", "Year 2 (FY27E)", "Year 3 (FY28E)", "Year 4 (FY29E)", "Year 5 (FY30E)"]
    for col_idx, text in enumerate(headers_s5, start=1):
        c = ws_dcf.cell(row=82, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 2 else align_left
        c.border = border_header
        
    # Row 83: Discount Period
    ws_dcf["A83"] = "Discount Period (Years - Mid-Year Convention)"
    ws_dcf["A83"].font = font_regular
    ws_dcf["A83"].border = border_cell
    ws_dcf["B83"] = ""
    ws_dcf["B83"].border = border_cell
    periods = [0.5, 1.5, 2.5, 3.5, 4.5]
    for c_idx, p in enumerate(periods, start=3):
        cell = ws_dcf.cell(row=83, column=c_idx, value=p)
        cell.font = font_regular
        cell.number_format = "0.0"
        cell.alignment = align_center
        cell.border = border_cell
        
    # Row 84: Discount Factor
    ws_dcf["A84"] = "Discount Factor [ 1 / (1 + WACC)^t ]"
    ws_dcf["A84"].font = font_regular
    ws_dcf["A84"].border = border_cell
    ws_dcf["B84"] = ""
    ws_dcf["B84"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=84, column=c_idx, value=f"=1/(1+$B$64)^{col_let}83")
        cell.font = font_regular
        cell.number_format = "0.0000"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 85: Present Value of UFCF
    ws_dcf["A85"] = "Present Value of UFCF ($M)"
    ws_dcf["A85"].font = font_bold
    ws_dcf["A85"].border = border_cell
    ws_dcf["B85"] = ""
    ws_dcf["B85"].border = border_cell
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws_dcf.cell(row=85, column=c_idx, value=f"={col_let}78*{col_let}84")
        cell.font = font_bold
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        
    # Row 86: Sum of PV of Explicit Cash Flows
    ws_dcf["A86"] = "Cumulative PV of 5-Year Explicit Cash Flows ($M)"
    ws_dcf["A86"].font = font_bold
    ws_dcf["A86"].fill = fill_soft_blue
    ws_dcf["A86"].border = border_total
    ws_dcf["B86"] = "=SUM(C85:G85)"
    ws_dcf["B86"].font = font_bold
    ws_dcf["B86"].fill = fill_soft_blue
    ws_dcf["B86"].number_format = "$#,##0.00;($#,##0.00);$0.00"
    ws_dcf["B86"].alignment = align_right
    ws_dcf["B86"].border = border_total
    ws_dcf.merge_cells("C86:G86")
    for c_idx in range(3, 8):
        ws_dcf.cell(row=86, column=c_idx).fill = fill_soft_blue
        ws_dcf.cell(row=86, column=c_idx).border = border_total
        
    # --- Section VI: Terminal Value & Valuation Summary ---
    add_section_header(ws_dcf, 88, "VI. TERMINAL VALUE & VALUATION SUMMARY", "H")
    
    val_summary_items = [
        ("Terminal Value Assumptions & Calculation:", None, None, None, True),
        ("  Final Year Projected UFCF (FY2030E) ($M)", "=G78", "$#,##0.00;($#,##0.00);$0.00", None, False),
        ("  Perpetual Terminal Growth Rate (g)", "=$H$44", "0.0%", None, False),
        ("  WACC (Discount Rate)", "=$B$64", "0.00%", None, False),
        ("  Terminal Year Cash Flow ($M)", "=B90*(1+B91)", "$#,##0.00", None, False),
        ("  Terminal Value at FY2030E ($M)", "=B93/(B92-B91)", "$#,##0.00", None, False),
        ("  PV of Terminal Value ($M)", "=B94/(1+B92)^4.5", "$#,##0.00", None, False),
        ("  Terminal Value as % of Enterprise Value", "=B95/B100", "0.0%", None, False),
        ("Enterprise Value to Equity Value Bridge ($M):", None, None, None, True),
        ("  PV of Explicit 5-Year Cash Flows ($M)", "=B86", "$#,##0.00;($#,##0.00);$0.00", None, False),
        ("  (+) PV of Terminal Value ($M)", "=B95", "$#,##0.00", None, False),
        ("Enterprise Value (EV) ($M)", "=B98+B99", "$#,##0.00;($#,##0.00);$0.00", None, False),
        ("  (-) Total Debt ($M)", "=$B$11", "($#,##0.00);($#,##0.00);$0.00", None, False),
        ("  (+) Total Cash & Equivalents ($M)", "=$B$10", "$#,##0.00", None, False),
        ("  (-) Net Debt ($M)", "=$B$12", "$#,##0.00;($#,##0.00);$0.00", None, False),
        ("Implied Equity Value ($M)", "=B100-B103", "$#,##0.00", None, False),
        ("  Diluted Shares Outstanding (M)", "=$B$8", "#,##0.00", None, False),
        ("Implied Price per Share ($)", "=B104/B105", "$#,##0.00", None, False),
        ("Current Market Share Price ($)", "=$B$7", "$#,##0.00", None, False),
        ("Implied Upside / (Downside) (%)", "=(B106/B107)-1", "+0.0%;-0.0%;0.0%", None, False),
    ]
    
    current_r = 89
    for item in val_summary_items:
        label, formula_val, num_fmt, comment_t, is_header = item
        if is_header:
            ws_dcf[f"A{current_r}"] = label
            ws_dcf[f"A{current_r}"].font = font_subhdr
            ws_dcf[f"A{current_r}"].fill = fill_soft_blue
            ws_dcf.merge_cells(f"A{current_r}:H{current_r}")
            for c_idx in range(1, 9):
                ws_dcf.cell(row=current_r, column=c_idx).fill = fill_soft_blue
                ws_dcf.cell(row=current_r, column=c_idx).border = border_cell
        else:
            ws_dcf[f"A{current_r}"] = label
            is_highlight = label in ["Enterprise Value (EV) ($M)", "Implied Equity Value ($M)", "Implied Price per Share ($)", "Implied Upside / (Downside) (%)"]
            ws_dcf[f"A{current_r}"].font = font_bold if is_highlight else font_regular
            ws_dcf[f"A{current_r}"].border = border_cell
            if is_highlight:
                ws_dcf[f"A{current_r}"].fill = fill_accent_blue
                
            ws_dcf[f"B{current_r}"] = formula_val
            ws_dcf[f"B{current_r}"].font = font_bold if is_highlight else font_regular
            ws_dcf[f"B{current_r}"].number_format = num_fmt
            ws_dcf[f"B{current_r}"].alignment = align_right
            ws_dcf[f"B{current_r}"].border = border_cell
            if is_highlight:
                ws_dcf[f"B{current_r}"].fill = fill_accent_blue
            if comment_t:
                ws_dcf[f"B{current_r}"].comment = Comment(comment_t, "Valuation Summary")
        current_r += 1
        
    # --- Section VII: Sensitivity Analysis (5x5 Institutional Grids) ---
    add_section_header(ws_dcf, 111, "VII. SENSITIVITY ANALYSIS (5x5 INSTITUTIONAL VALUATION GRIDS)", "H")
    
    # Table 1: WACC vs Terminal Growth Rate
    ws_dcf["A112"] = "Sensitivity Table 1: Implied Share Price ($) vs. WACC and Perpetual Terminal Growth Rate (g)"
    ws_dcf["A112"].font = font_subhdr
    ws_dcf.merge_cells("A112:H112")
    
    ws_dcf["A113"] = "WACC \\ g"
    ws_dcf["A113"].font = font_tbl_hdr
    ws_dcf["A113"].fill = fill_soft_blue
    ws_dcf["A113"].alignment = align_center
    ws_dcf["A113"].border = border_header
    
    t1_g_values = [0.015, 0.020, 0.025, 0.030, 0.035]
    t1_wacc_values = [0.1000, 0.1100, 0.1196, 0.1300, 0.1400]
    
    for col_idx, g_val in enumerate(t1_g_values, start=2):
        c = ws_dcf.cell(row=113, column=col_idx, value=g_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "0.0%"
        c.alignment = align_center
        c.border = border_header
        
    for row_idx, w_val in enumerate(t1_wacc_values, start=114):
        ws_dcf.cell(row=row_idx, column=1, value=w_val)
        ws_dcf.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws_dcf.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 116 else fill_soft_blue
        ws_dcf.cell(row=row_idx, column=1).number_format = "0.00%"
        ws_dcf.cell(row=row_idx, column=1).alignment = align_center
        ws_dcf.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            # Full DCF recalculation formula:
            formula = (
                f"=(($C$78/(1+$A{row_idx})^0.5 + $D$78/(1+$A{row_idx})^1.5 + $E$78/(1+$A{row_idx})^2.5 + "
                f"$F$78/(1+$A{row_idx})^3.5 + $G$78/(1+$A{row_idx})^4.5 + "
                f"($G$78*(1+{col_let}$113)/($A{row_idx}-{col_let}$113))/(1+$A{row_idx})^4.5) - $B$12)/$B$8"
            )
            cell = ws_dcf.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 116 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 116 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!
                
    # Table 2: FY26-30 Revenue Growth Delta vs Terminal EBIT Margin
    ws_dcf["A121"] = "Sensitivity Table 2: Implied Share Price ($) vs. Revenue Growth Delta & Terminal EBIT Margin"
    ws_dcf["A121"].font = font_subhdr
    ws_dcf.merge_cells("A121:H121")
    
    ws_dcf["A122"] = "EBIT Mgn \\ Rev Δ"
    ws_dcf["A122"].font = font_tbl_hdr
    ws_dcf["A122"].fill = fill_soft_blue
    ws_dcf["A122"].alignment = align_center
    ws_dcf["A122"].border = border_header
    
    t2_rev_delta = [-0.100, -0.050, 0.000, 0.050, 0.100]
    t2_ebit_margins = [0.050, 0.060, 0.070, 0.080, 0.090]
    
    for col_idx, g_val in enumerate(t2_rev_delta, start=2):
        c = ws_dcf.cell(row=122, column=col_idx, value=g_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "+0.0%;-0.0%;0.0%"
        c.alignment = align_center
        c.border = border_header
        
    for row_idx, ebit_mgn in enumerate(t2_ebit_margins, start=123):
        ws_dcf.cell(row=row_idx, column=1, value=ebit_mgn)
        ws_dcf.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws_dcf.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 125 else fill_soft_blue
        ws_dcf.cell(row=row_idx, column=1).number_format = "0.0%"
        ws_dcf.cell(row=row_idx, column=1).alignment = align_center
        ws_dcf.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            # Scaled DCF directly with live Excel formula:
            term_fcf_expr = f"($G$70*(1+{col_let}$122)*($A{row_idx}*(1-$B$14)+$G$46-$G$47))"
            formula = (
                f"=((($C$78*(1+{col_let}$122))/(1+$B$64)^0.5 + "
                f"($D$78*(1+{col_let}$122))/(1+$B$64)^1.5 + "
                f"($E$78*(1+{col_let}$122))/(1+$B$64)^2.5 + "
                f"($F$78*(1+{col_let}$122))/(1+$B$64)^3.5 + "
                f"({term_fcf_expr})/(1+$B$64)^4.5 + "
                f"(({term_fcf_expr}*(1+$H$44)/($B$64-$H$44))/(1+$B$64)^4.5)) - $B$12)/$B$8"
            )
            cell = ws_dcf.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 125 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 125 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!
                
    # Table 3: Beta vs Risk-Free Rate
    ws_dcf["A130"] = "Sensitivity Table 3: Implied Share Price ($) vs. Beta & 10-Year Treasury Yield (Risk-Free Rate)"
    ws_dcf["A130"].font = font_subhdr
    ws_dcf.merge_cells("A130:H130")
    
    ws_dcf["A131"] = "Beta \\ Risk-Free Rate"
    ws_dcf["A131"].font = font_tbl_hdr
    ws_dcf["A131"].fill = fill_soft_blue
    ws_dcf["A131"].alignment = align_center
    ws_dcf["A131"].border = border_header
    
    t3_rf_values = [0.0364, 0.0414, 0.0464, 0.0514, 0.0564]
    t3_beta_values = [1.13, 1.23, 1.33, 1.43, 1.53]
    
    for col_idx, rf_val in enumerate(t3_rf_values, start=2):
        c = ws_dcf.cell(row=131, column=col_idx, value=rf_val)
        c.font = font_tbl_hdr
        c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
        c.number_format = "0.00%"
        c.alignment = align_center
        c.border = border_header
        
    for row_idx, beta_val in enumerate(t3_beta_values, start=132):
        ws_dcf.cell(row=row_idx, column=1, value=beta_val)
        ws_dcf.cell(row=row_idx, column=1).font = font_tbl_hdr
        ws_dcf.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 134 else fill_soft_blue
        ws_dcf.cell(row=row_idx, column=1).number_format = "0.00"
        ws_dcf.cell(row=row_idx, column=1).alignment = align_center
        ws_dcf.cell(row=row_idx, column=1).border = border_cell
        
        for col_idx in range(2, 7):
            col_let = get_column_letter(col_idx)
            # Derived WACC = Ke = Rf + Beta * ERP = {col_let}$131 + $A{row_idx} * 0.0550
            wacc_expr = f"({col_let}$131 + $A{row_idx}*0.055)"
            formula = (
                f"=(($C$78/(1+{wacc_expr})^0.5 + $D$78/(1+{wacc_expr})^1.5 + $E$78/(1+{wacc_expr})^2.5 + "
                f"$F$78/(1+{wacc_expr})^3.5 + $G$78/(1+{wacc_expr})^4.5 + "
                f"($G$78*(1+$H$44)/({wacc_expr}-$H$44))/(1+{wacc_expr})^4.5) - $B$12)/$B$8"
            )
            cell = ws_dcf.cell(row=row_idx, column=col_idx, value=formula)
            cell.font = font_bold if (row_idx == 134 and col_idx == 4) else font_regular
            cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
            cell.alignment = align_right
            cell.border = border_cell
            if row_idx == 134 and col_idx == 4:
                cell.fill = fill_accent_blue  # Center cell = Base Case!
                
    # --- Column Widths Auto-fit ---
    col_widths = {
        "A": 48,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 20
    }
    for col_let, width in col_widths.items():
        ws_dcf.column_dimensions[col_let].width = width
        
    wb.save(output_path)
    print(f"Model saved successfully to {output_path}")

if __name__ == "__main__":
    build_krus_dcf_model()
