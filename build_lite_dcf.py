import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

def build_lite_dcf_model(output_path="LITE_DCF_Model_Gemini-3.7-Flash_20260818.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: DCF Valuation
    ws_dcf = wb.active
    ws_dcf.title = "DCF Valuation"
    ws_dcf.views.sheetView[0].showGridLines = True
    
    # Sheet 2: WACC Build
    ws_wacc = wb.create_sheet(title="WACC Build")
    ws_wacc.views.sheetView[0].showGridLines = True
    
    # Palette definition (Institutional Classic Navy & Slate Blue)
    NAVY = "1F4E79"        # Primary headers
    MED_BLUE = "2F5597"    # Secondary headers
    SOFT_BLUE = "D9E1F2"   # Table headers
    ACCENT_BLUE = "BDD7EE" # Highlight / Base Case
    LIGHT_GRAY = "F2F2F2"  # Input fill
    ZEBRA_FILL = "F9FBFD"  # Alternating row
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"
    INPUT_BLUE = "0000FF"  # Hardcoded font
    BLACK = "000000"       # Formula font
    LINK_GREEN = "008000"  # Sheet link font
    
    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_regular = Font(name="Calibri", size=10, bold=False, color=BLACK)
    font_input = Font(name="Calibri", size=10, bold=False, color=INPUT_BLUE)
    font_input_bold = Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE)
    font_link = Font(name="Calibri", size=10, bold=False, color=LINK_GREEN)
    font_link_bold = Font(name="Calibri", size=10, bold=True, color=LINK_GREEN)
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_navy = Side(border_style="medium", color=NAVY)
    double_navy = Side(border_style="double", color=NAVY)
    top_thin_navy = Side(border_style="thin", color=NAVY)
    
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_navy)
    border_total = Border(top=top_thin_navy, bottom=double_navy, left=thin_line, right=thin_line)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Helper for merged section header
    def add_section_header(ws, row, title, max_col="J"):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_sec_hdr
        ws[f"A{row}"].fill = fill_navy
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_navy
            c.border = Border(top=thick_navy, bottom=thick_navy, left=thin_line, right=thin_line)
    
    # ==========================================
    # SHEET 2: WACC Build FIRST (for clean linking)
    # ==========================================
    ws_wacc["A1"] = "Lumentum Holdings Inc. (NASDAQ: LITE)"
    ws_wacc["A1"].font = font_title
    ws_wacc["A2"] = "Weighted Average Cost of Capital (WACC) Schedule | CAPM Methodology & Capital Structure"
    ws_wacc["A2"].font = font_subtitle
    
    add_section_header(ws_wacc, 4, "I. COST OF EQUITY (CAPM METHODOLOGY)", "E")
    
    wacc_equity_inputs = [
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0474, "0.00%", "US 10-Year Treasury Yield benchmark as of August 18, 2026", "B5"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.50, "0.00", "5-Year Monthly Equity Beta vs S&P 500 index (Reflects AI infrastructure volatility)", "B6"),
        ("Market Equity Risk Premium (ERP)", 0.0550, "0.00%", "Standard institutional equity risk premium (5.50%)", "B7"),
        ("Cost of Equity (Ke = Rf + Beta * ERP)", "=B5+B6*B7", "0.00%", None, "B8"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_equity_inputs, start=5):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model")
    
    add_section_header(ws_wacc, 10, "II. COST OF DEBT & CAPITAL STRUCTURE WEIGHTS", "E")
    
    wacc_debt_inputs = [
        ("Pre-Tax Cost of Debt (Kd)", 0.0550, "0.00%", "Weighted average coupon and effective borrowing rate on convertible senior notes", "B11"),
        ("Effective Corporate Tax Rate (t)", 0.2100, "0.0%", "Statutory US federal & state marginal corporate tax rate", "B12"),
        ("After-Tax Cost of Debt (Kd * (1 - t))", "=B11*(1-B12)", "0.00%", None, "B13"),
        ("Current Share Price ($)", "='DCF Valuation'!B7", "$#,##0.00", None, "B14"),
        ("Diluted Shares Outstanding (M)", "='DCF Valuation'!B8", "#,##0.00", None, "B15"),
        ("Market Capitalization ($M)", "=B14*B15", "$#,##0.00", None, "B16"),
        ("Total Debt (Convertible Senior Notes) ($M)", 2020.00, "$#,##0.00", "Carrying value of convertible senior notes following $1.1B equitization in Q4 FY2026", "B17"),
        ("Cash and Cash Equivalents ($M)", 2740.00, "$#,##0.00", "Cash, cash equivalents & short-term investments as of June 27, 2026 Form 10-K", "B18"),
        ("Net Debt ($M) [Net Cash if negative]", "=B17-B18", "$#,##0.00", None, "B19"),
        ("Enterprise Value ($M)", "=B16+B19", "$#,##0.00", None, "B20"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_debt_inputs, start=11):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "='DCF Valuation'" in str(val):
            ws_wacc[f"B{idx}"].font = font_link
        elif "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model")
    
    add_section_header(ws_wacc, 22, "III. WACC CALCULATION SUMMARY", "E")
    
    wacc_summary_headers = ["Component", "Market Value ($M)", "Capital Weight (%)", "Cost of Capital (%)", "Weighted Contribution (%)"]
    for c_idx, h_text in enumerate(wacc_summary_headers, start=1):
        c = ws_wacc.cell(row=23, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    ws_wacc["A24"] = "Common Equity (Market Cap)"
    ws_wacc["B24"] = "=B16"
    ws_wacc["C24"] = "=B24/$B$20"
    ws_wacc["D24"] = "=B8"
    ws_wacc["E24"] = "=C24*D24"
    
    ws_wacc["A25"] = "Net Debt (Net Cash Adjustment)"
    ws_wacc["B25"] = "=B19"
    ws_wacc["C25"] = "=B25/$B$20"
    ws_wacc["D25"] = "=B13"
    ws_wacc["E25"] = "=C25*D25"
    
    ws_wacc["A26"] = "Total Capitalization / Base WACC"
    ws_wacc["B26"] = "=B20"
    ws_wacc["C26"] = "=SUM(C24:C25)"
    ws_wacc["D26"] = "-"
    ws_wacc["E26"] = "=SUM(E24:E25)"
    
    for r in range(24, 27):
        ws_wacc[f"A{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"A{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"B{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"B{r}"].number_format = "$#,##0.00"
        ws_wacc[f"B{r}"].alignment = align_right
        ws_wacc[f"B{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"C{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"C{r}"].number_format = "0.0%"
        ws_wacc[f"C{r}"].alignment = align_right
        ws_wacc[f"C{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"D{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"D{r}"].number_format = "0.00%" if r != 26 else "@"
        ws_wacc[f"D{r}"].alignment = align_right
        ws_wacc[f"D{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"E{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"E{r}"].number_format = "0.00%"
        ws_wacc[f"E{r}"].alignment = align_right
        ws_wacc[f"E{r}"].border = border_total if r == 26 else border_cell
        if r == 26:
            ws_wacc[f"E{r}"].fill = fill_accent_blue
            
    ws_wacc.column_dimensions["A"].width = 44
    ws_wacc.column_dimensions["B"].width = 20
    ws_wacc.column_dimensions["C"].width = 18
    ws_wacc.column_dimensions["D"].width = 18
    ws_wacc.column_dimensions["E"].width = 24
    
    # ==========================================
    # SHEET 1: DCF Valuation
    # ==========================================
    
    # --- Title Block ---
    ws_dcf["A1"] = "Lumentum Holdings Inc. (NASDAQ: LITE)"
    ws_dcf["A1"].font = font_title
    ws_dcf["A2"] = "Institutional DCF Valuation Model | Report Date: August 18, 2026 | Financial Source: SEC Form 10-K & FY26 Results"
    ws_dcf["A2"].font = font_subtitle
    
    # --- Active Scenario Selector ---
    ws_dcf["A4"] = "Active Scenario Selector:"
    ws_dcf["A4"].font = font_bold
    ws_dcf["B4"] = 2
    ws_dcf["B4"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    ws_dcf["B4"].fill = fill_navy
    ws_dcf["B4"].alignment = align_center
    ws_dcf["B4"].border = border_cell
    ws_dcf["B4"].comment = Comment("Case Selector: 1 = Bear Case, 2 = Base Case (Consensus AI Ramp), 3 = Bull Case (Supercycle)", "Model Builder")
    
    ws_dcf["C4"] = '=IF(B4=1,"[1] BEAR CASE (AI Digestion / Slowdown)",IF(B4=2,"[2] BASE CASE (Consensus AI / 1.6T Ramp)","[3] BULL CASE (Accelerated AI Supercycle)"))'
    ws_dcf["C4"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws_dcf["C4"].alignment = align_left
    
    # --- Section I: Market Data & Capital Structure Inputs ---
    add_section_header(ws_dcf, 6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS", "J")
    
    market_inputs_dcf = [
        ("Current Stock Price ($)", 968.90, "$#,##0.00", "Market close price as of August 18, 2026", "B7"),
        ("Diluted Shares Outstanding (M)", 89.70, "#,##0.00", "SEC Form 10-K FY2026 Diluted share count as of August 14, 2026 (89.7M shares)", "B8"),
        ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
        ("Cash and Cash Equivalents ($M)", 2740.00, "$#,##0.00", "SEC Form 10-K Balance sheet as of June 27, 2026 ($2.74B in Cash & Short-Term Investments)", "B10"),
        ("Total Debt (Convertible Senior Notes) ($M)", 2020.00, "$#,##0.00", "SEC Form 10-K Balance sheet (Reflects $1.1B convertible note equitization in Q4 FY26)", "B11"),
        ("Net Debt ($M) [Net Cash if negative]", "=B11-B10", "$#,##0.00", None, "B12"),
        ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
        ("Effective Corporate Tax Rate (%)", 0.210, "0.0%", "US Federal & State statutory corporate tax rate", "B14"),
        ("Risk-Free Rate (10-Yr US Treasury) (%)", "='WACC Build'!B5", "0.00%", None, "B15"),
        ("Equity Beta (5-Year Monthly)", "='WACC Build'!B6", "0.00", None, "B16"),
        ("Market Equity Risk Premium (ERP) (%)", "='WACC Build'!B7", "0.00%", None, "B17"),
        ("Base Case WACC (%)", "='WACC Build'!E26", "0.00%", None, "B18"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs_dcf, start=7):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        if "='WACC Build'" in str(val):
            ws_dcf[f"B{idx}"].font = font_link_bold
        elif "=" in str(val):
            ws_dcf[f"B{idx}"].font = font_bold
        else:
            ws_dcf[f"B{idx}"].font = font_input
            ws_dcf[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_dcf[f"B{idx}"].comment = Comment(comment_text, "Lumentum 10-K / Market Data")
                
    # --- Section II: Scenario Assumptions & Driver Blocks ---
    add_section_header(ws_dcf, 20, "II. SCENARIO ASSUMPTIONS & DRIVER BLOCKS", "J")
    
    headers_s2 = ["Driver / Assumption", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E", "Terminal / WACC"]
    for col_idx, text in enumerate(headers_s2, start=1):
        c = ws_dcf.cell(row=21, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header
    
    # 3 Scenario blocks
    scenario_blocks = [
        ("Bear Case Assumptions (Selector = 1)", 22, [
            ("  Revenue Growth (%)", [0.750, 0.250, 0.150, 0.080, 0.050], 0.025, "0.0%"),
            ("  Gross Margin (%)", [0.460, 0.470, 0.475, 0.480, 0.480], None, "0.0%"),
            ("  EBIT Margin (%)", [0.280, 0.300, 0.310, 0.315, 0.320], 0.1350, "0.0%"),
            ("  D&A (% of Revenue)", [0.070, 0.065, 0.060, 0.055, 0.050], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.085, 0.075, 0.070, 0.060, 0.055], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ]),
        ("Base Case Assumptions (Selector = 2)", 29, [
            ("  Revenue Growth (%)", [1.000, 0.500, 0.300, 0.180, 0.100], 0.030, "0.0%"),
            ("  Gross Margin (%)", [0.505, 0.520, 0.530, 0.535, 0.540], None, "0.0%"),
            ("  EBIT Margin (%)", [0.325, 0.360, 0.385, 0.400, 0.415], "='WACC Build'!E26", "0.0%"),
            ("  D&A (% of Revenue)", [0.065, 0.060, 0.055, 0.050, 0.045], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.075, 0.070, 0.065, 0.055, 0.050], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ]),
        ("Bull Case Assumptions (Selector = 3)", 36, [
            ("  Revenue Growth (%)", [1.200, 0.650, 0.400, 0.250, 0.150], 0.035, "0.0%"),
            ("  Gross Margin (%)", [0.520, 0.540, 0.555, 0.565, 0.570], None, "0.0%"),
            ("  EBIT Margin (%)", [0.350, 0.390, 0.420, 0.435, 0.445], 0.1250, "0.0%"),
            ("  D&A (% of Revenue)", [0.060, 0.055, 0.050, 0.045, 0.040], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.070, 0.065, 0.060, 0.050, 0.045], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ]),
    ]
    
    for b_title, start_r, rows_data in scenario_blocks:
        ws_dcf[f"A{start_r}"] = b_title
        ws_dcf[f"A{start_r}"].font = font_bold_navy
        ws_dcf[f"A{start_r}"].fill = fill_soft_blue
        ws_dcf.merge_cells(f"A{start_r}:H{start_r}")
        for col_idx in range(1, 9):
            ws_dcf.cell(row=start_r, column=col_idx).fill = fill_soft_blue
            ws_dcf.cell(row=start_r, column=col_idx).border = border_header
            
        for r_offset, (label, val_list, term_val, num_fmt) in enumerate(rows_data, start=1):
            curr_r = start_r + r_offset
            ws_dcf[f"A{curr_r}"] = label
            ws_dcf[f"A{curr_r}"].font = font_regular
            ws_dcf[f"A{curr_r}"].border = border_cell
            ws_dcf[f"B{curr_r}"] = "-"
            ws_dcf[f"B{curr_r}"].alignment = align_center
            ws_dcf[f"B{curr_r}"].border = border_cell
            
            for c_idx, val in enumerate(val_list, start=3):
                col_let = get_column_letter(c_idx)
                ws_dcf[f"{col_let}{curr_r}"] = val
                ws_dcf[f"{col_let}{curr_r}"].font = font_input
                ws_dcf[f"{col_let}{curr_r}"].fill = fill_input
                ws_dcf[f"{col_let}{curr_r}"].number_format = num_fmt
                ws_dcf[f"{col_let}{curr_r}"].alignment = align_right
                ws_dcf[f"{col_let}{curr_r}"].border = border_cell
                
            ws_dcf[f"H{curr_r}"] = term_val if term_val is not None else "-"
            ws_dcf[f"H{curr_r}"].border = border_cell
            if term_val is not None:
                if isinstance(term_val, str) and term_val.startswith("="):
                    ws_dcf[f"H{curr_r}"].font = font_link_bold
                else:
                    ws_dcf[f"H{curr_r}"].font = font_input_bold
                    ws_dcf[f"H{curr_r}"].fill = fill_input
                ws_dcf[f"H{curr_r}"].number_format = num_fmt
                ws_dcf[f"H{curr_r}"].alignment = align_right
            else:
                ws_dcf[f"H{curr_r}"].alignment = align_center
                
    # --- Consolidated Active Driver Table (Rows 43-49) ---
    ws_dcf["A43"] = "ACTIVE CONSOLIDATED DRIVERS (PULLED VIA INDEX FORMULAS)"
    ws_dcf["A43"].font = font_bold_navy
    ws_dcf["A43"].fill = fill_accent_blue
    ws_dcf.merge_cells("A43:H43")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=43, column=col_idx).fill = fill_accent_blue
        ws_dcf.cell(row=43, column=col_idx).border = border_header
        
    driver_rows_map = [
        ("  Active Revenue Growth (%)", [23, 30, 37], 44, "0.0%"),
        ("  Active Gross Margin (%)", [24, 31, 38], 45, "0.0%"),
        ("  Active EBIT Margin (%)", [25, 32, 39], 46, "0.0%"),
        ("  Active D&A (% of Revenue)", [26, 33, 40], 47, "0.0%"),
        ("  Active CapEx (% of Revenue)", [27, 34, 41], 48, "0.0%"),
        ("  Active Δ NWC (% of Δ Revenue)", [28, 35, 42], 49, "0.0%"),
    ]
    
    for label, src_rows, target_r, num_fmt in driver_rows_map:
        ws_dcf[f"A{target_r}"] = label
        ws_dcf[f"A{target_r}"].font = font_bold
        ws_dcf[f"A{target_r}"].border = border_cell
        ws_dcf[f"B{target_r}"] = "-"
        ws_dcf[f"B{target_r}"].alignment = align_center
        ws_dcf[f"B{target_r}"].border = border_cell
        
        # FY2027E to FY2031E (Cols C to G)
        for col_idx in range(3, 8):
            c_let = get_column_letter(col_idx)
            formula = f"=INDEX(({c_let}${src_rows[0]},{c_let}${src_rows[1]},{c_let}${src_rows[2]}),1,1,$B$4)"
            ws_dcf[f"{c_let}{target_r}"] = formula
            ws_dcf[f"{c_let}{target_r}"].font = font_bold
            ws_dcf[f"{c_let}{target_r}"].number_format = num_fmt
            ws_dcf[f"{c_let}{target_r}"].alignment = align_right
            ws_dcf[f"{c_let}{target_r}"].border = border_cell
            
        # Terminal / WACC column (Col H)
        if target_r in [44, 46]: # Terminal growth or WACC
            formula_h = f"=INDEX((H${src_rows[0]},H${src_rows[1]},H${src_rows[2]}),1,1,$B$4)"
            ws_dcf[f"H{target_r}"] = formula_h
            ws_dcf[f"H{target_r}"].font = font_bold
            ws_dcf[f"H{target_r}"].number_format = num_fmt
            ws_dcf[f"H{target_r}"].alignment = align_right
            ws_dcf[f"H{target_r}"].border = border_cell
        else:
            ws_dcf[f"H{target_r}"] = "-"
            ws_dcf[f"H{target_r}"].alignment = align_center
            ws_dcf[f"H{target_r}"].border = border_cell
            
    # --- Section III: Historical & Projected Income Statement ---
    add_section_header(ws_dcf, 51, "III. FINANCIAL PERFORMANCE & 5-YEAR PROJECTIONS ($M)", "J")
    
    headers_is = ["Line Item ($M)", "FY2023A", "FY2024A", "FY2025A", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
    for c_idx, h_text in enumerate(headers_is, start=1):
        c = ws_dcf.cell(row=52, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if c_idx > 1 else align_left
        c.border = border_header
        
    # Row 53: Net Revenue
    ws_dcf["A53"] = "Net Revenue"
    ws_dcf["B53"] = 1541.6
    ws_dcf["C53"] = 1359.2
    ws_dcf["D53"] = 1645.0
    ws_dcf["E53"] = 3014.0
    ws_dcf["F53"] = "=E53*(1+C44)"
    ws_dcf["G53"] = "=F53*(1+D44)"
    ws_dcf["H53"] = "=G53*(1+E44)"
    ws_dcf["I53"] = "=H53*(1+F44)"
    ws_dcf["J53"] = "=I53*(1+G44)"
    
    # Row 54: YoY Revenue Growth (%)
    ws_dcf["A54"] = "  YoY Growth (%)"
    ws_dcf["B54"] = "-"
    ws_dcf["C54"] = "=C53/B53-1"
    ws_dcf["D54"] = "=D53/C53-1"
    ws_dcf["E54"] = "=E53/D53-1"
    ws_dcf["F54"] = "=F53/E53-1"
    ws_dcf["G54"] = "=G53/F53-1"
    ws_dcf["H54"] = "=H53/G53-1"
    ws_dcf["I54"] = "=I53/H53-1"
    ws_dcf["J54"] = "=J53/I53-1"
    
    # Row 55: Gross Profit
    ws_dcf["A55"] = "Gross Profit"
    ws_dcf["B55"] = 665.4
    ws_dcf["C55"] = 447.8
    ws_dcf["D55"] = 687.2
    ws_dcf["E55"] = 1438.0
    ws_dcf["F55"] = "=F53*C45"
    ws_dcf["G55"] = "=G53*D45"
    ws_dcf["H55"] = "=H53*E45"
    ws_dcf["I55"] = "=I53*F45"
    ws_dcf["J55"] = "=J53*G45"
    
    # Row 56: Gross Margin (%)
    ws_dcf["A56"] = "  Gross Margin (%)"
    ws_dcf["B56"] = "=B55/B53"
    ws_dcf["C56"] = "=C55/C53"
    ws_dcf["D56"] = "=D55/D53"
    ws_dcf["E56"] = "=E55/E53"
    ws_dcf["F56"] = "=F55/F53"
    ws_dcf["G56"] = "=G55/G53"
    ws_dcf["H56"] = "=H55/H53"
    ws_dcf["I56"] = "=I55/I53"
    ws_dcf["J56"] = "=J55/J53"
    
    # Row 57: Research & Development (R&D)
    ws_dcf["A57"] = "  Research & Development (R&D)"
    ws_dcf["B57"] = 274.2
    ws_dcf["C57"] = 289.5
    ws_dcf["D57"] = 320.0
    ws_dcf["E57"] = 356.5
    ws_dcf["F57"] = "=F53*0.105"
    ws_dcf["G57"] = "=G53*0.095"
    ws_dcf["H57"] = "=H53*0.090"
    ws_dcf["I57"] = "=I53*0.085"
    ws_dcf["J57"] = "=J53*0.080"
    
    # Row 58: Selling, General & Administrative (SG&A)
    ws_dcf["A58"] = "  Selling, General & Administrative (SG&A)"
    ws_dcf["B58"] = 248.5
    ws_dcf["C58"] = 258.0
    ws_dcf["D58"] = 286.0
    ws_dcf["E58"] = 289.5
    ws_dcf["F58"] = "=F55-F57-F60"
    ws_dcf["G58"] = "=G55-G57-G60"
    ws_dcf["H58"] = "=H55-H57-H60"
    ws_dcf["I58"] = "=I55-I57-I60"
    ws_dcf["J58"] = "=J55-J57-J60"
    
    # Row 59: Total Operating Expenses
    ws_dcf["A59"] = "Total Operating Expenses"
    ws_dcf["B59"] = "=B57+B58"
    ws_dcf["C59"] = "=C57+C58"
    ws_dcf["D59"] = "=D57+D58"
    ws_dcf["E59"] = "=E57+E58"
    ws_dcf["F59"] = "=F57+F58"
    ws_dcf["G59"] = "=G57+G58"
    ws_dcf["H59"] = "=H57+H58"
    ws_dcf["I59"] = "=I57+I58"
    ws_dcf["J59"] = "=J57+J58"
    
    # Row 60: Operating Income (EBIT)
    ws_dcf["A60"] = "Operating Income (EBIT)"
    ws_dcf["B60"] = 142.7
    ws_dcf["C60"] = -99.7
    ws_dcf["D60"] = 81.2
    ws_dcf["E60"] = 792.0
    ws_dcf["F60"] = "=F53*C46"
    ws_dcf["G60"] = "=G53*D46"
    ws_dcf["H60"] = "=H53*E46"
    ws_dcf["I60"] = "=I53*F46"
    ws_dcf["J60"] = "=J53*G46"
    
    # Row 61: EBIT Margin (%)
    ws_dcf["A61"] = "  EBIT Margin (%)"
    ws_dcf["B61"] = "=B60/B53"
    ws_dcf["C61"] = "=C60/C53"
    ws_dcf["D61"] = "=D60/D53"
    ws_dcf["E61"] = "=E60/E53"
    ws_dcf["F61"] = "=F60/F53"
    ws_dcf["G61"] = "=G60/G53"
    ws_dcf["H61"] = "=H60/H53"
    ws_dcf["I61"] = "=I60/I53"
    ws_dcf["J61"] = "=J60/J53"
    
    # Row 62: Taxes on Operating Income (EBIT * t)
    ws_dcf["A62"] = "(-) Taxes on EBIT"
    ws_dcf["B62"] = "=IF(B60>0,B60*$B$14,0)"
    ws_dcf["C62"] = "=IF(C60>0,C60*$B$14,0)"
    ws_dcf["D62"] = "=IF(D60>0,D60*$B$14,0)"
    ws_dcf["E62"] = "=IF(E60>0,E60*$B$14,0)"
    ws_dcf["F62"] = "=IF(F60>0,F60*$B$14,0)"
    ws_dcf["G62"] = "=IF(G60>0,G60*$B$14,0)"
    ws_dcf["H62"] = "=IF(H60>0,H60*$B$14,0)"
    ws_dcf["I62"] = "=IF(I60>0,I60*$B$14,0)"
    ws_dcf["J62"] = "=IF(J60>0,J60*$B$14,0)"
    
    # Row 63: Net Operating Profit After Tax (NOPAT)
    ws_dcf["A63"] = "Net Operating Profit After Tax (NOPAT)"
    ws_dcf["B63"] = "=B60-B62"
    ws_dcf["C63"] = "=C60-C62"
    ws_dcf["D63"] = "=D60-D62"
    ws_dcf["E63"] = "=E60-E62"
    ws_dcf["F63"] = "=F60-F62"
    ws_dcf["G63"] = "=G60-G62"
    ws_dcf["H63"] = "=H60-H62"
    ws_dcf["I63"] = "=I60-I62"
    ws_dcf["J63"] = "=J60-J62"
    
    # Formatting for IS block
    for r in range(53, 64):
        is_bold_row = r in [53, 55, 59, 60, 63]
        ws_dcf[f"A{r}"].font = font_bold if is_bold_row else font_regular
        ws_dcf[f"A{r}"].border = border_cell
        
        for col_idx in range(2, 11):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_cell
            cell.alignment = align_right
            
            if r in [54, 56, 61]:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                
            if str(cell.value).startswith("="):
                cell.font = font_bold if is_bold_row else font_regular
            elif cell.value == "-":
                cell.font = font_regular
                cell.alignment = align_center
            else:
                cell.font = font_input_bold if is_bold_row else font_input
                cell.fill = fill_input
                cell.comment = Comment(f"Source: Form 10-K FY2026 / Historic Financials", "DCF Model")
                
    # --- Section IV: Free Cash Flow Schedule (Rows 65-73) ---
    add_section_header(ws_dcf, 65, "IV. UNLEVERED FREE CASH FLOW (UFCF) BUILD ($M)", "J")
    
    headers_fcf = ["Cash Flow Line Item ($M)", "FY2023A", "FY2024A", "FY2025A", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
    for c_idx, h_text in enumerate(headers_fcf, start=1):
        c = ws_dcf.cell(row=66, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if c_idx > 1 else align_left
        c.border = border_header
        
    # Row 67: NOPAT
    ws_dcf["A67"] = "Net Operating Profit After Tax (NOPAT)"
    for col_idx in range(2, 11):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}67"] = f"={c_let}63"
        
    # Row 68: (+) Depreciation & Amortization (D&A)
    ws_dcf["A68"] = "(+) Depreciation & Amortization (D&A)"
    ws_dcf["B68"] = 138.0
    ws_dcf["C68"] = 165.0
    ws_dcf["D68"] = 182.0
    ws_dcf["E68"] = 215.0
    ws_dcf["F68"] = "=F53*C47"
    ws_dcf["G68"] = "=G53*D47"
    ws_dcf["H68"] = "=H53*E47"
    ws_dcf["I68"] = "=I53*F47"
    ws_dcf["J68"] = "=J53*G47"
    
    # Row 69: (-) Capital Expenditures (CapEx)
    ws_dcf["A69"] = "(-) Capital Expenditures (CapEx)"
    ws_dcf["B69"] = 145.0
    ws_dcf["C69"] = 168.0
    ws_dcf["D69"] = 195.0
    ws_dcf["E69"] = 245.0
    ws_dcf["F69"] = "=F53*C48"
    ws_dcf["G69"] = "=G53*D48"
    ws_dcf["H69"] = "=H53*E48"
    ws_dcf["I69"] = "=I53*F48"
    ws_dcf["J69"] = "=J53*G48"
    
    # Row 70: (-) Change in Net Working Capital (Δ NWC)
    ws_dcf["A70"] = "(-) Change in Net Working Capital (Δ NWC)"
    ws_dcf["B70"] = -12.0
    ws_dcf["C70"] = -25.0
    ws_dcf["D70"] = 18.0
    ws_dcf["E70"] = 45.0
    ws_dcf["F70"] = "=(F53-E53)*C49"
    ws_dcf["G70"] = "=(G53-F53)*D49"
    ws_dcf["H70"] = "=(H53-G53)*E49"
    ws_dcf["I70"] = "=(I53-H53)*F49"
    ws_dcf["J70"] = "=(J53-I53)*G49"
    
    # Row 71: Unlevered Free Cash Flow (UFCF)
    ws_dcf["A71"] = "Unlevered Free Cash Flow (UFCF)"
    for col_idx in range(2, 11):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}71"] = f"={c_let}67+{c_let}68-{c_let}69-{c_let}70"
        
    # Row 72: UFCF Conversion Margin (% of Revenue)
    ws_dcf["A72"] = "  FCF Conversion Margin (% of Revenue)"
    for col_idx in range(2, 11):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}72"] = f"={c_let}71/{c_let}53"
        
    for r in range(67, 73):
        is_bold_row = r in [67, 71]
        ws_dcf[f"A{r}"].font = font_bold if is_bold_row else font_regular
        ws_dcf[f"A{r}"].border = border_total if r == 71 else border_cell
        
        for col_idx in range(2, 11):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_total if r == 71 else border_cell
            cell.alignment = align_right
            
            if r == 72:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                
            if str(cell.value).startswith("="):
                cell.font = font_bold if is_bold_row else font_regular
            else:
                cell.font = font_input_bold if is_bold_row else font_input
                cell.fill = fill_input
                cell.comment = Comment(f"Source: Form 10-K FY2026 Cash Flow Statement", "DCF Model")
                
            if r == 71:
                cell.fill = fill_accent_blue
                
    # --- Section V: Discounting Schedule & Valuation Summary (Rows 74-96) ---
    add_section_header(ws_dcf, 74, "V. DISCOUNTING SCHEDULE & VALUATION SUMMARY ($M)", "J")
    
    headers_val = ["DCF Component ($M)", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E", "Terminal Year", "", "", ""]
    for c_idx, h_text in enumerate(headers_val, start=1):
        c = ws_dcf.cell(row=75, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if 1 < c_idx <= 7 else align_left
        c.border = border_header
        
    # Row 76: Explicit UFCF
    ws_dcf["A76"] = "Unlevered Free Cash Flow (UFCF)"
    ws_dcf["B76"] = "=F71"
    ws_dcf["C76"] = "=G71"
    ws_dcf["D76"] = "=H71"
    ws_dcf["E76"] = "=I71"
    ws_dcf["F76"] = "=J71"
    ws_dcf["G76"] = "=F76*(1+$H$44)" # Terminal UFCF = Final Year UFCF * (1 + g)
    
    # Row 77: Mid-Year Discount Period (t)
    ws_dcf["A77"] = "Mid-Year Discount Period (t)"
    ws_dcf["B77"] = 0.5
    ws_dcf["C77"] = 1.5
    ws_dcf["D77"] = 2.5
    ws_dcf["E77"] = 3.5
    ws_dcf["F77"] = 4.5
    ws_dcf["G77"] = 4.5 # For terminal value discount
    
    # Row 78: Discount Factor (1 / (1 + WACC)^t)
    ws_dcf["A78"] = "Discount Factor (WACC Discounting)"
    ws_dcf["B78"] = "=1/(1+$H$46)^B77"
    ws_dcf["C78"] = "=1/(1+$H$46)^C77"
    ws_dcf["D78"] = "=1/(1+$H$46)^D77"
    ws_dcf["E78"] = "=1/(1+$H$46)^E77"
    ws_dcf["F78"] = "=1/(1+$H$46)^F77"
    ws_dcf["G78"] = "=1/(1+$H$46)^G77"
    
    # Row 79: Present Value of Explicit FCF (PV)
    ws_dcf["A79"] = "Present Value of Cash Flow (PV of UFCF)"
    ws_dcf["B79"] = "=B76*B78"
    ws_dcf["C79"] = "=C76*C78"
    ws_dcf["D79"] = "=D76*D78"
    ws_dcf["E79"] = "=E76*E78"
    ws_dcf["F79"] = "=F76*F78"
    ws_dcf["G79"] = "-"
    
    # Row 80: Terminal Value calculation
    ws_dcf["A80"] = "Nominal Terminal Value (Perpetuity Growth Method)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}80"] = "-"
    ws_dcf["G80"] = "=G76/($H$46-$H$44)" # TV = Terminal FCF / (WACC - g)
    
    # Row 81: PV of Terminal Value
    ws_dcf["A81"] = "Present Value of Terminal Value (PV of TV)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}81"] = "-"
    ws_dcf["G81"] = "=G80*G78"
    
    for r in range(76, 82):
        ws_dcf[f"A{r}"].font = font_bold if r in [76, 79, 80, 81] else font_regular
        ws_dcf[f"A{r}"].border = border_cell
        for col_idx in range(2, 8):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_cell
            cell.alignment = align_right
            if r == 77:
                cell.number_format = "0.0"
                cell.font = font_regular
            elif r == 78:
                cell.number_format = "0.0000"
                cell.font = font_regular
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                cell.font = font_bold if r in [76, 79, 81] else font_regular
            if str(cell.value) == "-":
                cell.alignment = align_center
                
    # --- Equity Value Bridge (Rows 83-94) ---
    ws_dcf["A83"] = "VALUATION BRIDGE & PER-SHARE IMPLICATION"
    ws_dcf["A83"].font = font_bold_navy
    ws_dcf["A83"].fill = fill_soft_blue
    ws_dcf.merge_cells("A83:C83")
    for col_idx in range(1, 4):
        ws_dcf.cell(row=83, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=83, column=col_idx).border = border_header
        
    equity_bridge = [
        ("Cumulative PV of Explicit 5-Year FCFs ($M)", "=SUM(B79:F79)", "$#,##0.00", "B84", False),
        ("Present Value of Terminal Value ($M)", "=G81", "$#,##0.00", "B85", False),
        ("Enterprise Value ($M)", "=B84+B85", "$#,##0.00", "B86", True),
        ("(-) Total Debt ($M)", "=-B11", "$#,##0.00", "B87", False),
        ("(+) Cash and Short-Term Investments ($M)", "=B10", "$#,##0.00", "B88", False),
        ("Net Debt Adjustment ($M) [Net Cash = +$]", "=B87+B88", "$#,##0.00", "B89", False),
        ("Implied Equity Value ($M)", "=B86+B89", "$#,##0.00", "B90", True),
        ("Diluted Shares Outstanding (M)", "=B8", "#,##0.00", "B91", False),
        ("IMPLIED INTRINSIC VALUE PER SHARE ($)", "=B90/B91", "$#,##0.00", "B92", True),
        ("Current Market Stock Price ($)", "=B7", "$#,##0.00", "B93", False),
        ("Implied Upside / (Downside) (%)", "=B92/B93-1", "+0.0%;-0.0%;0.0%", "B94", True),
    ]
    
    for idx, (label, formula_val, num_fmt, cell_ref, is_highlight) in enumerate(equity_bridge, start=84):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"A{idx}"].border = border_total if idx in [86, 90, 92, 94] else border_cell
        
        ws_dcf[f"B{idx}"] = formula_val
        ws_dcf[f"B{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_total if idx in [86, 90, 92, 94] else border_cell
        
        if is_highlight:
            ws_dcf[f"A{idx}"].fill = fill_accent_blue
            ws_dcf[f"B{idx}"].fill = fill_accent_blue
            
    # Add Terminal Value Sanity Metrics
    ws_dcf["D84"] = "Terminal Value Metrics & Checks:"
    ws_dcf["D84"].font = font_bold_navy
    ws_dcf["D85"] = "PV of TV % of Enterprise Value"
    ws_dcf["E85"] = "=B85/B86"
    ws_dcf["E85"].number_format = "0.0%"
    ws_dcf["E85"].font = font_bold
    ws_dcf["E85"].alignment = align_right
    ws_dcf["E85"].border = border_cell
    
    ws_dcf["D86"] = "Implied Exit Multiple (EV / FY31E EBITDA)"
    ws_dcf["E86"] = "=B86/(J60+J68)"
    ws_dcf["E86"].number_format = "0.0x"
    ws_dcf["E86"].font = font_bold
    ws_dcf["E86"].alignment = align_right
    ws_dcf["E86"].border = border_cell
    
    for r_chk in [85, 86]:
        ws_dcf[f"D{r_chk}"].border = border_cell
        ws_dcf[f"D{r_chk}"].font = font_regular
        
    # =========================================================================
    # SECTION VI: THREE INSTITUTIONAL 5X5 SENSITIVITY TABLES (Rows 98-148)
    # =========================================================================
    add_section_header(ws_dcf, 98, "VI. INSTITUTIONAL SENSITIVITY MATRICES (5x5 GRIDS WITH FULL DCF RECALCULATION)", "H")
    
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 1: WACC vs Perpetual Growth Rate (g)
    # -------------------------------------------------------------------------
    ws_dcf["A100"] = "TABLE 1: IMPLIED SHARE PRICE ($) - WACC vs. TERMINAL GROWTH RATE (g)"
    ws_dcf["A100"].font = font_bold_navy
    ws_dcf["A100"].fill = fill_soft_blue
    ws_dcf.merge_cells("A100:F100")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=100, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=100, column=col_idx).border = border_header
        
    ws_dcf["A101"] = "WACC \\ Terminal g"
    ws_dcf["A101"].font = font_tbl_hdr
    ws_dcf["A101"].fill = fill_soft_blue
    ws_dcf["A101"].border = border_header
    ws_dcf["A101"].alignment = align_center
    
    # Axis ranges centered on Base Case (WACC = 12.92%, g = 3.00%)
    g_cols = [0.020, 0.025, 0.030, 0.035, 0.040]
    wacc_rows = [0.1192, 0.1242, 0.1292, 0.1342, 0.1392]
    
    for c_idx, g_val in enumerate(g_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}101"] = g_val
        ws_dcf[f"{c_let}101"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}101"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}101"].number_format = "0.0%"
        ws_dcf[f"{c_let}101"].alignment = align_center
        ws_dcf[f"{c_let}101"].border = border_header
        
    for r_idx, w_val in enumerate(wacc_rows, start=102):
        ws_dcf[f"A{r_idx}"] = w_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 104 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 104 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        # Populate 5 columns of formulas programmatically
        for c_idx, g_val in enumerate(g_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t1 = (
                f"=(($B$76/(1+$A{r_idx})^0.5 + $C$76/(1+$A{r_idx})^1.5 + $D$76/(1+$A{r_idx})^2.5 + $E$76/(1+$A{r_idx})^3.5 + $F$76/(1+$A{r_idx})^4.5) + "
                f"(($F$76*(1+{c_let}$101)/($A{r_idx}-{c_let}$101))/(1+$A{r_idx})^4.5) + $B$89)/$B$91"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t1
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 104 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 2: Revenue Growth Scale vs. Target EBIT Margin
    # -------------------------------------------------------------------------
    ws_dcf["A110"] = "TABLE 2: IMPLIED SHARE PRICE ($) - REVENUE GROWTH MULTIPLIER vs. FY31E EBIT MARGIN"
    ws_dcf["A110"].font = font_bold_navy
    ws_dcf["A110"].fill = fill_soft_blue
    ws_dcf.merge_cells("A110:F110")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=110, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=110, column=col_idx).border = border_header
        
    ws_dcf["A111"] = "Growth Index \\ EBIT Margin"
    ws_dcf["A111"].font = font_tbl_hdr
    ws_dcf["A111"].fill = fill_soft_blue
    ws_dcf["A111"].border = border_header
    ws_dcf["A111"].alignment = align_center
    
    ebit_cols = [0.355, 0.385, 0.415, 0.445, 0.475]
    growth_rows = [0.80, 0.90, 1.00, 1.10, 1.20]
    
    for c_idx, em_val in enumerate(ebit_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}111"] = em_val
        ws_dcf[f"{c_let}111"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}111"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}111"].number_format = "0.0%"
        ws_dcf[f"{c_let}111"].alignment = align_center
        ws_dcf[f"{c_let}111"].border = border_header
        
    for r_idx, gr_val in enumerate(growth_rows, start=112):
        ws_dcf[f"A{r_idx}"] = gr_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 114 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 114 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, em_val in enumerate(ebit_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t2 = (
                f"=((($B$76*$A{r_idx}*({c_let}$111/$G$46)/(1+$H$46)^0.5 + "
                f"$C$76*$A{r_idx}*({c_let}$111/$G$46)/(1+$H$46)^1.5 + "
                f"$D$76*$A{r_idx}*({c_let}$111/$G$46)/(1+$H$46)^2.5 + "
                f"$E$76*$A{r_idx}*({c_let}$111/$G$46)/(1+$H$46)^3.5 + "
                f"$F$76*$A{r_idx}*({c_let}$111/$G$46)/(1+$H$46)^4.5) + "
                f"(($F$76*$A{r_idx}*({c_let}$111/$G$46)*(1+$H$44)/($H$46-$H$44))/(1+$H$46)^4.5) + $B$89)/$B$91)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t2
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 114 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 3: Equity Beta vs. Risk-Free Rate (Rf)
    # -------------------------------------------------------------------------
    ws_dcf["A120"] = "TABLE 3: IMPLIED SHARE PRICE ($) - EQUITY BETA vs. RISK-FREE RATE (Rf)"
    ws_dcf["A120"].font = font_bold_navy
    ws_dcf["A120"].fill = fill_soft_blue
    ws_dcf.merge_cells("A120:F120")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=120, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=120, column=col_idx).border = border_header
        
    ws_dcf["A121"] = "Beta \\ Risk-Free Rate"
    ws_dcf["A121"].font = font_tbl_hdr
    ws_dcf["A121"].fill = fill_soft_blue
    ws_dcf["A121"].border = border_header
    ws_dcf["A121"].alignment = align_center
    
    rf_cols = [0.0374, 0.0424, 0.0474, 0.0524, 0.0574]
    beta_rows = [1.20, 1.35, 1.50, 1.65, 1.80]
    
    for c_idx, rf_val in enumerate(rf_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}121"] = rf_val
        ws_dcf[f"{c_let}121"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}121"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}121"].number_format = "0.00%"
        ws_dcf[f"{c_let}121"].alignment = align_center
        ws_dcf[f"{c_let}121"].border = border_header
        
    for r_idx, b_val in enumerate(beta_rows, start=122):
        ws_dcf[f"A{r_idx}"] = b_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 124 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 124 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, rf_val in enumerate(rf_cols, start=2):
            c_let = get_column_letter(c_idx)
            wacc_formula_snippet = f"({c_let}$121+$A{r_idx}*$B$17)*('WACC Build'!$C$24) + ('WACC Build'!$B$13)*('WACC Build'!$C$25)"
            
            formula_t3 = (
                f"=((($B$76/(1+({wacc_formula_snippet}))^0.5 + "
                f"$C$76/(1+({wacc_formula_snippet}))^1.5 + "
                f"$D$76/(1+({wacc_formula_snippet}))^2.5 + "
                f"$E$76/(1+({wacc_formula_snippet}))^3.5 + "
                f"$F$76/(1+({wacc_formula_snippet}))^4.5) + "
                f"(($F$76*(1+$H$44)/(({wacc_formula_snippet})-$H$44))/(1+({wacc_formula_snippet}))^4.5) + $B$89)/$B$91)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t3
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 124 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # Adjust Column Widths on DCF Sheet
    ws_dcf.column_dimensions["A"].width = 46
    ws_dcf.column_dimensions["B"].width = 16
    ws_dcf.column_dimensions["C"].width = 16
    ws_dcf.column_dimensions["D"].width = 16
    ws_dcf.column_dimensions["E"].width = 16
    ws_dcf.column_dimensions["F"].width = 16
    ws_dcf.column_dimensions["G"].width = 16
    ws_dcf.column_dimensions["H"].width = 16
    ws_dcf.column_dimensions["I"].width = 16
    ws_dcf.column_dimensions["J"].width = 16
    
    wb.save(output_path)
    print(f"Successfully generated institutional DCF model at: {output_path}")

if __name__ == "__main__":
    build_lite_dcf_model()
