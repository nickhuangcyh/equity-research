#!/usr/bin/env python3
"""
build_meta_dcf.py - Institutional DCF Valuation Model for Meta Platforms, Inc. (NASDAQ: META)
Follows Investment Banking financial modeling standards and strict openpyxl formatting rules.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

def build_meta_dcf_model(output_path="META_DCF_Model_Gemini-3.7-Flash_20260819.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: DCF Valuation
    ws_dcf = wb.active
    ws_dcf.title = "DCF Valuation"
    ws_dcf.views.sheetView[0].showGridLines = True
    
    # Sheet 2: WACC Build
    ws_wacc = wb.create_sheet(title="WACC Build")
    ws_wacc.views.sheetView[0].showGridLines = True
    
    # Palette definition (Institutional Classic Navy & Slate Blue)
    NAVY = "1F4E79"        # Primary headers
    MED_BLUE = "2F5597"    # Secondary headers
    SOFT_BLUE = "D9E1F2"   # Table headers
    ACCENT_BLUE = "BDD7EE" # Highlight / Base Case
    LIGHT_GRAY = "F2F2F2"  # Input fill
    ZEBRA_FILL = "F9FBFD"  # Alternating row
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"
    INPUT_BLUE = "0000FF"  # Hardcoded font
    BLACK = "000000"       # Formula font
    LINK_GREEN = "008000"  # Sheet link font
    
    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_regular = Font(name="Calibri", size=10, bold=False, color=BLACK)
    font_input = Font(name="Calibri", size=10, bold=False, color=INPUT_BLUE)
    font_input_bold = Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE)
    font_link = Font(name="Calibri", size=10, bold=False, color=LINK_GREEN)
    font_link_bold = Font(name="Calibri", size=10, bold=True, color=LINK_GREEN)
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_navy = Side(border_style="medium", color=NAVY)
    double_navy = Side(border_style="double", color=NAVY)
    top_thin_navy = Side(border_style="thin", color=NAVY)
    
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_navy)
    border_total = Border(top=top_thin_navy, bottom=double_navy, left=thin_line, right=thin_line)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Helper for merged section header
    def add_section_header(ws, row, title, max_col="H"):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_sec_hdr
        ws[f"A{row}"].fill = fill_navy
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_navy
            c.border = Border(top=thick_navy, bottom=thick_navy, left=thin_line, right=thin_line)
            
    # ==========================================
    # SHEET 2: WACC Build FIRST (for clean linking)
    # ==========================================
    ws_wacc["A1"] = "Meta Platforms, Inc. (NASDAQ: META)"
    ws_wacc["A1"].font = font_title
    ws_wacc["A2"] = "Weighted Average Cost of Capital (WACC) Schedule | CAPM Methodology & Capital Structure"
    ws_wacc["A2"].font = font_subtitle
    
    add_section_header(ws_wacc, 4, "I. COST OF EQUITY (CAPM METHODOLOGY)", "E")
    
    wacc_equity_inputs = [
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0470, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 19, 2026", "B5"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.24, "0.00", "Source: Market data 5-year monthly regression beta (Reflects AI investments & digital ad cycle)", "B6"),
        ("Market Equity Risk Premium (ERP)", 0.0550, "0.00%", "Source: Institutional consensus equity risk premium (5.50%)", "B7"),
        ("Cost of Equity (Ke = Rf + Beta * ERP)", "=B5+B6*B7", "0.00%", None, "B8"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_equity_inputs, start=5):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 10, "II. COST OF DEBT & CAPITAL STRUCTURE WEIGHTS", "E")
    
    wacc_debt_inputs = [
        ("Pre-Tax Cost of Debt (Kd)", 0.0520, "0.00%", "Source: Meta Senior Unsecured Notes yield to maturity schedule", "B11"),
        ("Effective Corporate Tax Rate (t)", 0.1600, "0.0%", "Source: Normalized corporate effective tax rate reflecting global operations", "B12"),
        ("After-Tax Cost of Debt (Kd * (1 - t))", "=B11*(1-B12)", "0.00%", None, "B13"),
        ("Current Share Price ($)", "='DCF Valuation'!B7", "$#,##0.00", None, "B14"),
        ("Diluted Shares Outstanding (M)", "='DCF Valuation'!B8", "#,##0.00", None, "B15"),
        ("Market Capitalization ($M)", "=B14*B15", "$#,##0.00", None, "B16"),
        ("Total Debt (Current + Noncurrent) ($M)", 83660.00, "$#,##0.00", "Source: SEC Form 10-Q Q2 2026 Balance Sheet total carrying debt ($83.66B)", "B17"),
        ("Cash and Liquid Investments ($M)", 90260.00, "$#,##0.00", "Source: SEC Form 10-Q Q2 2026 Balance Sheet cash & marketable securities ($90.26B)", "B18"),
        ("Net Debt ($M)", "=B17-B18", "$#,##0.00", None, "B19"),
        ("Enterprise Value ($M)", "=B16+B19", "$#,##0.00", None, "B20"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_debt_inputs, start=11):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "='DCF Valuation'" in str(val):
            ws_wacc[f"B{idx}"].font = font_link
        elif "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 22, "III. WACC CALCULATION SUMMARY", "E")
    
    wacc_summary_headers = ["Component", "Market Value ($M)", "Capital Weight (%)", "Cost of Capital (%)", "Weighted Contribution (%)"]
    for c_idx, h_text in enumerate(wacc_summary_headers, start=1):
        c = ws_wacc.cell(row=23, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.border = border_header
        c.alignment = align_left if c_idx == 1 else align_right
        
    ws_wacc["A24"] = "Common Equity (Market Cap)"
    ws_wacc["B24"] = "=B16"
    ws_wacc["C24"] = "=B24/$B$20"
    ws_wacc["D24"] = "=B8"
    ws_wacc["E24"] = "=C24*D24"
    
    ws_wacc["A25"] = "Net Debt (Debt Weight Adjustment)"
    ws_wacc["B25"] = "=B19"
    ws_wacc["C25"] = "=B25/$B$20"
    ws_wacc["D25"] = "=B13"
    ws_wacc["E25"] = "=C25*D25"
    
    ws_wacc["A26"] = "Total Capitalization / Base WACC"
    ws_wacc["B26"] = "=B20"
    ws_wacc["C26"] = "=SUM(C24:C25)"
    ws_wacc["D26"] = "-"
    ws_wacc["E26"] = "=SUM(E24:E25)"
    
    for r in range(24, 27):
        ws_wacc[f"A{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"A{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"B{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"B{r}"].number_format = "$#,##0.00"
        ws_wacc[f"B{r}"].alignment = align_right
        ws_wacc[f"B{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"C{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"C{r}"].number_format = "0.0%"
        ws_wacc[f"C{r}"].alignment = align_right
        ws_wacc[f"C{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"D{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"D{r}"].number_format = "0.00%" if r != 26 else "@"
        ws_wacc[f"D{r}"].alignment = align_right
        ws_wacc[f"D{r}"].border = border_total if r == 26 else border_cell
        
        ws_wacc[f"E{r}"].font = font_bold if r == 26 else font_regular
        ws_wacc[f"E{r}"].number_format = "0.00%"
        ws_wacc[f"E{r}"].alignment = align_right
        ws_wacc[f"E{r}"].border = border_total if r == 26 else border_cell
        if r == 26:
            ws_wacc[f"E{r}"].fill = fill_accent_blue
            
    ws_wacc.column_dimensions["A"].width = 44
    ws_wacc.column_dimensions["B"].width = 20
    ws_wacc.column_dimensions["C"].width = 18
    ws_wacc.column_dimensions["D"].width = 18
    ws_wacc.column_dimensions["E"].width = 24
    
    # ==========================================
    # SHEET 1: DCF Valuation
    # ==========================================
    
    # --- Title Block ---
    ws_dcf["A1"] = "Meta Platforms, Inc. (NASDAQ: META)"
    ws_dcf["A1"].font = font_title
    ws_dcf["A2"] = "Institutional DCF Valuation Model | Report Date: August 19, 2026 | Financial Source: SEC Form 10-K, 10-Q & Consensus Estimates"
    ws_dcf["A2"].font = font_subtitle
    
    # --- Active Scenario Selector ---
    ws_dcf["A4"] = "Active Scenario Selector:"
    ws_dcf["A4"].font = font_bold
    ws_dcf["B4"] = 2
    ws_dcf["B4"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    ws_dcf["B4"].fill = fill_navy
    ws_dcf["B4"].alignment = align_center
    ws_dcf["B4"].border = border_cell
    ws_dcf["B4"].comment = Comment("Case Selector: 1 = Bear Case (Ad Slowdown / Capex Drag), 2 = Base Case (AI Monetization & Operating Leverage), 3 = Bull Case (GenAI & Smart Hardware Boom)", "DCF Model Builder")
    
    ws_dcf["C4"] = '=IF(B4=1,"[1] BEAR CASE (Macro Headwinds & Heavy AI Capex Drag)",IF(B4=2,"[2] BASE CASE (AI-Driven Ad Efficiency & Cloud Normalization)","[3] BULL CASE (GenAI Ecosystem & Hardware Breakthrough)"))'
    ws_dcf["C4"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws_dcf["C4"].alignment = align_left
    
    # --- Section I: Market Data & Capital Structure Inputs ---
    add_section_header(ws_dcf, 6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS", "H")
    
    market_inputs_dcf = [
        ("Current Stock Price ($)", 543.67, "$#,##0.00", "Source: Market close price as of August 19, 2026", "B7"),
        ("Diluted Shares Outstanding (M)", 2570.00, "#,##0.00", "Source: SEC Form 10-Q Q2 2026 diluted share count (2,570M shares)", "B8"),
        ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
        ("Cash, Cash Equivalents & Liquid Assets ($M)", 90260.00, "$#,##0.00", "Source: SEC Form 10-Q Q2 2026 cash & marketable securities ($90.26B)", "B10"),
        ("Total Debt (Current + Noncurrent) ($M)", 83660.00, "$#,##0.00", "Source: SEC Form 10-Q Q2 2026 total carrying debt ($83.66B)", "B11"),
        ("Net Debt ($M)", "=B11-B10", "$#,##0.00", None, "B12"),
        ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
        ("Effective Corporate Tax Rate (%)", 0.160, "0.0%", "Source: Normalized corporate effective tax rate", "B14"),
        ("Risk-Free Rate (10-Yr US Treasury) (%)", "='WACC Build'!B5", "0.00%", None, "B15"),
        ("Equity Beta (5-Year Monthly)", "='WACC Build'!B6", "0.00", None, "B16"),
        ("Market Equity Risk Premium (ERP) (%)", "='WACC Build'!B7", "0.00%", None, "B17"),
        ("Base Case WACC (%)", "='WACC Build'!E26", "0.00%", None, "B18"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs_dcf, start=7):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        if "='WACC Build'" in str(val):
            ws_dcf[f"B{idx}"].font = font_link_bold
        elif "=" in str(val):
            ws_dcf[f"B{idx}"].font = font_bold
        else:
            ws_dcf[f"B{idx}"].font = font_input
            ws_dcf[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_dcf[f"B{idx}"].comment = Comment(comment_text, "Meta Platforms 10-Q / Market Data")
                
    # --- Section II: Scenario Assumptions & Driver Blocks ---
    add_section_header(ws_dcf, 20, "II. SCENARIO ASSUMPTIONS & DRIVER BLOCKS", "H")
    
    headers_s2 = ["Driver / Assumption", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal / WACC"]
    for col_idx, text in enumerate(headers_s2, start=1):
        c = ws_dcf.cell(row=21, column=col_idx, value=text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if col_idx > 1 else align_left
        c.border = border_header
        
    scenario_blocks = [
        ("Bear Case Assumptions (Selector = 1)", 22, [
            ("  Revenue Growth (%)", [0.120, 0.100, 0.080, 0.070, 0.060], 0.025, "0.0%"),
            ("  EBIT Margin (%)", [0.380, 0.370, 0.360, 0.350, 0.350], 0.1200, "0.0%"),
            ("  D&A (% of Revenue)", [0.115, 0.120, 0.125, 0.125, 0.120], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.330, 0.300, 0.270, 0.250, 0.240], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
        ("Base Case Assumptions (Selector = 2)", 29, [
            ("  Revenue Growth (%)", [0.160, 0.140, 0.120, 0.100, 0.080], 0.030, "0.0%"),
            ("  EBIT Margin (%)", [0.415, 0.420, 0.425, 0.430, 0.435], "='WACC Build'!E26", "0.0%"),
            ("  D&A (% of Revenue)", [0.110, 0.120, 0.125, 0.125, 0.120], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.320, 0.280, 0.250, 0.230, 0.220], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
        ("Bull Case Assumptions (Selector = 3)", 36, [
            ("  Revenue Growth (%)", [0.200, 0.170, 0.150, 0.130, 0.110], 0.035, "0.0%"),
            ("  EBIT Margin (%)", [0.430, 0.440, 0.450, 0.455, 0.460], 0.1000, "0.0%"),
            ("  D&A (% of Revenue)", [0.105, 0.115, 0.120, 0.120, 0.115], None, "0.0%"),
            ("  CapEx (% of Revenue)", [0.300, 0.260, 0.230, 0.210, 0.200], None, "0.0%"),
            ("  Δ NWC (% of Δ Revenue)", [0.010, 0.010, 0.010, 0.010, 0.010], None, "0.0%"),
        ]),
    ]
    
    for b_title, start_r, rows_data in scenario_blocks:
        ws_dcf[f"A{start_r}"] = b_title
        ws_dcf[f"A{start_r}"].font = font_bold_navy
        ws_dcf[f"A{start_r}"].fill = fill_soft_blue
        ws_dcf.merge_cells(f"A{start_r}:H{start_r}")
        for col_idx in range(1, 9):
            ws_dcf.cell(row=start_r, column=col_idx).fill = fill_soft_blue
            ws_dcf.cell(row=start_r, column=col_idx).border = border_header
            
        for r_offset, (label, val_list, term_val, num_fmt) in enumerate(rows_data, start=1):
            curr_r = start_r + r_offset
            ws_dcf[f"A{curr_r}"] = label
            ws_dcf[f"A{curr_r}"].font = font_regular
            ws_dcf[f"A{curr_r}"].border = border_cell
            ws_dcf[f"B{curr_r}"] = "-"
            ws_dcf[f"B{curr_r}"].alignment = align_center
            ws_dcf[f"B{curr_r}"].border = border_cell
            
            for c_idx, val in enumerate(val_list, start=3):
                col_let = get_column_letter(c_idx)
                ws_dcf[f"{col_let}{curr_r}"] = val
                ws_dcf[f"{col_let}{curr_r}"].font = font_input
                ws_dcf[f"{col_let}{curr_r}"].fill = fill_input
                ws_dcf[f"{col_let}{curr_r}"].number_format = num_fmt
                ws_dcf[f"{col_let}{curr_r}"].alignment = align_right
                ws_dcf[f"{col_let}{curr_r}"].border = border_cell
                
            ws_dcf[f"H{curr_r}"] = term_val if term_val is not None else "-"
            ws_dcf[f"H{curr_r}"].border = border_cell
            if term_val is not None:
                if isinstance(term_val, str) and term_val.startswith("="):
                    ws_dcf[f"H{curr_r}"].font = font_link_bold
                else:
                    ws_dcf[f"H{curr_r}"].font = font_input_bold
                    ws_dcf[f"H{curr_r}"].fill = fill_input
                ws_dcf[f"H{curr_r}"].number_format = num_fmt
                ws_dcf[f"H{curr_r}"].alignment = align_right
            else:
                ws_dcf[f"H{curr_r}"].alignment = align_center
                
    # --- Consolidated Active Driver Table (Rows 43-48) ---
    ws_dcf["A43"] = "ACTIVE CONSOLIDATED DRIVERS (PULLED VIA INDEX FORMULAS)"
    ws_dcf["A43"].font = font_bold_navy
    ws_dcf["A43"].fill = fill_accent_blue
    ws_dcf.merge_cells("A43:H43")
    for col_idx in range(1, 9):
        ws_dcf.cell(row=43, column=col_idx).fill = fill_accent_blue
        ws_dcf.cell(row=43, column=col_idx).border = border_header
        
    driver_rows_map = [
        ("  Active Revenue Growth (%)", [23, 30, 37], 44, "0.0%"),
        ("  Active EBIT Margin (%)", [24, 31, 38], 45, "0.0%"),
        ("  Active D&A (% of Revenue)", [25, 32, 39], 46, "0.0%"),
        ("  Active CapEx (% of Revenue)", [26, 33, 40], 47, "0.0%"),
        ("  Active Δ NWC (% of Δ Revenue)", [27, 34, 41], 48, "0.0%"),
    ]
    
    for label, src_rows, target_r, num_fmt in driver_rows_map:
        ws_dcf[f"A{target_r}"] = label
        ws_dcf[f"A{target_r}"].font = font_bold
        ws_dcf[f"A{target_r}"].border = border_cell
        ws_dcf[f"B{target_r}"] = "-"
        ws_dcf[f"B{target_r}"].alignment = align_center
        ws_dcf[f"B{target_r}"].border = border_cell
        
        # FY2026E to FY2030E (Cols C to G)
        for col_idx in range(3, 8):
            c_let = get_column_letter(col_idx)
            formula = f"=INDEX(({c_let}${src_rows[0]},{c_let}${src_rows[1]},{c_let}${src_rows[2]}),1,1,$B$4)"
            ws_dcf[f"{c_let}{target_r}"] = formula
            ws_dcf[f"{c_let}{target_r}"].font = font_bold
            ws_dcf[f"{c_let}{target_r}"].number_format = num_fmt
            ws_dcf[f"{c_let}{target_r}"].alignment = align_right
            ws_dcf[f"{c_let}{target_r}"].border = border_cell
            
        # Terminal / WACC column (Col H)
        if target_r in [44, 45]:  # Terminal growth or WACC
            formula_h = f"=INDEX((H${src_rows[0]},H${src_rows[1]},H${src_rows[2]}),1,1,$B$4)"
            ws_dcf[f"H{target_r}"] = formula_h
            ws_dcf[f"H{target_r}"].font = font_bold
            ws_dcf[f"H{target_r}"].number_format = num_fmt
            ws_dcf[f"H{target_r}"].alignment = align_right
            ws_dcf[f"H{target_r}"].border = border_cell
        else:
            ws_dcf[f"H{target_r}"] = "-"
            ws_dcf[f"H{target_r}"].alignment = align_center
            ws_dcf[f"H{target_r}"].border = border_cell
            
    # --- Section III: Financial Performance & 5-Year Projections (Rows 51-58) ---
    add_section_header(ws_dcf, 51, "III. FINANCIAL PERFORMANCE & 5-YEAR PROJECTIONS ($M)", "H")
    
    headers_is = ["Line Item ($M)", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal Year"]
    for c_idx, h_text in enumerate(headers_is, start=1):
        c = ws_dcf.cell(row=52, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if c_idx > 1 else align_left
        c.border = border_header
        
    # Row 53: Total Revenue
    ws_dcf["A53"] = "Total Revenue"
    ws_dcf["B53"] = 200966.0
    ws_dcf["C53"] = "=B53*(1+C44)"
    ws_dcf["D53"] = "=C53*(1+D44)"
    ws_dcf["E53"] = "=D53*(1+E44)"
    ws_dcf["F53"] = "=E53*(1+F44)"
    ws_dcf["G53"] = "=F53*(1+G44)"
    ws_dcf["H53"] = "=G53*(1+H44)"
    
    # Row 54: YoY Revenue Growth (%)
    ws_dcf["A54"] = "  YoY Revenue Growth (%)"
    ws_dcf["B54"] = 0.222
    ws_dcf["C54"] = "=C53/B53-1"
    ws_dcf["D54"] = "=D53/C53-1"
    ws_dcf["E54"] = "=E53/D53-1"
    ws_dcf["F54"] = "=F53/E53-1"
    ws_dcf["G54"] = "=G53/F53-1"
    ws_dcf["H54"] = "=H53/G53-1"
    
    # Row 55: Operating Income (EBIT)
    ws_dcf["A55"] = "Operating Income (EBIT)"
    ws_dcf["B55"] = 83276.0
    ws_dcf["C55"] = "=C53*C45"
    ws_dcf["D55"] = "=D53*D45"
    ws_dcf["E55"] = "=E53*E45"
    ws_dcf["F55"] = "=F53*F45"
    ws_dcf["G55"] = "=G53*G45"
    ws_dcf["H55"] = "=H53*G45"
    
    # Row 56: EBIT Margin (%)
    ws_dcf["A56"] = "  EBIT Margin (%)"
    ws_dcf["B56"] = "=B55/B53"
    ws_dcf["C56"] = "=C55/C53"
    ws_dcf["D56"] = "=D55/D53"
    ws_dcf["E56"] = "=E55/E53"
    ws_dcf["F56"] = "=F55/F53"
    ws_dcf["G56"] = "=G55/G53"
    ws_dcf["H56"] = "=H55/H53"
    
    # Row 57: Taxes on Operating Income (EBIT * t)
    ws_dcf["A57"] = "(-) Taxes on EBIT"
    ws_dcf["B57"] = "=IF(B55>0,B55*$B$14,0)"
    ws_dcf["C57"] = "=IF(C55>0,C55*$B$14,0)"
    ws_dcf["D57"] = "=IF(D55>0,D55*$B$14,0)"
    ws_dcf["E57"] = "=IF(E55>0,E55*$B$14,0)"
    ws_dcf["F57"] = "=IF(F55>0,F55*$B$14,0)"
    ws_dcf["G57"] = "=IF(G55>0,G55*$B$14,0)"
    ws_dcf["H57"] = "=IF(H55>0,H55*$B$14,0)"
    
    # Row 58: Net Operating Profit After Tax (NOPAT)
    ws_dcf["A58"] = "Net Operating Profit After Tax (NOPAT)"
    ws_dcf["B58"] = "=B55-B57"
    ws_dcf["C58"] = "=C55-C57"
    ws_dcf["D58"] = "=D55-D57"
    ws_dcf["E58"] = "=E55-E57"
    ws_dcf["F58"] = "=F55-F57"
    ws_dcf["G58"] = "=G55-G57"
    ws_dcf["H58"] = "=H55-H57"
    
    for r in range(53, 59):
        is_bold_row = r in [53, 55, 58]
        ws_dcf[f"A{r}"].font = font_bold if is_bold_row else font_regular
        ws_dcf[f"A{r}"].border = border_cell
        
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_cell
            cell.alignment = align_right
            
            if r in [54, 56]:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                
            if str(cell.value).startswith("="):
                cell.font = font_bold if is_bold_row else font_regular
            elif cell.value == "-":
                cell.font = font_regular
                cell.alignment = align_center
            else:
                cell.font = font_input_bold if is_bold_row else font_input
                cell.fill = fill_input
                cell.comment = Comment("Source: SEC Form 10-K FY2025 Consolidated Financial Statements", "DCF Model Builder")
                
    # --- Section IV: Free Cash Flow Schedule (Rows 60-67) ---
    add_section_header(ws_dcf, 60, "IV. UNLEVERED FREE CASH FLOW (UFCF) BUILD ($M)", "H")
    
    headers_fcf = ["Cash Flow Line Item ($M)", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal Year"]
    for c_idx, h_text in enumerate(headers_fcf, start=1):
        c = ws_dcf.cell(row=61, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if c_idx > 1 else align_left
        c.border = border_header
        
    # Row 62: NOPAT
    ws_dcf["A62"] = "Net Operating Profit After Tax (NOPAT)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}62"] = f"={c_let}58"
        
    # Row 63: (+) Depreciation & Amortization (D&A)
    ws_dcf["A63"] = "(+) Depreciation & Amortization (D&A)"
    ws_dcf["B63"] = 18616.0
    ws_dcf["C63"] = "=C53*C46"
    ws_dcf["D63"] = "=D53*D46"
    ws_dcf["E63"] = "=E53*E46"
    ws_dcf["F63"] = "=F53*F46"
    ws_dcf["G63"] = "=G53*G46"
    ws_dcf["H63"] = "=H53*G46"
    
    # Row 64: (-) Capital Expenditures (CapEx)
    ws_dcf["A64"] = "(-) Capital Expenditures (CapEx)"
    ws_dcf["B64"] = 69691.0
    ws_dcf["C64"] = "=C53*C47"
    ws_dcf["D64"] = "=D53*D47"
    ws_dcf["E64"] = "=E53*E47"
    ws_dcf["F64"] = "=F53*F47"
    ws_dcf["G64"] = "=G53*G47"
    ws_dcf["H64"] = "=H53*G47"
    
    # Row 65: (-) Change in Net Working Capital (Δ NWC)
    ws_dcf["A65"] = "(-) Change in Net Working Capital (Δ NWC)"
    ws_dcf["B65"] = 0.0
    ws_dcf["C65"] = "=(C53-B53)*C48"
    ws_dcf["D65"] = "=(D53-C53)*D48"
    ws_dcf["E65"] = "=(E53-D53)*E48"
    ws_dcf["F65"] = "=(F53-E53)*F48"
    ws_dcf["G65"] = "=(G53-F53)*G48"
    ws_dcf["H65"] = "=(H53-G53)*G48"
    
    # Row 66: Unlevered Free Cash Flow (UFCF)
    ws_dcf["A66"] = "Unlevered Free Cash Flow (UFCF)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}66"] = f"={c_let}62+{c_let}63-{c_let}64-{c_let}65"
        
    # Row 67: FCF Conversion Margin (% of Revenue)
    ws_dcf["A67"] = "  FCF Conversion Margin (% of Revenue)"
    for col_idx in range(2, 9):
        c_let = get_column_letter(col_idx)
        ws_dcf[f"{c_let}67"] = f"={c_let}66/{c_let}53"
        
    for r in range(62, 68):
        is_bold_row = r in [62, 66]
        ws_dcf[f"A{r}"].font = font_bold if is_bold_row else font_regular
        ws_dcf[f"A{r}"].border = border_total if r == 66 else border_cell
        
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_total if r == 66 else border_cell
            cell.alignment = align_right
            
            if r == 67:
                cell.number_format = "0.0%"
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                
            if str(cell.value).startswith("="):
                cell.font = font_bold if is_bold_row else font_regular
            else:
                cell.font = font_input_bold if is_bold_row else font_input
                cell.fill = fill_input
                cell.comment = Comment("Source: SEC Form 10-K FY2025 Statement of Cash Flows", "DCF Model Builder")
                
            if r == 66:
                cell.fill = fill_accent_blue
                
    # --- Section V: Discounting Schedule & Valuation Summary (Rows 69-90) ---
    add_section_header(ws_dcf, 69, "V. DISCOUNTING SCHEDULE & VALUATION SUMMARY ($M)", "H")
    
    headers_val = ["DCF Component ($M)", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal Year", ""]
    for c_idx, h_text in enumerate(headers_val, start=1):
        c = ws_dcf.cell(row=70, column=c_idx, value=h_text)
        c.font = font_tbl_hdr
        c.fill = fill_soft_blue
        c.alignment = align_center if 1 < c_idx <= 7 else align_left
        c.border = border_header
        
    # Row 71: Explicit UFCF
    ws_dcf["A71"] = "Unlevered Free Cash Flow (UFCF)"
    ws_dcf["B71"] = "=C66"
    ws_dcf["C71"] = "=D66"
    ws_dcf["D71"] = "=E66"
    ws_dcf["E71"] = "=F66"
    ws_dcf["F71"] = "=G66"
    ws_dcf["G71"] = "=F71*(1+$H$44)"  # Terminal UFCF = Final Year UFCF * (1 + g)
    ws_dcf["H71"] = "-"
    
    # Row 72: Mid-Year Discount Period (t)
    ws_dcf["A72"] = "Mid-Year Discount Period (t)"
    ws_dcf["B72"] = 0.5
    ws_dcf["C72"] = 1.5
    ws_dcf["D72"] = 2.5
    ws_dcf["E72"] = 3.5
    ws_dcf["F72"] = 4.5
    ws_dcf["G72"] = 4.5  # Mid-year convention for terminal value
    ws_dcf["H72"] = "-"
    
    # Row 73: Discount Factor (1 / (1 + WACC)^t)
    ws_dcf["A73"] = "Discount Factor (WACC Discounting)"
    ws_dcf["B73"] = "=1/(1+$H$45)^B72"
    ws_dcf["C73"] = "=1/(1+$H$45)^C72"
    ws_dcf["D73"] = "=1/(1+$H$45)^D72"
    ws_dcf["E73"] = "=1/(1+$H$45)^E72"
    ws_dcf["F73"] = "=1/(1+$H$45)^F72"
    ws_dcf["G73"] = "=1/(1+$H$45)^G72"
    ws_dcf["H73"] = "-"
    
    # Row 74: Present Value of Explicit FCF (PV)
    ws_dcf["A74"] = "Present Value of Cash Flow (PV of UFCF)"
    ws_dcf["B74"] = "=B71*B73"
    ws_dcf["C74"] = "=C71*C73"
    ws_dcf["D74"] = "=D71*D73"
    ws_dcf["E74"] = "=E71*E73"
    ws_dcf["F74"] = "=F71*F73"
    ws_dcf["G74"] = "-"
    ws_dcf["H74"] = "-"
    
    # Row 75: Terminal Value calculation
    ws_dcf["A75"] = "Nominal Terminal Value (Perpetuity Growth Method)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}75"] = "-"
    ws_dcf["G75"] = "=G71/($H$45-$H$44)"  # TV = Terminal FCF / (WACC - g)
    ws_dcf["H75"] = "-"
    
    # Row 76: PV of Terminal Value
    ws_dcf["A76"] = "Present Value of Terminal Value (PV of TV)"
    for col_idx in range(2, 7):
        ws_dcf[f"{get_column_letter(col_idx)}76"] = "-"
    ws_dcf["G76"] = "=G75*G73"
    ws_dcf["H76"] = "-"
    
    for r in range(71, 77):
        ws_dcf[f"A{r}"].font = font_bold if r in [71, 74, 75, 76] else font_regular
        ws_dcf[f"A{r}"].border = border_cell
        for col_idx in range(2, 9):
            c_let = get_column_letter(col_idx)
            cell = ws_dcf[f"{c_let}{r}"]
            cell.border = border_cell
            cell.alignment = align_right
            if r == 72:
                cell.number_format = "0.0"
                cell.font = font_regular
            elif r == 73:
                cell.number_format = "0.0000"
                cell.font = font_regular
            else:
                cell.number_format = "$#,##0.00;($#,##0.00);-"
                cell.font = font_bold if r in [71, 74, 76] else font_regular
            if str(cell.value) == "-":
                cell.alignment = align_center
                
    # --- Equity Value Bridge (Rows 78-89) ---
    ws_dcf["A78"] = "VALUATION BRIDGE & PER-SHARE IMPLICATION"
    ws_dcf["A78"].font = font_bold_navy
    ws_dcf["A78"].fill = fill_soft_blue
    ws_dcf.merge_cells("A78:C78")
    for col_idx in range(1, 4):
        ws_dcf.cell(row=78, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=78, column=col_idx).border = border_header
        
    equity_bridge = [
        ("Cumulative PV of Explicit 5-Year FCFs ($M)", "=SUM(B74:F74)", "$#,##0.00", "B79", False),
        ("Present Value of Terminal Value ($M)", "=G76", "$#,##0.00", "B80", False),
        ("Enterprise Value ($M)", "=B79+B80", "$#,##0.00", "B81", True),
        ("(-) Total Debt ($M)", "=-B11", "$#,##0.00", "B82", False),
        ("(+) Cash and Short-Term Investments ($M)", "=B10", "$#,##0.00", "B83", False),
        ("Net Debt / (Net Cash) Adjustment ($M)", "=B82+B83", "$#,##0.00", "B84", False),
        ("Implied Equity Value ($M)", "=B81+B84", "$#,##0.00", "B85", True),
        ("Diluted Shares Outstanding (M)", "=B8", "#,##0.00", "B86", False),
        ("IMPLIED INTRINSIC VALUE PER SHARE ($)", "=B85/B86", "$#,##0.00", "B87", True),
        ("Current Market Stock Price ($)", "=B7", "$#,##0.00", "B88", False),
        ("Implied Upside / (Downside) (%)", "=B87/B88-1", "+0.0%;-0.0%;0.0%", "B89", True),
    ]
    
    for idx, (label, formula_val, num_fmt, cell_ref, is_highlight) in enumerate(equity_bridge, start=79):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"A{idx}"].border = border_total if idx in [81, 85, 87, 89] else border_cell
        
        ws_dcf[f"B{idx}"] = formula_val
        ws_dcf[f"B{idx}"].font = font_bold if is_highlight else font_regular
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_total if idx in [81, 85, 87, 89] else border_cell
        
        if is_highlight:
            ws_dcf[f"A{idx}"].fill = fill_accent_blue
            ws_dcf[f"B{idx}"].fill = fill_accent_blue
            
    # Terminal Value Sanity Metrics (Cols D & E)
    ws_dcf["D79"] = "Terminal Value Metrics & Checks:"
    ws_dcf["D79"].font = font_bold_navy
    ws_dcf["D80"] = "PV of TV % of Enterprise Value"
    ws_dcf["E80"] = "=B80/B81"
    ws_dcf["E80"].number_format = "0.0%"
    ws_dcf["E80"].font = font_bold
    ws_dcf["E80"].alignment = align_right
    ws_dcf["E80"].border = border_cell
    
    ws_dcf["D81"] = "Implied Exit Multiple (EV / FY30E EBITDA)"
    ws_dcf["E81"] = "=B81/(G55+G63)"
    ws_dcf["E81"].number_format = "0.0x"
    ws_dcf["E81"].font = font_bold
    ws_dcf["E81"].alignment = align_right
    ws_dcf["E81"].border = border_cell
    
    for r_chk in [80, 81]:
        ws_dcf[f"D{r_chk}"].border = border_cell
        ws_dcf[f"D{r_chk}"].font = font_regular
        
    # =========================================================================
    # SECTION VI: THREE INSTITUTIONAL 5X5 SENSITIVITY TABLES (Rows 92-132)
    # =========================================================================
    add_section_header(ws_dcf, 92, "VI. INSTITUTIONAL SENSITIVITY MATRICES (5x5 GRIDS WITH FULL DCF RECALCULATION)", "H")
    
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 1: WACC vs Perpetual Growth Rate (g)
    # -------------------------------------------------------------------------
    ws_dcf["A94"] = "TABLE 1: IMPLIED SHARE PRICE ($) - WACC vs. TERMINAL GROWTH RATE (g)"
    ws_dcf["A94"].font = font_bold_navy
    ws_dcf["A94"].fill = fill_soft_blue
    ws_dcf.merge_cells("A94:F94")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=94, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=94, column=col_idx).border = border_header
        
    ws_dcf["A95"] = "WACC \\ Terminal g"
    ws_dcf["A95"].font = font_tbl_hdr
    ws_dcf["A95"].fill = fill_soft_blue
    ws_dcf["A95"].border = border_header
    ws_dcf["A95"].alignment = align_center
    
    # Axis ranges centered on Base Case (WACC = 11.00%, g = 3.00%)
    g_cols = [0.020, 0.025, 0.030, 0.035, 0.040]
    wacc_rows = [0.0900, 0.1000, 0.1100, 0.1200, 0.1300]
    
    for c_idx, g_val in enumerate(g_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}95"] = g_val
        ws_dcf[f"{c_let}95"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}95"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}95"].number_format = "0.0%"
        ws_dcf[f"{c_let}95"].alignment = align_center
        ws_dcf[f"{c_let}95"].border = border_header
        
    for r_idx, w_val in enumerate(wacc_rows, start=96):
        ws_dcf[f"A{r_idx}"] = w_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 98 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 98 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, g_val in enumerate(g_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t1 = (
                f"=(($B$71/(1+$A{r_idx})^0.5 + $C$71/(1+$A{r_idx})^1.5 + $D$71/(1+$A{r_idx})^2.5 + $E$71/(1+$A{r_idx})^3.5 + $F$71/(1+$A{r_idx})^4.5) + "
                f"(($F$71*(1+{c_let}$95)/($A{r_idx}-{c_let}$95))/(1+$A{r_idx})^4.5) + $B$84)/$B$86"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t1
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 98 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 2: Revenue Growth Scale vs. Target EBIT Margin
    # -------------------------------------------------------------------------
    ws_dcf["A104"] = "TABLE 2: IMPLIED SHARE PRICE ($) - REVENUE GROWTH MULTIPLIER vs. FY30E EBIT MARGIN"
    ws_dcf["A104"].font = font_bold_navy
    ws_dcf["A104"].fill = fill_soft_blue
    ws_dcf.merge_cells("A104:F104")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=104, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=104, column=col_idx).border = border_header
        
    ws_dcf["A105"] = "Growth Multiplier \\ EBIT Margin"
    ws_dcf["A105"].font = font_tbl_hdr
    ws_dcf["A105"].fill = fill_soft_blue
    ws_dcf["A105"].border = border_header
    ws_dcf["A105"].alignment = align_center
    
    ebit_cols = [0.395, 0.415, 0.435, 0.455, 0.475]
    growth_rows = [0.80, 0.90, 1.00, 1.10, 1.20]
    
    for c_idx, em_val in enumerate(ebit_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}105"] = em_val
        ws_dcf[f"{c_let}105"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}105"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}105"].number_format = "0.0%"
        ws_dcf[f"{c_let}105"].alignment = align_center
        ws_dcf[f"{c_let}105"].border = border_header
        
    for r_idx, gr_val in enumerate(growth_rows, start=106):
        ws_dcf[f"A{r_idx}"] = gr_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 108 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 108 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0%"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, em_val in enumerate(ebit_cols, start=2):
            c_let = get_column_letter(c_idx)
            formula_t2 = (
                f"=((($B$71*$A{r_idx}*({c_let}$105/$G$45)/(1+$H$45)^0.5 + "
                f"$C$71*$A{r_idx}*({c_let}$105/$G$45)/(1+$H$45)^1.5 + "
                f"$D$71*$A{r_idx}*({c_let}$105/$G$45)/(1+$H$45)^2.5 + "
                f"$E$71*$A{r_idx}*({c_let}$105/$G$45)/(1+$H$45)^3.5 + "
                f"$F$71*$A{r_idx}*({c_let}$105/$G$45)/(1+$H$45)^4.5) + "
                f"(($F$71*$A{r_idx}*({c_let}$105/$G$45)*(1+$H$44)/($H$45-$H$44))/(1+$H$45)^4.5) + $B$84)/$B$86)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t2
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 108 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # -------------------------------------------------------------------------
    # SENSITIVITY TABLE 3: Equity Beta vs. Risk-Free Rate (Rf)
    # -------------------------------------------------------------------------
    ws_dcf["A114"] = "TABLE 3: IMPLIED SHARE PRICE ($) - EQUITY BETA vs. RISK-FREE RATE (Rf)"
    ws_dcf["A114"].font = font_bold_navy
    ws_dcf["A114"].fill = fill_soft_blue
    ws_dcf.merge_cells("A114:F114")
    for col_idx in range(1, 7):
        ws_dcf.cell(row=114, column=col_idx).fill = fill_soft_blue
        ws_dcf.cell(row=114, column=col_idx).border = border_header
        
    ws_dcf["A115"] = "Beta \\ Risk-Free Rate"
    ws_dcf["A115"].font = font_tbl_hdr
    ws_dcf["A115"].fill = fill_soft_blue
    ws_dcf["A115"].border = border_header
    ws_dcf["A115"].alignment = align_center
    
    rf_cols = [0.0370, 0.0420, 0.0470, 0.0520, 0.0570]
    beta_rows = [1.04, 1.14, 1.24, 1.34, 1.44]
    
    for c_idx, rf_val in enumerate(rf_cols, start=2):
        c_let = get_column_letter(c_idx)
        ws_dcf[f"{c_let}115"] = rf_val
        ws_dcf[f"{c_let}115"].font = font_bold if c_idx == 4 else font_tbl_hdr
        ws_dcf[f"{c_let}115"].fill = fill_accent_blue if c_idx == 4 else fill_soft_blue
        ws_dcf[f"{c_let}115"].number_format = "0.00%"
        ws_dcf[f"{c_let}115"].alignment = align_center
        ws_dcf[f"{c_let}115"].border = border_header
        
    for r_idx, b_val in enumerate(beta_rows, start=116):
        ws_dcf[f"A{r_idx}"] = b_val
        ws_dcf[f"A{r_idx}"].font = font_bold if r_idx == 118 else font_tbl_hdr
        ws_dcf[f"A{r_idx}"].fill = fill_accent_blue if r_idx == 118 else fill_soft_blue
        ws_dcf[f"A{r_idx}"].number_format = "0.00"
        ws_dcf[f"A{r_idx}"].alignment = align_right
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        for c_idx, rf_val in enumerate(rf_cols, start=2):
            c_let = get_column_letter(c_idx)
            wacc_formula_snippet = f"({c_let}$115+$A{r_idx}*$B$17)*('WACC Build'!$C$24) + ('WACC Build'!$B$13)*('WACC Build'!$C$25)"
            
            formula_t3 = (
                f"=((($B$71/(1+({wacc_formula_snippet}))^0.5 + "
                f"$C$71/(1+({wacc_formula_snippet}))^1.5 + "
                f"$D$71/(1+({wacc_formula_snippet}))^2.5 + "
                f"$E$71/(1+({wacc_formula_snippet}))^3.5 + "
                f"$F$71/(1+({wacc_formula_snippet}))^4.5) + "
                f"(($F$71*(1+$H$44)/(({wacc_formula_snippet})-$H$44))/(1+({wacc_formula_snippet}))^4.5) + $B$84)/$B$86)"
            )
            cell = ws_dcf[f"{c_let}{r_idx}"]
            cell.value = formula_t3
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_idx == 118 and c_idx == 4:
                cell.font = font_bold
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular
                
    # Adjust Column Widths on DCF Sheet
    ws_dcf.column_dimensions["A"].width = 46
    ws_dcf.column_dimensions["B"].width = 16
    ws_dcf.column_dimensions["C"].width = 16
    ws_dcf.column_dimensions["D"].width = 16
    ws_dcf.column_dimensions["E"].width = 16
    ws_dcf.column_dimensions["F"].width = 16
    ws_dcf.column_dimensions["G"].width = 16
    ws_dcf.column_dimensions["H"].width = 16
    
    wb.save(output_path)
    print(f"Successfully generated institutional DCF model at: {output_path}")

if __name__ == "__main__":
    build_meta_dcf_model()
