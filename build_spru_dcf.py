#!/usr/bin/env python3
"""
SPRU DCF Model Builder - Spruce Power Holding Corporation
Date: 2026-08-17
7-year projection 2026E-2032E, 3 scenarios (Bear/Base/Bull)
Two sheets: DCF (with sensitivity tables), WACC
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Output path ──────────────────────────────────────────────────────────────
OUT_DIR = "./reports/dcf"
os.makedirs(OUT_DIR, exist_ok=True)
OUTPUT_PATH = f"{OUT_DIR}/SPRU_DCF_Model_2026-08-17.xlsx"

# ── Colour palette ───────────────────────────────────────────────────────────
C_HDR_DARK   = "1F4E79"   # Dark blue  – section headers
C_HDR_LIGHT  = "D9E1F2"   # Light blue – column headers / sub-headers
C_INPUT_FILL = "F2F2F2"   # Light grey – input cells
C_OUTPUT_FILL= "BDD7EE"   # Medium blue– key output rows
C_WHITE      = "FFFFFF"
C_BLUE_FONT  = "0000FF"   # Blue text  – hardcoded inputs
C_BLACK_FONT = "000000"   # Black text – formulas
C_GREEN_FONT = "008000"   # Green text – cross-sheet links

# ── Font helpers ─────────────────────────────────────────────────────────────
def fnt(bold=False, color=C_BLACK_FONT, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(left=True, right=True, top=True, bottom=True):
    s = Side(style="thin")
    n = Side(style=None)
    return Border(
        left=s if left else n,
        right=s if right else n,
        top=s if top else n,
        bottom=s if bottom else n,
    )

def medium_border(left=False, right=False, top=False, bottom=False):
    s = Side(style="medium")
    n = Side(style=None)
    return Border(
        left=s if left else n,
        right=s if right else n,
        top=s if top else n,
        bottom=s if bottom else n,
    )

def thick_box(ws, min_row, max_row, min_col, max_col):
    """Draw a thick outer border around a rectangular range."""
    tk = Side(style="medium")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            l = tk if cell.column == min_col else Side(style=None)
            r = tk if cell.column == max_col else Side(style=None)
            t = tk if cell.row == min_row else Side(style=None)
            b = tk if cell.row == max_row else Side(style=None)
            cell.border = Border(left=l, right=r, top=t, bottom=b)

def section_header(ws, row, col_start, col_end, text, merge=True):
    """Dark blue merged section header."""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = fnt(bold=True, color=C_WHITE, size=10)
    cell.fill = fill(C_HDR_DARK)
    cell.alignment = align("center")
    if merge and col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end
        )

def col_header(ws, row, col, text, center=True):
    """Light blue column header."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = fnt(bold=True)
    cell.fill = fill(C_HDR_LIGHT)
    cell.alignment = align("center" if center else "left")

def input_cell(ws, row, col, value, fmt=None, comment_text=None):
    """Blue-font input cell on light grey fill."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = fnt(color=C_BLUE_FONT)
    cell.fill = fill(C_INPUT_FILL)
    cell.alignment = align("right")
    if fmt:
        cell.number_format = fmt
    if comment_text:
        from openpyxl.comments import Comment
        cell.comment = Comment(comment_text, "Model")
    return cell

def formula_cell(ws, row, col, formula, fmt=None):
    """Black-font formula cell."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = fnt(color=C_BLACK_FONT)
    cell.alignment = align("right")
    if fmt:
        cell.number_format = fmt
    return cell

def label_cell(ws, row, col, text, indent=0, bold=False):
    """Row label (left-aligned)."""
    cell = ws.cell(row=row, column=col, value=("  " * indent) + text)
    cell.font = fnt(bold=bold)
    cell.alignment = align("left")
    return cell

def output_row_cell(ws, row, col, value_or_formula, fmt=None, is_formula=False):
    """Medium-blue fill output row."""
    cell = ws.cell(row=row, column=col, value=value_or_formula)
    cell.fill = fill(C_OUTPUT_FILL)
    cell.font = fnt(bold=True, color=C_BLACK_FONT if is_formula else C_BLUE_FONT)
    cell.alignment = align("right")
    if fmt:
        cell.number_format = fmt
    return cell

# ── Number formats ────────────────────────────────────────────────────────────
FMT_USD   = '#,##0.0;(#,##0.0);"-"'
FMT_USD2  = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT   = '0.0%'
FMT_PCT2  = '0.00%'
FMT_PRICE = '$#,##0.00'
FMT_INT   = '#,##0'
FMT_MULT  = '0.0x'

# ── Years ─────────────────────────────────────────────────────────────────────
HIST_YEARS = [2023, 2024, 2025]
PROJ_YEARS = [2026, 2027, 2028, 2029, 2030, 2031, 2032]
ALL_YEARS  = HIST_YEARS + PROJ_YEARS

# ── Historical data ───────────────────────────────────────────────────────────
HIST = {
    2023: {"rev": 79.9,  "da": 21.6, "ebit": -13.4, "ebit_m": None},
    2024: {"rev": 82.1,  "da": 20.3, "ebit": -16.8, "ebit_m": None},
    2025: {"rev": 111.8, "da": 26.1, "ebit":  17.9, "ebit_m": 0.160},
}

# ── Scenario assumptions ──────────────────────────────────────────────────────
# Growth rates by scenario: index 0=2026E ... 6=2032E
REV_GROWTH = {
    "Bear": [0.000, 0.020, 0.030, 0.020, 0.020, 0.010, 0.010],
    "Base": [0.030, 0.050, 0.060, 0.050, 0.040, 0.030, 0.030],
    "Bull": [0.050, 0.080, 0.100, 0.080, 0.070, 0.050, 0.040],
}
EBIT_MARGIN = {
    "Bear": [0.150, 0.150, 0.150, 0.150, 0.150, 0.140, 0.140],
    "Base": [0.200, 0.210, 0.220, 0.220, 0.220, 0.230, 0.230],
    "Bull": [0.260, 0.270, 0.280, 0.280, 0.280, 0.300, 0.300],
}
DA_PCT = {"Bear": 0.25, "Base": 0.24, "Bull": 0.22}
CAPEX_PCT = {"Bear": 0.005, "Base": 0.003, "Bull": 0.002}
NWC_PCT   = {"Bear": 0.020, "Base": 0.010, "Bull": 0.000}
TAX_RATE  = {"Bear": 0.00,  "Base": 0.00,  "Bull": 0.00}
TERM_G    = {"Bear": 0.015, "Base": 0.025, "Bull": 0.035}
WACC_S    = {"Bear": 0.073, "Base": 0.072, "Bull": 0.070}

# ── Market / balance sheet inputs ────────────────────────────────────────────
STOCK_PRICE    = 1.84
SHARES_BASIC   = 19.25    # M
SHARES_DILUTED = 23.82    # M
NET_DEBT       = 598.0    # M  (679.5 debt - 81.5 cash, Q2 2026)
TOTAL_DEBT     = 679.5    # M
TOTAL_CASH     = 81.5     # M
RF_RATE        = 0.0468
BETA           = 1.20
ERP            = 0.0550
CSRP           = 0.015


# ════════════════════════════════════════════════════════════════════════════
# WACC SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_wacc_sheet(wb):
    ws = wb.create_sheet("WACC")

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "SPRU – Weighted Average Cost of Capital (WACC)"
    c.font = fnt(bold=True, size=13, color=C_WHITE)
    c.fill = fill(C_HDR_DARK)
    c.alignment = align("center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "Spruce Power Holding Corporation  |  As of 2026-08-17  |  All $ in millions"
    c.font = fnt(italic=True, size=9)
    c.alignment = align("center")

    # ── Scenario headers row 4 ────────────────────────────────────────────
    row = 4
    ws.cell(row=row, column=1, value="Input / Calculation").font = fnt(bold=True)
    for i, s in enumerate(["Bear", "Base", "Bull"], start=3):
        col_header(ws, row, i, s)

    # ── COST OF EQUITY ────────────────────────────────────────────────────
    section_header(ws, 5, 1, 5, "COST OF EQUITY")

    def wacc_input(r, label, bear, base, bull, fmt=FMT_PCT2, comment=""):
        label_cell(ws, r, 1, label)
        for col, val, scen in [(3, bear, "Bear"), (4, base, "Base"), (5, bull, "Bull")]:
            src = f"Source: Model assumptions, 2026-08-17. {comment}" if comment else "Source: Model assumptions, 2026-08-17."
            input_cell(ws, r, col, val, fmt=fmt, comment_text=src)

    wacc_input(6,  "Risk-Free Rate (10Y Treasury)",
               RF_RATE, RF_RATE, RF_RATE,
               comment="FRED DGS10, 2026-08-14: 4.68%")
    wacc_input(7,  "Beta (5-Year Monthly)",
               BETA, BETA, BETA,
               comment="StockAnalysis.com, 2026-08-14, 5Y monthly vs S&P 500")
    wacc_input(8,  "Equity Risk Premium (ERP)",
               ERP, ERP, ERP,
               comment="Damodaran ERP estimate 2026")
    wacc_input(9,  "Company-Specific Risk Premium (CSRP)",
               0.020, 0.015, 0.010,
               comment="Governance dispute + illiquidity + small-cap; Bear=2.0%, Base=1.5%, Bull=1.0%")

    # Ke formula: =B6 + B7*B8 + B9
    label_cell(ws, 10, 1, "Cost of Equity (Ke)", bold=True)
    for col, rf, beta_v, erp_v, csrp_r in [
        (3, "C6", "C7", "C8", "C9"),
        (4, "D6", "D7", "D8", "D9"),
        (5, "E6", "E7", "E8", "E9"),
    ]:
        formula_cell(ws, 10, col,
                     f"={rf}+{beta_v}*{erp_v}+{csrp_r}", fmt=FMT_PCT2)
        ws.cell(row=10, column=col).font = fnt(bold=True, color=C_BLACK_FONT)
        ws.cell(row=10, column=col).fill = fill(C_OUTPUT_FILL)

    # ── COST OF DEBT ──────────────────────────────────────────────────────
    section_header(ws, 12, 1, 5, "COST OF DEBT")
    wacc_input(13, "Pre-Tax Cost of Debt (Kd)",
               0.065, 0.062, 0.058,
               comment="Q2 2026 blended rate 6.2%; Bear premium +0.3%, Bull discount -0.4%")
    wacc_input(14, "Effective Tax Rate",
               0.00, 0.00, 0.00,
               comment="NOL carryforward shields all taxable income in forecast period")

    label_cell(ws, 15, 1, "After-Tax Cost of Debt", bold=True)
    for col, kd, tr in [(3,"C13","C14"), (4,"D13","D14"), (5,"E13","E14")]:
        formula_cell(ws, 15, col, f"={kd}*(1-{tr})", fmt=FMT_PCT2)
        ws.cell(row=15, column=col).font = fnt(bold=True, color=C_BLACK_FONT)
        ws.cell(row=15, column=col).fill = fill(C_OUTPUT_FILL)

    # ── CAPITAL STRUCTURE ─────────────────────────────────────────────────
    section_header(ws, 17, 1, 5, "CAPITAL STRUCTURE (Normalised)")

    wacc_input(18, "Total Debt ($M)",
               TOTAL_DEBT, TOTAL_DEBT, TOTAL_DEBT,
               fmt=FMT_USD,
               comment=f"Q2 2026 10-Q balance sheet: non-recourse project debt ${TOTAL_DEBT}M")
    wacc_input(19, "Cash & Equivalents incl. Restricted ($M)",
               TOTAL_CASH, TOTAL_CASH, TOTAL_CASH,
               fmt=FMT_USD,
               comment=f"Q2 2026 press release: total cash ${TOTAL_CASH}M ($4.24/share)")

    label_cell(ws, 20, 1, "Net Debt ($M)", bold=True)
    for col, d, c_ in [(3,"C18","C19"),(4,"D18","D19"),(5,"E18","E19")]:
        formula_cell(ws, 20, col, f"={d}-{c_}", fmt=FMT_USD)
        ws.cell(row=20, column=col).fill = fill(C_OUTPUT_FILL)
        ws.cell(row=20, column=col).font = fnt(bold=True)

    wacc_input(21, "Equity Weight (E/EV) – Normalised",
               0.10, 0.15, 0.20,
               comment="Normalised: Bear=10%, Base=15%, Bull=20% (accounts for depressed market cap vs book)")
    label_cell(ws, 22, 1, "Debt Weight (D/EV) – Normalised")
    for col, ew in [(3,"C21"),(4,"D21"),(5,"E21")]:
        formula_cell(ws, 22, col, f"=1-{ew}", fmt=FMT_PCT2)

    # ── WACC SUMMARY ──────────────────────────────────────────────────────
    section_header(ws, 24, 1, 5, "WACC CALCULATION")
    label_cell(ws, 25, 1, "  Ke × E/EV")
    label_cell(ws, 26, 1, "  Kd(1-T) × D/EV")

    for col, ke, ew, kd_at, dw in [
        (3,"C10","C21","C15","C22"),
        (4,"D10","D21","D15","D22"),
        (5,"E10","E21","E15","E22"),
    ]:
        formula_cell(ws, 25, col, f"={ke}*{ew}", fmt=FMT_PCT2)
        formula_cell(ws, 26, col, f"={kd_at}*{dw}", fmt=FMT_PCT2)

    label_cell(ws, 27, 1, "WACC", bold=True)
    ws.row_dimensions[27].height = 16
    for col, r25, r26 in [(3,"C25","C26"),(4,"D25","D26"),(5,"E25","E26")]:
        c2 = ws.cell(row=27, column=col,
                     value=f"={r25}+{r26}")
        c2.font = fnt(bold=True, size=11, color=C_BLACK_FONT)
        c2.fill = fill(C_OUTPUT_FILL)
        c2.number_format = FMT_PCT2
        c2.alignment = align("center")

    thick_box(ws, 24, 27, 1, 5)

    # ── Scenario labels row 4 ─────────────────────────────────────────────
    ws.row_dimensions[4].height = 15

    return ws


# ════════════════════════════════════════════════════════════════════════════
# DCF SHEET – helpers used in next sections
# ════════════════════════════════════════════════════════════════════════════

# Column layout on DCF sheet
# A=labels, B=2023A, C=2024A, D=2025A,
# E=2026E, F=2027E, G=2028E, H=2029E, I=2030E, J=2031E, K=2032E
# L = "Selected" consolidation column (INDEX formula)
# M = Bear, N = Base, O = Bull  (assumption blocks, further right)
YEAR_COLS  = {y: 2+i for i, y in enumerate(ALL_YEARS)}  # B=2, C=3 ...
PROJ_COLS  = {y: YEAR_COLS[y] for y in PROJ_YEARS}       # E..K
COL_SEL    = 12  # L – consolidation (INDEX)
COL_BEAR   = 14  # N
COL_BASE   = 15  # O
COL_BULL   = 16  # P
COL_LABELS = 1   # A
LAST_DATA_COL = 11  # K (2032E)


def col(y):
    """Return column index for year y."""
    return YEAR_COLS[y]


if __name__ == "__main__":
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # ── Build WACC sheet first ───────────────────────────────────────────
    build_wacc_sheet(wb)

    print("Part 1 (helpers + WACC sheet) ready – DCF sheet will be added by next script.")
    wb.save(OUTPUT_PATH)
    print(f"Saved interim: {OUTPUT_PATH}")
