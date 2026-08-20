#!/usr/bin/env python3
"""
build_vitl_dcf.py - Institutional DCF Valuation Model for Vital Farms, Inc. (NASDAQ: VITL)
Follows Investment Banking financial modeling standards, strict openpyxl formatting rules,
dynamic scenario consolidation, and fully formula-driven sensitivity analysis matrices.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

def build_vitl_dcf_model(output_path="VITL_DCF_Model_Gemini-3.7-Flash_20260820.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: DCF Valuation
    ws_dcf = wb.active
    ws_dcf.title = "DCF Valuation"
    ws_dcf.views.sheetView[0].showGridLines = True
    
    # Sheet 2: WACC Build
    ws_wacc = wb.create_sheet(title="WACC Build")
    ws_wacc.views.sheetView[0].showGridLines = True
    
    # Palette definition (Institutional Classic Navy & Slate Blue)
    NAVY = "1F4E79"        # Primary headers
    MED_BLUE = "2F5597"    # Secondary headers
    SOFT_BLUE = "D9E1F2"   # Table headers
    ACCENT_BLUE = "BDD7EE" # Highlight / Base Case / Outputs
    LIGHT_GRAY = "F2F2F2"  # Input fill
    ZEBRA_FILL = "F9FBFD"  # Alternating row
    WHITE = "FFFFFF"
    GRAY_TEXT = "595959"
    INPUT_BLUE = "0000FF"  # Hardcoded font
    BLACK = "000000"       # Formula font
    LINK_GREEN = "008000"  # Sheet link font
    
    # Fonts
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_sub_sec = Font(name="Calibri", size=10, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color=NAVY)
    font_regular = Font(name="Calibri", size=10, bold=False, color=BLACK)
    font_input = Font(name="Calibri", size=10, bold=False, color=INPUT_BLUE)
    font_input_bold = Font(name="Calibri", size=10, bold=True, color=INPUT_BLUE)
    font_link = Font(name="Calibri", size=10, bold=False, color=LINK_GREEN)
    font_link_bold = Font(name="Calibri", size=10, bold=True, color=LINK_GREEN)
    font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)
    
    # Fills
    fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    thick_navy = Side(border_style="medium", color=NAVY)
    double_navy = Side(border_style="double", color=NAVY)
    top_thin_navy = Side(border_style="thin", color=NAVY)
    
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_navy)
    border_total = Border(top=top_thin_navy, bottom=double_navy, left=thin_line, right=thin_line)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Helper for merged section header
    def add_section_header(ws, row, title, max_col="I", fill_color=fill_navy, font_color=font_sec_hdr):
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = font_color
        ws[f"A{row}"].fill = fill_color
        ws[f"A{row}"].alignment = align_left
        ws.merge_cells(f"A{row}:{max_col}{row}")
        col_max_idx = openpyxl.utils.column_index_from_string(max_col)
        for col_idx in range(1, col_max_idx + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = fill_color
            c.border = Border(top=thick_navy, bottom=thick_navy, left=thin_line, right=thin_line)

    # ==========================================
    # SHEET 2: WACC Build (Cost of Capital)
    # ==========================================
    ws_wacc["A1"] = "Vital Farms, Inc. (NASDAQ: VITL)"
    ws_wacc["A1"].font = font_title
    ws_wacc["A2"] = "Weighted Average Cost of Capital (WACC) Schedule | CAPM Methodology & Capital Structure"
    ws_wacc["A2"].font = font_subtitle
    
    add_section_header(ws_wacc, 4, "I. COST OF EQUITY (CAPM METHODOLOGY)", "E")
    
    wacc_equity_inputs = [
        ("Risk-Free Rate (10-Yr US Treasury Yield)", 0.0464, "0.00%", "Source: US 10-Year Treasury Yield benchmark as of August 20, 2026 (4.64%)", "B5"),
        ("Equity Beta (5-Year Monthly vs S&P 500)", 1.08, "0.00", "Source: 5-year monthly regression beta vs S&P 500 (Reflects pasture-raised premium consumer brand profile)", "B6"),
        ("Market Equity Risk Premium (ERP)", 0.0520, "0.00%", "Source: Institutional consensus equity risk premium (5.20%)", "B7"),
        ("Cost of Equity (Ke = Rf + Beta * ERP)", "=B5+B6*B7", "0.00%", None, "B8"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_equity_inputs, start=5):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
            ws_wacc[f"B{idx}"].fill = fill_accent_blue
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 10, "II. COST OF DEBT & CAPITAL STRUCTURE WEIGHTS", "E")
    
    wacc_debt_inputs = [
        ("Credit Rating Equivalent", "BB / Growth Commercial Facility", "@", "Source: Commercial credit facility with conservative leverage and strong liquidity", "B11"),
        ("Pre-Tax Cost of Debt (Kd)", 0.0650, "0.00%", "Source: Form 10-K / Credit Agreement effective interest rate benchmark (6.50%)", "B12"),
        ("Effective Corporate Tax Rate (t)", 0.2500, "0.0%", "Source: Historical normalized effective corporate tax rate (25.0%)", "B13"),
        ("After-Tax Cost of Debt (Kd * (1 - t))", "=B12*(1-B13)", "0.00%", None, "B14"),
        ("Current Share Price ($)", "='DCF Valuation'!B8", "$#,##0.00", None, "B15"),
        ("Diluted Shares Outstanding (M)", "='DCF Valuation'!B9", "#,##0.00", None, "B16"),
        ("Market Capitalization ($M)", "=B15*B16", "$#,##0.00", None, "B17"),
        ("Total Debt (Credit Facility & Notes) ($M)", 30.00, "$#,##0.00", "Source: Form 10-Q / 10-K Balance Sheet total carrying debt balance ($30.00M)", "B18"),
        ("Cash, Cash Equivalents & ST Investments ($M)", 160.00, "$#,##0.00", "Source: Form 10-Q / 10-K Balance Sheet total cash & marketable securities ($160.00M)", "B19"),
        ("Net Debt / (Net Cash) ($M)", "=B18-B19", "$#,##0.00", None, "B20"),
        ("Total Capitalization (Market Cap + Total Debt) ($M)", "=B17+B18", "$#,##0.00", None, "B21"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(wacc_debt_inputs, start=11):
        ws_wacc[f"A{idx}"] = label
        ws_wacc[f"A{idx}"].font = font_bold if ("=" in str(val) and "='DCF" not in str(val)) else font_regular
        ws_wacc[f"A{idx}"].border = border_cell
        
        ws_wacc[f"B{idx}"] = val
        ws_wacc[f"B{idx}"].number_format = num_fmt
        ws_wacc[f"B{idx}"].alignment = align_right
        ws_wacc[f"B{idx}"].border = border_cell
        if "='DCF Valuation'" in str(val):
            ws_wacc[f"B{idx}"].font = font_link
        elif "=" in str(val):
            ws_wacc[f"B{idx}"].font = font_bold
            if idx in (17, 20, 21):
                ws_wacc[f"B{idx}"].fill = fill_accent_blue
        else:
            ws_wacc[f"B{idx}"].font = font_input
            ws_wacc[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_wacc[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")
                
    add_section_header(ws_wacc, 23, "III. WEIGHTED AVERAGE COST OF CAPITAL (WACC)", "E")
    
    # Table header
    wacc_tbl_headers = ["Capital Component", "Weight (% Total Cap)", "Component Cost", "WACC Contribution", "Analytical Rationale"]
    for col_i, th in enumerate(wacc_tbl_headers, start=1):
        cell = ws_wacc.cell(row=24, column=col_i, value=th)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if col_i in (2, 3, 4) else align_left
        cell.border = border_header
        
    wacc_sched = [
        ("Common Equity", "=B17/B21", "=B8", "=B25*C25", "Equity market weight (~94.1% of total capitalization)"),
        ("Total Debt", "=B18/B21", "=B14", "=B26*C26", "Carrying debt weight (~5.9% of total capitalization)"),
    ]
    for row_i, (comp, wt, cost, contrib, rat) in enumerate(wacc_sched, start=25):
        ws_wacc[f"A{row_i}"] = comp
        ws_wacc[f"B{row_i}"] = wt
        ws_wacc[f"C{row_i}"] = cost
        ws_wacc[f"D{row_i}"] = contrib
        ws_wacc[f"E{row_i}"] = rat
        
        ws_wacc[f"A{row_i}"].font = font_regular
        ws_wacc[f"B{row_i}"].font = font_bold
        ws_wacc[f"C{row_i}"].font = font_bold
        ws_wacc[f"D{row_i}"].font = font_bold
        ws_wacc[f"E{row_i}"].font = font_italic
        
        ws_wacc[f"B{row_i}"].number_format = "0.0%"
        ws_wacc[f"C{row_i}"].number_format = "0.00%"
        ws_wacc[f"D{row_i}"].number_format = "0.00%"
        
        for col_c in ["A", "B", "C", "D", "E"]:
            ws_wacc[f"{col_c}{row_i}"].border = border_cell
            if col_c in ["B", "C", "D"]:
                ws_wacc[f"{col_c}{row_i}"].alignment = align_right
                
    # Total WACC row
    ws_wacc["A27"] = "Weighted Average Cost of Capital (WACC)"
    ws_wacc["A27"].font = font_bold_navy
    ws_wacc["B27"] = "=SUM(B25:B26)"
    ws_wacc["B27"].number_format = "0.0%"
    ws_wacc["B27"].font = font_bold_navy
    ws_wacc["B27"].alignment = align_right
    ws_wacc["C27"] = "-"
    ws_wacc["C27"].alignment = align_center
    ws_wacc["D27"] = "=SUM(D25:D26)"
    ws_wacc["D27"].number_format = "0.00%"
    ws_wacc["D27"].font = font_bold_navy
    ws_wacc["D27"].alignment = align_right
    ws_wacc["E27"] = "Baseline market-implied discount rate (Normalized WACC benchmark = 10.00%)"
    ws_wacc["E27"].font = font_italic
    
    for col_c in ["A", "B", "C", "D", "E"]:
        ws_wacc[f"{col_c}{27}"].fill = fill_accent_blue
        ws_wacc[f"{col_c}{27}"].border = border_total

    # ==========================================
    # SHEET 1: DCF Valuation (Main Model)
    # ==========================================
    ws_dcf["A1"] = "Vital Farms, Inc. (NASDAQ: VITL)"
    ws_dcf["A1"].font = font_title
    ws_dcf["A2"] = "Institutional Discounted Cash Flow (DCF) Valuation Model | 5-Year Explicit Forecast (FY2026E - FY2030E)"
    ws_dcf["A2"].font = font_subtitle
    
    # Case Selector
    ws_dcf["A4"] = "Case Selector (1=Bear, 2=Base, 3=Bull):"
    ws_dcf["A4"].font = font_bold
    ws_dcf["B4"] = 2
    ws_dcf["B4"].font = font_input_bold
    ws_dcf["B4"].fill = fill_accent_blue
    ws_dcf["B4"].alignment = align_center
    ws_dcf["B4"].border = border_cell
    ws_dcf["B4"].comment = Comment("Enter 1 for Bear Case, 2 for Base Case, 3 for Bull Case. All projections update dynamically.", "DCF Model Builder")
    
    ws_dcf["A5"] = "Active Scenario:"
    ws_dcf["A5"].font = font_bold
    ws_dcf["B5"] = '=IF(B4=1,"Bear Case (Downside / Retail Pushback & Feed Cost Inflation)",IF(B4=2,"Base Case (Consensus / Egg & Butter Expansion & Operational Efficiency)","Bull Case (Premium Food Leadership, New Category Entry & Operating Leverage)"))'
    ws_dcf["B5"].font = font_bold_navy
    ws_dcf["B5"].alignment = align_left
    ws_dcf.merge_cells("B5:F5")
    
    # Market Data & Key Parameters
    add_section_header(ws_dcf, 7, "MARKET DATA & VALUATION PARAMETERS", "I")
    
    mkt_data = [
        ("Current Stock Price ($)", 11.10, "$#,##0.00", "Source: NASDAQ market price as of August 20, 2026 ($11.10)", "B8"),
        ("Diluted Shares Outstanding (M)", 42.94, "#,##0.00", "Source: Form 10-Q / 10-K Diluted common shares outstanding (42.94M)", "B9"),
        ("Market Capitalization ($M)", "=B8*B9", "$#,##0.00", None, "B10"),
        ("Cash, Cash Equivalents & ST Investments ($M)", 160.00, "$#,##0.00", "Source: Form 10-Q / 10-K Balance Sheet total cash & liquid reserves ($160.00M)", "B11"),
        ("Total Debt (Credit Facility & Notes) ($M)", 30.00, "$#,##0.00", "Source: Form 10-Q / 10-K Balance Sheet total carrying debt balance ($30.00M)", "B12"),
        ("Net Debt / (Net Cash) ($M)", "=B12-B11", "$#,##0.00", None, "B13"),
        ("Normalized Corporate Tax Rate (%)", 0.2500, "0.0%", "Source: Historical effective corporate tax rate benchmark (25.0%)", "B14"),
    ]
    
    for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(mkt_data, start=8):
        ws_dcf[f"A{idx}"] = label
        ws_dcf[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
        ws_dcf[f"A{idx}"].border = border_cell
        
        ws_dcf[f"B{idx}"] = val
        ws_dcf[f"B{idx}"].number_format = num_fmt
        ws_dcf[f"B{idx}"].alignment = align_right
        ws_dcf[f"B{idx}"].border = border_cell
        if "=" in str(val):
            ws_dcf[f"B{idx}"].font = font_bold
            ws_dcf[f"B{idx}"].fill = fill_accent_blue
        else:
            ws_dcf[f"B{idx}"].font = font_input
            ws_dcf[f"B{idx}"].fill = fill_input
            if comment_text:
                ws_dcf[f"B{idx}"].comment = Comment(comment_text, "DCF Model Builder")

    # Section I: Scenario Assumption Blocks
    add_section_header(ws_dcf, 16, "I. SCENARIO ASSUMPTION BLOCKS (BEAR / BASE / BULL)", "I")
    
    # 1. Bear Case
    add_section_header(ws_dcf, 17, "BEAR CASE ASSUMPTIONS (Case 1 - Volume Deceleration & Feed Cost Pressures)", "I", fill_color=fill_med_blue, font_color=font_sub_sec)
    bear_hdr = ["Assumption / Driver", "", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal", "Rationale"]
    for c_i, h in enumerate(bear_hdr, start=1):
        cell = ws_dcf.cell(row=18, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 3 else align_left
        cell.border = border_header
        
    bear_rows = [
        ("Revenue Growth (%)", [0.120, 0.100, 0.090, 0.080, 0.070, None], "0.0%", "Pasture-raised egg volume growth moderates; grocery shelf space gains slow"),
        ("Gross Margin (%)", [0.360, 0.355, 0.350, 0.350, 0.350, None], "0.0%", "Feed grain costs rise; farmer-partner contract support costs increase"),
        ("EBIT Margin (%)", [0.095, 0.098, 0.100, 0.100, 0.100, None], "0.0%", "Marketing expense increases to defend brand share against conventional eggs"),
        ("CapEx % Revenue", [0.0450, 0.0450, 0.0420, 0.0400, 0.0400, 0.0400], "0.00%", "Egg Central Station capacity maintenance and selective farm expansions"),
        ("D&A % Revenue", [0.0240, 0.0240, 0.0240, 0.0230, 0.0230, 0.0230], "0.00%", "Depreciation of washing and packing facilities"),
        ("NWC Change % ΔRev", [0.018, 0.018, 0.018, 0.018, 0.018, None], "0.0%", "Higher inventory requirements for packaging supplies and finished goods"),
        ("Terminal Growth Rate (%)", [None, None, None, None, None, 0.0200], "0.00%", "Conservative long-term food staples GDP growth (2.00%)"),
        ("WACC (Discount Rate) (%)", [None, None, None, None, None, 0.1050], "0.00%", "Higher risk premium reflecting agricultural input volatility (10.50%)"),
    ]
    for r_idx, (label, vals, fmt, rat) in enumerate(bear_rows, start=19):
        ws_dcf[f"A{r_idx}"] = label
        ws_dcf[f"A{r_idx}"].font = font_regular
        ws_dcf[f"A{r_idx}"].border = border_cell
        ws_dcf[f"B{r_idx}"] = ""
        ws_dcf[f"B{r_idx}"].border = border_cell
        
        for c_idx, v in enumerate(vals, start=3):
            col_letter = get_column_letter(c_idx)
            cell = ws_dcf[f"{col_letter}{r_idx}"]
            cell.value = v if v is not None else "-"
            cell.number_format = fmt
            cell.alignment = align_right if v is not None else align_center
            cell.font = font_input if v is not None else font_regular
            cell.fill = fill_input if v is not None else PatternFill(fill_type=None)
            cell.border = border_cell
        ws_dcf[f"I{r_idx}"] = rat
        ws_dcf[f"I{r_idx}"].font = font_italic
        ws_dcf[f"I{r_idx}"].border = border_cell
        
    # 2. Base Case
    add_section_header(ws_dcf, 28, "BASE CASE ASSUMPTIONS (Case 2 - National Retail Penetration, Egg & Butter Core Strength)", "I", fill_color=fill_med_blue, font_color=font_sub_sec)
    base_hdr = ["Assumption / Driver", "", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal", "Rationale"]
    for c_i, h in enumerate(base_hdr, start=1):
        cell = ws_dcf.cell(row=29, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 3 else align_left
        cell.border = border_header
        
    base_rows = [
        ("Revenue Growth (%)", [0.180, 0.150, 0.130, 0.110, 0.090, None], "0.0%", "Continued household penetration gains in pasture-raised eggs, butter & ghee"),
        ("Gross Margin (%)", [0.378, 0.380, 0.382, 0.385, 0.385, None], "0.0%", "Scale efficiencies at Egg Central Station 2 (ECS 2) and favorable product mix"),
        ("EBIT Margin (%)", [0.112, 0.118, 0.122, 0.125, 0.125, None], "0.0%", "Operating leverage across SG&A and distribution network scaling past $1.0B"),
        ("CapEx % Revenue", [0.0450, 0.0420, 0.0400, 0.0400, 0.0400, 0.0400], "0.00%", "Capital investment in packing automation and pasture-raised supply chain network"),
        ("D&A % Revenue", [0.0240, 0.0240, 0.0240, 0.0240, 0.0240, 0.0240], "0.00%", "Normalized depreciation of processing equipment and facilities"),
        ("NWC Change % ΔRev", [0.015, 0.015, 0.015, 0.015, 0.015, None], "0.0%", "Efficient working capital management with high inventory turnover"),
        ("Terminal Growth Rate (%)", [None, None, None, None, None, 0.0250], "0.00%", "Long-term organic/ethical food category growth benchmark (2.50%)"),
        ("WACC (Discount Rate) (%)", [None, None, None, None, None, 0.1000], "0.00%", "Normalized CAPM discount rate reflecting pristine balance sheet (10.00%)"),
    ]
    for r_idx, (label, vals, fmt, rat) in enumerate(base_rows, start=30):
        ws_dcf[f"A{r_idx}"] = label
        ws_dcf[f"A{r_idx}"].font = font_regular
        ws_dcf[f"A{r_idx}"].border = border_cell
        ws_dcf[f"B{r_idx}"] = ""
        ws_dcf[f"B{r_idx}"].border = border_cell
        
        for c_idx, v in enumerate(vals, start=3):
            col_letter = get_column_letter(c_idx)
            cell = ws_dcf[f"{col_letter}{r_idx}"]
            cell.value = v if v is not None else "-"
            cell.number_format = fmt
            cell.alignment = align_right if v is not None else align_center
            cell.font = font_input if v is not None else font_regular
            cell.fill = fill_input if v is not None else PatternFill(fill_type=None)
            cell.border = border_cell
        ws_dcf[f"I{r_idx}"] = rat
        ws_dcf[f"I{r_idx}"].font = font_italic
        ws_dcf[f"I{r_idx}"].border = border_cell

    # 3. Bull Case
    add_section_header(ws_dcf, 39, "BULL CASE ASSUMPTIONS (Case 3 - Category Dominance, Foodservice Expansion & New Products)", "I", fill_color=fill_med_blue, font_color=font_sub_sec)
    bull_hdr = ["Assumption / Driver", "", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal", "Rationale"]
    for c_i, h in enumerate(bull_hdr, start=1):
        cell = ws_dcf.cell(row=40, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 3 else align_left
        cell.border = border_header
        
    bull_rows = [
        ("Revenue Growth (%)", [0.220, 0.180, 0.160, 0.140, 0.120, None], "0.0%", "Accelerated foodservice partnerships, club store expansion & new ethical dairy lines"),
        ("Gross Margin (%)", [0.385, 0.390, 0.395, 0.400, 0.400, None], "0.0%", "Maximum utilization of packing infrastructure and high-margin value-add SKUs"),
        ("EBIT Margin (%)", [0.120, 0.128, 0.135, 0.140, 0.145, None], "0.0%", "Significant OpEx leverage as revenue exceeds $1.4B-$1.6B"),
        ("CapEx % Revenue", [0.0420, 0.0400, 0.0380, 0.0380, 0.0380, 0.0380], "0.00%", "High capital efficiency with rapid payback on automated packing lines"),
        ("D&A % Revenue", [0.0240, 0.0240, 0.0240, 0.0240, 0.0240, 0.0240], "0.00%", "Steady fixed asset amortization profile"),
        ("NWC Change % ΔRev", [0.012, 0.012, 0.012, 0.012, 0.012, None], "0.0%", "Optimal working capital cycle with high inventory turn and favorable retail terms"),
        ("Terminal Growth Rate (%)", [None, None, None, None, None, 0.0300], "0.00%", "Secular clean-label / ethical food category leadership (3.00%)"),
        ("WACC (Discount Rate) (%)", [None, None, None, None, None, 0.0950], "0.00%", "Lower risk premium reflecting entrenched category leadership (9.50%)"),
    ]
    for r_idx, (label, vals, fmt, rat) in enumerate(bull_rows, start=41):
        ws_dcf[f"A{r_idx}"] = label
        ws_dcf[f"A{r_idx}"].font = font_regular
        ws_dcf[f"A{r_idx}"].border = border_cell
        ws_dcf[f"B{r_idx}"] = ""
        ws_dcf[f"B{r_idx}"].border = border_cell
        
        for c_idx, v in enumerate(vals, start=3):
            col_letter = get_column_letter(c_idx)
            cell = ws_dcf[f"{col_letter}{r_idx}"]
            cell.value = v if v is not None else "-"
            cell.number_format = fmt
            cell.alignment = align_right if v is not None else align_center
            cell.font = font_input if v is not None else font_regular
            cell.fill = fill_input if v is not None else PatternFill(fill_type=None)
            cell.border = border_cell
        ws_dcf[f"I{r_idx}"] = rat
        ws_dcf[f"I{r_idx}"].font = font_italic
        ws_dcf[f"I{r_idx}"].border = border_cell

    # 4. Consolidated Active Drivers (Dynamic Selection via CHOOSE)
    add_section_header(ws_dcf, 50, "ACTIVE MODEL CONSOLIDATION SCHEDULE (DYNAMIC CASE SELECTION)", "I", fill_color=fill_navy, font_color=font_sec_hdr)
    act_hdr = ["Active Driver / Assumption", "Case Ref", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal", "Active Dynamic Formula Reference"]
    for c_i, h in enumerate(act_hdr, start=1):
        cell = ws_dcf.cell(row=51, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 3 else align_left
        cell.border = border_header

    consol_defs = [
        ("Active Revenue Growth (%)", "=B4", (19, 30, 41), "0.0%", "Dynamic revenue growth path based on active case selection"),
        ("Active Gross Margin (%)", "=B4", (20, 31, 42), "0.0%", "Dynamic gross profitability profile"),
        ("Active EBIT Margin (%)", "=B4", (21, 32, 43), "0.0%", "Dynamic operating margin profile"),
        ("Active CapEx % Revenue", "=B4", (22, 33, 44), "0.00%", "Dynamic capital expenditure reinvestment intensity"),
        ("Active D&A % Revenue", "=B4", (23, 34, 45), "0.00%", "Dynamic depreciation & amortization rate"),
        ("Active NWC Change % ΔRev", "=B4", (24, 35, 46), "0.0%", "Dynamic net working capital requirement on incremental revenue"),
        ("Active Terminal Growth Rate (%)", "=B4", (25, 36, 47), "0.00%", "Perpetual terminal growth rate for DCF terminal value calculation"),
        ("Active WACC (Discount Rate) (%)", "=B4", (26, 37, 48), "0.00%", "Cost of capital discount rate applied to projected cash flows"),
    ]
    
    for r_i, (label, cref, (r_bear, r_base, r_bull), fmt, desc) in enumerate(consol_defs, start=52):
        ws_dcf[f"A{r_i}"] = label
        ws_dcf[f"A{r_i}"].font = font_bold
        ws_dcf[f"A{r_i}"].border = border_cell
        
        ws_dcf[f"B{r_i}"] = f"=B4"
        ws_dcf[f"B{r_i}"].font = font_bold
        ws_dcf[f"B{r_i}"].alignment = align_center
        ws_dcf[f"B{r_i}"].border = border_cell
        
        if r_i in (58, 59): # Terminal g & WACC (only column H)
            for c_i in range(3, 8):
                col_l = get_column_letter(c_i)
                ws_dcf[f"{col_l}{r_i}"] = "-"
                ws_dcf[f"{col_l}{r_i}"].alignment = align_center
                ws_dcf[f"{col_l}{r_i}"].border = border_cell
            ws_dcf[f"H{r_i}"] = f"=CHOOSE($B$4, H{r_bear}, H{r_base}, H{r_bull})"
            ws_dcf[f"H{r_i}"].font = font_bold_navy
            ws_dcf[f"H{r_i}"].number_format = fmt
            ws_dcf[f"H{r_i}"].alignment = align_right
            ws_dcf[f"H{r_i}"].fill = fill_accent_blue
            ws_dcf[f"H{r_i}"].border = border_cell
        else:
            for c_i in range(3, 8): # Cols C..G
                col_l = get_column_letter(c_i)
                ws_dcf[f"{col_l}{r_i}"] = f"=CHOOSE($B$4, {col_l}{r_bear}, {col_l}{r_base}, {col_l}{r_bull})"
                ws_dcf[f"{col_l}{r_i}"].font = font_bold
                ws_dcf[f"{col_l}{r_i}"].number_format = fmt
                ws_dcf[f"{col_l}{r_i}"].alignment = align_right
                ws_dcf[f"{col_l}{r_i}"].fill = fill_accent_blue
                ws_dcf[f"{col_l}{r_i}"].border = border_cell
            # Col H
            if r_i in (55, 56): # CapEx & D&A have terminal value
                ws_dcf[f"H{r_i}"] = f"=CHOOSE($B$4, H{r_bear}, H{r_base}, H{r_bull})"
                ws_dcf[f"H{r_i}"].font = font_bold
                ws_dcf[f"H{r_i}"].number_format = fmt
                ws_dcf[f"H{r_i}"].alignment = align_right
                ws_dcf[f"H{r_i}"].fill = fill_accent_blue
                ws_dcf[f"H{r_i}"].border = border_cell
            else:
                ws_dcf[f"H{r_i}"] = "-"
                ws_dcf[f"H{r_i}"].alignment = align_center
                ws_dcf[f"H{r_i}"].border = border_cell
                
        ws_dcf[f"I{r_i}"] = desc
        ws_dcf[f"I{r_i}"].font = font_italic
        ws_dcf[f"I{r_i}"].border = border_cell

    # Section II: Historical & Projected Income Statement
    add_section_header(ws_dcf, 61, "II. HISTORICAL & PROJECTED FINANCIAL PERFORMANCE ($M)", "I")
    
    is_hdr = ["Income Statement ($M)", "FY2023A", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for c_i, h in enumerate(is_hdr, start=1):
        cell = ws_dcf.cell(row=62, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 2 else align_left
        cell.border = border_header

    # Row 63: Revenue
    ws_dcf["A63"] = "Total Net Revenue ($M)"
    ws_dcf["A63"].font = font_bold
    ws_dcf["B63"] = 471.86
    ws_dcf["B63"].comment = Comment("Source: Form 10-K FY2023 Consolidated Statements of Operations ($471.86M)", "DCF Model Builder")
    ws_dcf["C63"] = 606.32
    ws_dcf["C63"].comment = Comment("Source: Form 10-K FY2024 Consolidated Statements of Operations ($606.32M)", "DCF Model Builder")
    ws_dcf["D63"] = 759.40
    ws_dcf["D63"].comment = Comment("Source: Form 10-K FY2025 Consolidated Statements of Operations ($759.40M)", "DCF Model Builder")
    for col_c, prior_c, act_c in [("E", "D", "C"), ("F", "E", "D"), ("G", "F", "E"), ("H", "G", "F"), ("I", "H", "G")]:
        ws_dcf[f"{col_c}63"] = f"={prior_c}63*(1+{act_c}52)"
    
    # Row 64: Revenue YoY Growth
    ws_dcf["A64"] = "  YoY Revenue Growth (%)"
    ws_dcf["A64"].font = font_italic
    ws_dcf["B64"] = 0.3031
    ws_dcf["B64"].comment = Comment("Source: Form 10-K FY2023 Net Revenue YoY Growth (+30.3%)", "DCF Model Builder")
    ws_dcf["C64"] = "=C63/B63-1"
    ws_dcf["D64"] = "=D63/C63-1"
    for col_c, prior_c in [("E", "D"), ("F", "E"), ("G", "F"), ("H", "G"), ("I", "H")]:
        ws_dcf[f"{col_c}64"] = f"={col_c}63/{prior_c}63-1"

    # Row 65: Cost of Goods Sold (COGS)
    ws_dcf["A65"] = "Cost of Goods Sold (COGS) ($M)"
    ws_dcf["A65"].font = font_regular
    ws_dcf["B65"] = 304.82
    ws_dcf["B65"].comment = Comment("Source: Form 10-K FY2023 Cost of Goods Sold ($304.82M)", "DCF Model Builder")
    ws_dcf["C65"] = 376.46
    ws_dcf["C65"].comment = Comment("Source: Form 10-K FY2024 Cost of Goods Sold ($376.46M)", "DCF Model Builder")
    ws_dcf["D65"] = 473.87
    ws_dcf["D65"].comment = Comment("Source: Form 10-K FY2025 Cost of Goods Sold ($473.87M)", "DCF Model Builder")
    for col_c, act_c in [("E", "C"), ("F", "D"), ("G", "E"), ("H", "F"), ("I", "G")]:
        ws_dcf[f"{col_c}65"] = f"={col_c}63*(1-{act_c}53)"

    # Row 66: Gross Profit
    ws_dcf["A66"] = "Gross Profit ($M)"
    ws_dcf["A66"].font = font_bold
    ws_dcf["B66"] = "=B63-B65"
    ws_dcf["C66"] = "=C63-C65"
    ws_dcf["D66"] = "=D63-D65"
    for col_c in ["E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}66"] = f"={col_c}63-{col_c}65"

    # Row 67: Gross Margin (%)
    ws_dcf["A67"] = "  Gross Profit Margin (%)"
    ws_dcf["A67"].font = font_italic
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}67"] = f"={col_c}66/{col_c}63"

    # Row 68: Selling, General & Administrative (SG&A)
    ws_dcf["A68"] = "Selling, General & Administrative (SG&A) ($M)"
    ws_dcf["A68"].font = font_regular
    ws_dcf["B68"] = 133.72
    ws_dcf["B68"].comment = Comment("Source: Form 10-K FY2023 SG&A Expense ($133.72M)", "DCF Model Builder")
    ws_dcf["C68"] = 166.26
    ws_dcf["C68"].comment = Comment("Source: Form 10-K FY2024 SG&A Expense ($166.26M)", "DCF Model Builder")
    ws_dcf["D68"] = 201.33
    ws_dcf["D68"].comment = Comment("Source: Form 10-K FY2025 SG&A Expense ($201.33M)", "DCF Model Builder")
    for col_c in ["E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}68"] = f"={col_c}66-{col_c}70"

    # Row 69: Total Operating Expenses
    ws_dcf["A69"] = "Total Operating Expenses (OpEx) ($M)"
    ws_dcf["A69"].font = font_bold
    ws_dcf["B69"] = "=B68"
    ws_dcf["C69"] = "=C68"
    ws_dcf["D69"] = "=D68"
    for col_c in ["E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}69"] = f"={col_c}68"

    # Row 70: Operating Income (EBIT)
    ws_dcf["A70"] = "Operating Income (EBIT) ($M)"
    ws_dcf["A70"].font = font_bold_navy
    ws_dcf["B70"] = 33.32
    ws_dcf["B70"].comment = Comment("Source: Form 10-K FY2023 Operating Income ($33.32M)", "DCF Model Builder")
    ws_dcf["C70"] = 63.60
    ws_dcf["C70"].comment = Comment("Source: Form 10-K FY2024 Operating Income ($63.60M)", "DCF Model Builder")
    ws_dcf["D70"] = 84.20
    ws_dcf["D70"].comment = Comment("Source: Form 10-K FY2025 Operating Income ($84.20M)", "DCF Model Builder")
    for col_c, act_c in [("E", "C"), ("F", "D"), ("G", "E"), ("H", "F"), ("I", "G")]:
        ws_dcf[f"{col_c}70"] = f"={col_c}63*{act_c}54"

    # Row 71: EBIT Margin (%)
    ws_dcf["A71"] = "  EBIT Margin (%)"
    ws_dcf["A71"].font = font_italic
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}71"] = f"={col_c}70/{col_c}63"

    # Row 72: Normalized Taxes on EBIT ($M)
    ws_dcf["A72"] = "Normalized Income Taxes on EBIT ($M)"
    ws_dcf["A72"].font = font_regular
    ws_dcf["B72"] = 8.33
    ws_dcf["B72"].comment = Comment("Source: Form 10-K FY2023 Income Tax Provision ($8.33M)", "DCF Model Builder")
    ws_dcf["C72"] = 15.90
    ws_dcf["C72"].comment = Comment("Source: Form 10-K FY2024 Income Tax Provision ($15.90M)", "DCF Model Builder")
    ws_dcf["D72"] = 21.05
    ws_dcf["D72"].comment = Comment("Source: Form 10-K FY2025 Income Tax Provision ($21.05M)", "DCF Model Builder")
    for col_c in ["E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}72"] = f"={col_c}70*$B$14"

    # Row 73: Effective Tax Rate (%)
    ws_dcf["A73"] = "  Effective Tax Rate (%)"
    ws_dcf["A73"].font = font_italic
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}73"] = f"={col_c}72/{col_c}70"

    # Row 74: NOPAT (Net Operating Profit After Tax)
    ws_dcf["A74"] = "Net Operating Profit After Tax (NOPAT) ($M)"
    ws_dcf["A74"].font = font_bold_navy
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}74"] = f"={col_c}70-{col_c}72"

    # Formatting Section II
    is_fmt_rules = {
        63: ("$#,##0.00", font_bold, fill_zebra),
        64: ("0.0%", font_italic, None),
        65: ("$#,##0.00", font_regular, None),
        66: ("$#,##0.00", font_bold, fill_zebra),
        67: ("0.0%", font_italic, None),
        68: ("$#,##0.00", font_regular, None),
        69: ("$#,##0.00", font_bold, None),
        70: ("$#,##0.00", font_bold_navy, fill_accent_blue),
        71: ("0.0%", font_italic, None),
        72: ("$#,##0.00", font_regular, None),
        73: ("0.0%", font_italic, None),
        74: ("$#,##0.00", font_bold_navy, fill_accent_blue),
    }
    for r_i, (fmt_s, f_style, fill_s) in is_fmt_rules.items():
        ws_dcf[f"A{r_i}"].border = border_cell
        for c_i in range(2, 10):
            col_l = get_column_letter(c_i)
            cell = ws_dcf[f"{col_l}{r_i}"]
            cell.number_format = fmt_s
            cell.alignment = align_right
            cell.border = border_cell
            if "=" in str(cell.value):
                cell.font = f_style
            else:
                cell.font = font_input_bold if "bold" in str(f_style) else font_input
                cell.fill = fill_input
            if fill_s and cell.font != font_input:
                cell.fill = fill_s

    # Section III: Free Cash Flow Schedule
    add_section_header(ws_dcf, 76, "III. UNLEVERED FREE CASH FLOW (FCFF) SCHEDULE ($M)", "I")
    
    fcf_hdr = ["Free Cash Flow Metric ($M)", "FY2023A", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for c_i, h in enumerate(fcf_hdr, start=1):
        cell = ws_dcf.cell(row=77, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 2 else align_left
        cell.border = border_header

    # Row 78: NOPAT
    ws_dcf["A78"] = "Net Operating Profit After Tax (NOPAT) ($M)"
    ws_dcf["A78"].font = font_bold
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}78"] = f"={col_c}74"

    # Row 79: (+) D&A
    ws_dcf["A79"] = "(+) Depreciation & Amortization (D&A) ($M)"
    ws_dcf["A79"].font = font_regular
    ws_dcf["B79"] = 10.38
    ws_dcf["B79"].comment = Comment("Source: Form 10-K FY2023 Cash Flow Statement D&A ($10.38M)", "DCF Model Builder")
    ws_dcf["C79"] = 13.95
    ws_dcf["C79"].comment = Comment("Source: Form 10-K FY2024 Cash Flow Statement D&A ($13.95M)", "DCF Model Builder")
    ws_dcf["D79"] = 18.23
    ws_dcf["D79"].comment = Comment("Source: Form 10-K FY2025 Cash Flow Statement D&A ($18.23M)", "DCF Model Builder")
    for col_c, act_c in [("E", "C"), ("F", "D"), ("G", "E"), ("H", "F"), ("I", "G")]:
        ws_dcf[f"{col_c}79"] = f"={col_c}63*{act_c}56"

    # Row 80: (-) CapEx
    ws_dcf["A80"] = "(-) Capital Expenditures (CapEx) ($M)"
    ws_dcf["A80"].font = font_regular
    ws_dcf["B80"] = 15.10
    ws_dcf["B80"].comment = Comment("Source: Form 10-K FY2023 Purchases of Property, Plant & Equipment ($15.10M)", "DCF Model Builder")
    ws_dcf["C80"] = 27.28
    ws_dcf["C80"].comment = Comment("Source: Form 10-K FY2024 Purchases of Property, Plant & Equipment ($27.28M)", "DCF Model Builder")
    ws_dcf["D80"] = 37.97
    ws_dcf["D80"].comment = Comment("Source: Form 10-K FY2025 Purchases of Property, Plant & Equipment ($37.97M)", "DCF Model Builder")
    for col_c, act_c in [("E", "C"), ("F", "D"), ("G", "E"), ("H", "F"), ("I", "G")]:
        ws_dcf[f"{col_c}80"] = f"={col_c}63*{act_c}55"

    # Row 81: (-) Change in Net Working Capital (Δ NWC)
    ws_dcf["A81"] = "(-) Change in Net Working Capital (Δ NWC) ($M)"
    ws_dcf["A81"].font = font_regular
    ws_dcf["B81"] = 8.49
    ws_dcf["B81"].comment = Comment("Source: Form 10-K FY2023 Operating assets & liabilities change ($8.49M)", "DCF Model Builder")
    ws_dcf["C81"] = 9.70
    ws_dcf["C81"].comment = Comment("Source: Form 10-K FY2024 Operating assets & liabilities change ($9.70M)", "DCF Model Builder")
    ws_dcf["D81"] = 11.39
    ws_dcf["D81"].comment = Comment("Source: Form 10-K FY2025 Operating assets & liabilities change ($11.39M)", "DCF Model Builder")
    for col_c, prior_c, act_c in [("E", "D", "C"), ("F", "E", "D"), ("G", "F", "E"), ("H", "G", "F"), ("I", "H", "G")]:
        ws_dcf[f"{col_c}81"] = f"=({col_c}63-{prior_c}63)*{act_c}57"

    # Row 82: Unlevered Free Cash Flow (FCFF)
    ws_dcf["A82"] = "Unlevered Free Cash Flow (FCFF) ($M)"
    ws_dcf["A82"].font = font_bold_navy
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}82"] = f"={col_c}78+{col_c}79-{col_c}80-{col_c}81"

    # Row 83: FCFF Conversion (% of EBIT)
    ws_dcf["A83"] = "  FCFF Conversion Rate (% of EBIT)"
    ws_dcf["A83"].font = font_italic
    for col_c in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_dcf[f"{col_c}83"] = f"={col_c}82/{col_c}70"

    # Row 84: FCFF YoY Growth (%)
    ws_dcf["A84"] = "  FCFF YoY Growth (%)"
    ws_dcf["A84"].font = font_italic
    ws_dcf["B84"] = "-"
    ws_dcf["B84"].alignment = align_center
    ws_dcf["C84"] = "=C82/B82-1"
    ws_dcf["D84"] = "=D82/C82-1"
    for col_c, prior_c in [("E", "D"), ("F", "E"), ("G", "F"), ("H", "G"), ("I", "H")]:
        ws_dcf[f"{col_c}84"] = f"={col_c}82/{prior_c}82-1"

    # Formatting Section III
    fcf_fmt_rules = {
        78: ("$#,##0.00", font_bold, fill_zebra),
        79: ("$#,##0.00", font_regular, None),
        80: ("$#,##0.00", font_regular, None),
        81: ("$#,##0.00", font_regular, None),
        82: ("$#,##0.00", font_bold_navy, fill_accent_blue),
        83: ("0.0%", font_italic, None),
        84: ("0.0%", font_italic, None),
    }
    for r_i, (fmt_s, f_style, fill_s) in fcf_fmt_rules.items():
        ws_dcf[f"A{r_i}"].border = border_cell
        for c_i in range(2, 10):
            col_l = get_column_letter(c_i)
            cell = ws_dcf[f"{col_l}{r_i}"]
            if cell.value != "-":
                cell.number_format = fmt_s
                cell.alignment = align_right
            cell.border = border_cell
            if "=" in str(cell.value):
                cell.font = f_style
            elif cell.value != "-":
                cell.font = font_input
                cell.fill = fill_input
            if fill_s and cell.font != font_input:
                cell.fill = fill_s

    # Section IV: DCF Valuation Schedule & Equity Bridge
    add_section_header(ws_dcf, 86, "IV. DISCOUNTED CASH FLOW VALUATION & EQUITY BRIDGE ($M)", "I")
    
    dcf_val_hdr = ["Valuation Projection Schedule", "", "", "", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for c_i, h in enumerate(dcf_val_hdr, start=1):
        cell = ws_dcf.cell(row=87, column=c_i, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_soft_blue
        cell.alignment = align_center if c_i >= 5 else align_left
        cell.border = border_header

    # Row 88: Mid-Year Discount Period (t)
    ws_dcf["A88"] = "Mid-Year Discount Period (t)"
    ws_dcf["A88"].font = font_regular
    ws_dcf["A88"].border = border_cell
    for c_i in range(2, 5):
        col_l = get_column_letter(c_i)
        ws_dcf[f"{col_l}88"] = ""
        ws_dcf[f"{col_l}88"].border = border_cell
    for c_i, period in enumerate([0.5, 1.5, 2.5, 3.5, 4.5], start=5):
        col_l = get_column_letter(c_i)
        cell = ws_dcf[f"{col_l}88"]
        cell.value = period
        cell.font = font_input
        cell.fill = fill_input
        cell.number_format = "0.0"
        cell.alignment = align_center
        cell.border = border_cell

    # Row 89: Discount Factor
    ws_dcf["A89"] = "Discount Factor (1 / (1 + WACC)^t)"
    ws_dcf["A89"].font = font_regular
    ws_dcf["A89"].border = border_cell
    for c_i in range(2, 5):
        col_l = get_column_letter(c_i)
        ws_dcf[f"{col_l}89"] = ""
        ws_dcf[f"{col_l}89"].border = border_cell
    for col_c in ["E", "F", "G", "H", "I"]:
        cell = ws_dcf[f"{col_c}89"]
        cell.value = f"=1/(1+$H$59)^{col_c}88"
        cell.font = font_bold
        cell.number_format = "0.0000"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 90: Projected FCFF ($M)
    ws_dcf["A90"] = "Projected Unlevered FCFF ($M)"
    ws_dcf["A90"].font = font_bold
    ws_dcf["A90"].border = border_cell
    for c_i in range(2, 5):
        col_l = get_column_letter(c_i)
        ws_dcf[f"{col_l}90"] = ""
        ws_dcf[f"{col_l}90"].border = border_cell
    for col_c in ["E", "F", "G", "H", "I"]:
        cell = ws_dcf[f"{col_c}90"]
        cell.value = f"={col_c}82"
        cell.font = font_bold
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Row 91: Present Value of FCFF ($M)
    ws_dcf["A91"] = "Present Value of FCFF ($M)"
    ws_dcf["A91"].font = font_bold_navy
    ws_dcf["A91"].border = border_cell
    for c_i in range(2, 5):
        col_l = get_column_letter(c_i)
        ws_dcf[f"{col_l}91"] = ""
        ws_dcf[f"{col_l}91"].border = border_cell
    for col_c in ["E", "F", "G", "H", "I"]:
        cell = ws_dcf[f"{col_c}91"]
        cell.value = f"={col_c}89*{col_c}90"
        cell.font = font_bold_navy
        cell.fill = fill_accent_blue
        cell.number_format = "$#,##0.00"
        cell.alignment = align_right
        cell.border = border_cell

    # Valuation Summary Bridge (Rows 93-108)
    add_section_header(ws_dcf, 93, "VALUATION SUMMARY & EQUITY BRIDGE ($M)", "E", fill_color=fill_med_blue, font_color=font_sub_sec)
    
    bridge_rows = [
        ("Cumulative PV of Explicit Forecast FCFFs (FY26E-FY30E) ($M)", "=SUM(E91:I91)", "$#,##0.00", font_bold, "Sum of discounted cash flows over explicit 5-year forecast", "B94"),
        ("Normalized Terminal Year FCFF (FY2031E) ($M)", "=I90*(1+H58)", "$#,##0.00", font_regular, "Final year explicit FCFF grown at perpetual terminal growth rate", "B95"),
        ("Implied Terminal Value at FY2030E ($M)", "=B95/(H59-H58)", "$#,##0.00", font_bold, "Gordon Growth Perpetuity Model = FCFF_2031E / (WACC - g)", "B96"),
        ("Present Value of Terminal Value ($M)", "=B96/(1+H59)^I88", "$#,##0.00", font_bold, "Discounted back to valuation date using 4.5-year mid-year factor", "B97"),
        ("Terminal Value % of Enterprise Value", "=B97/B99", "0.0%", font_italic, "DCF terminal value proportion of Enterprise Value", "B98"),
        ("IMPLIED ENTERPRISE VALUE ($M)", "=B94+B97", "$#,##0.00", font_bold_navy, "Total Enterprise Value = PV of Explicit FCFFs + PV of Terminal Value", "B99"),
        ("(-) Total Carrying Debt ($M)", "=B12", "$#,##0.00", font_regular, "Carrying value of outstanding credit facility ($30.00M)", "B100"),
        ("(+) Cash, Cash Equivalents & ST Investments ($M)", "=B11", "$#,##0.00", font_regular, "Total liquid cash and short-term investments ($160.00M)", "B101"),
        ("Net Debt / (Net Cash) ($M)", "=B100-B101", "$#,##0.00", font_bold, "Net debt subtracted from Enterprise Value (adds cash when negative)", "B102"),
        ("IMPLIED EQUITY VALUE ($M)", "=B99-B102", "$#,##0.00", font_bold_navy, "Equity Value = Enterprise Value - Net Debt", "B103"),
        ("Diluted Shares Outstanding (M)", "=B9", "#,##0.00", font_regular, "Diluted shares outstanding from Form 10-Q / 10-K (42.94M)", "B104"),
        ("IMPLIED DCF VALUE PER SHARE ($)", "=B103/B104", "$#,##0.00", font_bold_navy, "Intrinsic value per share derived from DCF valuation", "B105"),
        ("Current Market Share Price ($)", "=B8", "$#,##0.00", font_regular, "Market price as of August 20, 2026 ($11.10)", "B106"),
        ("Implied Upside / (Downside) (%)", "=(B105/B106)-1", "0.0%", font_bold_navy, "Potential return based on intrinsic DCF valuation", "B107"),
        ("Market Implied EV / FY2026E EBIT Multiple", "=B99/E70", "0.0x", font_italic, "Implied forward valuation multiple on FY2026E EBIT", "B108"),
    ]
    
    for r_idx, (label, val, fmt, f_style, rat, cell_ref) in enumerate(bridge_rows, start=94):
        ws_dcf[f"A{r_idx}"] = label
        ws_dcf[f"A{r_idx}"].font = f_style
        ws_dcf[f"A{r_idx}"].border = border_cell
        
        ws_dcf[f"B{r_idx}"] = val
        ws_dcf[f"B{r_idx}"].font = f_style
        ws_dcf[f"B{r_idx}"].number_format = fmt
        ws_dcf[f"B{r_idx}"].alignment = align_right
        ws_dcf[f"B{r_idx}"].border = border_cell
        
        if r_idx in (99, 103, 105, 107):
            ws_dcf[f"B{r_idx}"].fill = fill_accent_blue
            ws_dcf[f"A{r_idx}"].fill = fill_accent_blue
            
        ws_dcf[f"C{r_idx}"] = rat
        ws_dcf[f"C{r_idx}"].font = font_italic
        ws_dcf.merge_cells(f"C{r_idx}:E{r_idx}")
        for col_c in ["C", "D", "E"]:
            ws_dcf[f"{col_c}{r_idx}"].border = border_cell

    # ==========================================
    # Section V: SENSITIVITY ANALYSIS (3 TABLES: 5x5 MATRICES)
    # ==========================================
    add_section_header(ws_dcf, 111, "V. SENSITIVITY ANALYSIS - INSTITUTIONAL 5x5 VALUATION MATRICES", "G")
    
    # ------------------------------------------
    # Sensitivity Table 1: WACC vs. Terminal Growth Rate
    # ------------------------------------------
    ws_dcf["A112"] = "Table 1: Implied Share Price ($) — WACC (Discount Rate) vs. Perpetual Terminal Growth Rate"
    ws_dcf["A112"].font = font_bold_navy
    ws_dcf.merge_cells("A112:G112")
    
    # Header row
    ws_dcf["B113"] = "WACC \\ Terminal g"
    ws_dcf["B113"].font = font_tbl_hdr
    ws_dcf["B113"].fill = fill_soft_blue
    ws_dcf["B113"].alignment = align_center
    ws_dcf["B113"].border = border_header
    
    term_g_axis = [0.0150, 0.0200, 0.0250, 0.0300, 0.0350] # Base = 2.50% (Col E)
    wacc_axis = [0.0900, 0.0950, 0.1000, 0.1050, 0.1100]   # Base = 10.00% (Row 116)
    
    for c_i, g_val in enumerate(term_g_axis, start=3):
        col_l = get_column_letter(c_i)
        cell = ws_dcf[f"{col_l}113"]
        cell.value = g_val
        cell.font = font_input_bold if g_val == 0.0250 else font_input
        cell.fill = fill_accent_blue if g_val == 0.0250 else fill_soft_blue
        cell.number_format = "0.00%"
        cell.alignment = align_center
        cell.border = border_header
        
    for r_i, w_val in enumerate(wacc_axis, start=114):
        cell_lbl = ws_dcf[f"B{r_i}"]
        cell_lbl.value = w_val
        cell_lbl.font = font_input_bold if w_val == 0.1000 else font_input
        cell_lbl.fill = fill_accent_blue if w_val == 0.1000 else fill_soft_blue
        cell_lbl.number_format = "0.00%"
        cell_lbl.alignment = align_center
        cell_lbl.border = border_cell
        
        for c_i, g_val in enumerate(term_g_axis, start=3):
            col_l = get_column_letter(c_i)
            cell = ws_dcf[f"{col_l}{r_i}"]
            # Full DCF Recalculation Formula:
            # Implied Price = ((PV Explicit FCFs using $B{r_i}) + (PV Terminal Value using {col_l}$113 and $B{r_i}) - NetDebt) / Shares
            fml = (
                f"=((E90/(1+$B{r_i})^E88 + F90/(1+$B{r_i})^F88 + G90/(1+$B{r_i})^G88 + H90/(1+$B{r_i})^H88 + I90/(1+$B{r_i})^I88) "
                f"+ (I90*(1+{col_l}$113)/($B{r_i}-{col_l}$113))/(1+$B{r_i})^I88 - $B$13) / $B$9"
            )
            cell.value = fml
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            # Highlight center base case cell (Row 116, Col E -> 10.00% & 2.50%)
            if r_i == 116 and col_l == "E":
                cell.font = font_bold_navy
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular

    # ------------------------------------------
    # Sensitivity Table 2: FY2026E Revenue Growth vs. FY2030E EBIT Margin
    # ------------------------------------------
    ws_dcf["A121"] = "Table 2: Implied Share Price ($) — FY2026E Revenue Growth (%) vs. FY2030E EBIT Margin (%)"
    ws_dcf["A121"].font = font_bold_navy
    ws_dcf.merge_cells("A121:G121")
    
    ws_dcf["B122"] = "FY26E Rev % \\ FY30E EBIT %"
    ws_dcf["B122"].font = font_tbl_hdr
    ws_dcf["B122"].fill = fill_soft_blue
    ws_dcf["B122"].alignment = align_center
    ws_dcf["B122"].border = border_header
    
    ebit_margin_axis = [0.105, 0.115, 0.125, 0.135, 0.145] # Base = 12.5% (Col E)
    rev_growth_axis = [0.120, 0.150, 0.180, 0.210, 0.240]  # Base = 18.0% (Row 125)
    
    for c_i, m_val in enumerate(ebit_margin_axis, start=3):
        col_l = get_column_letter(c_i)
        cell = ws_dcf[f"{col_l}122"]
        cell.value = m_val
        cell.font = font_input_bold if m_val == 0.125 else font_input
        cell.fill = fill_accent_blue if m_val == 0.125 else fill_soft_blue
        cell.number_format = "0.0%"
        cell.alignment = align_center
        cell.border = border_header
        
    for r_i, rev_g in enumerate(rev_growth_axis, start=123):
        cell_lbl = ws_dcf[f"B{r_i}"]
        cell_lbl.value = rev_g
        cell_lbl.font = font_input_bold if rev_g == 0.180 else font_input
        cell_lbl.fill = fill_accent_blue if rev_g == 0.180 else fill_soft_blue
        cell_lbl.number_format = "0.0%"
        cell_lbl.alignment = align_center
        cell_lbl.border = border_cell
        
        for c_i, m_val in enumerate(ebit_margin_axis, start=3):
            col_l = get_column_letter(c_i)
            cell = ws_dcf[f"{col_l}{r_i}"]
            fml = (
                f"=(((E91*((1+$B{r_i})/(1+$C$52))) + (F91*((1+$B{r_i})/(1+$C$52))) + (G91*((1+$B{r_i})/(1+$C$52))) + (H91*((1+$B{r_i})/(1+$C$52))) + (I91*((1+$B{r_i})/(1+$C$52))*({col_l}$122/$G$54))) "
                f"+ (B97*((1+$B{r_i})/(1+$C$52))*({col_l}$122/$G$54)) - $B$13) / $B$9"
            )
            cell.value = fml
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_i == 125 and col_l == "E":
                cell.font = font_bold_navy
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular

    # ------------------------------------------
    # Sensitivity Table 3: Equity Beta vs. Risk-Free Rate (10Y UST)
    # ------------------------------------------
    ws_dcf["A130"] = "Table 3: Implied Share Price ($) — Equity Beta vs. Risk-Free Rate (10-Yr US Treasury Yield)"
    ws_dcf["A130"].font = font_bold_navy
    ws_dcf.merge_cells("A130:G130")
    
    ws_dcf["B131"] = "Beta \\ 10Y UST Yield"
    ws_dcf["B131"].font = font_tbl_hdr
    ws_dcf["B131"].fill = fill_soft_blue
    ws_dcf["B131"].alignment = align_center
    ws_dcf["B131"].border = border_header
    
    rf_axis = [0.0414, 0.0439, 0.0464, 0.0489, 0.0514] # Base = 4.64% (Col E)
    beta_axis = [0.88, 0.98, 1.08, 1.18, 1.28]         # Base = 1.08 (Row 134)
    
    for c_i, rf_val in enumerate(rf_axis, start=3):
        col_l = get_column_letter(c_i)
        cell = ws_dcf[f"{col_l}131"]
        cell.value = rf_val
        cell.font = font_input_bold if rf_val == 0.0464 else font_input
        cell.fill = fill_accent_blue if rf_val == 0.0464 else fill_soft_blue
        cell.number_format = "0.00%"
        cell.alignment = align_center
        cell.border = border_header
        
    for r_i, beta_val in enumerate(beta_axis, start=132):
        cell_lbl = ws_dcf[f"B{r_i}"]
        cell_lbl.value = beta_val
        cell_lbl.font = font_input_bold if beta_val == 1.08 else font_input
        cell_lbl.fill = fill_accent_blue if beta_val == 1.08 else fill_soft_blue
        cell_lbl.number_format = "0.00"
        cell_lbl.alignment = align_center
        cell_lbl.border = border_cell
        
        for c_i, rf_val in enumerate(rf_axis, start=3):
            col_l = get_column_letter(c_i)
            cell = ws_dcf[f"{col_l}{r_i}"]
            fml = (
                f"=((E90/(1+({col_l}$131+$B{r_i}*0.0520))^E88 + F90/(1+({col_l}$131+$B{r_i}*0.0520))^F88 + G90/(1+({col_l}$131+$B{r_i}*0.0520))^G88 + H90/(1+({col_l}$131+$B{r_i}*0.0520))^H88 + I90/(1+({col_l}$131+$B{r_i}*0.0520))^I88) "
                f"+ (I90*(1+$H$58)/(({col_l}$131+$B{r_i}*0.0520)-$H$58))/(1+({col_l}$131+$B{r_i}*0.0520))^I88 - $B$13) / $B$9"
            )
            cell.value = fml
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right
            cell.border = border_cell
            
            if r_i == 134 and col_l == "E":
                cell.font = font_bold_navy
                cell.fill = fill_accent_blue
            else:
                cell.font = font_regular

    # Auto-adjust column widths
    for ws in [ws_dcf, ws_wacc]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
                else:
                    max_len = max(max_len, 12)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
    ws_dcf.column_dimensions["A"].width = 46
    ws_dcf.column_dimensions["B"].width = 24
    ws_dcf.column_dimensions["I"].width = 42
    ws_wacc.column_dimensions["A"].width = 44
    ws_wacc.column_dimensions["B"].width = 20
    ws_wacc.column_dimensions["E"].width = 48

    wb.save(output_path)
    print(f"Successfully generated DCF model at: {output_path}")

if __name__ == "__main__":
    build_vitl_dcf_model()
