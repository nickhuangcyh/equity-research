import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DCF Valuation"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Color Palette (Institutional Classic Navy)
NAVY = "1F4E79"       # Primary Headers
MED_BLUE = "2F5597"   # Secondary Headers
SOFT_BLUE = "D9E1F2"  # Subheaders / Table Headers
ACCENT_BLUE = "BDD7EE"# Active Case / Base Case highlight
ZEBRA_FILL = "F9FBFD" # Alternating row fill
WHITE = "FFFFFF"
GRAY_TEXT = "595959"

# Font definitions
font_title = Font(name="Calibri", size=15, bold=True, color="1F4E79")
font_subtitle = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
font_sec_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_tbl_hdr = Font(name="Calibri", size=10, bold=True, color="000000")
font_subhdr = Font(name="Calibri", size=10, bold=True, color="1F4E79")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_regular = Font(name="Calibri", size=10, bold=False, color="000000")
font_italic = Font(name="Calibri", size=9, italic=True, color=GRAY_TEXT)

# Fills
fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
fill_soft_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
fill_accent_blue = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")

# Borders
thin_line = Side(border_style="thin", color="D9D9D9")
thick_bottom = Side(border_style="medium", color="1F4E79")
double_bottom = Side(border_style="double", color="1F4E79")
top_thin = Side(border_style="thin", color="1F4E79")

border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)
border_total = Border(top=top_thin, bottom=double_bottom, left=thin_line, right=thin_line)

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# --- Title Block ---
ws["A1"] = "Wave Life Sciences Ltd. (NASDAQ: WVE)"
ws["A1"].font = font_title
ws["A2"] = "Institutional DCF Valuation Model | Report Date: August 19, 2026 | Source: SEC Form 10-K & Q2 2026 10-Q Disclosures"
ws["A2"].font = font_subtitle

# --- Scenario Selector ---
ws["A4"] = "Active Scenario Selector:"
ws["A4"].font = font_bold
ws["B4"] = 2
ws["B4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws["B4"].fill = fill_navy
ws["B4"].alignment = align_center
ws["B4"].comment = Comment("Case Selector: 1 = Bear Case, 2 = Base Case, 3 = Bull Case", "Model Builder")

ws["C4"] = '=IF(B4=1,"[1] BEAR CASE",IF(B4=2,"[2] BASE CASE (Active)","[3] BULL CASE"))'
ws["C4"].font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
ws["C4"].alignment = align_left

# Helper for section headers
def add_section_header(row, title, max_col="H"):
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = font_sec_hdr
    ws[f"A{row}"].fill = fill_navy
    ws[f"A{row}"].alignment = align_left
    ws.merge_cells(f"A{row}:{max_col}{row}")
    col_max_idx = openpyxl.utils.column_index_from_string(max_col)
    for col_idx in range(1, col_max_idx + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill = fill_navy

# --- Section I: Market Data & Capital Structure Inputs ---
add_section_header(6, "I. MARKET DATA & CAPITAL STRUCTURE INPUTS")

market_inputs = [
    ("Current Stock Price ($)", 5.22, "$#,##0.00", "Market close price as of August 19, 2026 ($5.22)", "B7"),
    ("Diluted Shares Outstanding (M)", 200.17, "#,##0.00", "SEC Form 10-Q Q2 2026 Diluted share count (200.17M shares)", "B8"),
    ("Implied Equity Market Capitalization ($M)", "=B7*B8", "$#,##0.00", None, "B9"),
    ("Cash and Cash Equivalents ($M)", 490.60, "$#,##0.00", "SEC Form 10-Q Balance sheet as of June 30, 2026 ($490.60M)", "B10"),
    ("Total Debt ($M)", 43.59, "$#,##0.00", "SEC Form 10-Q Balance sheet as of June 30, 2026 ($43.59M)", "B11"),
    ("Net Debt ($M)", "=B11-B10", "$#,##0.00;($#,##0.00);$0.00", None, "B12"),
    ("Implied Enterprise Value ($M)", "=B9+B12", "$#,##0.00", None, "B13"),
    ("Effective Corporate Tax Rate (%)", 0.210, "0.0%", "US Federal statutory corporate tax rate (21.0%)", "B14"),
    ("Risk-Free Rate (10-Yr US Treasury) (%)", 0.0469, "0.00%", "10-Year US Treasury Benchmark Yield (August 2026, 4.69%)", "B15"),
    ("Equity Beta (5-Year / 1-Year Monthly)", 1.69, "0.00", "Beta vs S&P 500 reflecting clinical biopharma volatility (1.69)", "B16"),
    ("Market Equity Risk Premium (ERP) (%)", 0.0550, "0.00%", "Standard institutional market equity risk premium (5.50%)", "B17"),
    ("Pre-Tax Cost of Debt (%)", 0.0700, "0.00%", "Estimated corporate borrowing rate (~7.00%)", "B18"),
]

for idx, (label, val, num_fmt, comment_text, cell_ref) in enumerate(market_inputs, start=7):
    ws[f"A{idx}"] = label
    ws[f"A{idx}"].font = font_bold if "=" in str(val) else font_regular
    ws[f"A{idx}"].border = border_cell
    
    ws[f"B{idx}"] = val
    ws[f"B{idx}"].font = font_bold if "=" in str(val) else font_regular
    ws[f"B{idx}"].number_format = num_fmt
    ws[f"B{idx}"].alignment = align_right
    ws[f"B{idx}"].border = border_cell
    if comment_text:
        ws[f"B{idx}"].comment = Comment(comment_text, "Wave Life Sciences 10-Q / Market Data")

# --- Section II: Scenario Assumptions & Driver Blocks ---
add_section_header(20, "II. SCENARIO ASSUMPTIONS & DRIVER BLOCKS")

headers_s2 = ["Driver / Assumption", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal / WACC"]
for col_idx, text in enumerate(headers_s2, start=1):
    c = ws.cell(row=21, column=col_idx, value=text)
    c.font = font_tbl_hdr
    c.fill = fill_soft_blue
    c.alignment = align_center if col_idx > 1 else align_left
    c.border = border_header

scenario_blocks = [
    ("Bear Case Assumptions (Selector = 1)", 22, [
        ("  Revenue Growth (%)", [0.054, 0.333, 0.500, 0.556, 0.571], 0.020, "0.0%"),
        ("  Gross Margin (%)", [0.800, 0.810, 0.820, 0.830, 0.840], 0.1350, "0.0%"),
        ("  R&D Expenses ($M)", [195.0, 205.0, 215.0, 225.0, 235.0], None, "$#,##0.00"),
        ("  SG&A Expenses ($M)", [45.0, 50.0, 60.0, 72.0, 85.0], None, "$#,##0.00"),
        ("  D&A (% of Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ("  CapEx (% of Revenue)", [0.050, 0.050, 0.050, 0.050, 0.050], None, "0.0%"),
        ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
    ]),
    ("Base Case Assumptions (Selector = 2)", 30, [
        ("  Revenue Growth (%)", [0.218, 0.635, 0.765, 0.867, 0.714], 0.030, "0.0%"),
        ("  Gross Margin (%)", [0.820, 0.830, 0.840, 0.850, 0.860], 0.1350, "0.0%"),
        ("  R&D Expenses ($M)", [195.0, 210.0, 225.0, 240.0, 255.0], None, "$#,##0.00"),
        ("  SG&A Expenses ($M)", [45.0, 52.0, 65.0, 80.0, 95.0], None, "$#,##0.00"),
        ("  D&A (% of Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ("  CapEx (% of Revenue)", [0.050, 0.050, 0.050, 0.050, 0.050], None, "0.0%"),
        ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
    ]),
    ("Bull Case Assumptions (Selector = 3)", 38, [
        ("  Revenue Growth (%)", [0.522, 0.846, 1.083, 0.920, 0.667], 0.035, "0.0%"),
        ("  Gross Margin (%)", [0.840, 0.850, 0.860, 0.870, 0.880], 0.1350, "0.0%"),
        ("  R&D Expenses ($M)", [195.0, 215.0, 235.0, 255.0, 275.0], None, "$#,##0.00"),
        ("  SG&A Expenses ($M)", [45.0, 55.0, 75.0, 95.0, 115.0], None, "$#,##0.00"),
        ("  D&A (% of Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
        ("  CapEx (% of Revenue)", [0.050, 0.050, 0.050, 0.050, 0.050], None, "0.0%"),
        ("  Δ NWC (% of Δ Revenue)", [0.030, 0.030, 0.030, 0.030, 0.030], None, "0.0%"),
    ]),
]

for title, sub_hdr_row, rows in scenario_blocks:
    ws[f"A{sub_hdr_row}"] = title
    ws[f"A{sub_hdr_row}"].font = font_subhdr
    ws[f"A{sub_hdr_row}"].fill = fill_soft_blue
    ws.merge_cells(f"A{sub_hdr_row}:H{sub_hdr_row}")
    for c_idx in range(1, 9):
        ws.cell(row=sub_hdr_row, column=c_idx).fill = fill_soft_blue
        ws.cell(row=sub_hdr_row, column=c_idx).border = border_cell
        
    for r_offset, (label, vals, term_val, num_fmt) in enumerate(rows, start=1):
        curr_r = sub_hdr_row + r_offset
        ws[f"A{curr_r}"] = label
        ws[f"A{curr_r}"].font = font_regular
        ws[f"A{curr_r}"].border = border_cell
        # Col B (Historical FY2025A base reference)
        ws.cell(row=curr_r, column=2, value="—").alignment = align_center
        ws.cell(row=curr_r, column=2).border = border_cell
        
        for c_idx, val in enumerate(vals, start=3):
            cell = ws.cell(row=curr_r, column=c_idx, value=val)
            cell.font = font_regular
            cell.number_format = num_fmt
            cell.alignment = align_right
            cell.border = border_cell
        if term_val is not None:
            cell_term = ws.cell(row=curr_r, column=8, value=term_val)
            cell_term.font = font_regular
            cell_term.number_format = "0.0%" if r_offset == 1 else "0.00%"
            cell_term.alignment = align_right
            cell_term.border = border_cell
        else:
            cell_term = ws.cell(row=curr_r, column=8, value="")
            cell_term.border = border_cell

# Consolidated Active Drivers (Dynamic via CHOOSE formulas)
ws["A46"] = "Active Scenario Consolidated Drivers (Dynamic)"
ws["A46"].font = font_bold
ws["A46"].fill = fill_accent_blue
ws.merge_cells("A46:H46")
for c_idx in range(1, 9):
    ws.cell(row=46, column=c_idx).fill = fill_accent_blue
    ws.cell(row=46, column=c_idx).border = border_cell

active_driver_rows = [
    ("  Active Revenue Growth (%)", 23, 31, 39, "0.0%"),
    ("  Active Gross Margin (%)", 24, 32, 40, "0.0%"),
    ("  Active R&D Expenses ($M)", 25, 33, 41, "$#,##0.00"),
    ("  Active SG&A Expenses ($M)", 26, 34, 42, "$#,##0.00"),
    ("  Active D&A (% of Revenue)", 27, 35, 43, "0.0%"),
    ("  Active CapEx (% of Revenue)", 28, 36, 44, "0.0%"),
    ("  Active Δ NWC (% of Δ Rev)", 29, 37, 45, "0.0%"),
]

for idx, (label, r_bear, r_base, r_bull, num_fmt) in enumerate(active_driver_rows, start=47):
    ws[f"A{idx}"] = label
    ws[f"A{idx}"].font = font_bold
    ws[f"A{idx}"].border = border_cell
    ws.cell(row=idx, column=2, value="—").alignment = align_center
    ws.cell(row=idx, column=2).border = border_cell
    
    for c_idx in range(3, 8):
        col_let = get_column_letter(c_idx)
        cell = ws.cell(row=idx, column=c_idx)
        cell.value = f"=CHOOSE($B$4, {col_let}{r_bear}, {col_let}{r_base}, {col_let}{r_bull})"
        cell.font = font_bold
        cell.number_format = num_fmt
        cell.alignment = align_right
        cell.border = border_cell
    
    # Terminal / WACC in column H (Col 8)
    if idx == 47: # Terminal g
        cell_g = ws.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H23, H31, H39)")
        cell_g.font = font_bold
        cell_g.number_format = "0.0%"
        cell_g.alignment = align_right
        cell_g.border = border_cell
    elif idx == 48: # Target WACC
        cell_g = ws.cell(row=idx, column=8, value=f"=CHOOSE($B$4, H24, H32, H40)")
        cell_g.font = font_bold
        cell_g.number_format = "0.00%"
        cell_g.alignment = align_right
        cell_g.border = border_cell
    else:
        cell_g = ws.cell(row=idx, column=8, value="")
        cell_g.border = border_cell

# --- Section III: Cost of Capital (WACC) Schedule ---
add_section_header(56, "III. COST OF CAPITAL (WACC) PARAMETERS & SCHEDULE")

wacc_items = [
    ("Risk-Free Rate (Rf)", "=$B$15", "0.00%", "Row 15: 10-Yr US Treasury Yield", False),
    ("Equity Beta (β)", "=$B$16", "0.00", "Row 16: 5-Year Monthly Beta", False),
    ("Equity Risk Premium (ERP)", "=$B$17", "0.00%", "Row 17: Market Equity Risk Premium", False),
    ("Cost of Equity (Ke) [Rf + β * ERP]", "=B57+B58*B59", "0.00%", "CAPM Formula: Cost of Equity", True),
    ("Pre-Tax Cost of Debt (Kd)", "=$B$18", "0.00%", "Row 18: Borrowing rate (~7.00%)", False),
    ("Effective Corporate Tax Rate (t)", "=$B$14", "0.0%", "Row 14: Corporate tax rate", False),
    ("After-Tax Cost of Debt [Kd * (1 - t)]", "=B61*(1-B62)", "0.00%", "Effective after-tax cost of debt", False),
    ("Market Capitalization (Equity Value) ($M)", "=$B$9", "$#,##0.00", "Row 9: Diluted Market Cap", False),
    ("Total Debt ($M)", "=$B$11", "$#,##0.00", "Row 11: Total Debt", False),
    ("Cash and Cash Equivalents ($M)", "=$B$10", "$#,##0.00", "Row 10: Cash and Marketable Securities", False),
    ("Net Debt ($M)", "=B65-B66", "$#,##0.00;($#,##0.00);$0.00", "Row 12: Total Debt - Cash", False),
    ("Enterprise Value (EV) ($M)", "=B64+B67", "$#,##0.00", "Row 13: Market Cap + Net Debt", False),
    ("Selected DCF Discount Rate (WACC)", "=$H$48", "0.00%", "Discount rate aligned with biopharma stage", True),
]

for idx, (label, val, num_fmt, comment_t, is_key) in enumerate(wacc_items, start=57):
    ws[f"A{idx}"] = label
    ws[f"A{idx}"].font = font_bold if is_key else font_regular
    ws[f"A{idx}"].border = border_cell
    if is_key:
        ws[f"A{idx}"].fill = fill_accent_blue
        
    ws[f"B{idx}"] = val
    ws[f"B{idx}"].font = font_bold if is_key else font_regular
    ws[f"B{idx}"].number_format = num_fmt
    ws[f"B{idx}"].alignment = align_right
    ws[f"B{idx}"].border = border_cell
    if is_key:
        ws[f"B{idx}"].fill = fill_accent_blue
    if comment_t:
        ws[f"B{idx}"].comment = Comment(comment_t, "WACC Schedule")

# --- Section IV: Unlevered Free Cash Flow (UFCF) Projections ---
add_section_header(72, "IV. UNLEVERED FREE CASH FLOW (UFCF) PROJECTIONS")

headers_s4 = ["Line Item ($M)", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
for col_idx, text in enumerate(headers_s4, start=1):
    c = ws.cell(row=73, column=col_idx, value=text)
    c.font = font_tbl_hdr
    c.fill = fill_soft_blue
    c.alignment = align_center if col_idx > 1 else align_left
    c.border = border_header

# Row 74: Revenue Growth Rate
ws["A74"] = "Revenue Growth Rate (%)"
ws["A74"].font = font_regular
ws["A74"].border = border_cell
ws["B74"] = -0.606
ws["B74"].font = font_regular
ws["B74"].number_format = "-0.0%;0.0%;0.0%"
ws["B74"].alignment = align_right
ws["B74"].border = border_cell
ws["B74"].comment = Comment("FY2025 Actual Net Revenue YoY Decline (-60.6%) due to Takeda collaboration conclusion", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=74, column=c_idx, value=f"={col_let}47")
    cell.font = font_regular
    cell.number_format = "0.0%"
    cell.alignment = align_right
    cell.border = border_cell

# Row 75: Total Revenue ($M)
ws["A75"] = "Total Revenue ($M)"
ws["A75"].font = font_bold
ws["A75"].border = border_cell
ws["B75"] = 42.70
ws["B75"].font = font_bold
ws["B75"].number_format = "$#,##0.00"
ws["B75"].alignment = align_right
ws["B75"].border = border_cell
ws["B75"].comment = Comment("FY2025 Actual Total Revenue ($42.70M)", "SEC Form 10-K")

ws["C75"] = "=B75*(1+C74)"
ws["D75"] = "=C75*(1+D74)"
ws["E75"] = "=D75*(1+E74)"
ws["F75"] = "=E75*(1+F74)"
ws["G75"] = "=F75*(1+G74)"
for c_idx in range(3, 8):
    c = ws.cell(row=75, column=c_idx)
    c.font = font_bold
    c.number_format = "$#,##0.00"
    c.alignment = align_right
    c.border = border_cell

# Row 76: Gross Margin (%)
ws["A76"] = "Gross Margin (%)"
ws["A76"].font = font_regular
ws["A76"].border = border_cell
ws["B76"] = 0.848
ws["B76"].font = font_regular
ws["B76"].number_format = "0.0%"
ws["B76"].alignment = align_right
ws["B76"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=76, column=c_idx, value=f"={col_let}48")
    cell.font = font_regular
    cell.number_format = "0.0%"
    cell.alignment = align_right
    cell.border = border_cell

# Row 77: Gross Profit ($M)
ws["A77"] = "Gross Profit ($M)"
ws["A77"].font = font_bold
ws["A77"].border = border_cell
ws["B77"] = 36.20
ws["B77"].font = font_bold
ws["B77"].number_format = "$#,##0.00"
ws["B77"].alignment = align_right
ws["B77"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=77, column=c_idx, value=f"={col_let}75*{col_let}76")
    cell.font = font_bold
    cell.number_format = "$#,##0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 78: (-) R&D Expenses ($M)
ws["A78"] = "(-) Research & Development (R&D) ($M)"
ws["A78"].font = font_regular
ws["A78"].border = border_cell
ws["B78"] = 182.80
ws["B78"].font = font_regular
ws["B78"].number_format = "($#,##0.00);($#,##0.00);$0.00"
ws["B78"].alignment = align_right
ws["B78"].border = border_cell
ws["B78"].comment = Comment("FY2025 Actual R&D Expenses ($182.80M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=78, column=c_idx, value=f"={col_let}49")
    cell.font = font_regular
    cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 79: (-) SG&A Expenses ($M)
ws["A79"] = "(-) General & Administrative (SG&A) ($M)"
ws["A79"].font = font_regular
ws["A79"].border = border_cell
ws["B79"] = 41.50
ws["B79"].font = font_regular
ws["B79"].number_format = "($#,##0.00);($#,##0.00);$0.00"
ws["B79"].alignment = align_right
ws["B79"].border = border_cell
ws["B79"].comment = Comment("FY2025 Actual G&A Expenses ($41.50M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=79, column=c_idx, value=f"={col_let}50")
    cell.font = font_regular
    cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 80: Operating Income / (Loss) (EBIT) ($M)
ws["A80"] = "Operating Income / (Loss) (EBIT) ($M)"
ws["A80"].font = font_bold
ws["A80"].border = border_cell
ws["B80"] = -188.10
ws["B80"].font = font_bold
ws["B80"].number_format = "$#,##0.00;($#,##0.00);$0.00"
ws["B80"].alignment = align_right
ws["B80"].border = border_cell
ws["B80"].comment = Comment("FY2025 Actual Operating Loss ($(188.10)M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=80, column=c_idx, value=f"={col_let}77-{col_let}78-{col_let}79")
    cell.font = font_bold
    cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 81: EBIT Margin (%)
ws["A81"] = "EBIT Margin (%)"
ws["A81"].font = font_regular
ws["A81"].border = border_cell
ws["B81"] = "=B80/B75"
ws["B81"].font = font_regular
ws["B81"].number_format = "-0.0%;+0.0%;0.0%"
ws["B81"].alignment = align_right
ws["B81"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=81, column=c_idx, value=f"={col_let}80/{col_let}75")
    cell.font = font_regular
    cell.number_format = "-0.0%;+0.0%;0.0%"
    cell.alignment = align_right
    cell.border = border_cell

# Row 82: (-) Taxes ($M)
ws["A82"] = "(-) Provision for Taxes ($M)"
ws["A82"].font = font_regular
ws["A82"].border = border_cell
ws["B82"] = 0.00
ws["B82"].font = font_regular
ws["B82"].number_format = "($#,##0.00);($#,##0.00);$0.00"
ws["B82"].alignment = align_right
ws["B82"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=82, column=c_idx, value=f"=MAX(0, {col_let}80*$B$14)")
    cell.font = font_regular
    cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 83: NOPAT ($M)
ws["A83"] = "Net Operating Profit After Tax (NOPAT) ($M)"
ws["A83"].font = font_bold
ws["A83"].border = border_cell
ws["B83"] = -188.10
ws["B83"].font = font_bold
ws["B83"].number_format = "$#,##0.00;($#,##0.00);$0.00"
ws["B83"].alignment = align_right
ws["B83"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=83, column=c_idx, value=f"={col_let}80-{col_let}82")
    cell.font = font_bold
    cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 84: (+) D&A ($M)
ws["A84"] = "(+) Depreciation & Amortization (D&A) ($M)"
ws["A84"].font = font_regular
ws["A84"].border = border_cell
ws["B84"] = 2.50
ws["B84"].font = font_regular
ws["B84"].number_format = "$#,##0.00"
ws["B84"].alignment = align_right
ws["B84"].border = border_cell
ws["B84"].comment = Comment("FY2025 Actual D&A (~$2.50M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=84, column=c_idx, value=f"={col_let}75*{col_let}51")
    cell.font = font_regular
    cell.number_format = "$#,##0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 85: (-) CapEx ($M)
ws["A85"] = "(-) Capital Expenditures (CapEx) ($M)"
ws["A85"].font = font_regular
ws["A85"].border = border_cell
ws["B85"] = 9.20
ws["B85"].font = font_regular
ws["B85"].number_format = "($#,##0.00);($#,##0.00);$0.00"
ws["B85"].alignment = align_right
ws["B85"].border = border_cell
ws["B85"].comment = Comment("FY2025 Actual CapEx ($9.20M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=85, column=c_idx, value=f"={col_let}75*{col_let}52")
    cell.font = font_regular
    cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 86: (-) Change in NWC ($M)
ws["A86"] = "(-) Change in Net Working Capital (Δ NWC) ($M)"
ws["A86"].font = font_regular
ws["A86"].border = border_cell
ws["B86"] = 1.00
ws["B86"].font = font_regular
ws["B86"].number_format = "($#,##0.00);($#,##0.00);$0.00"
ws["B86"].alignment = align_right
ws["B86"].border = border_cell
ws["B86"].comment = Comment("FY2025 Working Capital change (~$1.00M)", "SEC Form 10-K")
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    prev_col = get_column_letter(c_idx - 1)
    cell = ws.cell(row=86, column=c_idx, value=f"=({col_let}75-{prev_col}75)*{col_let}53")
    cell.font = font_regular
    cell.number_format = "($#,##0.00);($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 87: Unlevered Free Cash Flow (UFCF) ($M)
ws["A87"] = "Unlevered Free Cash Flow (UFCF) ($M)"
ws["A87"].font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
ws["A87"].fill = fill_accent_blue
ws["A87"].border = border_total
ws["B87"] = -195.80
ws["B87"].font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
ws["B87"].alignment = align_right
ws["B87"].number_format = "$#,##0.00;($#,##0.00);$0.00"
ws["B87"].fill = fill_accent_blue
ws["B87"].border = border_total
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=87, column=c_idx, value=f"={col_let}83+{col_let}84-{col_let}85-{col_let}86")
    cell.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    cell.fill = fill_accent_blue
    cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_total

# --- Section V: Discounting & Present Value ---
add_section_header(89, "V. DISCOUNTING & PRESENT VALUE (MID-YEAR CONVENTION)")

headers_s5 = ["Discount Parameters", "", "Year 1 (FY26E)", "Year 2 (FY27E)", "Year 3 (FY28E)", "Year 4 (FY29E)", "Year 5 (FY30E)"]
for col_idx, text in enumerate(headers_s5, start=1):
    c = ws.cell(row=90, column=col_idx, value=text)
    c.font = font_tbl_hdr
    c.fill = fill_soft_blue
    c.alignment = align_center if col_idx > 2 else align_left
    c.border = border_header

# Row 91: Discount Period
ws["A91"] = "Discount Period (Years - Mid-Year Convention)"
ws["A91"].font = font_regular
ws["A91"].border = border_cell
ws["B91"] = ""
ws["B91"].border = border_cell
periods = [0.5, 1.5, 2.5, 3.5, 4.5]
for c_idx, p in enumerate(periods, start=3):
    cell = ws.cell(row=91, column=c_idx, value=p)
    cell.font = font_regular
    cell.number_format = "0.0"
    cell.alignment = align_center
    cell.border = border_cell

# Row 92: Discount Factor
ws["A92"] = "Discount Factor [ 1 / (1 + WACC)^t ]"
ws["A92"].font = font_regular
ws["A92"].border = border_cell
ws["B92"] = ""
ws["B92"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=92, column=c_idx, value=f"=1/(1+$B$69)^{col_let}91")
    cell.font = font_regular
    cell.number_format = "0.0000"
    cell.alignment = align_right
    cell.border = border_cell

# Row 93: Present Value of UFCF
ws["A93"] = "Present Value of UFCF ($M)"
ws["A93"].font = font_bold
ws["A93"].border = border_cell
ws["B93"] = ""
ws["B93"].border = border_cell
for c_idx in range(3, 8):
    col_let = get_column_letter(c_idx)
    cell = ws.cell(row=93, column=c_idx, value=f"={col_let}87*{col_let}92")
    cell.font = font_bold
    cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
    cell.alignment = align_right
    cell.border = border_cell

# Row 94: Sum of PV of Explicit Cash Flows
ws["A94"] = "Cumulative PV of 5-Year Explicit Cash Flows ($M)"
ws["A94"].font = font_bold
ws["A94"].fill = fill_soft_blue
ws["A94"].border = border_total
ws["B94"] = "=SUM(C93:G93)"
ws["B94"].font = font_bold
ws["B94"].fill = fill_soft_blue
ws["B94"].number_format = "$#,##0.00;($#,##0.00);$0.00"
ws["B94"].alignment = align_right
ws["B94"].border = border_total
ws.merge_cells("C94:G94")
for c_idx in range(3, 8):
    ws.cell(row=94, column=c_idx).fill = fill_soft_blue
    ws.cell(row=94, column=c_idx).border = border_total

# --- Section VI: Terminal Value & Valuation Summary ---
add_section_header(96, "VI. TERMINAL VALUE & VALUATION SUMMARY")

val_summary_items = [
    ("Terminal Value Assumptions & Calculation:", None, None, None, True),
    ("  Final Year Projected UFCF (FY2030E) ($M)", "=G87", "$#,##0.00;($#,##0.00);$0.00", None, False),
    ("  Normalized Mature Terminal Cash Flow ($M)", "=IF(B4=1,40.0,IF(B4=2,110.0,240.0))", "$#,##0.00", "Normalized mature biotech cash flow post-commercial ramp", False),
    ("  Perpetual Terminal Growth Rate (g)", "=$H$47", "0.0%", None, False),
    ("  WACC (Discount Rate)", "=$B$69", "0.00%", None, False),
    ("  Terminal Year Cash Flow ($M)", "=B99*(1+B100)", "$#,##0.00", None, False),
    ("  Terminal Value at FY2030E ($M)", "=B102/(B101-B100)", "$#,##0.00", None, False),
    ("  PV of Terminal Value ($M)", "=B103/(1+B101)^4.5", "$#,##0.00", None, False),
    ("  Terminal Value as % of Enterprise Value", "=B104/B109", "0.0%", None, False),
    ("Enterprise Value to Equity Value Bridge ($M):", None, None, None, True),
    ("  PV of Explicit 5-Year Cash Flows ($M)", "=B94", "$#,##0.00;($#,##0.00);$0.00", None, False),
    ("  (+) PV of Terminal Value ($M)", "=B104", "$#,##0.00", None, False),
    ("Enterprise Value (EV) ($M)", "=B107+B108", "$#,##0.00;($#,##0.00);$0.00", None, False),
    ("  (-) Total Debt ($M)", "=$B$11", "($#,##0.00);($#,##0.00);$0.00", None, False),
    ("  (+) Total Cash & Equivalents ($M)", "=$B$10", "$#,##0.00", None, False),
    ("  (-) Net Debt ($M)", "=$B$12", "$#,##0.00;($#,##0.00);$0.00", None, False),
    ("Implied Equity Value ($M)", "=B109-B112", "$#,##0.00", None, False),
    ("  Diluted Shares Outstanding (M)", "=$B$8", "#,##0.00", None, False),
    ("Implied Price per Share ($)", "=B113/B114", "$#,##0.00", None, False),
    ("Current Market Share Price ($)", "=$B$7", "$#,##0.00", None, False),
    ("Implied Upside / (Downside) (%)", "=(B115/B116)-1", "+0.0%;-0.0%;0.0%", None, False),
]

current_r = 97
for item in val_summary_items:
    label, formula_val, num_fmt, comment_t, is_header = item
    if is_header:
        ws[f"A{current_r}"] = label
        ws[f"A{current_r}"].font = font_subhdr
        ws[f"A{current_r}"].fill = fill_soft_blue
        ws.merge_cells(f"A{current_r}:H{current_r}")
        for c_idx in range(1, 9):
            ws.cell(row=current_r, column=c_idx).fill = fill_soft_blue
            ws.cell(row=current_r, column=c_idx).border = border_cell
    else:
        ws[f"A{current_r}"] = label
        is_highlight = label in ["Enterprise Value (EV) ($M)", "Implied Equity Value ($M)", "Implied Price per Share ($)", "Implied Upside / (Downside) (%)"]
        ws[f"A{current_r}"].font = font_bold if is_highlight else font_regular
        ws[f"A{current_r}"].border = border_cell
        if is_highlight:
            ws[f"A{current_r}"].fill = fill_accent_blue
            
        ws[f"B{current_r}"] = formula_val
        ws[f"B{current_r}"].font = font_bold if is_highlight else font_regular
        ws[f"B{current_r}"].number_format = num_fmt
        ws[f"B{current_r}"].alignment = align_right
        ws[f"B{current_r}"].border = border_cell
        if is_highlight:
            ws[f"B{current_r}"].fill = fill_accent_blue
        if comment_t:
            ws[f"B{current_r}"].comment = Comment(comment_t, "Valuation Summary")
    current_r += 1

# --- Section VII: Sensitivity Analysis (5x5 Institutional Grids) ---
add_section_header(120, "VII. SENSITIVITY ANALYSIS (5x5 INSTITUTIONAL VALUATION GRIDS)")

# Table 1: WACC vs Terminal Growth Rate
ws["A121"] = "Sensitivity Table 1: Implied Share Price ($) vs. WACC and Perpetual Terminal Growth Rate (g)"
ws["A121"].font = font_subhdr
ws.merge_cells("A121:H121")

ws["A122"] = "WACC \\ g"
ws["A122"].font = font_tbl_hdr
ws["A122"].fill = fill_soft_blue
ws["A122"].alignment = align_center
ws["A122"].border = border_header

t1_g_values = [0.020, 0.025, 0.030, 0.035, 0.040]
t1_wacc_values = [0.1150, 0.1250, 0.1350, 0.1450, 0.1550]

for col_idx, g_val in enumerate(t1_g_values, start=2):
    c = ws.cell(row=122, column=col_idx, value=g_val)
    c.font = font_tbl_hdr
    c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
    c.number_format = "0.0%"
    c.alignment = align_center
    c.border = border_header

for row_idx, w_val in enumerate(t1_wacc_values, start=123):
    ws.cell(row=row_idx, column=1, value=w_val)
    ws.cell(row=row_idx, column=1).font = font_tbl_hdr
    ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 125 else fill_soft_blue
    ws.cell(row=row_idx, column=1).number_format = "0.00%"
    ws.cell(row=row_idx, column=1).alignment = align_center
    ws.cell(row=row_idx, column=1).border = border_cell
    
    for col_idx in range(2, 7):
        col_let = get_column_letter(col_idx)
        # Full DCF recalculation formula:
        formula = (
            f"=(($C$87/(1+$A{row_idx})^0.5 + $D$87/(1+$A{row_idx})^1.5 + $E$87/(1+$A{row_idx})^2.5 + "
            f"$F$87/(1+$A{row_idx})^3.5 + $G$87/(1+$A{row_idx})^4.5 + "
            f"($B$99*(1+{col_let}$122)/($A{row_idx}-{col_let}$122))/(1+$A{row_idx})^4.5) - $B$12)/$B$8"
        )
        cell = ws.cell(row=row_idx, column=col_idx, value=formula)
        cell.font = font_bold if (row_idx == 125 and col_idx == 4) else font_regular
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        if row_idx == 125 and col_idx == 4:
            cell.fill = fill_accent_blue  # Center cell = Base Case!

# Table 2: FY26-30 Revenue Growth Delta vs Normalized Terminal Cash Flow ($M)
ws["A130"] = "Sensitivity Table 2: Implied Share Price ($) vs. Revenue Growth Delta & Normalized Terminal FCF ($M)"
ws["A130"].font = font_subhdr
ws.merge_cells("A130:H130")

ws["A131"] = "Term FCF \\ Rev Δ"
ws["A131"].font = font_tbl_hdr
ws["A131"].fill = fill_soft_blue
ws["A131"].alignment = align_center
ws["A131"].border = border_header

t2_rev_delta = [-0.200, -0.100, 0.000, 0.100, 0.200]
t2_terminal_fcf = [70.0, 90.0, 110.0, 130.0, 150.0]

for col_idx, g_val in enumerate(t2_rev_delta, start=2):
    c = ws.cell(row=131, column=col_idx, value=g_val)
    c.font = font_tbl_hdr
    c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
    c.number_format = "+0.0%;-0.0%;0.0%"
    c.alignment = align_center
    c.border = border_header

for row_idx, fcf_val in enumerate(t2_terminal_fcf, start=132):
    ws.cell(row=row_idx, column=1, value=fcf_val)
    ws.cell(row=row_idx, column=1).font = font_tbl_hdr
    ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 134 else fill_soft_blue
    ws.cell(row=row_idx, column=1).number_format = "$#,##0.0"
    ws.cell(row=row_idx, column=1).alignment = align_center
    ws.cell(row=row_idx, column=1).border = border_cell
    
    for col_idx in range(2, 7):
        col_let = get_column_letter(col_idx)
        # Scaled DCF directly with live Excel formula:
        formula = (
            f"=((($C$87*(1+{col_let}$131))/(1+$B$69)^0.5 + "
            f"($D$87*(1+{col_let}$131))/(1+$B$69)^1.5 + "
            f"($E$87*(1+{col_let}$131))/(1+$B$69)^2.5 + "
            f"($F$87*(1+{col_let}$131))/(1+$B$69)^3.5 + "
            f"($G$87*(1+{col_let}$131))/(1+$B$69)^4.5 + "
            f"(($A{row_idx}*(1+$H$47)/($B$69-$H$47))/(1+$B$69)^4.5)) - $B$12)/$B$8"
        )
        cell = ws.cell(row=row_idx, column=col_idx, value=formula)
        cell.font = font_bold if (row_idx == 134 and col_idx == 4) else font_regular
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        if row_idx == 134 and col_idx == 4:
            cell.fill = fill_accent_blue  # Center cell = Base Case!

# Table 3: Beta vs Risk-Free Rate
ws["A139"] = "Sensitivity Table 3: Implied Share Price ($) vs. Beta & 10-Year Treasury Yield (Risk-Free Rate)"
ws["A139"].font = font_subhdr
ws.merge_cells("A139:H139")

ws["A140"] = "Beta \\ Risk-Free Rate"
ws["A140"].font = font_tbl_hdr
ws["A140"].fill = fill_soft_blue
ws["A140"].alignment = align_center
ws["A140"].border = border_header

t3_rf_values = [0.0369, 0.0419, 0.0469, 0.0519, 0.0569]
t3_beta_values = [1.29, 1.49, 1.69, 1.89, 2.09]

for col_idx, rf_val in enumerate(t3_rf_values, start=2):
    c = ws.cell(row=140, column=col_idx, value=rf_val)
    c.font = font_tbl_hdr
    c.fill = fill_accent_blue if col_idx == 4 else fill_soft_blue
    c.number_format = "0.00%"
    c.alignment = align_center
    c.border = border_header

for row_idx, beta_val in enumerate(t3_beta_values, start=141):
    ws.cell(row=row_idx, column=1, value=beta_val)
    ws.cell(row=row_idx, column=1).font = font_tbl_hdr
    ws.cell(row=row_idx, column=1).fill = fill_accent_blue if row_idx == 143 else fill_soft_blue
    ws.cell(row=row_idx, column=1).number_format = "0.00"
    ws.cell(row=row_idx, column=1).alignment = align_center
    ws.cell(row=row_idx, column=1).border = border_cell
    
    for col_idx in range(2, 7):
        col_let = get_column_letter(col_idx)
        # Derived WACC = Ke = Rf + Beta * ERP = {col_let}$140 + $A{row_idx} * 0.055
        wacc_expr = f"({col_let}$140 + $A{row_idx}*0.055)"
        formula = (
            f"=(($C$87/(1+{wacc_expr})^0.5 + $D$87/(1+{wacc_expr})^1.5 + $E$87/(1+{wacc_expr})^2.5 + "
            f"$F$87/(1+{wacc_expr})^3.5 + $G$87/(1+{wacc_expr})^4.5 + "
            f"($B$99*(1+$H$47)/({wacc_expr}-$H$47))/(1+{wacc_expr})^4.5) - $B$12)/$B$8"
        )
        cell = ws.cell(row=row_idx, column=col_idx, value=formula)
        cell.font = font_bold if (row_idx == 143 and col_idx == 4) else font_regular
        cell.number_format = "$#,##0.00;($#,##0.00);$0.00"
        cell.alignment = align_right
        cell.border = border_cell
        if row_idx == 143 and col_idx == 4:
            cell.fill = fill_accent_blue  # Center cell = Base Case!

# --- Column Widths Auto-fit ---
col_widths = {
    "A": 48,
    "B": 18,
    "C": 18,
    "D": 18,
    "E": 18,
    "F": 18,
    "G": 18,
    "H": 20
}
for col_let, width in col_widths.items():
    ws.column_dimensions[col_let].width = width

output_filename = "WVE_DCF_Model_Gemini-3.7-Flash_20260819.xlsx"
wb.save(output_filename)
print(f"Saved {output_filename} successfully.")
