#!/usr/bin/env python3
"""
generate_ingr_dcf.py - Creates institutional DCF model for Ingredion (INGR)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

def build_dcf_model(output_path="INGR_DCF_Valuation_Model.xlsx"):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    ws = wb.create_sheet(title="DCF Valuation")
    wb.remove(default_sheet)
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Palettes
    navy_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sub_navy_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    bear_fill = PatternFill(start_color="4A607A", end_color="4A607A", fill_type="solid")
    base_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    bull_fill = PatternFill(start_color="1E4D6B", end_color="1E4D6B", fill_type="solid")
    active_scenario_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    
    table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    light_blue_input_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    selector_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_cell_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    target_price_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    tbl_header_font = Font(name="Calibri", size=10, bold=True, color="000000")
    bold_font = Font(name="Calibri", size=10, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=10, color="000000")
    italic_font = Font(name="Calibri", size=9, italic=True, color="595959")
    target_price_font = Font(name="Calibri", size=12, bold=True, color="276A3C")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    dark_border_side = Side(border_style="thin", color="000000")
    double_bottom_side = Side(border_style="double", color="000000")
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    top_thin_bottom_double = Border(top=dark_border_side, bottom=double_bottom_side)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Formats
    CURR_M_FMT = "$#,##0.0;($#,##0.0);\"-\""
    PRICE_FMT = "$#,##0.00;($#,##0.00);\"-\""
    PCT_1_FMT = "0.0%"
    PCT_2_FMT = "0.00%"
    DEC_2_FMT = "0.00"

    # Set Column widths
    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16

    # Helper function for merged section header
    def make_section_header(row, text, fill=navy_header_fill, font=section_font, cols=9):
        col_end = get_column_letter(cols)
        ws[f"A{row}"] = text
        ws.merge_cells(f"A{row}:{col_end}{row}")
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = left_align
            ws.row_dimensions[row].height = 24

    # 1. Main Title
    ws["A1"] = "INGREDION INCORPORATED (NYSE: INGR) — 5-YEAR DCF VALUATION MODEL"
    ws.merge_cells("A1:I1")
    for c in range(1, 10):
        cell = ws.cell(row=1, column=c)
        cell.fill = navy_header_fill
        cell.font = title_font
        cell.alignment = left_align
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Institutional Equity Research | Discounted Cash Flow Valuation with Scenario Switches & Sensitivity Analysis"
    ws.merge_cells("A2:I2")
    for c in range(1, 10):
        cell = ws.cell(row=2, column=c)
        cell.fill = sub_navy_fill
        cell.font = subtitle_font
        cell.alignment = left_align
    ws.row_dimensions[2].height = 18

    # 2. Scenario Control
    make_section_header(4, "1. SCENARIO CONTROL & SELECTION")
    ws["A5"] = "Active Scenario Selector [1=Bear, 2=Base, 3=Bull]:"
    ws["A5"].font = bold_font
    ws["B5"] = 2
    ws["B5"].font = Font(name="Calibri", size=11, bold=True, color="B25900")
    ws["B5"].fill = selector_fill
    ws["B5"].alignment = center_align
    ws["B5"].comment = Comment("Source: User Scenario Switch (1 = Bear Case, 2 = Base Case, 3 = Bull Case)", "Antigravity DCF Builder")
    
    ws["C5"] = "Active Scenario Name:"
    ws["C5"].font = bold_font
    ws["C5"].alignment = right_align
    ws["D5"] = '=IF(B5=1,"Bear Case",IF(B5=2,"Base Case","Bull Case"))'
    ws["D5"].font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    ws["D5"].alignment = left_align

    ws["A6"] = "Scenario Narrative Summary:"
    ws["A6"].font = italic_font
    ws["B6"] = '=IF(B5=1,"Conservative: 1.0-2.0% top-line growth, 11.0-11.5% EBIT margin, terminal g=2.0%",IF(B5=2,"Base Case: 3.0-4.0% revenue growth, specialty mix expansion to 12.5% EBIT margin, terminal g=2.5%","Optimistic: 4.0-6.0% growth on clean-label demand, 13.5% EBIT margin, terminal g=3.0%"))'
    ws["B6"].font = italic_font
    ws.merge_cells("B6:I6")

    # 3. Market Data & Capital Structure
    make_section_header(8, "2. MARKET DATA & CAPITAL STRUCTURE")
    headers_8 = ["Metric", "Value", "Unit", "Reference / Logic", "Data Source & Filing Notes"]
    for idx, h in enumerate(headers_8, start=1):
        cell = ws.cell(row=9, column=idx, value=h)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx in [1, 4, 5] else center_align

    market_data = [
        (10, "Current Share Price", 105.19, "USD ($)", "Market Close Price", "Source: NYSE Market Data Aug 2026", PRICE_FMT, "Source: Market Data close price $105.19"),
        (11, "Diluted Shares Outstanding", 63.20, "Million (M)", "Filing Diluted Count", "Source: 2024 10-K & Guidance (63.0M - 64.0M)", DEC_2_FMT, "Source: 2024 Form 10-K Note on EPS and Diluted Shares"),
        (12, "Market Capitalization", "=B10*B11", "$ Millions", "Price * Diluted Shares", "Calculated Equity Value", CURR_M_FMT, None),
        (13, "Total Debt (Short + Long-Term)", 1831.0, "$ Millions", "Balance Sheet Debt", "Source: 2024 10-K Consolidated Balance Sheet", CURR_M_FMT, "Source: 2024 Form 10-K Total Debt $1,831M"),
        (14, "Less: Cash & Cash Equivalents", 997.0, "$ Millions", "Balance Sheet Cash", "Source: 2024 10-K Consolidated Balance Sheet", CURR_M_FMT, "Source: 2024 Form 10-K Cash & Cash Equivalents $997M"),
        (15, "Net Debt", "=B13-B14", "$ Millions", "Total Debt - Cash", "Calculated Net Debt", CURR_M_FMT, None),
        (16, "Enterprise Value (Market EV)", "=B12+B15", "$ Millions", "Market Cap + Net Debt", "Calculated Market Enterprise Value", CURR_M_FMT, None),
    ]

    for row_num, label, val, unit, ref, src, fmt, cmt in market_data:
        ws.cell(row=row_num, column=1, value=label).font = bold_font if "Net Debt" in label or "Enterprise Value" in label or "Capitalization" in label else regular_font
        c2 = ws.cell(row=row_num, column=2, value=val)
        c2.number_format = fmt
        c2.font = bold_font if str(val).startswith("=") else regular_font
        c2.alignment = right_align
        if cmt:
            c2.comment = Comment(cmt, "Antigravity DCF Builder")
            c2.fill = light_blue_input_fill
        ws.cell(row=row_num, column=3, value=unit).alignment = center_align
        ws.cell(row=row_num, column=4, value=ref).alignment = left_align
        ws.cell(row=row_num, column=5, value=src).alignment = left_align
        for col_idx in range(1, 6):
            ws.cell(row=row_num, column=col_idx).border = cell_border

    # 4. WACC Parameters
    make_section_header(18, "3. COST OF CAPITAL (WACC) PARAMETERS")
    headers_18 = ["Parameter", "Value", "Unit / Metric", "Formula / Methodology", "Source & Industry Benchmark"]
    for idx, h in enumerate(headers_18, start=1):
        cell = ws.cell(row=19, column=idx, value=h)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx in [1, 4, 5] else center_align

    wacc_data = [
        (20, "Risk-Free Rate (Rf)", 0.0450, "%", "10-Year US Treasury Benchmark", "Source: US Department of the Treasury Yield Curve", PCT_2_FMT, "Source: US 10-Year Treasury Yield Benchmark 4.50%"),
        (21, "Equity Beta (β)", 0.65, "Multiple", "5-Year Monthly Regression Beta", "Source: Bloomberg / S&P Market Intelligence", DEC_2_FMT, "Source: 5-Year Monthly Beta 0.65"),
        (22, "Equity Risk Premium (ERP)", 0.0550, "%", "US Market Risk Premium", "Source: Damodaran ERP (2025/2026 Standard)", PCT_2_FMT, "Source: Damodaran US Equity Risk Premium 5.50%"),
        (23, "Cost of Equity (Ke)", "=B20+B21*B22", "%", "CAPM: Rf + Beta * ERP", "Estimated Cost of Equity", PCT_2_FMT, None),
        (24, "Pre-Tax Cost of Debt (Kd)", 0.0530, "%", "Effective Borrowing Cost", "Source: 2024 10-K Interest Expense / Avg Debt", PCT_2_FMT, "Source: 2024 10-K Effective Borrowing Rate ~5.30%"),
        (25, "Marginal Effective Tax Rate (t)", 0.2500, "%", "Statutory Federal + State/Intl", "Source: Historical Normalized Effective Tax Rate", PCT_2_FMT, "Source: Historical Normalized Tax Rate 25.0%"),
        (26, "After-Tax Cost of Debt [Kd*(1-t)]", "=B24*(1-B25)", "%", "Kd * (1 - t)", "Calculated After-Tax Cost of Debt", PCT_2_FMT, None),
        (27, "Weight of Equity (We)", "=B12/B16", "%", "Market Cap / Enterprise Value", "Market Value Capital Weight", PCT_1_FMT, None),
        (28, "Weight of Debt (Wd)", "=B15/B16", "%", "Net Debt / Enterprise Value", "Market Value Debt Weight", PCT_1_FMT, None),
        (29, "Weighted Average Cost of Capital (WACC)", "=B23*B27+B26*B28", "%", "(Ke * We) + (Kd*(1-t) * Wd)", "Calculated Model Discount Rate", PCT_2_FMT, None),
        (30, "Terminal Growth Rate (g)", 0.0250, "%", "Long-term Sustainable GDP Growth", "Base Perpetuity Growth Assumption", PCT_2_FMT, "Source: Long-term Global GDP Growth Target 2.50%"),
    ]

    for row_num, label, val, unit, ref, src, fmt, cmt in wacc_data:
        is_highlight = "WACC" in label or "Cost of Equity" in label
        c1 = ws.cell(row=row_num, column=1, value=label)
        c1.font = bold_font if is_highlight else regular_font
        c2 = ws.cell(row=row_num, column=2, value=val)
        c2.number_format = fmt
        c2.font = bold_font if is_highlight or str(val).startswith("=") else regular_font
        c2.alignment = right_align
        if is_highlight:
            c1.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        if cmt:
            c2.comment = Comment(cmt, "Antigravity DCF Builder")
            c2.fill = light_blue_input_fill
        ws.cell(row=row_num, column=3, value=unit).alignment = center_align
        ws.cell(row=row_num, column=4, value=ref).alignment = left_align
        ws.cell(row=row_num, column=5, value=src).alignment = left_align
        for col_idx in range(1, 6):
            ws.cell(row=row_num, column=col_idx).border = cell_border

    # 5. Scenario Assumption Blocks
    make_section_header(32, "4. SCENARIO ASSUMPTIONS BLOCK (BEAR / BASE / BULL)")
    
    # Bear Case (Rows 33-40)
    make_section_header(33, "BEAR CASE ASSUMPTIONS (Scenario 1)", fill=bear_fill, cols=9)
    scenario_years = ["Assumption Metric", "FY2022", "FY2023", "FY2024A", "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for idx, y in enumerate(scenario_years, start=1):
        cell = ws.cell(row=34, column=idx, value=y)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx == 1 else center_align

    bear_assumptions = [
        (35, "Revenue Growth Rate (%)", [None, None, None, 0.010, 0.015, 0.020, 0.020, 0.020], PCT_1_FMT),
        (36, "EBIT Margin (%)", [None, None, None, 0.110, 0.112, 0.115, 0.115, 0.115], PCT_1_FMT),
        (37, "D&A (% of Revenue)", [None, None, None, 0.028, 0.028, 0.028, 0.028, 0.028], PCT_1_FMT),
        (38, "CapEx (% of Revenue)", [None, None, None, 0.038, 0.038, 0.038, 0.038, 0.038], PCT_1_FMT),
        (39, "NWC Change (% of ΔRev)", [None, None, None, 0.015, 0.015, 0.015, 0.015, 0.015], PCT_1_FMT),
        (40, "Terminal Growth Rate (g)", [None, None, None, 0.020, 0.020, 0.020, 0.020, 0.020], PCT_1_FMT),
    ]
    for row_num, label, vals, fmt in bear_assumptions:
        ws.cell(row=row_num, column=1, value=label).font = regular_font
        for c_idx, val in enumerate(vals, start=2):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.number_format = fmt
            cell.alignment = right_align
            if val is not None:
                cell.fill = light_blue_input_fill
                cell.comment = Comment(f"Bear Case Assumption: {label} = {val:.1%}", "Antigravity DCF Builder")
            cell.border = cell_border

    # Base Case (Rows 41-48)
    make_section_header(41, "BASE CASE ASSUMPTIONS (Scenario 2 — Consensus / Most Likely)", fill=base_fill, cols=9)
    for idx, y in enumerate(scenario_years, start=1):
        cell = ws.cell(row=42, column=idx, value=y)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx == 1 else center_align

    base_assumptions = [
        (43, "Revenue Growth Rate (%)", [None, None, None, 0.030, 0.035, 0.040, 0.035, 0.030], PCT_1_FMT),
        (44, "EBIT Margin (%)", [None, None, None, 0.120, 0.122, 0.124, 0.125, 0.125], PCT_1_FMT),
        (45, "D&A (% of Revenue)", [None, None, None, 0.028, 0.028, 0.028, 0.028, 0.028], PCT_1_FMT),
        (46, "CapEx (% of Revenue)", [None, None, None, 0.038, 0.037, 0.036, 0.035, 0.035], PCT_1_FMT),
        (47, "NWC Change (% of ΔRev)", [None, None, None, 0.015, 0.015, 0.015, 0.015, 0.015], PCT_1_FMT),
        (48, "Terminal Growth Rate (g)", [None, None, None, 0.025, 0.025, 0.025, 0.025, 0.025], PCT_1_FMT),
    ]
    for row_num, label, vals, fmt in base_assumptions:
        ws.cell(row=row_num, column=1, value=label).font = regular_font
        for c_idx, val in enumerate(vals, start=2):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.number_format = fmt
            cell.alignment = right_align
            if val is not None:
                cell.fill = light_blue_input_fill
                cell.comment = Comment(f"Base Case Assumption: {label} = {val:.1%}", "Antigravity DCF Builder")
            cell.border = cell_border

    # Bull Case (Rows 49-56)
    make_section_header(49, "BULL CASE ASSUMPTIONS (Scenario 3 — High Growth / Specialty Expansion)", fill=bull_fill, cols=9)
    for idx, y in enumerate(scenario_years, start=1):
        cell = ws.cell(row=50, column=idx, value=y)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx == 1 else center_align

    bull_assumptions = [
        (51, "Revenue Growth Rate (%)", [None, None, None, 0.050, 0.055, 0.060, 0.050, 0.040], PCT_1_FMT),
        (52, "EBIT Margin (%)", [None, None, None, 0.125, 0.128, 0.132, 0.135, 0.135], PCT_1_FMT),
        (53, "D&A (% of Revenue)", [None, None, None, 0.028, 0.028, 0.028, 0.028, 0.028], PCT_1_FMT),
        (54, "CapEx (% of Revenue)", [None, None, None, 0.038, 0.036, 0.035, 0.034, 0.034], PCT_1_FMT),
        (55, "NWC Change (% of ΔRev)", [None, None, None, 0.015, 0.015, 0.015, 0.015, 0.015], PCT_1_FMT),
        (56, "Terminal Growth Rate (g)", [None, None, None, 0.030, 0.030, 0.030, 0.030, 0.030], PCT_1_FMT),
    ]
    for row_num, label, vals, fmt in bull_assumptions:
        ws.cell(row=row_num, column=1, value=label).font = regular_font
        for c_idx, val in enumerate(vals, start=2):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.number_format = fmt
            cell.alignment = right_align
            if val is not None:
                cell.fill = light_blue_input_fill
                cell.comment = Comment(f"Bull Case Assumption: {label} = {val:.1%}", "Antigravity DCF Builder")
            cell.border = cell_border

    # Active Consolidated Assumptions (Rows 57-64)
    make_section_header(57, "ACTIVE CONSOLIDATED ASSUMPTIONS (Dynamically Linked via Selector B5)", fill=active_scenario_fill, cols=9)
    for idx, y in enumerate(scenario_years, start=1):
        cell = ws.cell(row=58, column=idx, value=y)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx == 1 else center_align

    active_items = [
        (59, "Revenue Growth Rate (%)", 35, 43, 51),
        (60, "EBIT Margin (%)", 36, 44, 52),
        (61, "D&A (% of Revenue)", 37, 45, 53),
        (62, "CapEx (% of Revenue)", 38, 46, 54),
        (63, "NWC Change (% of ΔRev)", 39, 47, 55),
        (64, "Terminal Growth Rate (g)", 40, 48, 56),
    ]
    for row_num, label, r_bear, r_base, r_bull in active_items:
        ws.cell(row=row_num, column=1, value=label).font = bold_font
        for c_idx in range(5, 10): # Cols E to I
            col_letter = get_column_letter(c_idx)
            fmla = f"=IF($B$5=1,{col_letter}{r_bear},IF($B$5=2,{col_letter}{r_base},{col_letter}{r_bull}))"
            cell = ws.cell(row=row_num, column=c_idx, value=fmla)
            cell.number_format = PCT_1_FMT
            cell.font = bold_font
            cell.alignment = right_align
        for col_idx in range(1, 10):
            ws.cell(row=row_num, column=col_idx).border = cell_border

    # 6. Financial Statements & Unlevered Free Cash Flow Schedule
    make_section_header(66, "5. HISTORICAL FINANCIALS & 5-YEAR PROJECTIONS ($M)")
    for idx, y in enumerate(scenario_years, start=1):
        cell = ws.cell(row=67, column=idx, value=y)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx == 1 else center_align

    # Row 68: Net Sales
    ws.cell(row=68, column=1, value="Net Sales (Revenue)").font = bold_font
    ws.cell(row=68, column=2, value=7946.0).comment = Comment("Source: 2024 10-K Consolidated Statements of Income (FY2022)", "Antigravity DCF Builder")
    ws.cell(row=68, column=3, value=8160.0).comment = Comment("Source: 2024 10-K Consolidated Statements of Income (FY2023)", "Antigravity DCF Builder")
    ws.cell(row=68, column=4, value=7430.0).comment = Comment("Source: 2024 10-K Consolidated Statements of Income (FY2024)", "Antigravity DCF Builder")
    ws.cell(row=68, column=5, value="=D68*(1+E59)")
    ws.cell(row=68, column=6, value="=E68*(1+F59)")
    ws.cell(row=68, column=7, value="=F68*(1+G59)")
    ws.cell(row=68, column=8, value="=G68*(1+H59)")
    ws.cell(row=68, column=9, value="=H68*(1+I59)")
    for c in range(2, 10):
        cell = ws.cell(row=68, column=c)
        cell.number_format = CURR_M_FMT
        cell.font = bold_font
        cell.alignment = right_align

    # Row 69: YoY Growth
    ws.cell(row=69, column=1, value="  YoY Revenue Growth (%)").font = italic_font
    ws.cell(row=69, column=2, value=0.152)
    ws.cell(row=69, column=3, value="=(C68-B68)/B68")
    ws.cell(row=69, column=4, value="=(D68-C68)/C68")
    for c in range(5, 10):
        prev = get_column_letter(c-1)
        curr = get_column_letter(c)
        ws.cell(row=69, column=c, value=f"=({curr}68-{prev}68)/{prev}68")
    for c in range(2, 10):
        cell = ws.cell(row=69, column=c)
        cell.number_format = PCT_1_FMT
        cell.font = italic_font
        cell.alignment = right_align

    # Row 70: Cost of Sales
    ws.cell(row=70, column=1, value="Cost of Sales (COGS)").font = regular_font
    ws.cell(row=70, column=2, value=6452.0).comment = Comment("Source: 2024 10-K COGS FY2022", "Antigravity DCF Builder")
    ws.cell(row=70, column=3, value=6411.0).comment = Comment("Source: 2024 10-K COGS FY2023", "Antigravity DCF Builder")
    ws.cell(row=70, column=4, value=5639.0).comment = Comment("Source: 2024 10-K COGS FY2024", "Antigravity DCF Builder")
    for c in range(5, 10):
        curr = get_column_letter(c)
        ws.cell(row=70, column=c, value=f"={curr}68-{curr}71")
    for c in range(2, 10):
        cell = ws.cell(row=70, column=c)
        cell.number_format = CURR_M_FMT
        cell.alignment = right_align

    # Row 71: Gross Profit
    ws.cell(row=71, column=1, value="Gross Profit").font = bold_font
    ws.cell(row=71, column=2, value="=B68-B70")
    ws.cell(row=71, column=3, value="=C68-C70")
    ws.cell(row=71, column=4, value="=D68-D70")
    ws.cell(row=71, column=5, value="=E68*0.245")
    ws.cell(row=71, column=6, value="=F68*0.248")
    ws.cell(row=71, column=7, value="=G68*0.250")
    ws.cell(row=71, column=8, value="=H68*0.252")
    ws.cell(row=71, column=9, value="=I68*0.255")
    for c in range(2, 10):
        cell = ws.cell(row=71, column=c)
        cell.number_format = CURR_M_FMT
        cell.font = bold_font
        cell.alignment = right_align

    # Row 72: Gross Margin
    ws.cell(row=72, column=1, value="  Gross Profit Margin (%)").font = italic_font
    for c in range(2, 10):
        curr = get_column_letter(c)
        cell = ws.cell(row=72, column=c, value=f"={curr}71/{curr}68")
        cell.number_format = PCT_1_FMT
        cell.font = italic_font
        cell.alignment = right_align

    # Row 73: Operating Income (EBIT)
    ws.cell(row=73, column=1, value="Operating Income (EBIT)").font = bold_font
    ws.cell(row=73, column=2, value=762.0).comment = Comment("Source: 2024 10-K Operating Income FY2022", "Antigravity DCF Builder")
    ws.cell(row=73, column=3, value=957.0).comment = Comment("Source: 2024 10-K Operating Income FY2023", "Antigravity DCF Builder")
    ws.cell(row=73, column=4, value=883.0).comment = Comment("Source: 2024 10-K Operating Income FY2024", "Antigravity DCF Builder")
    for c in range(5, 10):
        curr = get_column_letter(c)
        ws.cell(row=73, column=c, value=f"={curr}68*{curr}60")
    for c in range(2, 10):
        cell = ws.cell(row=73, column=c)
        cell.number_format = CURR_M_FMT
        cell.font = bold_font
        cell.alignment = right_align

    # Row 74: EBIT Margin
    ws.cell(row=74, column=1, value="  Operating EBIT Margin (%)").font = italic_font
    for c in range(2, 10):
        curr = get_column_letter(c)
        cell = ws.cell(row=74, column=c, value=f"={curr}73/{curr}68")
        cell.number_format = PCT_1_FMT
        cell.font = italic_font
        cell.alignment = right_align

    # Row 75: Taxes
    ws.cell(row=75, column=1, value="Less: Taxes on Operating Profit (t = 25.0%)").font = regular_font
    for c in range(2, 10):
        curr = get_column_letter(c)
        cell = ws.cell(row=75, column=c, value=f"={curr}73*$B$25")
        cell.number_format = CURR_M_FMT
        cell.alignment = right_align

    # Row 76: NOPAT
    ws.cell(row=76, column=1, value="Net Operating Profit After Tax (NOPAT)").font = bold_font
    for c in range(2, 10):
        curr = get_column_letter(c)
        cell = ws.cell(row=76, column=c, value=f"={curr}73-{curr}75")
        cell.number_format = CURR_M_FMT
        cell.font = bold_font
        cell.alignment = right_align

    # Row 77: D&A
    ws.cell(row=77, column=1, value="Plus: Depreciation & Amortization (D&A)").font = regular_font
    ws.cell(row=77, column=2, value=215.0).comment = Comment("Source: 2024 10-K Cash Flow D&A FY2022", "Antigravity DCF Builder")
    ws.cell(row=77, column=3, value=219.0).comment = Comment("Source: 2024 10-K Cash Flow D&A FY2023", "Antigravity DCF Builder")
    ws.cell(row=77, column=4, value=214.0).comment = Comment("Source: 2024 10-K Cash Flow D&A FY2024", "Antigravity DCF Builder")
    for c in range(5, 10):
        curr = get_column_letter(c)
        ws.cell(row=77, column=c, value=f"={curr}68*{curr}61")
    for c in range(2, 10):
        cell = ws.cell(row=77, column=c)
        cell.number_format = CURR_M_FMT
        cell.alignment = right_align

    # Row 78: CapEx
    ws.cell(row=78, column=1, value="Less: Capital Expenditures (CapEx)").font = regular_font
    ws.cell(row=78, column=2, value=300.0).comment = Comment("Source: 2024 10-K Cash Flow CapEx FY2022", "Antigravity DCF Builder")
    ws.cell(row=78, column=3, value=316.0).comment = Comment("Source: 2024 10-K Cash Flow CapEx FY2023", "Antigravity DCF Builder")
    ws.cell(row=78, column=4, value=301.0).comment = Comment("Source: 2024 10-K Cash Flow CapEx FY2024", "Antigravity DCF Builder")
    for c in range(5, 10):
        curr = get_column_letter(c)
        ws.cell(row=78, column=c, value=f"={curr}68*{curr}62")
    for c in range(2, 10):
        cell = ws.cell(row=78, column=c)
        cell.number_format = CURR_M_FMT
        cell.alignment = right_align

    # Row 79: Change in NWC
    ws.cell(row=79, column=1, value="Less: Change in Net Working Capital (ΔNWC)").font = regular_font
    ws.cell(row=79, column=2, value=-42.0).comment = Comment("Source: 2024 10-K Historical NWC change FY2022", "Antigravity DCF Builder")
    ws.cell(row=79, column=3, value=125.0).comment = Comment("Source: 2024 10-K Historical NWC change FY2023", "Antigravity DCF Builder")
    ws.cell(row=79, column=4, value=35.0).comment = Comment("Source: 2024 10-K Historical NWC change FY2024", "Antigravity DCF Builder")
    for c in range(5, 10):
        curr = get_column_letter(c)
        prev = get_column_letter(c-1)
        ws.cell(row=79, column=c, value=f"=({curr}68-{prev}68)*{curr}63")
    for c in range(2, 10):
        cell = ws.cell(row=79, column=c)
        cell.number_format = CURR_M_FMT
        cell.alignment = right_align

    # Row 80: Unlevered FCF
    ws.cell(row=80, column=1, value="Unlevered Free Cash Flow (UFCF)").font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    for c in range(2, 10):
        curr = get_column_letter(c)
        cell = ws.cell(row=80, column=c, value=f"={curr}76+{curr}77-{curr}78-{curr}79")
        cell.number_format = CURR_M_FMT
        cell.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
        cell.fill = table_header_fill
        cell.alignment = right_align
        cell.border = top_thin_bottom_double

    for r in range(68, 80):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = cell_border

    # 7. Discounting & Valuation Summary Bridge
    make_section_header(82, "6. DCF VALUATION & EQUITY VALUE BRIDGE")
    
    val_headers = ["Valuation Metric / Schedule", None, None, None, "FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]
    for idx, h in enumerate(val_headers, start=1):
        if h:
            cell = ws.cell(row=83, column=idx, value=h)
            cell.fill = table_header_fill
            cell.font = tbl_header_font
            cell.alignment = left_align if idx == 1 else center_align

    # Row 84: Mid-Year Discount Period (t)
    ws.cell(row=84, column=1, value="Discount Period (t - Mid-Year)").font = regular_font
    periods = [0.5, 1.5, 2.5, 3.5, 4.5]
    for idx, p in enumerate(periods, start=5):
        cell = ws.cell(row=84, column=idx, value=p)
        cell.number_format = DEC_2_FMT
        cell.alignment = center_align
        cell.border = cell_border

    # Row 85: Discount Factor
    ws.cell(row=85, column=1, value="Discount Factor [1 / (1 + WACC)^t]").font = regular_font
    for idx in range(5, 10):
        curr = get_column_letter(idx)
        cell = ws.cell(row=85, column=idx, value=f"=1/(1+$B$29)^{curr}84")
        cell.number_format = "0.0000"
        cell.alignment = right_align
        cell.border = cell_border

    # Row 86: PV of FCF
    ws.cell(row=86, column=1, value="Present Value of UFCF (PV of FCF)").font = bold_font
    for idx in range(5, 10):
        curr = get_column_letter(idx)
        cell = ws.cell(row=86, column=idx, value=f"={curr}80*{curr}85")
        cell.number_format = CURR_M_FMT
        cell.font = bold_font
        cell.alignment = right_align
        cell.fill = table_header_fill
        cell.border = cell_border

    # Valuation Summary Bridge Table
    bridge_headers = ["Equity Bridge Component", "Amount ($M)", "Unit / Metric", "Formula Reference", "Notes & Description"]
    for idx, h in enumerate(bridge_headers, start=1):
        cell = ws.cell(row=87, column=idx, value=h)
        cell.fill = table_header_fill
        cell.font = tbl_header_font
        cell.alignment = left_align if idx in [1, 4, 5] else center_align

    bridge_rows = [
        (88, "Cumulative PV of 5-Year Explicit FCFs", "=SUM(E86:I86)", "$ Millions", "SUM(E86:I86)", "Sum of discounted cash flows (FY25E - FY29E)", CURR_M_FMT, bold_font, None),
        (89, "Terminal Year Normalized FCF", "=I80*(1+E64)", "$ Millions", "FY2029E FCF * (1 + g)", "Normalized ongoing cash flow at horizon", CURR_M_FMT, regular_font, None),
        (90, "Terminal Value at Horizon (Perpetuity Growth)", "=B89/($B$29-E64)", "$ Millions", "Terminal FCF / (WACC - g)", "Implied undiscounted terminal value", CURR_M_FMT, regular_font, None),
        (91, "Terminal Value Discount Factor (Year 5)", "=1/(1+$B$29)^5", "Multiple", "1 / (1 + WACC)^5", "Discount factor to PV at horizon", "0.0000", regular_font, None),
        (92, "Present Value of Terminal Value (PV of TV)", "=B90*B91", "$ Millions", "Terminal Value * TV Discount Factor", "Discounted terminal value", CURR_M_FMT, bold_font, None),
        (93, "Terminal Value as % of Enterprise Value", "=B92/B94", "%", "PV of TV / Enterprise Value", "Sanity check on terminal weighting (~60-80%)", PCT_1_FMT, italic_font, None),
        (94, "Implied Enterprise Value (EV)", "=B88+B92", "$ Millions", "PV of Explicit FCFs + PV of TV", "Total firm operating enterprise value", CURR_M_FMT, Font(name="Calibri", size=11, bold=True, color="002060"), table_header_fill),
        (95, "Less: Net Debt (Total Debt - Cash)", "=$B$15", "$ Millions", "Balance Sheet Net Debt (Total Debt - Cash)", "Deduction for net debt obligations", CURR_M_FMT, regular_font, None),
        (96, "Implied Equity Value", "=B94-B95", "$ Millions", "Enterprise Value - Net Debt", "Value of equity attributable to common shareholders", CURR_M_FMT, Font(name="Calibri", size=11, bold=True, color="002060"), table_header_fill),
        (97, "Diluted Shares Outstanding (M)", "=$B$11", "Million shares", "Diluted Share Count", "Diluted common share count", DEC_2_FMT, regular_font, None),
        (98, "IMPLIED TARGET PRICE PER SHARE ($)", "=B96/B97", "USD ($)", "Equity Value / Diluted Shares", "Primary Intrinsic Valuation Target", PRICE_FMT, target_price_font, target_price_fill),
        (99, "Current Market Share Price ($)", "=$B$10", "USD ($)", "Current Market Quote", "Current NYSE market price", PRICE_FMT, regular_font, None),
        (100, "Implied Upside / (Downside) (%)", "=(B98-B99)/B99", "%", "(Target Price - Current Price) / Current Price", "Estimated total capital return potential", PCT_1_FMT, Font(name="Calibri", size=11, bold=True, color="276A3C"), target_price_fill),
    ]

    for row_num, label, val, unit, ref, desc, fmt, fnt, fill in bridge_rows:
        c1 = ws.cell(row=row_num, column=1, value=label)
        c1.font = fnt
        c2 = ws.cell(row=row_num, column=2, value=val)
        c2.number_format = fmt
        c2.font = fnt
        c2.alignment = right_align
        if fill:
            c1.fill = fill
            c2.fill = fill
        if row_num == 98:
            c1.border = top_thin_bottom_double
            c2.border = top_thin_bottom_double
        else:
            c1.border = cell_border
            c2.border = cell_border
        
        ws.cell(row=row_num, column=3, value=unit).alignment = center_align
        ws.cell(row=row_num, column=3).border = cell_border
        ws.cell(row=row_num, column=4, value=ref).alignment = left_align
        ws.cell(row=row_num, column=4).border = cell_border
        ws.cell(row=row_num, column=5, value=desc).alignment = left_align
        ws.cell(row=row_num, column=5).border = cell_border

    # 8. Sensitivity Tables (3 Matrices - 5x5 Grids)
    make_section_header(102, "7. VALUATION SENSITIVITY ANALYSIS (3 INSTITUTIONAL 5x5 GRIDS)")

    # -------------------------------------------------------------
    # Table 1: WACC vs Terminal Growth Rate (g) (Rows 103 - 110)
    # -------------------------------------------------------------
    make_section_header(103, "TABLE 1: IMPLIED SHARE PRICE ($) — WACC vs. TERMINAL PERPETUITY GROWTH (g)", fill=sub_navy_fill, cols=7)
    
    ws.cell(row=104, column=2, value="WACC \\ g").font = tbl_header_font
    ws.cell(row=104, column=2).fill = table_header_fill
    ws.cell(row=104, column=2).alignment = center_align
    ws.cell(row=104, column=2).border = cell_border

    t1_g_formulas = ["=$E$64-0.010", "=$E$64-0.005", "=$E$64", "=$E$64+0.005", "=$E$64+0.010"]
    for c_idx, g_fmla in enumerate(t1_g_formulas, start=3):
        cell = ws.cell(row=104, column=c_idx, value=g_fmla)
        cell.number_format = PCT_2_FMT
        cell.font = bold_font if c_idx == 5 else tbl_header_font
        cell.fill = center_cell_fill if c_idx == 5 else table_header_fill
        cell.alignment = center_align
        cell.border = cell_border

    t1_wacc_formulas = ["=$B$29-0.010", "=$B$29-0.005", "=$B$29", "=$B$29+0.005", "=$B$29+0.010"]
    for r_offset, wacc_fmla in enumerate(t1_wacc_formulas):
        r_num = 105 + r_offset
        rh_cell = ws.cell(row=r_num, column=2, value=wacc_fmla)
        rh_cell.number_format = PCT_2_FMT
        rh_cell.font = bold_font if r_num == 107 else tbl_header_font
        rh_cell.fill = center_cell_fill if r_num == 107 else table_header_fill
        rh_cell.alignment = center_align
        rh_cell.border = cell_border

        for c_idx in range(3, 8):
            c_letter = get_column_letter(c_idx)
            fmla = f"=(( $E$80/(1+$B{r_num})^0.5 + $F$80/(1+$B{r_num})^1.5 + $G$80/(1+$B{r_num})^2.5 + $H$80/(1+$B{r_num})^3.5 + $I$80/(1+$B{r_num})^4.5 ) + (($I$80*(1+{c_letter}$104))/($B{r_num}-{c_letter}$104))/(1+$B{r_num})^5 - $B$15) / $B$11"
            cell = ws.cell(row=r_num, column=c_idx, value=fmla)
            cell.number_format = PRICE_FMT
            cell.alignment = right_align
            cell.border = cell_border
            if r_num == 107 and c_idx == 5:
                cell.fill = center_cell_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="002060")
                cell.comment = Comment("BASE CASE VALUE: WACC = $B$29, g = $E$64 (Matches Base DCF Valuation)", "Antigravity DCF Builder")
            else:
                cell.font = regular_font

    # -------------------------------------------------------------
    # Table 2: Revenue Growth Delta vs EBIT Margin Delta (Rows 112 - 119)
    # -------------------------------------------------------------
    make_section_header(112, "TABLE 2: IMPLIED SHARE PRICE ($) — REVENUE GROWTH DELTA vs. EBIT MARGIN DELTA", fill=sub_navy_fill, cols=7)

    ws.cell(row=113, column=2, value="Margin Δ \\ Rev Δ").font = tbl_header_font
    ws.cell(row=113, column=2).fill = table_header_fill
    ws.cell(row=113, column=2).alignment = center_align
    ws.cell(row=113, column=2).border = cell_border

    t2_rev_deltas = [-0.0200, -0.0100, 0.0000, 0.0100, 0.0200]
    for c_idx, rev_d in enumerate(t2_rev_deltas, start=3):
        cell = ws.cell(row=113, column=c_idx, value=rev_d)
        cell.number_format = "+0.0%;-0.0%;\"Base (0.0%)\""
        cell.font = bold_font if c_idx == 5 else tbl_header_font
        cell.fill = center_cell_fill if c_idx == 5 else table_header_fill
        cell.alignment = center_align
        cell.border = cell_border

    t2_margin_deltas = [-0.0200, -0.0100, 0.0000, 0.0100, 0.0200]
    for r_offset, m_delta in enumerate(t2_margin_deltas):
        r_num = 114 + r_offset
        rh_cell = ws.cell(row=r_num, column=2, value=m_delta)
        rh_cell.number_format = "+0.0%;-0.0%;\"Base (0.0%)\""
        rh_cell.font = bold_font if r_num == 116 else tbl_header_font
        rh_cell.fill = center_cell_fill if r_num == 116 else table_header_fill
        rh_cell.alignment = center_align
        rh_cell.border = cell_border

        for c_idx in range(3, 8):
            c_letter = get_column_letter(c_idx)
            fmla = (
                f"=(( ($E$76*(1+{c_letter}$113)*(1+$B{r_num}/$E$60) + $E$77*(1+{c_letter}$113) - $E$78*(1+{c_letter}$113) - $E$79*(1+{c_letter}$113))/(1+$B$29)^0.5 "
                f"+ ($F$76*(1+{c_letter}$113)^2*(1+$B{r_num}/$F$60) + $F$77*(1+{c_letter}$113)^2 - $F$78*(1+{c_letter}$113)^2 - $F$79*(1+{c_letter}$113))/(1+$B$29)^1.5 "
                f"+ ($G$76*(1+{c_letter}$113)^3*(1+$B{r_num}/$G$60) + $G$77*(1+{c_letter}$113)^3 - $G$78*(1+{c_letter}$113)^3 - $G$79*(1+{c_letter}$113))/(1+$B$29)^2.5 "
                f"+ ($H$76*(1+{c_letter}$113)^4*(1+$B{r_num}/$H$60) + $H$77*(1+{c_letter}$113)^4 - $H$78*(1+{c_letter}$113)^4 - $H$79*(1+{c_letter}$113))/(1+$B$29)^3.5 "
                f"+ ($I$76*(1+{c_letter}$113)^5*(1+$B{r_num}/$I$60) + $I$77*(1+{c_letter}$113)^5 - $I$78*(1+{c_letter}$113)^5 - $I$79*(1+{c_letter}$113))/(1+$B$29)^4.5 ) "
                f"+ ( (($I$76*(1+{c_letter}$113)^5*(1+$B{r_num}/$I$60) + $I$77*(1+{c_letter}$113)^5 - $I$78*(1+{c_letter}$113)^5 - $I$79*(1+{c_letter}$113))*(1+$E$64))/($B$29-$E$64) )/(1+$B$29)^5 - $B$15) / $B$11"
            )
            cell = ws.cell(row=r_num, column=c_idx, value=fmla)
            cell.number_format = PRICE_FMT
            cell.alignment = right_align
            cell.border = cell_border
            if r_num == 116 and c_idx == 5:
                cell.fill = center_cell_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="002060")
                cell.comment = Comment("BASE CASE VALUE: Rev Delta = 0.0%, Margin Delta = 0.0%", "Antigravity DCF Builder")
            else:
                cell.font = regular_font

    # -------------------------------------------------------------
    # Table 3: Beta (β) vs Risk-Free Rate (Rf) (Rows 121 - 128)
    # -------------------------------------------------------------
    make_section_header(121, "TABLE 3: IMPLIED SHARE PRICE ($) — EQUITY BETA (β) vs. RISK-FREE RATE (Rf)", fill=sub_navy_fill, cols=7)

    ws.cell(row=122, column=2, value="Beta (β) \\ Rf").font = tbl_header_font
    ws.cell(row=122, column=2).fill = table_header_fill
    ws.cell(row=122, column=2).alignment = center_align
    ws.cell(row=122, column=2).border = cell_border

    t3_rf_formulas = ["=$B$20-0.010", "=$B$20-0.005", "=$B$20", "=$B$20+0.005", "=$B$20+0.010"]
    for c_idx, rf_fmla in enumerate(t3_rf_formulas, start=3):
        cell = ws.cell(row=122, column=c_idx, value=rf_fmla)
        cell.number_format = PCT_2_FMT
        cell.font = bold_font if c_idx == 5 else tbl_header_font
        cell.fill = center_cell_fill if c_idx == 5 else table_header_fill
        cell.alignment = center_align
        cell.border = cell_border

    t3_beta_formulas = ["=$B$21-0.20", "=$B$21-0.10", "=$B$21", "=$B$21+0.10", "=$B$21+0.20"]
    for r_offset, beta_fmla in enumerate(t3_beta_formulas):
        r_num = 123 + r_offset
        rh_cell = ws.cell(row=r_num, column=2, value=beta_fmla)
        rh_cell.number_format = DEC_2_FMT
        rh_cell.font = bold_font if r_num == 125 else tbl_header_font
        rh_cell.fill = center_cell_fill if r_num == 125 else table_header_fill
        rh_cell.alignment = center_align
        rh_cell.border = cell_border

        for c_idx in range(3, 8):
            c_letter = get_column_letter(c_idx)
            wacc_expr = f"(({c_letter}$122+$B{r_num}*$B$22)*$B$27+$B$26*$B$28)"
            fmla = (
                f"=(( $E$80/(1+{wacc_expr})^0.5 + $F$80/(1+{wacc_expr})^1.5 + $G$80/(1+{wacc_expr})^2.5 + $H$80/(1+{wacc_expr})^3.5 + $I$80/(1+{wacc_expr})^4.5 ) "
                f"+ (($I$80*(1+$E$64))/({wacc_expr}-$E$64))/(1+{wacc_expr})^5 - $B$15) / $B$11"
            )
            cell = ws.cell(row=r_num, column=c_idx, value=fmla)
            cell.number_format = PRICE_FMT
            cell.alignment = right_align
            cell.border = cell_border
            if r_num == 125 and c_idx == 5:
                cell.fill = center_cell_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="002060")
                cell.comment = Comment("BASE CASE VALUE: Beta = $B$21, Rf = $B$20", "Antigravity DCF Builder")
            else:
                cell.font = regular_font

    wb.save(output_path)
    print(f"Successfully generated DCF model at: {output_path}")

if __name__ == "__main__":
    build_dcf_model()
