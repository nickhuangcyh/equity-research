#!/usr/bin/env python3
"""
recalc.py - Excel formula error scanner (openpyxl-based, no LibreOffice required)
Usage: python recalc.py <model.xlsx> [timeout_seconds]

Scans all cells for Excel error values:
  #REF!, #DIV/0!, #VALUE!, #NAME?, #NULL!, #NUM!, #N/A
Returns JSON with status, total_errors, total_formulas, error_summary.
"""

import sys
import json
from pathlib import Path

EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def scan_workbook(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}

    if not Path(path).exists():
        return {"status": "error", "message": f"File not found: {path}"}

    try:
        wb_values = openpyxl.load_workbook(path, data_only=True)
        wb_formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        return {"status": "error", "message": f"Cannot open file: {e}"}

    error_summary: dict = {}
    total_errors = 0
    total_formulas = 0

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                # Count formula cells
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    total_formulas += 1
                    # Check value for error
                    val = str(cell_v.value) if cell_v.value is not None else ""
                    for err in EXCEL_ERRORS:
                        if val == err:
                            total_errors += 1
                            loc = f"{sheet_name}!{cell_f.coordinate}"
                            if err not in error_summary:
                                error_summary[err] = {"count": 0, "locations": []}
                            error_summary[err]["count"] += 1
                            if len(error_summary[err]["locations"]) < 20:
                                error_summary[err]["locations"].append(loc)
                            break

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
    }
    if error_summary:
        result["error_summary"] = error_summary

    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <model.xlsx> [timeout_seconds]")
        sys.exit(1)

    path = sys.argv[1]
    result = scan_workbook(path)
    print(json.dumps(result, indent=2))

    if result.get("status") == "errors_found":
        sys.exit(1)
    elif result.get("status") == "error":
        sys.exit(2)
    else:
        sys.exit(0)
