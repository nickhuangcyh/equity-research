#!/usr/bin/env python3
"""
VITL (Vital Farms) DCF Model Builder
Two sheets: DCF (main model + sensitivity) and WACC
5-year forecast FY2026-FY2030, Bear/Base/Bull scenarios
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import date

OUTPUT_PATH = "/Users/nickhuang/Documents/personal/nickhuangcyh/equity-research/reports/dcf/VITL_DCF_Model_2026-08-20.xlsx"

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"   # section headers bg
LIGHT_BLUE  = "D9E1F2"   # sub-headers / col headers bg
MED_BLUE    = "BDD7EE"   # output rows / sensitivity center
LIGHT_GREY  = "F2F2F2"   # input cell bg
WHITE       = "FFFFFF"
FONT_BLUE   = "0000FF"   # hardcoded inputs
FONT_BLACK  = "000000"   # formulas
FONT_WHITE  = "FFFFFF"   # text on dark headers
FONT_GREEN  = "008000"   # sheet cross-references

# ── Number formats ───────────────────────────────────────────────────────────
FMT_USD     = '#,##0.0;(#,##0.0);"-"'
FMT_USD2    = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT     = '0.0%;(0.0%);"-"'
FMT_PCT1    = '0.0%'
FMT_MULT    = '0.0x'
FMT_INT     = '#,##0;(#,##0);"-"'

# ── Style helpers ────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color, bold=False, sz=10, italic=False):
    return Font(name="Calibri", color=hex_color, bold=bold, size=sz, italic=italic)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style="thin")
    return Border(bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def add_comment(cell, text):
    c = Comment(text, "Model Builder")
    c.width  = 300
    c.height = 80
    cell.comment = c

def style_section_header(ws, row, col_start, col_end, text):
    """Dark blue merged section header."""
    ws.cell(row=row, column=col_start).value = text
    ws.cell(row=row, column=col_start).font      = font(FONT_WHITE, bold=True, sz=10)
    ws.cell(row=row, column=col_start).fill      = fill(DARK_BLUE)
    ws.cell(row=row, column=col_start).alignment = align("left")
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    # border around merged range
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border()

def style_col_header(ws, row, col_start, col_end):
    """Light blue column header row."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(LIGHT_BLUE)
        cell.font      = font(FONT_BLACK, bold=True, sz=9)
        cell.alignment = align("center")
        cell.border    = thin_border()

def style_output_row(ws, row, col_start, col_end):
    """Medium-blue output row."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill  = fill(MED_BLUE)
        cell.font  = font(FONT_BLACK, bold=True, sz=10)
        cell.border = thin_border()

def apply_section_border(ws, row_start, row_end, col_start, col_end):
    """Thick border around a section block."""
    med = Side(style="medium")
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            left   = med if c == col_start else cell.border.left
            right  = med if c == col_end   else cell.border.right
            top    = med if r == row_start else cell.border.top
            bottom = med if r == row_end   else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# ── Row / column layout constants (DCF sheet) ───────────────────────────────
# Col A = label, B = hist FY2021, C = FY2022, D = FY2023, E = FY2024, F = FY2025
# G = FY2026E, H = FY2027E, I = FY2028E, J = FY2029E, K = FY2030E
# Col M = Consolidation column ("Selected" scenario values)

COL_LABEL  = 1   # A
COL_2021   = 2   # B
COL_2022   = 3   # C
COL_2023   = 4   # D
COL_2024   = 5   # E
COL_2025   = 6   # F
COL_2026E  = 7   # G
COL_2027E  = 8   # H
COL_2028E  = 9   # I
COL_2029E  = 10  # J
COL_2030E  = 11  # K
COL_BLANK  = 12  # L  (spacer)
COL_SEL    = 13  # M  (consolidation / selected-case values)
COL_END    = 14  # N  (right edge)

PROJ_COLS  = [COL_2026E, COL_2027E, COL_2028E, COL_2029E, COL_2030E]
HIST_COLS  = [COL_2021, COL_2022, COL_2023, COL_2024, COL_2025]
ALL_COLS   = HIST_COLS + PROJ_COLS

YEAR_LABELS_HIST = ["FY2021A", "FY2022A", "FY2023A", "FY2024A", "FY2025A"]
YEAR_LABELS_PROJ = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
YEAR_LABELS_ALL  = YEAR_LABELS_HIST + YEAR_LABELS_PROJ

# ── DCF sheet row layout ─────────────────────────────────────────────────────
# Section 1: Header
R_TITLE      = 1
R_SUBTITLE   = 2
R_BLANK1     = 3
R_CASE_HDR   = 4   # "Case Selector"
R_CASE_SEL   = 5   # dropdown  1=Bear 2=Base 3=Bull
R_CASE_NAME  = 6   # =IF(B5=1,"Bear",IF(B5=2,"Base","Bull"))
R_BLANK2     = 7

# Section 2: Market Data
R_MKT_HDR    = 8
R_MKT_COL    = 9
R_PRICE      = 10
R_SHARES     = 11
R_MKTCAP     = 12
R_NET_DEBT   = 13
R_BLANK3     = 14

# Section 3: Bear Case Assumptions
R_BEAR_HDR   = 15
R_BEAR_COL   = 16
R_BEAR_REV   = 17
R_BEAR_GM    = 18
R_BEAR_EBIT  = 19
R_BEAR_TAX   = 20
R_BEAR_DA    = 21
R_BEAR_CAPEX = 22
R_BEAR_NWC   = 23
R_BEAR_TG    = 24
R_BEAR_WACC  = 25
R_BLANK4     = 26

# Section 4: Base Case Assumptions
R_BASE_HDR   = 27
R_BASE_COL   = 28
R_BASE_REV   = 29
R_BASE_GM    = 30
R_BASE_EBIT  = 31
R_BASE_TAX   = 32
R_BASE_DA    = 33
R_BASE_CAPEX = 34
R_BASE_NWC   = 35
R_BASE_TG    = 36
R_BASE_WACC  = 37
R_BLANK5     = 38

# Section 5: Bull Case Assumptions
R_BULL_HDR   = 39
R_BULL_COL   = 40
R_BULL_REV   = 41
R_BULL_GM    = 42
R_BULL_EBIT  = 43
R_BULL_TAX   = 44
R_BULL_DA    = 45
R_BULL_CAPEX = 46
R_BULL_NWC   = 47
R_BULL_TG    = 48
R_BULL_WACC  = 49
R_BLANK6     = 50

# Section 6: Consolidation (Selected Case) - single-year values in col M
R_SEL_HDR    = 51
R_SEL_REVG   = 52   # revenue growth %  per year → 5 values in cols G:K mapped via INDEX
R_SEL_GM     = 53   # gross margin
R_SEL_EBIT   = 54   # EBIT margin
R_SEL_TAX    = 55   # tax rate
R_SEL_DA     = 56   # D&A % rev
R_SEL_CAPEX  = 57   # CapEx % rev
R_SEL_NWC    = 58   # NWC change % delta-rev
R_SEL_TG     = 59   # terminal growth
R_SEL_WACC   = 60   # WACC (linked from WACC sheet)
R_BLANK7     = 61

# Section 7: Income Statement
R_IS_HDR     = 62
R_IS_COL     = 63
R_REV        = 64
R_REV_GR     = 65
R_BLANK_IS1  = 66
R_GP         = 67
R_GM_PCT     = 68
R_BLANK_IS2  = 69
R_SGA        = 70
R_SGA_PCT    = 71
R_BLANK_IS3  = 72
R_EBIT       = 73
R_EBIT_PCT   = 74
R_BLANK_IS4  = 75
R_TAX        = 76
R_TAX_RATE   = 77
R_BLANK_IS5  = 78
R_NOPAT      = 79
R_BLANK_IS6  = 80

# Section 8: FCF Build
R_FCF_HDR    = 81
R_FCF_COL    = 82
R_FCF_NOPAT  = 83
R_DA         = 84
R_DA_PCT     = 85
R_CAPEX      = 86
R_CAPEX_PCT  = 87
R_NWC        = 88
R_NWC_PCT    = 89
R_BLANK_FCF  = 90
R_UFCF       = 91
R_BLANK_FCF2 = 92

# Section 9: Discount & Terminal Value
R_DCF_HDR    = 93
R_DCF_COL    = 94
R_PERIOD     = 95
R_DISC_FACT  = 96
R_PV_FCF     = 97
R_BLANK_D1   = 98
R_TV_HDR     = 99
R_TERM_FCF   = 100
R_TERM_VAL   = 101
R_PV_TV      = 102
R_BLANK_D2   = 103

# Section 10: Valuation Summary
R_VAL_HDR    = 104
R_SUM_PV_FCF = 105
R_SUM_PV_TV  = 106
R_EV         = 107
R_LESS_DEBT  = 108
R_EQ_VAL     = 109
R_BLANK_V1   = 110
R_SHARES_OUT = 111
R_IMPL_PRICE = 112
R_CUR_PRICE  = 113
R_UPSIDE     = 114
R_BLANK_V2   = 115

# Section 11: Sensitivity tables start at row 117
R_SENS1_HDR  = 117
R_SENS1_TLBL = 118   # title row (axis labels)
R_SENS1_DATA = 119   # first data row (5 rows × 5 cols)
# rows 119-123

R_SENS2_HDR  = 126
R_SENS2_TLBL = 127
R_SENS2_DATA = 128
# rows 128-132

R_SENS3_HDR  = 135
R_SENS3_TLBL = 136
R_SENS3_DATA = 137
# rows 137-141

print("Part 1: constants defined")

# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws_dcf  = wb.active
ws_dcf.title = "DCF"
ws_wacc = wb.create_sheet("WACC")

# ── Column widths (DCF sheet) ────────────────────────────────────────────────
ws_dcf.column_dimensions["A"].width = 34
for col_idx in range(2, 12):   # B–K  (hist + proj)
    ws_dcf.column_dimensions[get_column_letter(col_idx)].width = 12
ws_dcf.column_dimensions["L"].width = 2    # spacer
ws_dcf.column_dimensions["M"].width = 13   # consolidation
ws_dcf.column_dimensions["N"].width = 5

# ════════════════════════════════════════════════════════════════════════════
# SECTION 1: HEADER
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[R_TITLE].height = 20
c = ws_dcf.cell(R_TITLE, COL_LABEL, "Vital Farms, Inc. (VITL) – DCF Valuation Model")
c.font = font(DARK_BLUE, bold=True, sz=14)

c = ws_dcf.cell(R_SUBTITLE, COL_LABEL,
    f"NASDAQ: VITL  |  Model Date: {date.today().strftime('%B %d, %Y')}  |  Fiscal Year End: December")
c.font = font(FONT_BLACK, italic=True, sz=9)

# Case selector
ws_dcf.cell(R_CASE_HDR, COL_LABEL, "Scenario Selector (1=Bear / 2=Base / 3=Bull):")
ws_dcf.cell(R_CASE_HDR, COL_LABEL).font = font(FONT_BLACK, bold=True)

ws_dcf.cell(R_CASE_SEL, COL_LABEL, "Selected Case:")
c = ws_dcf.cell(R_CASE_SEL, 2, 2)          # default = Base
c.font      = font(FONT_BLUE, bold=True, sz=12)
c.fill      = fill(LIGHT_GREY)
c.border    = thin_border()
c.alignment = align("center")
add_comment(c, "Source: User Input — 1=Bear, 2=Base, 3=Bull")

c = ws_dcf.cell(R_CASE_NAME, 2)
c.value     = '=IF(B5=1,"Bear Case",IF(B5=2,"Base Case","Bull Case"))'
c.font      = font(FONT_GREEN, bold=True, sz=11)
c.alignment = align("center")

# ════════════════════════════════════════════════════════════════════════════
# SECTION 2: MARKET DATA
# ════════════════════════════════════════════════════════════════════════════
style_section_header(ws_dcf, R_MKT_HDR, COL_LABEL, COL_END, "MARKET DATA & KEY INPUTS")
style_col_header(ws_dcf, R_MKT_COL, COL_LABEL, COL_END)
ws_dcf.cell(R_MKT_COL, COL_LABEL, "Item")
ws_dcf.cell(R_MKT_COL, 2, "Value")

# Stock price
ws_dcf.cell(R_PRICE, COL_LABEL, "Current Stock Price")
c = ws_dcf.cell(R_PRICE, 2, 10.87)
c.font        = font(FONT_BLUE)
c.fill        = fill(LIGHT_GREY)
c.number_format = FMT_USD2
c.border      = thin_border()
add_comment(c, "Source: Market Data, 2026-08-19, NASDAQ Close Price")

# Shares outstanding
ws_dcf.cell(R_SHARES, COL_LABEL, "Diluted Shares Outstanding (M)")
c = ws_dcf.cell(R_SHARES, 2, 42.9)
c.font        = font(FONT_BLUE)
c.fill        = fill(LIGHT_GREY)
c.number_format = FMT_USD
c.border      = thin_border()
add_comment(c, "Source: Vital Farms Q2 2026 10-Q, Jun 28 2026, Basic shares 42,928,046 + dilution est.")

# Market cap (formula)
ws_dcf.cell(R_MKTCAP, COL_LABEL, "Market Capitalization ($M)")
c = ws_dcf.cell(R_MKTCAP, 2)
c.value         = "=B10*B11"
c.font          = font(FONT_BLACK)
c.number_format = FMT_USD
c.border        = thin_border()

# Net debt
ws_dcf.cell(R_NET_DEBT, COL_LABEL, "Net Debt ($M)  [Debt – Cash]")
c = ws_dcf.cell(R_NET_DEBT, 2, 8.8)
c.font        = font(FONT_BLUE)
c.fill        = fill(LIGHT_GREY)
c.number_format = FMT_USD
c.border      = thin_border()
add_comment(c, "Source: Vital Farms Q2 2026 Press Release, Jun 28 2026 — Cash $21.2M, LT Debt $30.0M → Net Debt = $8.8M")

apply_section_border(ws_dcf, R_MKT_HDR, R_NET_DEBT, COL_LABEL, COL_END)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 3–5: SCENARIO ASSUMPTION BLOCKS
# ════════════════════════════════════════════════════════════════════════════
# Each block: header, col-header row, 8 assumption rows (rev growth, GM, EBIT, tax, D&A, CapEx, NWC, terminal g/WACC)
# Rev growth / GM / EBIT / D&A / CapEx / NWC have values for each of the 5 projection years (cols G–K)
# Tax / terminal-g / WACC are scalar (single value in col G, rest blank)

def write_scenario_block(ws, hdr_row, col_row, rev_row, gm_row, ebit_row,
                          tax_row, da_row, capex_row, nwc_row, tg_row, wacc_row,
                          label, color,
                          rev_vals, gm_vals, ebit_vals, tax_val,
                          da_vals, capex_vals, nwc_vals, tg_val, wacc_val,
                          source_prefix):
    """Write one scenario assumption block."""
    style_section_header(ws, hdr_row, COL_LABEL, COL_END, label)

    # Col header row
    style_col_header(ws, col_row, COL_LABEL, COL_END)
    ws.cell(col_row, COL_LABEL, "Assumption")
    for i, proj_col in enumerate(PROJ_COLS):
        ws.cell(col_row, proj_col, YEAR_LABELS_PROJ[i])

    def write_row(row, row_label, values, fmt, comment_src, scalar=False):
        ws.cell(row, COL_LABEL, row_label).font = font(FONT_BLACK, sz=9)
        if scalar:
            c = ws.cell(row, COL_2026E, values[0])
            c.font = font(FONT_BLUE, sz=9); c.fill = fill(LIGHT_GREY)
            c.number_format = fmt; c.border = thin_border()
            add_comment(c, comment_src)
        else:
            for i, proj_col in enumerate(PROJ_COLS):
                c = ws.cell(row, proj_col, values[i])
                c.font = font(FONT_BLUE, sz=9); c.fill = fill(LIGHT_GREY)
                c.number_format = fmt; c.border = thin_border()
                add_comment(c, comment_src)

    write_row(rev_row,   "Revenue Growth (%)",          rev_vals,   FMT_PCT1, f"Source: {source_prefix}, revenue growth assumption")
    write_row(gm_row,    "Gross Margin (%)",             gm_vals,    FMT_PCT1, f"Source: {source_prefix}, gross margin assumption")
    write_row(ebit_row,  "EBIT Margin (%)",              ebit_vals,  FMT_PCT1, f"Source: {source_prefix}, EBIT margin assumption")
    write_row(tax_row,   "Tax Rate (%)",                 [tax_val],  FMT_PCT1, f"Source: {source_prefix}, normalized tax rate", scalar=True)
    write_row(da_row,    "D&A (% of Revenue)",           da_vals,    FMT_PCT1, f"Source: {source_prefix}, D&A % revenue assumption")
    write_row(capex_row, "CapEx (% of Revenue)",         capex_vals, FMT_PCT1, f"Source: {source_prefix}, CapEx % revenue assumption")
    write_row(nwc_row,   "NWC Change (% of Δ Revenue)",  nwc_vals,   FMT_PCT1, f"Source: {source_prefix}, NWC change assumption")
    write_row(tg_row,    "Terminal Growth Rate (%)",     [tg_val],   FMT_PCT1, f"Source: {source_prefix}, terminal growth rate", scalar=True)
    write_row(wacc_row,  "WACC (%)",                     [wacc_val], FMT_PCT1, f"Source: {source_prefix}, WACC assumption", scalar=True)

    apply_section_border(ws, hdr_row, wacc_row, COL_LABEL, COL_END)

# ── BEAR CASE ────────────────────────────────────────────────────────────────
write_scenario_block(
    ws_dcf,
    R_BEAR_HDR, R_BEAR_COL, R_BEAR_REV, R_BEAR_GM, R_BEAR_EBIT,
    R_BEAR_TAX, R_BEAR_DA, R_BEAR_CAPEX, R_BEAR_NWC, R_BEAR_TG, R_BEAR_WACC,
    "BEAR CASE ASSUMPTIONS", DARK_BLUE,
    rev_vals   = [0.03, 0.08, 0.10, 0.10, 0.09],   # FY26–FY30 revenue growth
    gm_vals    = [0.26, 0.30, 0.32, 0.33, 0.34],   # gross margin
    ebit_vals  = [-0.02, 0.02, 0.05, 0.07, 0.08],  # EBIT margin (FY26 low watermark)
    tax_val    = 0.24,
    da_vals    = [0.040, 0.038, 0.035, 0.033, 0.030],
    capex_vals = [0.095, 0.085, 0.070, 0.060, 0.050],
    nwc_vals   = [0.03, 0.03, 0.03, 0.02, 0.02],
    tg_val     = 0.025,
    wacc_val   = 0.115,
    source_prefix = "Analyst Estimate – Bear Case, 2026-08-20"
)

# ── BASE CASE ────────────────────────────────────────────────────────────────
write_scenario_block(
    ws_dcf,
    R_BASE_HDR, R_BASE_COL, R_BASE_REV, R_BASE_GM, R_BASE_EBIT,
    R_BASE_TAX, R_BASE_DA, R_BASE_CAPEX, R_BASE_NWC, R_BASE_TG, R_BASE_WACC,
    "BASE CASE ASSUMPTIONS", DARK_BLUE,
    rev_vals   = [0.035, 0.15, 0.18, 0.16, 0.13],  # FY26 guided midpoint ~3.7%, then recovery
    gm_vals    = [0.28, 0.34, 0.36, 0.37, 0.38],
    ebit_vals  = [0.00, 0.06, 0.09, 0.11, 0.13],
    tax_val    = 0.24,
    da_vals    = [0.038, 0.035, 0.032, 0.030, 0.028],
    capex_vals = [0.092, 0.075, 0.060, 0.050, 0.045],
    nwc_vals   = [0.025, 0.025, 0.020, 0.020, 0.015],
    tg_val     = 0.030,
    wacc_val   = 0.105,
    source_prefix = "Analyst Estimate – Base Case, 2026-08-20"
)

# ── BULL CASE ────────────────────────────────────────────────────────────────
write_scenario_block(
    ws_dcf,
    R_BULL_HDR, R_BULL_COL, R_BULL_REV, R_BULL_GM, R_BULL_EBIT,
    R_BULL_TAX, R_BULL_DA, R_BULL_CAPEX, R_BULL_NWC, R_BULL_TG, R_BULL_WACC,
    "BULL CASE ASSUMPTIONS", DARK_BLUE,
    rev_vals   = [0.04, 0.20, 0.22, 0.20, 0.16],
    gm_vals    = [0.30, 0.36, 0.38, 0.39, 0.40],
    ebit_vals  = [0.02, 0.09, 0.13, 0.15, 0.17],
    tax_val    = 0.24,
    da_vals    = [0.036, 0.032, 0.030, 0.028, 0.026],
    capex_vals = [0.090, 0.068, 0.055, 0.047, 0.042],
    nwc_vals   = [0.020, 0.020, 0.018, 0.015, 0.012],
    tg_val     = 0.035,
    wacc_val   = 0.095,
    source_prefix = "Analyst Estimate – Bull Case, 2026-08-20"
)

print("Part 2: market data + scenario blocks written")

# ════════════════════════════════════════════════════════════════════════════
# SECTION 6: CONSOLIDATION (Selected-Case) COLUMN
# ════════════════════════════════════════════════════════════════════════════
# Each cell in col M uses INDEX to pick Bear/Base/Bull based on B5
# For per-year assumptions (5 values), we embed per-year INDEX in each proj col

style_section_header(ws_dcf, R_SEL_HDR, COL_LABEL, COL_END,
    "SELECTED CASE ASSUMPTIONS  (auto-updates with Scenario Selector above)")

# Row labels in col A
sel_row_labels = {
    R_SEL_REVG:  "Revenue Growth — Selected (%)",
    R_SEL_GM:    "Gross Margin — Selected (%)",
    R_SEL_EBIT:  "EBIT Margin — Selected (%)",
    R_SEL_TAX:   "Tax Rate — Selected (%)",
    R_SEL_DA:    "D&A % Rev — Selected",
    R_SEL_CAPEX: "CapEx % Rev — Selected",
    R_SEL_NWC:   "NWC Δ% — Selected",
    R_SEL_TG:    "Terminal Growth Rate — Selected (%)",
    R_SEL_WACC:  "WACC — Selected",
}
for r, lbl in sel_row_labels.items():
    ws_dcf.cell(r, COL_LABEL, lbl).font = font(FONT_BLACK, sz=9, italic=True)

# Per-year INDEX formulas for rows that have 5 values  (REV, GM, EBIT, DA, CAPEX, NWC)
# Bear data cols: G–K in their respective bear rows
# INDEX(array_of_3_values, 1, selector) — horizontal range bear/base/bull for each year
def sel_formula_peryear(bear_row, base_row, bull_row, proj_col):
    """INDEX formula picking the correct scenario cell for a given projection year col."""
    c_letter = get_column_letter(proj_col)
    return (f"=INDEX({c_letter}{bear_row}:{c_letter}{bull_row},1,$B$5)")

def sel_formula_scalar(bear_row, base_row, bull_row):
    """INDEX formula for scalar assumption (value in col G only)."""
    return f"=INDEX(G{bear_row}:G{bull_row},1,$B$5)"

multi_year_rows = [
    (R_SEL_REVG,  R_BEAR_REV,  R_BASE_REV,  R_BULL_REV),
    (R_SEL_GM,    R_BEAR_GM,   R_BASE_GM,   R_BULL_GM),
    (R_SEL_EBIT,  R_BEAR_EBIT, R_BASE_EBIT, R_BULL_EBIT),
    (R_SEL_DA,    R_BEAR_DA,   R_BASE_DA,   R_BULL_DA),
    (R_SEL_CAPEX, R_BEAR_CAPEX,R_BASE_CAPEX,R_BULL_CAPEX),
    (R_SEL_NWC,   R_BEAR_NWC,  R_BASE_NWC,  R_BULL_NWC),
]

for sel_r, bear_r, base_r, bull_r in multi_year_rows:
    for proj_col in PROJ_COLS:
        c = ws_dcf.cell(sel_r, proj_col)
        c.value         = sel_formula_peryear(bear_r, base_r, bull_r, proj_col)
        c.font          = font(FONT_GREEN, sz=9)
        c.number_format = FMT_PCT1
        c.border        = thin_border()

# Scalar rows (tax, terminal g, WACC) — value in col G only
scalar_rows = [
    (R_SEL_TAX,  R_BEAR_TAX,  R_BASE_TAX,  R_BULL_TAX),
    (R_SEL_TG,   R_BEAR_TG,   R_BASE_TG,   R_BULL_TG),
    (R_SEL_WACC, R_BEAR_WACC, R_BASE_WACC, R_BULL_WACC),
]
for sel_r, bear_r, base_r, bull_r in scalar_rows:
    c = ws_dcf.cell(sel_r, COL_2026E)
    c.value         = sel_formula_scalar(bear_r, base_r, bull_r)
    c.font          = font(FONT_GREEN, sz=9)
    c.number_format = FMT_PCT1
    c.border        = thin_border()

apply_section_border(ws_dcf, R_SEL_HDR, R_SEL_WACC, COL_LABEL, COL_END)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 7: INCOME STATEMENT  (Historical + Projected)
# ════════════════════════════════════════════════════════════════════════════
style_section_header(ws_dcf, R_IS_HDR, COL_LABEL, COL_END,
    "INCOME STATEMENT  ($M)")
style_col_header(ws_dcf, R_IS_COL, COL_LABEL, COL_END)
ws_dcf.cell(R_IS_COL, COL_LABEL, "")
for i, col in enumerate(ALL_COLS):
    ws_dcf.cell(R_IS_COL, col, YEAR_LABELS_ALL[i])

# ── Historical revenue (hardcoded) ──────────────────────────────────────────
HIST_REV = {COL_2021: 260.9, COL_2022: 362.1, COL_2023: 471.9,
            COL_2024: 606.3, COL_2025: 759.4}
HIST_GP  = {COL_2021: 82.9,  COL_2022: 110.5, COL_2023: 162.3,
            COL_2024: 229.9, COL_2025: 285.7}
HIST_SGA = {COL_2021: 82.85, COL_2022: 106.15,COL_2023: 129.07,
            COL_2024: 166.37,COL_2025: 197.31}
HIST_EBIT= {COL_2021: 0.05,  COL_2022: 4.30,  COL_2023: 33.25,
            COL_2024: 63.55, COL_2025: 88.37}
HIST_TAX = {COL_2021: 0.0,   COL_2022: 1.60,  COL_2023: 6.64,
            COL_2024: 14.15, COL_2025: 24.98}
HIST_NOPAT={COL_2021: 0.05,  COL_2022: 2.70,  COL_2023: 26.61,
            COL_2024: 49.40, COL_2025: 63.39}

# Revenue row
ws_dcf.cell(R_REV, COL_LABEL, "Net Revenue ($M)").font = font(FONT_BLACK, bold=True)
for col, val in HIST_REV.items():
    c = ws_dcf.cell(R_REV, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Annual Reports / StockAnalysis.com, 2026-08-20")

# Projected revenue — formula: prev_rev * (1 + sel_growth)
# FY2026E: =F64*(1+G52)   FY2027E: =G64*(1+H52)  etc.
for i, proj_col in enumerate(PROJ_COLS):
    prev_col = get_column_letter(proj_col - 1)
    g_col    = get_column_letter(proj_col)
    c = ws_dcf.cell(R_REV, proj_col)
    c.value         = f"={prev_col}{R_REV}*(1+{g_col}{R_SEL_REVG})"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

# Revenue growth %
ws_dcf.cell(R_REV_GR, COL_LABEL, "  % Growth").font = font(FONT_BLACK, italic=True, sz=9)
for i, col in enumerate(ALL_COLS):
    if col == COL_2021:
        ws_dcf.cell(R_REV_GR, col, "—").font = font(FONT_BLACK, sz=9)
        continue
    prev = get_column_letter(col - 1)
    cur  = get_column_letter(col)
    c = ws_dcf.cell(R_REV_GR, col)
    c.value         = f"={cur}{R_REV}/{prev}{R_REV}-1"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT
    c.border        = bottom_border()

# ── Gross Profit ─────────────────────────────────────────────────────────────
ws_dcf.cell(R_GP, COL_LABEL, "Gross Profit ($M)").font = font(FONT_BLACK, bold=True)
for col, val in HIST_GP.items():
    c = ws_dcf.cell(R_GP, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Annual Reports / StockAnalysis.com, 2026-08-20")

for i, proj_col in enumerate(PROJ_COLS):
    g_col = get_column_letter(proj_col)
    c = ws_dcf.cell(R_GP, proj_col)
    c.value         = f"={g_col}{R_REV}*{g_col}{R_SEL_GM}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_GM_PCT, COL_LABEL, "  % Gross Margin").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_GM_PCT, col)
    c.value         = f"={g}{R_GP}/{g}{R_REV}"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# ── SG&A (Operating Expenses) ────────────────────────────────────────────────
ws_dcf.cell(R_SGA, COL_LABEL, "SG&A / Operating Expenses ($M)").font = font(FONT_BLACK)
for col, val in HIST_SGA.items():
    c = ws_dcf.cell(R_SGA, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Annual Reports / StockAnalysis.com, 2026-08-20")

# Projected SGA = GP - EBIT  (implied by EBIT margin assumption)
for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_SGA, proj_col)
    c.value         = f"={g}{R_GP}-{g}{R_EBIT}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_SGA_PCT, COL_LABEL, "  % Revenue").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_SGA_PCT, col)
    c.value         = f"={g}{R_SGA}/{g}{R_REV}"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# ── EBIT ─────────────────────────────────────────────────────────────────────
ws_dcf.cell(R_EBIT, COL_LABEL, "EBIT ($M)").font = font(FONT_BLACK, bold=True)
for col, val in HIST_EBIT.items():
    c = ws_dcf.cell(R_EBIT, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Annual Reports / StockAnalysis.com, 2026-08-20")

for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_EBIT, proj_col)
    c.value         = f"={g}{R_REV}*{g}{R_SEL_EBIT}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_EBIT_PCT, COL_LABEL, "  % EBIT Margin").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_EBIT_PCT, col)
    c.value         = f"=IF({g}{R_REV}=0,0,{g}{R_EBIT}/{g}{R_REV})"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# ── Taxes ────────────────────────────────────────────────────────────────────
ws_dcf.cell(R_TAX, COL_LABEL, "Income Tax ($M)").font = font(FONT_BLACK)
for col, val in HIST_TAX.items():
    c = ws_dcf.cell(R_TAX, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Annual Reports / StockAnalysis.com, 2026-08-20")

for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_TAX, proj_col)
    # Use selected tax rate from consolidation (scalar → use $G$55)
    c.value         = f"=MAX(0,{g}{R_EBIT}*$G${R_SEL_TAX})"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_TAX_RATE, COL_LABEL, "  Effective Tax Rate").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_TAX_RATE, col)
    c.value         = f"=IF({g}{R_EBIT}=0,0,{g}{R_TAX}/{g}{R_EBIT})"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# ── NOPAT ────────────────────────────────────────────────────────────────────
ws_dcf.cell(R_NOPAT, COL_LABEL, "NOPAT ($M)").font = font(FONT_BLACK, bold=True)
for col, val in HIST_NOPAT.items():
    c = ws_dcf.cell(R_NOPAT, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Derived — EBIT × (1 – Tax Rate), Annual Reports, 2026-08-20")

for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_NOPAT, proj_col)
    c.value         = f"={g}{R_EBIT}-{g}{R_TAX}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

apply_section_border(ws_dcf, R_IS_HDR, R_NOPAT, COL_LABEL, COL_END)

print("Part 3: income statement written")

# ════════════════════════════════════════════════════════════════════════════
# SECTION 8: FREE CASH FLOW BUILD
# ════════════════════════════════════════════════════════════════════════════
HIST_DA    = {COL_2021: 3.54, COL_2022: 5.44, COL_2023: 7.93,
              COL_2024: 13.09, COL_2025: 14.51}
HIST_CAPEX = {COL_2021: 16.71, COL_2022: 10.56, COL_2023: 11.54,
              COL_2024: 28.65, COL_2025: 81.95}
# NWC change (approx, from working capital change line)
HIST_NWC   = {COL_2021: -8.21, COL_2022: 24.18, COL_2023: -2.49,
              COL_2024: 16.22, COL_2025: 76.31}
HIST_UFCF  = {COL_2021: 1.18, COL_2022: -18.36, COL_2023: 30.80,
              COL_2024: 39.23, COL_2025: -22.78}

style_section_header(ws_dcf, R_FCF_HDR, COL_LABEL, COL_END,
    "FREE CASH FLOW BUILD  ($M)")
style_col_header(ws_dcf, R_FCF_COL, COL_LABEL, COL_END)
ws_dcf.cell(R_FCF_COL, COL_LABEL, "")
for i, col in enumerate(ALL_COLS):
    ws_dcf.cell(R_FCF_COL, col, YEAR_LABELS_ALL[i])

# NOPAT (linked)
ws_dcf.cell(R_FCF_NOPAT, COL_LABEL, "NOPAT ($M)").font = font(FONT_BLACK, bold=True)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_FCF_NOPAT, col)
    c.value         = f"={g}{R_NOPAT}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

# D&A
ws_dcf.cell(R_DA, COL_LABEL, "(+) D&A ($M)").font = font(FONT_BLACK)
for col, val in HIST_DA.items():
    c = ws_dcf.cell(R_DA, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Cash Flow Statement / StockAnalysis.com, 2026-08-20")

for i, proj_col in enumerate(PROJ_COLS):
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_DA, proj_col)
    c.value         = f"={g}{R_REV}*{g}{R_SEL_DA}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_DA_PCT, COL_LABEL, "  D&A % Revenue").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_DA_PCT, col)
    c.value         = f"={g}{R_DA}/{g}{R_REV}"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# CapEx
ws_dcf.cell(R_CAPEX, COL_LABEL, "(-) CapEx ($M)").font = font(FONT_BLACK)
for col, val in HIST_CAPEX.items():
    c = ws_dcf.cell(R_CAPEX, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Cash Flow Statement / StockAnalysis.com, 2026-08-20")

for i, proj_col in enumerate(PROJ_COLS):
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_CAPEX, proj_col)
    c.value         = f"={g}{R_REV}*{g}{R_SEL_CAPEX}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_CAPEX_PCT, COL_LABEL, "  CapEx % Revenue").font = font(FONT_BLACK, italic=True, sz=9)
for col in ALL_COLS:
    g = get_column_letter(col)
    c = ws_dcf.cell(R_CAPEX_PCT, col)
    c.value         = f"={g}{R_CAPEX}/{g}{R_REV}"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# NWC Change
ws_dcf.cell(R_NWC, COL_LABEL, "(-) Δ NWC ($M)").font = font(FONT_BLACK)
for col, val in HIST_NWC.items():
    c = ws_dcf.cell(R_NWC, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: Vital Farms Cash Flow — Change in Working Capital, StockAnalysis.com, 2026-08-20")

for i, proj_col in enumerate(PROJ_COLS):
    g    = get_column_letter(proj_col)
    prev = get_column_letter(proj_col - 1)
    c = ws_dcf.cell(R_NWC, proj_col)
    c.value         = f"=({g}{R_REV}-{prev}{R_REV})*{g}{R_SEL_NWC}"
    c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_NWC_PCT, COL_LABEL, "  NWC Δ% of Revenue Change").font = font(FONT_BLACK, italic=True, sz=9)
for i, proj_col in enumerate(PROJ_COLS):
    g    = get_column_letter(proj_col)
    prev = get_column_letter(proj_col - 1)
    c = ws_dcf.cell(R_NWC_PCT, proj_col)
    c.value         = f"=IF({g}{R_REV}-{prev}{R_REV}=0,0,{g}{R_NWC}/({g}{R_REV}-{prev}{R_REV}))"
    c.font          = font(FONT_BLACK, sz=9, italic=True)
    c.number_format = FMT_PCT; c.border = bottom_border()

# Unlevered FCF
ws_dcf.cell(R_UFCF, COL_LABEL, "Unlevered Free Cash Flow ($M)").font = font(FONT_BLACK, bold=True)
for col, val in HIST_UFCF.items():
    c = ws_dcf.cell(R_UFCF, col, val)
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = FMT_USD; c.border = thin_border()
    add_comment(c, "Source: StockAnalysis.com Unlevered FCF, 2026-08-20")

for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_UFCF, proj_col)
    c.value         = f"={g}{R_FCF_NOPAT}+{g}{R_DA}-{g}{R_CAPEX}-{g}{R_NWC}"
    c.font          = font(FONT_BLACK, bold=True); c.number_format = FMT_USD; c.border = thin_border()

style_output_row(ws_dcf, R_UFCF, COL_LABEL, COL_END)
apply_section_border(ws_dcf, R_FCF_HDR, R_UFCF, COL_LABEL, COL_END)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 9: DISCOUNTING & TERMINAL VALUE
# ════════════════════════════════════════════════════════════════════════════
style_section_header(ws_dcf, R_DCF_HDR, COL_LABEL, COL_END,
    "DCF DISCOUNTING  (Mid-Year Convention)")
style_col_header(ws_dcf, R_DCF_COL, COL_LABEL, COL_END)
ws_dcf.cell(R_DCF_COL, COL_LABEL, "")
for i, proj_col in enumerate(PROJ_COLS):
    ws_dcf.cell(R_DCF_COL, proj_col, YEAR_LABELS_PROJ[i])

# Discount periods (mid-year: 0.5, 1.5, 2.5, 3.5, 4.5)
ws_dcf.cell(R_PERIOD, COL_LABEL, "Discount Period (mid-year)").font = font(FONT_BLACK)
periods = [0.5, 1.5, 2.5, 3.5, 4.5]
for i, proj_col in enumerate(PROJ_COLS):
    c = ws_dcf.cell(R_PERIOD, proj_col, periods[i])
    c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
    c.number_format = "0.0"; c.border = thin_border()
    add_comment(c, "Source: Model Assumption — mid-year convention, 2026-08-20")

# Discount factors  = 1/(1+WACC)^period   WACC from WACC sheet (col WACC_COL)
# WACC is on WACC sheet; in formula we link it: WACC!B27
ws_dcf.cell(R_DISC_FACT, COL_LABEL, "Discount Factor").font = font(FONT_BLACK)
for i, proj_col in enumerate(PROJ_COLS):
    p_col = get_column_letter(proj_col)
    c = ws_dcf.cell(R_DISC_FACT, proj_col)
    c.value         = f"=1/(1+WACC!$B$27)^{p_col}{R_PERIOD}"
    c.font          = font(FONT_GREEN); c.number_format = "0.0000"; c.border = thin_border()

# PV of FCF
ws_dcf.cell(R_PV_FCF, COL_LABEL, "PV of FCF ($M)").font = font(FONT_BLACK, bold=True)
for proj_col in PROJ_COLS:
    g = get_column_letter(proj_col)
    c = ws_dcf.cell(R_PV_FCF, proj_col)
    c.value         = f"={g}{R_UFCF}*{g}{R_DISC_FACT}"
    c.font          = font(FONT_BLACK, bold=True); c.number_format = FMT_USD; c.border = thin_border()

# Terminal Value section
style_section_header(ws_dcf, R_TV_HDR, COL_LABEL, COL_END,
    "TERMINAL VALUE  (Perpetuity Growth Method)")

ws_dcf.cell(R_TERM_FCF, COL_LABEL, "Terminal Year FCF ($M)").font = font(FONT_BLACK)
k_col = get_column_letter(COL_2030E)
c = ws_dcf.cell(R_TERM_FCF, COL_2030E)
c.value         = f"={k_col}{R_UFCF}*(1+$G${R_SEL_TG})"
c.font          = font(FONT_BLACK); c.number_format = FMT_USD; c.border = thin_border()

ws_dcf.cell(R_TERM_VAL, COL_LABEL, "Terminal Value ($M)").font = font(FONT_BLACK)
c = ws_dcf.cell(R_TERM_VAL, COL_2030E)
c.value         = f"=K{R_TERM_FCF}/(WACC!$B$27-$G${R_SEL_TG})"
c.font          = font(FONT_GREEN); c.number_format = FMT_INT; c.border = thin_border()

ws_dcf.cell(R_PV_TV, COL_LABEL, "PV of Terminal Value ($M)").font = font(FONT_BLACK, bold=True)
c = ws_dcf.cell(R_PV_TV, COL_2030E)
c.value         = f"=K{R_TERM_VAL}/(1+WACC!$B$27)^4.5"
c.font          = font(FONT_GREEN, bold=True); c.number_format = FMT_INT; c.border = thin_border()

apply_section_border(ws_dcf, R_DCF_HDR, R_PV_TV, COL_LABEL, COL_END)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 10: VALUATION SUMMARY
# ════════════════════════════════════════════════════════════════════════════
style_section_header(ws_dcf, R_VAL_HDR, COL_LABEL, COL_END,
    "VALUATION SUMMARY")
style_col_header(ws_dcf, R_VAL_HDR + 0, COL_LABEL, COL_END)  # reuse header row

def val_row(row, label, formula, fmt, is_output=False):
    ws_dcf.cell(row, COL_LABEL, label).font = font(FONT_BLACK, bold=is_output)
    c = ws_dcf.cell(row, 2)
    c.value         = formula
    c.font          = font(FONT_GREEN if "!" in str(formula) else FONT_BLACK, bold=is_output)
    c.number_format = fmt
    c.border        = thin_border()
    if is_output:
        c.fill = fill(MED_BLUE)

val_row(R_SUM_PV_FCF, "Sum of PV of Projected FCFs ($M)",
        f"=SUM(G{R_PV_FCF}:K{R_PV_FCF})", FMT_INT)
val_row(R_SUM_PV_TV,  "PV of Terminal Value ($M)",
        f"=K{R_PV_TV}", FMT_INT)
val_row(R_EV,         "Enterprise Value ($M)",
        f"=B{R_SUM_PV_FCF}+B{R_SUM_PV_TV}", FMT_INT, is_output=True)
val_row(R_LESS_DEBT,  "(-) Net Debt ($M)  [+ if net debt, – if net cash]",
        f"=B{R_NET_DEBT}", FMT_INT)
val_row(R_EQ_VAL,     "Equity Value ($M)",
        f"=B{R_EV}-B{R_LESS_DEBT}", FMT_INT, is_output=True)

ws_dcf.cell(R_BLANK_V1, 1, "")

val_row(R_SHARES_OUT, "Diluted Shares Outstanding (M)",
        f"=B{R_SHARES}", FMT_USD)
val_row(R_IMPL_PRICE, "IMPLIED SHARE PRICE",
        f"=IF(B{R_SHARES_OUT}=0,0,B{R_EQ_VAL}/B{R_SHARES_OUT})",
        FMT_USD2, is_output=True)
val_row(R_CUR_PRICE,  "Current Stock Price",
        f"=B{R_PRICE}", FMT_USD2)
val_row(R_UPSIDE,     "Implied Upside / (Downside)",
        f"=B{R_IMPL_PRICE}/B{R_CUR_PRICE}-1", FMT_PCT, is_output=True)

# Large font for implied price
ws_dcf.cell(R_IMPL_PRICE, 1).font = font(FONT_BLACK, bold=True, sz=12)
ws_dcf.cell(R_IMPL_PRICE, 2).font = font(FONT_BLACK, bold=True, sz=12)
ws_dcf.row_dimensions[R_IMPL_PRICE].height = 22

apply_section_border(ws_dcf, R_VAL_HDR, R_UPSIDE, COL_LABEL, COL_END)

print("Part 4: FCF + discounting + valuation summary written")

# ════════════════════════════════════════════════════════════════════════════
# SECTION 11: SENSITIVITY TABLES
# ════════════════════════════════════════════════════════════════════════════
# Each table is a 5×5 grid. Each cell contains a FULL DCF recalculation
# using its row-header WACC and col-header terminal-g (or other axis).
# We build the formula inline rather than relying on Excel Data Tables.
#
# Core valuation formula (inline, no helper cells):
#   IMPLIED_PRICE(wacc, tg) =
#     [SUM of UFCF_t / (1+wacc)^period_t] + [UFCF_5*(1+tg)/(wacc-tg)] / (1+wacc)^4.5
#     − Net_Debt) / Shares
#
# UFCF_t is in cells G91:K91, periods 0.5..4.5, net debt in B13, shares in B11

def ufcf_pv_sum(wacc_ref):
    """Build SUM of PV of projected FCFs for a given WACC reference."""
    parts = []
    period_map = {COL_2026E: 0.5, COL_2027E: 1.5, COL_2028E: 2.5, COL_2029E: 3.5, COL_2030E: 4.5}
    for col, period in period_map.items():
        g = get_column_letter(col)
        parts.append(f"{g}{R_UFCF}/(1+{wacc_ref})^{period}")
    return "+".join(parts)

def tv_pv(wacc_ref, tg_ref):
    """Build PV of terminal value formula."""
    k = get_column_letter(COL_2030E)
    return (f"({k}{R_UFCF}*(1+{tg_ref})/({wacc_ref}-{tg_ref}))/(1+{wacc_ref})^4.5")

def full_dcf_formula(wacc_ref, tg_ref):
    """Full implied-price formula using substituted WACC and TG references."""
    pv_fcfs = ufcf_pv_sum(wacc_ref)
    tv      = tv_pv(wacc_ref, tg_ref)
    return f"=IF(B{R_SHARES}=0,0,({pv_fcfs}+{tv}-B{R_NET_DEBT})/B{R_SHARES})"

def full_dcf_formula_rev_ebit(rev_growth_5yr, ebit_margin_final, wacc_ref="WACC!$B$27", tg_ref=f"$G${R_SEL_TG}"):
    """
    Sensitivity: vary 5yr avg revenue growth and terminal EBIT margin.
    For simplicity: project revenue as FY2025 * (1+rev_g)^5 (compound),
    EBIT = rev * margin, NOPAT = EBIT*(1-tax), DA/CapEx/NWC from base assumptions,
    UFCF terminal = NOPAT_t + DA_t - CAPEX_t - NWC_t
    We embed a simplified single-terminal-year FCF using the axis values.
    """
    # Use terminal-year revenue: F64 * (1+rev_growth)^5
    # Terminal NOPAT = terminal_rev * ebit_margin * (1 - tax)
    # Terminal DA, CapEx, NWC from base (col K)
    k = get_column_letter(COL_2030E)
    f_col = get_column_letter(COL_2025)
    # Approximate: recalculate all 5 projected year FCFs with uniform rev growth and terminal ebit
    # Year-by-year revenue: F64*(1+rg)^n  for n=1..5
    # UFCF_n = rev_n * ebit_m * (1-tax) + rev_n*da_pct - rev_n*capex_pct - (rev_n - rev_{n-1})*nwc_pct
    pv_parts = []
    for n, period in enumerate([0.5, 1.5, 2.5, 3.5, 4.5], start=1):
        prev_n = n - 1
        rev_n   = f"({f_col}{R_REV}*{rev_growth_5yr}^{n})"
        rev_pm1 = f"({f_col}{R_REV}*{rev_growth_5yr}^{prev_n})"
        # Use base tax / DA / CapEx / NWC from consolidation col (col K for yr5 pcts as approximation)
        ufcf_n = (f"({rev_n}*{ebit_margin_final}*(1-$G${R_SEL_TAX})"
                  f"+{rev_n}*$K${R_SEL_DA}"
                  f"-{rev_n}*$K${R_SEL_CAPEX}"
                  f"-({rev_n}-{rev_pm1})*$K${R_SEL_NWC})")
        pv_parts.append(f"{ufcf_n}/(1+{wacc_ref})^{period}")
    pv_sum = "+".join(pv_parts)
    rev_5 = f"({f_col}{R_REV}*{rev_growth_5yr}^5)"
    term_fcf = (f"({rev_5}*{ebit_margin_final}*(1-$G${R_SEL_TAX})"
                f"+{rev_5}*$K${R_SEL_DA}"
                f"-{rev_5}*$K${R_SEL_CAPEX}"
                f"-({rev_5}-{f_col}{R_REV}*{rev_growth_5yr}^4)*$K${R_SEL_NWC})"
                f"*(1+{tg_ref})/({wacc_ref}-{tg_ref})")
    return f"=IF(B{R_SHARES}=0,0,({pv_sum}+{term_fcf}/(1+{wacc_ref})^4.5-B{R_NET_DEBT})/B{R_SHARES})"

def full_dcf_formula_beta_rf(beta_val, rf_val, erp=0.055):
    """Sensitivity: vary beta and risk-free rate → cost of equity → WACC."""
    # WACC = CoE * equity_weight + cost_debt_AT * debt_weight
    # Use current equity/debt weights from WACC sheet, substitute CoE
    # CoE = rf + beta * erp
    coe = f"({rf_val}+{beta_val}*{erp})"
    wacc_sub = (f"({coe}*WACC!$B$22/(WACC!$B$22+WACC!$B$24)"
                f"+WACC!$B$19*(1-WACC!$B$17)*WACC!$B$24/(WACC!$B$22+WACC!$B$24))")
    return full_dcf_formula(wacc_sub, f"$G${R_SEL_TG}")

# ─── TABLE 1: WACC vs Terminal Growth ───────────────────────────────────────
BASE_WACC = 0.105
BASE_TG   = 0.030
WACC_AXIS = [0.085, 0.095, BASE_WACC, 0.115, 0.125]   # 5 values, center = base
TG_AXIS   = [0.020, 0.025, BASE_TG,   0.035, 0.040]

style_section_header(ws_dcf, R_SENS1_HDR, COL_LABEL, 10,
    "SENSITIVITY 1 — Implied Price: WACC vs Terminal Growth Rate")

# Column headers (TG values)
ws_dcf.cell(R_SENS1_TLBL, COL_LABEL, "WACC \\ Term.Growth")
ws_dcf.cell(R_SENS1_TLBL, COL_LABEL).font = font(FONT_BLACK, bold=True, sz=9)
ws_dcf.cell(R_SENS1_TLBL, COL_LABEL).fill = fill(LIGHT_BLUE)
for j, tg_val in enumerate(TG_AXIS):
    c = ws_dcf.cell(R_SENS1_TLBL, COL_LABEL + 1 + j, tg_val)
    c.font = font(FONT_BLACK, bold=True, sz=9)
    c.fill = fill(LIGHT_BLUE)
    c.number_format = FMT_PCT1
    c.border = thin_border()
    c.alignment = align("center")

# Data cells
for i, wacc_val in enumerate(WACC_AXIS):
    row = R_SENS1_DATA + i
    # Row header
    c = ws_dcf.cell(row, COL_LABEL, wacc_val)
    c.font = font(FONT_BLACK, bold=True, sz=9); c.fill = fill(LIGHT_BLUE)
    c.number_format = FMT_PCT1; c.border = thin_border(); c.alignment = align("center")
    for j, tg_val in enumerate(TG_AXIS):
        col = COL_LABEL + 1 + j
        c = ws_dcf.cell(row, col)
        c.value         = full_dcf_formula(str(wacc_val), str(tg_val))
        c.number_format = FMT_USD2
        c.border        = thin_border()
        c.alignment     = align("center")
        # Center cell (base case) highlight
        if i == 2 and j == 2:
            c.fill = fill(MED_BLUE); c.font = font(FONT_BLACK, bold=True, sz=9)
        else:
            c.font = font(FONT_BLACK, sz=9)

apply_section_border(ws_dcf, R_SENS1_HDR, R_SENS1_DATA + 4, COL_LABEL, COL_LABEL + 5)

# ─── TABLE 2: Revenue Growth vs EBIT Margin ─────────────────────────────────
# 5yr avg revenue growth (compound factor = 1+g) vs terminal EBIT margin
REV_G_AXIS    = [1.08, 1.12, 1.15, 1.18, 1.22]   # (1+g) compound factor
EBIT_M_AXIS   = [0.08, 0.10, 0.13, 0.15, 0.17]
REV_G_LABELS  = ["8%", "12%", "15%", "18%", "22%"]

style_section_header(ws_dcf, R_SENS2_HDR, COL_LABEL, 10,
    "SENSITIVITY 2 — Implied Price: 5-Yr Rev CAGR vs Terminal EBIT Margin")

ws_dcf.cell(R_SENS2_TLBL, COL_LABEL, "RevCAGR \\ EBIT Mgn")
ws_dcf.cell(R_SENS2_TLBL, COL_LABEL).font = font(FONT_BLACK, bold=True, sz=9)
ws_dcf.cell(R_SENS2_TLBL, COL_LABEL).fill = fill(LIGHT_BLUE)
for j, em in enumerate(EBIT_M_AXIS):
    c = ws_dcf.cell(R_SENS2_TLBL, COL_LABEL + 1 + j, em)
    c.font = font(FONT_BLACK, bold=True, sz=9); c.fill = fill(LIGHT_BLUE)
    c.number_format = FMT_PCT1; c.border = thin_border(); c.alignment = align("center")

for i, (rg_factor, rg_label) in enumerate(zip(REV_G_AXIS, REV_G_LABELS)):
    row = R_SENS2_DATA + i
    c = ws_dcf.cell(row, COL_LABEL, rg_label)
    c.font = font(FONT_BLACK, bold=True, sz=9); c.fill = fill(LIGHT_BLUE)
    c.border = thin_border(); c.alignment = align("center")
    for j, em in enumerate(EBIT_M_AXIS):
        col = COL_LABEL + 1 + j
        c = ws_dcf.cell(row, col)
        c.value         = full_dcf_formula_rev_ebit(str(rg_factor), str(em))
        c.number_format = FMT_USD2
        c.border        = thin_border(); c.alignment = align("center")
        if i == 2 and j == 2:
            c.fill = fill(MED_BLUE); c.font = font(FONT_BLACK, bold=True, sz=9)
        else:
            c.font = font(FONT_BLACK, sz=9)

apply_section_border(ws_dcf, R_SENS2_HDR, R_SENS2_DATA + 4, COL_LABEL, COL_LABEL + 5)

# ─── TABLE 3: Beta vs Risk-Free Rate ────────────────────────────────────────
BETA_AXIS = [0.90, 1.10, 1.3756, 1.60, 1.80]
RF_AXIS   = [0.040, 0.045, 0.0468, 0.050, 0.055]

style_section_header(ws_dcf, R_SENS3_HDR, COL_LABEL, 10,
    "SENSITIVITY 3 — Implied Price: Beta vs Risk-Free Rate")

ws_dcf.cell(R_SENS3_TLBL, COL_LABEL, "Beta \\ Risk-Free")
ws_dcf.cell(R_SENS3_TLBL, COL_LABEL).font = font(FONT_BLACK, bold=True, sz=9)
ws_dcf.cell(R_SENS3_TLBL, COL_LABEL).fill = fill(LIGHT_BLUE)
for j, rf in enumerate(RF_AXIS):
    c = ws_dcf.cell(R_SENS3_TLBL, COL_LABEL + 1 + j, rf)
    c.font = font(FONT_BLACK, bold=True, sz=9); c.fill = fill(LIGHT_BLUE)
    c.number_format = FMT_PCT1; c.border = thin_border(); c.alignment = align("center")

for i, beta in enumerate(BETA_AXIS):
    row = R_SENS3_DATA + i
    c = ws_dcf.cell(row, COL_LABEL, beta)
    c.font = font(FONT_BLACK, bold=True, sz=9); c.fill = fill(LIGHT_BLUE)
    c.number_format = "0.00"; c.border = thin_border(); c.alignment = align("center")
    for j, rf in enumerate(RF_AXIS):
        col = COL_LABEL + 1 + j
        c = ws_dcf.cell(row, col)
        c.value         = full_dcf_formula_beta_rf(str(beta), str(rf))
        c.number_format = FMT_USD2
        c.border        = thin_border(); c.alignment = align("center")
        if i == 2 and j == 2:
            c.fill = fill(MED_BLUE); c.font = font(FONT_BLACK, bold=True, sz=9)
        else:
            c.font = font(FONT_BLACK, sz=9)

apply_section_border(ws_dcf, R_SENS3_HDR, R_SENS3_DATA + 4, COL_LABEL, COL_LABEL + 5)

print("Part 5: sensitivity tables written")

# ════════════════════════════════════════════════════════════════════════════
# WACC SHEET
# ════════════════════════════════════════════════════════════════════════════
ws_wacc.column_dimensions["A"].width = 36
ws_wacc.column_dimensions["B"].width = 16
ws_wacc.column_dimensions["C"].width = 18
ws_wacc.column_dimensions["D"].width = 14

# ── Header ───────────────────────────────────────────────────────────────────
c = ws_wacc.cell(1, 1, "Vital Farms (VITL) — WACC Calculation")
c.font = font(DARK_BLUE, bold=True, sz=13)
c = ws_wacc.cell(2, 1, f"Date: {date.today().strftime('%B %d, %Y')}  |  CAPM Methodology")
c.font = font(FONT_BLACK, italic=True, sz=9)

# ── SECTION A: Cost of Equity (CAPM) ────────────────────────────────────────
style_section_header(ws_wacc, 4, 1, 4, "COST OF EQUITY  —  CAPM")

ws_wacc.cell(5, 1, "Item"); ws_wacc.cell(5, 2, "Value"); ws_wacc.cell(5, 3, "Notes")
for col in [1, 2, 3]:
    ws_wacc.cell(5, col).fill = fill(LIGHT_BLUE)
    ws_wacc.cell(5, col).font = font(FONT_BLACK, bold=True, sz=9)
    ws_wacc.cell(5, col).border = thin_border()

# Row 6: Risk-Free Rate (B6)
ws_wacc.cell(6, 1, "Risk-Free Rate (10Y US Treasury)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(6, 2, 0.0468)
c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
c.number_format = FMT_PCT1; c.border = thin_border()
add_comment(c, "Source: US Federal Reserve / FRED DGS10, 2026-08-17, 4.68%")
ws_wacc.cell(6, 3, "FRED DGS10 — Aug 17 2026").font = font(FONT_BLACK, sz=8, italic=True)

# Row 7: Beta (B7)
ws_wacc.cell(7, 1, "Beta (5-Year)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(7, 2, 1.3756)
c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
c.number_format = "0.0000"; c.border = thin_border()
add_comment(c, "Source: GuruFocus.com, 2026-08-19, VITL 5-year beta = 1.3756")
ws_wacc.cell(7, 3, "GuruFocus, Aug 19 2026").font = font(FONT_BLACK, sz=8, italic=True)

# Row 8: Equity Risk Premium (B8)
ws_wacc.cell(8, 1, "Equity Risk Premium (ERP)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(8, 2, 0.055)
c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
c.number_format = FMT_PCT1; c.border = thin_border()
add_comment(c, "Source: Damodaran ERP estimate, 2026, 5.5% (US market)")
ws_wacc.cell(8, 3, "Damodaran, 2026 US ERP").font = font(FONT_BLACK, sz=8, italic=True)

# Row 9: Cost of Equity (B9) — FORMULA
ws_wacc.cell(9, 1, "Cost of Equity (CAPM)").font = font(FONT_BLACK, bold=True, sz=9)
c = ws_wacc.cell(9, 2)
c.value         = "=B6+B7*B8"
c.font          = font(FONT_BLACK, bold=True)
c.number_format = FMT_PCT1; c.border = thin_border()
ws_wacc.cell(9, 3, "= Rf + β × ERP").font = font(FONT_BLACK, sz=8, italic=True)
style_output_row(ws_wacc, 9, 1, 3)

# ── SECTION B: Cost of Debt ───────────────────────────────────────────────────
style_section_header(ws_wacc, 11, 1, 4, "COST OF DEBT")

ws_wacc.cell(12, 1, "Item"); ws_wacc.cell(12, 2, "Value")
for col in [1, 2]:
    ws_wacc.cell(12, col).fill = fill(LIGHT_BLUE)
    ws_wacc.cell(12, col).font = font(FONT_BLACK, bold=True, sz=9)
    ws_wacc.cell(12, col).border = thin_border()

# Row 13: Pre-tax cost of debt (B13)  — approximate via interest/debt
ws_wacc.cell(13, 1, "Pre-Tax Cost of Debt").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(13, 2, 0.085)
c.font = font(FONT_BLUE); c.fill = fill(LIGHT_GREY)
c.number_format = FMT_PCT1; c.border = thin_border()
add_comment(c, "Source: Vital Farms Q2 2026 — New $125M term loan, estimated rate ~8.5% (SOFR + spread)")
ws_wacc.cell(13, 3, "SOFR + spread estimate on new term loan").font = font(FONT_BLACK, sz=8, italic=True)

# Row 14: Tax rate for shield (B14)
ws_wacc.cell(14, 1, "Tax Rate (for debt shield)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(14, 2)
c.value         = f"=DCF!$G${R_SEL_TAX}"
c.font          = font(FONT_GREEN, sz=9)
c.number_format = FMT_PCT1; c.border = thin_border()
ws_wacc.cell(14, 3, "Linked from DCF sheet").font = font(FONT_BLACK, sz=8, italic=True)

# Row 15: After-tax cost of debt (B15)
ws_wacc.cell(15, 1, "After-Tax Cost of Debt").font = font(FONT_BLACK, bold=True, sz=9)
c = ws_wacc.cell(15, 2)
c.value         = "=B13*(1-B14)"
c.font          = font(FONT_BLACK, bold=True)
c.number_format = FMT_PCT1; c.border = thin_border()
style_output_row(ws_wacc, 15, 1, 3)

# ── SECTION C: Capital Structure ─────────────────────────────────────────────
style_section_header(ws_wacc, 17, 1, 4, "CAPITAL STRUCTURE")

ws_wacc.cell(18, 1, "Item"); ws_wacc.cell(18, 2, "Value")
for col in [1, 2]:
    ws_wacc.cell(18, col).fill = fill(LIGHT_BLUE)
    ws_wacc.cell(18, col).font = font(FONT_BLACK, bold=True, sz=9)
    ws_wacc.cell(18, col).border = thin_border()

# Row 19: Pre-tax cost of debt (same ref)
ws_wacc.cell(19, 1, "Pre-Tax Cost of Debt (ref)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(19, 2)
c.value = "=B13"
c.font = font(FONT_GREEN, sz=9); c.number_format = FMT_PCT1; c.border = thin_border()

# Row 20: Stock Price
ws_wacc.cell(20, 1, "Current Stock Price").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(20, 2)
c.value         = f"=DCF!$B${R_PRICE}"
c.font          = font(FONT_GREEN, sz=9)
c.number_format = FMT_USD2; c.border = thin_border()

# Row 21: Diluted Shares
ws_wacc.cell(21, 1, "Diluted Shares (M)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(21, 2)
c.value         = f"=DCF!$B${R_SHARES}"
c.font          = font(FONT_GREEN, sz=9)
c.number_format = FMT_USD; c.border = thin_border()

# Row 22: Market Cap (B22)
ws_wacc.cell(22, 1, "Market Capitalization ($M)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(22, 2)
c.value         = "=B20*B21"
c.font          = font(FONT_BLACK)
c.number_format = FMT_INT; c.border = thin_border()

# Row 23: Net Debt (B23)
ws_wacc.cell(23, 1, "Net Debt ($M)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(23, 2)
c.value         = f"=DCF!$B${R_NET_DEBT}"
c.font          = font(FONT_GREEN, sz=9)
c.number_format = FMT_INT; c.border = thin_border()

# Row 24: Enterprise Value (B24)
ws_wacc.cell(24, 1, "Enterprise Value ($M)").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(24, 2)
c.value         = "=B22+B23"
c.font          = font(FONT_BLACK)
c.number_format = FMT_INT; c.border = thin_border()

# Row 25: Equity weight
ws_wacc.cell(25, 1, "Equity Weight").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(25, 2)
c.value         = "=IF(B24=0,1,B22/B24)"
c.font          = font(FONT_BLACK)
c.number_format = FMT_PCT1; c.border = thin_border()

# Row 26: Debt weight
ws_wacc.cell(26, 1, "Debt Weight").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(26, 2)
c.value         = "=IF(B24=0,0,B23/B24)"
c.font          = font(FONT_BLACK)
c.number_format = FMT_PCT1; c.border = thin_border()

# ── SECTION D: WACC ───────────────────────────────────────────────────────────
style_section_header(ws_wacc, 28, 1, 4, "WACC CALCULATION")

ws_wacc.cell(29, 1, ""); ws_wacc.cell(29, 2, "Weight"); ws_wacc.cell(29, 3, "Cost"); ws_wacc.cell(29, 4, "Contribution")
for col in [1, 2, 3, 4]:
    ws_wacc.cell(29, col).fill = fill(LIGHT_BLUE)
    ws_wacc.cell(29, col).font = font(FONT_BLACK, bold=True, sz=9)
    ws_wacc.cell(29, col).border = thin_border()

ws_wacc.cell(30, 1, "Equity").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(30, 2); c.value = "=B25"; c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()
c = ws_wacc.cell(30, 3); c.value = "=B9";  c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()
c = ws_wacc.cell(30, 4); c.value = "=B30*C30"; c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()

ws_wacc.cell(31, 1, "Debt").font = font(FONT_BLACK, sz=9)
c = ws_wacc.cell(31, 2); c.value = "=B26"; c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()
c = ws_wacc.cell(31, 3); c.value = "=B15"; c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()
c = ws_wacc.cell(31, 4); c.value = "=B31*C31"; c.font = font(FONT_BLACK); c.number_format = FMT_PCT1; c.border = thin_border()

# Row 27 = WACC output (referenced in DCF sheet as WACC!$B$27) — BUT we placed WACC at row 27
# Let's use row 33 as the output (already defined label R_SEL_WACC etc. reference WACC!$B$27)
# We need WACC result at B27 → move section up so WACC result is exactly B27

# Since rows are already set, we'll put the actual WACC formula at row 27 
# (between capital structure and WACC calc section)
# Patch: write WACC formula at B27 which is referenced in DCF formulas
ws_wacc.cell(27, 1, "WEIGHTED AVG COST OF CAPITAL (WACC)").font = font(FONT_BLACK, bold=True, sz=11)
c = ws_wacc.cell(27, 2)
c.value         = "=D30+D31"
c.font          = font(FONT_BLACK, bold=True, sz=13)
c.fill          = fill(MED_BLUE)
c.number_format = FMT_PCT1
c.border        = thick_border()

apply_section_border(ws_wacc, 4,  9,  1, 4)
apply_section_border(ws_wacc, 11, 15, 1, 4)
apply_section_border(ws_wacc, 17, 26, 1, 4)
apply_section_border(ws_wacc, 27, 27, 1, 4)
apply_section_border(ws_wacc, 28, 31, 1, 4)

# ════════════════════════════════════════════════════════════════════════════
# FREEZE PANES & FINAL CLEANUP
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.freeze_panes  = "B64"   # freeze labels + year headers
ws_wacc.freeze_panes = "B5"

# Tab colours
ws_dcf.sheet_properties.tabColor  = "1F4E79"
ws_wacc.sheet_properties.tabColor = "2E75B6"

# ════════════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"✅ Saved: {OUTPUT_PATH}")
