#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook("/Users/nickhuang/Documents/personal/nickhuangcyh/equity-research/META_DCF_Model_Gemini-3.7-Flash_20260819.xlsx", data_only=True)

ws_dcf = wb["DCF Valuation"]
ws_wacc = wb["WACC Build"]

print("=== WACC BUILD ===")
print("Cost of Equity (B8):", ws_wacc["B8"].value)
print("After-tax Cost of Debt (B13):", ws_wacc["B13"].value)
print("Equity Weight (C24):", ws_wacc["C24"].value)
print("Debt Weight (C25):", ws_wacc["C25"].value)
print("Total WACC (E26):", ws_wacc["E26"].value)

print("\n=== DCF VALUATION ===")
print("Active Case (B4):", ws_dcf["B4"].value)
print("Active Case Title (C4):", ws_dcf["C4"].value)
print("Revenue FY25A - FY30E:")
for col in ["B", "C", "D", "E", "F", "G"]:
    print(f"  {col}53:", ws_dcf[f"{col}53"].value)
print("EBIT FY25A - FY30E:")
for col in ["B", "C", "D", "E", "F", "G"]:
    print(f"  {col}55:", ws_dcf[f"{col}55"].value)
print("UFCF FY25A - FY30E:")
for col in ["B", "C", "D", "E", "F", "G"]:
    print(f"  {col}66:", ws_dcf[f"{col}66"].value)

print("\n=== VALUATION BRIDGE ===")
print("Cumulative PV FCFs (B79):", ws_dcf["B79"].value)
print("PV Terminal Value (B80):", ws_dcf["B80"].value)
print("Enterprise Value (B81):", ws_dcf["B81"].value)
print("Total Debt (B82):", ws_dcf["B82"].value)
print("Cash (B83):", ws_dcf["B83"].value)
print("Net Debt Adjustment (B84):", ws_dcf["B84"].value)
print("Implied Equity Value (B85):", ws_dcf["B85"].value)
print("Diluted Shares (B86):", ws_dcf["B86"].value)
print("Implied Share Price (B87):", ws_dcf["B87"].value)
print("Current Share Price (B88):", ws_dcf["B88"].value)
print("Implied Upside/Downside (B89):", ws_dcf["B89"].value)

print("\n=== SENSITIVITY TABLE 1 (WACC vs g) ===")
for r in range(95, 101):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(1, 7)]
    print(f"Row {r}:", row_vals)

print("\n=== SENSITIVITY TABLE 2 (Growth vs EBIT Margin) ===")
for r in range(105, 111):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(1, 7)]
    print(f"Row {r}:", row_vals)

print("\n=== SENSITIVITY TABLE 3 (Beta vs Rf) ===")
for r in range(115, 121):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(1, 7)]
    print(f"Row {r}:", row_vals)
