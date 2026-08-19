#!/usr/bin/env python3
"""
verify_nflx_math.py - Inspect and verify calculated values in the recalculated NFLX DCF Excel model.
"""

import openpyxl

def inspect_model(file_path="NFLX_DCF_Model_Gemini-3.7-Flash_20260819.xlsx"):
    # Load data_only=True to get evaluated numbers
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_dcf = wb["DCF Valuation"]
    ws_wacc = wb["WACC Build"]
    
    print("=== WACC BUILD SHEET ===")
    for r in range(4, 27):
        print(f"Row {r:2d}: A={ws_wacc[f'A{r}'].value} | B={ws_wacc[f'B{r}'].value} | C={ws_wacc[f'C{r}'].value} | D={ws_wacc[f'D{r}'].value} | E={ws_wacc[f'E{r}'].value}")
        
    print("\n=== DCF VALUATION SUMMARY (ACTIVE CASE) ===")
    for r in range(70, 90):
        print(f"Row {r:2d}: A={ws_dcf[f'A{r}'].value} | B={ws_dcf[f'B{r}'].value} | C={ws_dcf[f'C{r}'].value} | D={ws_dcf[f'D{r}'].value} | E={ws_dcf[f'E{r}'].value} | F={ws_dcf[f'F{r}'].value} | G={ws_dcf[f'G{r}'].value} | H={ws_dcf[f'H{r}'].value}")

    print("\n=== SENSITIVITY TABLE 1: WACC vs Terminal g ===")
    for r in range(94, 101):
        row_vals = [str(ws_dcf.cell(row=r, column=c).value) for c in range(1, 7)]
        print(f"Row {r:3d}: {' | '.join(row_vals)}")
        
    print("\n=== SENSITIVITY TABLE 2: Growth Multiplier vs EBIT Margin ===")
    for r in range(104, 111):
        row_vals = [str(ws_dcf.cell(row=r, column=c).value) for c in range(1, 7)]
        print(f"Row {r:3d}: {' | '.join(row_vals)}")

    print("\n=== SENSITIVITY TABLE 3: Beta vs Risk-Free Rate ===")
    for r in range(114, 121):
        row_vals = [str(ws_dcf.cell(row=r, column=c).value) for c in range(1, 7)]
        print(f"Row {r:3d}: {' | '.join(row_vals)}")

if __name__ == "__main__":
    inspect_model()
