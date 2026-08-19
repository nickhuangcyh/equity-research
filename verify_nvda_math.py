import openpyxl

wb = openpyxl.load_workbook("NVDA_DCF_Model_Gemini-3.7-Flash_20260819.xlsx", data_only=True)

print("=== SHEET 2: WACC Build ===")
ws_wacc = wb["WACC Build"]
for r in range(4, 28):
    lbl = ws_wacc.cell(row=r, column=1).value
    val_b = ws_wacc.cell(row=r, column=2).value
    val_c = ws_wacc.cell(row=r, column=3).value
    val_d = ws_wacc.cell(row=r, column=4).value
    if lbl:
        print(f"Row {r:2d} | {str(lbl):45s} | B: {str(val_b):15s} | C: {str(val_c):10s} | D: {str(val_d):10s}")

print("\n=== SHEET 1: DCF Valuation ===")
ws_dcf = wb["DCF Valuation"]
print(f"Case Selector: {ws_dcf['B4'].value} -> Active Scenario: {ws_dcf['B5'].value}")
print("\n--- Key Outputs ---")
for r in range(95, 110):
    lbl = ws_dcf.cell(row=r, column=1).value
    val = ws_dcf.cell(row=r, column=2).value
    print(f"Row {r:3d} | {str(lbl):60s} | {str(val)}")

print("\n--- Sensitivity Table 1: WACC vs Terminal g (Center cell check) ---")
for r in range(114, 120):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(2, 8)]
    print(f"Row {r:3d} | " + " | ".join(f"{str(v):10s}" for v in row_vals))

print("\n--- Sensitivity Table 2: Rev Growth vs EBIT Margin (Center cell check) ---")
for r in range(123, 129):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(2, 8)]
    print(f"Row {r:3d} | " + " | ".join(f"{str(v):10s}" for v in row_vals))

print("\n--- Sensitivity Table 3: Beta vs Rf (Center cell check) ---")
for r in range(133, 139):
    row_vals = [ws_dcf.cell(row=r, column=c).value for c in range(2, 8)]
    print(f"Row {r:3d} | " + " | ".join(f"{str(v):10s}" for v in row_vals))
