import openpyxl

wb = openpyxl.load_workbook("WVE_DCF_Model_Gemini-3.7-Flash_20260819.xlsx", data_only=True)
ws = wb["DCF Valuation"]

print("=== BASE CASE VALUATION SUMMARY ===")
for r in range(97, 118):
    lbl = ws.cell(row=r, column=1).value
    val = ws.cell(row=r, column=2).value
    if lbl:
        print(f"Row {r:3d} | {lbl:<50} | {val}")

print("\n=== SENSITIVITY TABLE 1 (WACC vs Terminal g) ===")
g_hdrs = [ws.cell(row=122, column=c).value for c in range(2, 7)]
print("WACC \\ g  " + "  ".join([f"{g:7.1%}" for g in g_hdrs]))
for r in range(123, 128):
    w = ws.cell(row=r, column=1).value
    vals = [ws.cell(row=r, column=c).value for c in range(2, 7)]
    val_strs = [f"${v:6.2f}" if v is not None else "None" for v in vals]
    print(f"{w:8.2%}  " + "  ".join(val_strs))

print("\n=== SENSITIVITY TABLE 2 (Terminal FCF vs Rev Delta) ===")
rev_hdrs = [ws.cell(row=131, column=c).value for c in range(2, 7)]
print("FCF \\ Rev  " + "  ".join([f"{rd:+7.1%}" for rd in rev_hdrs]))
for r in range(132, 137):
    fcf = ws.cell(row=r, column=1).value
    vals = [ws.cell(row=r, column=c).value for c in range(2, 7)]
    val_strs = [f"${v:6.2f}" if v is not None else "None" for v in vals]
    print(f"${fcf:7.1f}M  " + "  ".join(val_strs))

print("\n=== SENSITIVITY TABLE 3 (Beta vs Rf) ===")
rf_hdrs = [ws.cell(row=140, column=c).value for c in range(2, 7)]
print("Beta \\ Rf  " + "  ".join([f"{rf:7.2%}" for rf in rf_hdrs]))
for r in range(141, 146):
    b = ws.cell(row=r, column=1).value
    vals = [ws.cell(row=r, column=c).value for c in range(2, 7)]
    val_strs = [f"${v:6.2f}" if v is not None else "None" for v in vals]
    print(f"{b:8.2f}  " + "  ".join(val_strs))
